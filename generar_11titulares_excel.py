#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_11titulares_excel.py
============================
Genera 03_OUTPUTS/11 titulares.xlsx

Una hoja por día de visita (Lu, Ma, Mi, Ju, Vi, Sa).
Columnas: Vendedor | Cód.Cliente | Cliente | [marca+cód_prod x11] | TOTAL

Fuentes
-------
  Ventas    : 01_INPUTS/ventas_acumulada.csv   (principal)
  Clientes  : 01_INPUTS/clientes.xlsx
  Objetivos : 01_INPUTS/objetivo 11T.xlsx

Excluye V2, V5, V20.
"""
import sys
from pathlib import Path
import pandas as pd

ROOT    = Path(__file__).parent
INPUTS  = ROOT / "01_INPUTS"
OUTPUTS = ROOT / "03_OUTPUTS"
OUTPUTS.mkdir(exist_ok=True)

EXCLUIDOS = {2, 5, 20}

DAY_ORDER  = ["Lu","Ma","Mi","Ju","Vi","Sa"]
DAY_LABELS = {
    "Lu":"LUNES","Ma":"MARTES","Mi":"MIÉRCOLES",
    "Ju":"JUEVES","Vi":"VIERNES","Sa":"SÁBADO",
}

# ── Alias de marcas ──────────────────────────────────────────────────────────
MARCA_LOOKUP = {
    "ALMA MORA":"ALMA MORA","ALARIS":"ALARIS","TRAPICHE ALARIS":"ALARIS",
    "DON DAVID":"DON DAVID","DADA":"DADA","LOS ARBOLES":"LOS ARBOLES",
    "FINCA LAS MORAS":"FINCA LAS MORAS","F LAS MORAS":"FINCA LAS MORAS",
    "TRAPICHE RESERVA":"TRAPICHE RESERVA",
    "FOND DE CAVE":"FOND DE CAVE","FOND CAVE":"FOND DE CAVE",
    "CAZADOR":"CAZADOR","ANTARES":"ANTARES",
    "GORDON'S FLAVOURS":"GORDON'S FLAVOURS","GORDONS FLAVOURS":"GORDON'S FLAVOURS",
    "GORDON'S":"GORDON'S FLAVOURS","GORDONS":"GORDON'S FLAVOURS","GORDON S":"GORDON'S FLAVOURS",
    "SMIRNOFF":"SMIRNOFF FLAVOURS","SMIRNOFF FLAVOURS":"SMIRNOFF FLAVOURS",
    "SMIRNOFF ICE FLAVOURS":"SMIRNOFF ICE","SMIRNOFF ICE":"SMIRNOFF ICE",
    "JW":"JW BLACK","JW BLACK":"JW BLACK","JW RED":"JW RED",
    "MASCOTA":"MASCOTA","LA MASCOTA":"MASCOTA",
    "NC ESPUMANTES":"NC ESPUMANTES","NAVARRO CORREAS":"NC ESPUMANTES",
    "TRAPICHE MEDALLA":"TRAPICHE MEDALLA","GRAN MEDALLA":"TRAPICHE MEDALLA",
}
ART_KW = [
    ("SMIRNOFF ICE","SMIRNOFF ICE"),("SMF ICE","SMIRNOFF ICE"),
    ("SMIRNOFF","SMIRNOFF FLAVOURS"),("GORDON","GORDON'S FLAVOURS"),
    ("ANTARES","ANTARES"),("CAZADOR","CAZADOR"),("FOND DE CAVE","FOND DE CAVE"),
    ("ALMA MORA","ALMA MORA"),("LOS ARBOLES","LOS ARBOLES"),("DADA","DADA"),
    ("FINCA LAS MORAS","FINCA LAS MORAS"),("F.LAS MORAS","FINCA LAS MORAS"),
    ("DON DAVID","DON DAVID"),("ALARIS","ALARIS"),("TRAPICHE RESERVA","TRAPICHE RESERVA"),
    ("JW BLACK","JW BLACK"),("JW RED","JW RED"),
]
OBJ_ALIAS = {
    "ALMA MORA":"ALMA MORA","TRAPICHE RESERVA":"TRAPICHE RESERVA",
    "FINCA LAS MORAS":"FINCA LAS MORAS","ALARIS":"ALARIS",
    "DON DAVID":"DON DAVID","DADA":"DADA",
    "SIMRNOFF FLAVORS":"SMIRNOFF FLAVOURS","SMIRNOFF FLAVORS":"SMIRNOFF FLAVOURS",
    "SMIRNOFF FLAVOURS":"SMIRNOFF FLAVOURS","LOS ARBOLES":"LOS ARBOLES",
    "ANTARES":"ANTARES","SMIRNOFF ICE":"SMIRNOFF ICE","SMF ICE":"SMIRNOFF ICE",
    "GORDONS FLAVOURS":"GORDON'S FLAVOURS","GORDONS FLAVORS":"GORDON'S FLAVOURS",
    "GORDON'S FLAVOURS":"GORDON'S FLAVOURS",
}


def normalize_marca(marca_raw, art_raw=""):
    m = str(marca_raw).upper().strip()
    if m in MARCA_LOOKUP:
        return MARCA_LOOKUP[m]
    art = str(art_raw).upper()
    for kw, mo in ART_KW:
        if kw in art:
            return mo
    return None


def load_marcas_objetivo():
    """Lista de marcas oficiales 11T desde objetivo 11T.xlsx."""
    try:
        df = pd.read_excel(INPUTS / "objetivo 11T.xlsx", header=1)
        df = df.dropna(subset=df.columns[1:2])
        marcas = []
        for _, row in df.iterrows():
            raw = str(row.iloc[1]).upper().strip()
            mk  = OBJ_ALIAS.get(raw, raw)
            if mk and mk not in marcas:
                marcas.append(mk)
        if marcas:
            return marcas
    except Exception as e:
        print(f"  AVISO: no se pudo leer objetivo 11T.xlsx ({e}). Usando lista por defecto.")
    return [
        "ALMA MORA","ALARIS","DON DAVID","DADA","LOS ARBOLES",
        "FINCA LAS MORAS","TRAPICHE RESERVA","FOND DE CAVE",
        "ANTARES","GORDON'S FLAVOURS","SMIRNOFF FLAVOURS","SMIRNOFF ICE",
    ]


def top_cod_producto(vac_norm, marcas_11t):
    """Para cada marca 11T, devuelve el código del producto más vendido (por CantBase)."""
    vac_norm["CantBase_n"] = pd.to_numeric(vac_norm.get("CantBase", 0), errors="coerce").fillna(0)
    cod_map = {}
    for marca in marcas_11t:
        sub = vac_norm[vac_norm["marca_norm"] == marca]
        if sub.empty:
            cod_map[marca] = "–"
            continue
        try:
            top_cod = (sub.groupby("Codigo")["CantBase_n"]
                       .sum().sort_values(ascending=False).index[0])
            cod_map[marca] = str(int(top_cod)) if str(top_cod).replace(".","").isdigit() else str(top_cod)
        except Exception:
            cod_map[marca] = "–"
    return cod_map


def main():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: instale openpyxl -> pip install openpyxl")
        sys.exit(1)

    print("=" * 54)
    print("  ORBIT · 11 TITULARES POR CLIENTE / VENDEDOR / DIA")
    print("=" * 54)

    # ── 1. Ventas acumuladas ─────────────────────────────────
    vac_path = INPUTS / "ventas_acumulada.csv"
    if not vac_path.exists():
        print(f"ERROR: {vac_path} no encontrado.")
        sys.exit(1)

    print("\n[1/4] Leyendo ventas_acumulada.csv ...")
    vac = pd.read_csv(vac_path, sep=";", encoding="latin1", low_memory=False)
    vac["ImporteNetoItem"] = pd.to_numeric(
        vac["ImporteNetoItem"].astype(str).str.replace(",",".",regex=False), errors="coerce")
    vac = vac[~vac["CodVendedor"].astype(float).astype(int).isin(EXCLUIDOS)]
    vac = vac[vac["ImporteNetoItem"] > 0]
    vac["marca_norm"] = vac.apply(
        lambda r: normalize_marca(r.get("Marca",""), r.get("Articulo","")), axis=1)

    # CCC: pares (cliente_id int, marca_norm str)
    vac_ok = vac[vac["marca_norm"].notna()][["Cliente","marca_norm"]].drop_duplicates()
    ccc_set = set(zip(vac_ok["Cliente"].astype(int), vac_ok["marca_norm"]))
    print(f"  {len(ccc_set):,} pares cliente-marca con compra (CCC)")

    # ── 2. Clientes ─────────────────────────────────────────
    print("[2/4] Leyendo clientes.xlsx ...")
    cli = pd.read_excel(INPUTS / "clientes.xlsx")

    cod_col  = next((c for c in cli.columns if c.lower() in ("codigo","cod","id")), cli.columns[0])
    nom_col  = next((c for c in cli.columns if "razon" in c.lower() or "nombre" in c.lower()), cli.columns[1])
    vend_col = next((c for c in cli.columns if c.lower() == "codven"), None)
    vnom_col = next((c for c in cli.columns if c.lower() == "vendedor"), None)
    dias_col = next((c for c in cli.columns if "diasvisita" in c.lower().replace(" ","")), None)
    seg_col  = next((c for c in cli.columns if "subsegmento" in c.lower()), None) or \
               next((c for c in cli.columns if "ramo" in c.lower()), None)

    if not dias_col:
        print("ERROR: columna DiasVisita no encontrada en clientes.xlsx")
        sys.exit(1)

    print(f"  {len(cli):,} clientes | cols: {cod_col}, {nom_col}, {vend_col}, {vnom_col}, {dias_col}, seg={seg_col}")

    if vend_col:
        cli = cli[~pd.to_numeric(cli[vend_col], errors="coerce").isin(EXCLUIDOS)]

    # ── 3. Marcas 11T + códigos de producto ─────────────────
    print("[3/4] Cargando marcas objetivo y códigos de producto ...")
    marcas_11t  = load_marcas_objetivo()
    cod_prod_map = top_cod_producto(vac, marcas_11t)
    print(f"  {len(marcas_11t)} marcas:")
    for m in marcas_11t:
        print(f"    {m:30s}  cod={cod_prod_map[m]}")

    # ── 4. Generar Excel ─────────────────────────────────────
    print("[4/4] Construyendo Excel ...")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Estilos ───────────────────────────────────────────────
    FILL_HDR   = PatternFill("solid", fgColor="1A1A2E")
    FILL_VEND  = PatternFill("solid", fgColor="0F0F2D")
    FILL_COD   = PatternFill("solid", fgColor="12122A")   # código cliente
    FILL_CLI   = PatternFill("solid", fgColor="16213E")
    FILL_OK    = PatternFill("solid", fgColor="155724")
    FILL_NO    = PatternFill("solid", fgColor="6B1A1A")
    FILL_TOT_A = PatternFill("solid", fgColor="0D4A1E")   # 100%
    FILL_TOT_M = PatternFill("solid", fgColor="0D3030")   # >=50%
    FILL_TOT_B = PatternFill("solid", fgColor="4A0D0D")   # <50%

    FNT_HDR  = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
    FNT_VEND = Font(bold=True, color="E2147A", name="Calibri", size=9)
    FNT_COD  = Font(color="AAAACC",            name="Calibri", size=9)
    FNT_CLI  = Font(color="CCCCCC",            name="Calibri", size=9)
    FNT_CHK  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    FNT_TOT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

    thin   = Side(border_style="thin", color="2A2A4A")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    AL_CTR = Alignment(horizontal="center", vertical="center")
    AL_LFT = Alignment(horizontal="left",   vertical="center")
    AL_HDR = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Columnas:  A=Vendedor  B=Cód.Cliente  C=Nombre  D=Segmento  E..O=marcas  P=TOTAL
    # Índices:   1           2              3          4            5..15        16
    COL_VEND = 1
    COL_COD  = 2   # código de cliente
    COL_NOM  = 3   # nombre cliente
    COL_SEG  = 4   # segmento ← NUEVO
    COL_MARC = 5   # marcas empiezan en col 5
    COL_TOT  = COL_MARC + len(marcas_11t)

    dias_presentes  = cli[dias_col].dropna().unique()
    dias_a_procesar = [d for d in DAY_ORDER if d in dias_presentes] + \
                      [d for d in dias_presentes if d not in DAY_ORDER]

    for dia in dias_a_procesar:
        label = DAY_LABELS.get(dia, dia.upper())
        ws = wb.create_sheet(title=label)

        mask    = cli[dias_col].astype(str).str.strip().str.lower() == dia.lower()
        cli_dia = cli[mask].copy()
        sort_cols = [c for c in [vend_col, nom_col] if c]
        if sort_cols:
            cli_dia = cli_dia.sort_values(sort_cols)

        # ── Encabezado ──────────────────────────────────────
        # Columnas fijas
        for ci, txt in [(COL_VEND,"Vendedor"), (COL_COD,"Cód.\nCliente"), (COL_NOM,"Cliente"), (COL_SEG,"Segmento")]:
            cell = ws.cell(row=1, column=ci, value=txt)
            cell.fill=FILL_HDR; cell.font=FNT_HDR; cell.alignment=AL_HDR; cell.border=BORDER

        # Columnas de marca (nombre + código de producto en dos líneas)
        for mi, marca in enumerate(marcas_11t):
            cod_p = cod_prod_map.get(marca,"–")
            hdr_txt = f"{marca[:20]}\n[{cod_p}]"
            ci = COL_MARC + mi
            cell = ws.cell(row=1, column=ci, value=hdr_txt)
            cell.fill=FILL_HDR; cell.font=FNT_HDR; cell.alignment=AL_HDR; cell.border=BORDER

        # Columna TOTAL
        cell = ws.cell(row=1, column=COL_TOT, value="TOTAL\n✓")
        cell.fill=FILL_HDR; cell.font=FNT_HDR; cell.alignment=AL_HDR; cell.border=BORDER

        ws.row_dimensions[1].height = 48
        ws.freeze_panes = "E2"   # congela A, B, C, D (Vendedor, Código, Cliente, Segmento)

        # ── Filas de datos ───────────────────────────────────
        for ri, (_, row) in enumerate(cli_dia.iterrows(), 2):
            cid  = int(row[cod_col]) if pd.notna(row[cod_col]) else -1
            cnom = str(row[nom_col]) if pd.notna(row[nom_col]) else "–"
            vend = int(row[vend_col]) if (vend_col and pd.notna(row[vend_col])) else 0
            vnom = str(row[vnom_col]) if (vnom_col and pd.notna(row[vnom_col])) else str(vend)
            label_v = f"V{vend} · {vnom}" if vnom != str(vend) else f"V{vend}"

            # Vendedor
            c = ws.cell(row=ri, column=COL_VEND, value=label_v)
            c.fill=FILL_VEND; c.font=FNT_VEND; c.alignment=AL_LFT; c.border=BORDER

            # Código cliente ← NUEVO
            c = ws.cell(row=ri, column=COL_COD, value=cid if cid >= 0 else "–")
            c.fill=FILL_COD; c.font=FNT_COD; c.alignment=AL_CTR; c.border=BORDER

            # Nombre cliente
            c = ws.cell(row=ri, column=COL_NOM, value=cnom)
            c.fill=FILL_CLI; c.font=FNT_CLI; c.alignment=AL_LFT; c.border=BORDER

            # Segmento
            seg_val = str(row[seg_col]) if (seg_col and pd.notna(row.get(seg_col))) else "–"
            c = ws.cell(row=ri, column=COL_SEG, value=seg_val)
            c.fill=FILL_CLI; c.font=FNT_CLI; c.alignment=AL_LFT; c.border=BORDER

            # Marcas
            total_ok = 0
            for mi, marca in enumerate(marcas_11t):
                tiene = (cid, marca) in ccc_set
                if tiene:
                    total_ok += 1
                ci = COL_MARC + mi
                c = ws.cell(row=ri, column=ci, value="✓" if tiene else "✗")
                c.fill=FILL_OK if tiene else FILL_NO
                c.font=FNT_CHK; c.alignment=AL_CTR; c.border=BORDER

            # Total
            pct = total_ok / len(marcas_11t)
            fgt = FILL_TOT_A if pct >= 1 else (FILL_TOT_M if pct >= 0.5 else FILL_TOT_B)
            c = ws.cell(row=ri, column=COL_TOT, value=total_ok)
            c.fill=fgt; c.font=FNT_TOT; c.alignment=AL_CTR; c.border=BORDER

        # ── Anchos de columna ────────────────────────────────
        ws.column_dimensions[get_column_letter(COL_VEND)].width = 28   # Vendedor
        ws.column_dimensions[get_column_letter(COL_COD)].width  = 9    # Cód. cliente
        ws.column_dimensions[get_column_letter(COL_NOM)].width  = 34   # Nombre cliente
        ws.column_dimensions[get_column_letter(COL_SEG)].width  = 22   # Segmento
        for i in range(len(marcas_11t)):
            ws.column_dimensions[get_column_letter(COL_MARC+i)].width = 14  # marcas
        ws.column_dimensions[get_column_letter(COL_TOT)].width  = 9    # TOTAL

        print(f"  Hoja {label:12s}: {len(cli_dia):3d} clientes")

    out = OUTPUTS / "11 titulares.xlsx"
    wb.save(out)
    print(f"\n{'='*54}")
    print(f"  GUARDADO: {out}")
    print(f"  Hojas   : {', '.join(ws.title for ws in wb.worksheets)}")
    print(f"{'='*54}\n")


if __name__ == "__main__":
    main()
