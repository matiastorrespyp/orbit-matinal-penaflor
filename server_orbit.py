"""
ORBIT Server v3 — Flask API con diagnóstico, CCC real, 11T real, sin mock
"""
from flask import Flask, jsonify, request, send_from_directory, make_response, send_file
import json, sqlite3, pandas as pd, math
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta, timezone

_ARG_TZ = timezone(timedelta(hours=-3))
def _now_ar():
    """Hora actual en Argentina (UTC-3) como string 'YYYY-MM-DD HH:MM:SS'."""
    return datetime.now(_ARG_TZ).strftime("%Y-%m-%d %H:%M:%S")

import os, shutil, csv as _csv, threading

# M1 (mount embebido en Orbit Home): con PENAFLOR_SKIP_BOOT=1, importar este módulo NO ejecuta
# el bloque STARTUP (backup/init/restore/export) ni lanza el hilo de warmup: sólo deja el objeto
# Flask `app` importable de forma segura para montarlo bajo /penaflor. El standalone (sin la
# variable) se comporta EXACTAMENTE igual que antes. No cambia lógica comercial ni endpoints.
_PENAFLOR_SKIP_BOOT = os.environ.get("PENAFLOR_SKIP_BOOT", "").strip() == "1"

app = Flask(__name__, static_folder=None)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0   # sin cache en portal.html

@app.after_request
def no_cache_portal(response):
    """Fuerza no-store en portal.html para que el browser siempre descargue la version nueva."""
    if response.content_type and "text/html" in response.content_type:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response
BASE = Path(__file__).parent
APP_DATA = BASE / "06_APP_DATA"
CONFIG = BASE / "09_CONFIG"
DATASETS = BASE / "04_DATASETS_ORBIT"
INPUTS = BASE / "01_INPUTS"
FRONTEND = BASE / "PAV MATINAL PE_A FLOR"
DB_PATH = Path(os.environ.get("ORBIT_DB_PATH", str(BASE / "orbit.db")))
if DB_PATH.parent != BASE:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
PLAN_BACKUP_DIR = Path(os.environ.get(
    "ORBIT_PLAN_BACKUP_DIR",
    str((BASE / "99_BACKUPS_ORBIT" / "planificacion") if DB_PATH.parent == BASE else (DB_PATH.parent / "planificacion"))
))
PLAN_CSV_LATEST = PLAN_BACKUP_DIR / "planificacion_latest.csv"

USERS = {
    "v3":{"pwd":"pav2026","rol":"vendedor","nombre":"Nadia Gambino"},
    "v4":{"pwd":"pav2026","rol":"vendedor","nombre":"Angel Gribaudo"},
    "v6":{"pwd":"pav2026","rol":"vendedor","nombre":"Andrea Peyronel"},
    "v7":{"pwd":"pav2026","rol":"vendedor","nombre":"Guillermo Jofre"},
    "v8":{"pwd":"pav2026","rol":"vendedor","nombre":"Vanesa Alvarez"},
    "v9":{"pwd":"pav2026","rol":"vendedor","nombre":"Fernando Sanchez"},
    "v10":{"pwd":"pav2026","rol":"vendedor","nombre":"Milagros Ortega"},
    "gerencia":{"pwd":"gerencia2026","rol":"gerencia","nombre":"Gerencia"},
}

def init_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    repo_db = BASE / "orbit.db"
    if DB_PATH != repo_db and not DB_PATH.exists() and repo_db.exists():
        try:
            shutil.copy2(str(repo_db), str(DB_PATH))
            print(f"[ORBIT] Seed DB persistente desde repo -> {DB_PATH}")
        except Exception as e:
            print(f"[WARN] seed orbit.db persistente: {e}")
    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS planificacion(
        id INTEGER PRIMARY KEY AUTOINCREMENT, fecha TEXT NOT NULL, vendedor_id TEXT NOT NULL,
        zona TEXT, dia_visita TEXT, venta_esperada REAL, ccc_tradicional INTEGER,
        ccc_autoservicio INTEGER, ccc_onpremise INTEGER, once_t INTEGER,
        marcas TEXT, clientes_clave TEXT, acciones TEXT, estado TEXT DEFAULT 'enviada',
        editado_por TEXT, comentario_gerencia TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    # Migración: agregar columnas si no existen (tabla puede ser anterior)
    for col, defn in [("editado_por", "TEXT"), ("comentario_gerencia", "TEXT")]:
        try:
            c.execute(f"ALTER TABLE planificacion ADD COLUMN {col} {defn}")
        except Exception:
            pass
    # Unique constraint para UPSERT por (fecha, vendedor_id)
    try:
        c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_plan_fecha_vend ON planificacion(fecha, vendedor_id)")
    except Exception:
        pass
    c.execute("""CREATE TABLE IF NOT EXISTS mensajes(
        id INTEGER PRIMARY KEY AUTOINCREMENT, vendedor_id TEXT NOT NULL,
        mensaje TEXT NOT NULL, leido INTEGER DEFAULT 0, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    # Seguimiento gerencial de alertas: nota por alerta (clave = vendedor|cliente|articulo)
    c.execute("""CREATE TABLE IF NOT EXISTS alerta_seguimiento(
        clave TEXT PRIMARY KEY, mensaje TEXT, autor TEXT, updated_at TEXT)""")
    # Alertas descartadas por gerencia: se ocultan en gerencia Y en el vendedor.
    # No se borra el dato de origen (la alerta se recalcula siempre desde ventas.csv);
    # esta tabla es el registro de "ya la vi, no me la muestres más". Ver _alerta_clave().
    c.execute("""CREATE TABLE IF NOT EXISTS alerta_descartada(
        clave TEXT PRIMARY KEY, autor TEXT, resumen TEXT, descartada_at TEXT)""")
    # Mensaje de seguimiento por punto de venta del Plan Cobertura (columna "Mensaje"
    # de la pantalla de gerencia). Clave = ID PUNTO DE VENTA del padrón (único en el
    # relevamiento); ver _plan_cob_clave(). Es carga manual: no toca el padrón.
    c.execute("""CREATE TABLE IF NOT EXISTS plan_cob_nota(
        clave TEXT PRIMARY KEY, mensaje TEXT, autor TEXT, updated_at TEXT)""")
    # Planificación semanal del mes (pantalla Semanal de gerencia): % del mes que
    # esperamos hacer en cada una de las 4 semanas, por KPI. Es carga manual de
    # gerencia — el real se calcula siempre desde las ventas, nunca desde acá.
    c.execute("""CREATE TABLE IF NOT EXISTS plan_semanal(
        periodo TEXT NOT NULL, kpi TEXT NOT NULL, semana INTEGER NOT NULL,
        pct REAL, editado_por TEXT, updated_at TEXT,
        PRIMARY KEY(periodo, kpi, semana))""")
    conn.commit()
    conn.close()

def backup_orbit_db():
    """Copia orbit.db con timestamp antes de cada arranque del servidor."""
    if not DB_PATH.exists():
        return
    try:
        PLAN_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        ts = _now_ar().replace("-","").replace(":","").replace(" ","_")
        dest = PLAN_BACKUP_DIR / f"orbit_{ts}.db"
        shutil.copy2(str(DB_PATH), str(dest))
        print(f"[ORBIT] Backup orbit.db -> {dest.name}")
    except Exception as e:
        print(f"[WARN] backup_orbit_db: {e}")

def export_planificacion_csv():
    """Exporta la tabla planificacion a CSV de seguridad tras cada escritura."""
    try:
        PLAN_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(str(DB_PATH))
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM planificacion ORDER BY fecha, vendedor_id").fetchall()
        conn.close()
        if not rows:
            return
        cols = rows[0].keys()
        with open(str(PLAN_CSV_LATEST), "w", newline="", encoding="utf-8") as f:
            w = _csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            w.writerows([dict(r) for r in rows])
        print(f"[ORBIT] Planificacion -> CSV seguridad ({len(rows)} registros)")
    except Exception as e:
        print(f"[WARN] export_planificacion_csv: {e}")

# ====== GOOGLE SHEETS — fuente de verdad de planificaciones ======
# SQLite es solo cache. La fuente de verdad es Google Sheets.
# Credenciales: NUNCA en Git. Se inyectan por variable de entorno en Render.
#   GSHEETS_SPREADSHEET_ID   = id del spreadsheet (cadena en la URL)
#   GSHEETS_SHEET_NAME       = nombre de la pestaña (default "planificaciones")
#   GSHEETS_CREDENTIALS_JSON = JSON del service account (o base64 del JSON)
#
# M1 — variables NAMESPACED por proveedor (aislamiento en proceso compartido tipo Orbit Home,
# para no mezclar la planilla de Peñaflor con la de otro módulo, p.ej. PepsiCo):
#   Preferencia:  PENAFLOR_GSHEETS_<X>  ->  GSHEETS_<X>  (fallback legacy, SOLO standalone).
#   Modo estricto (PENAFLOR_REQUIRE_NAMESPACED_GSHEETS=1): usa SOLO PENAFLOR_GSHEETS_*; NUNCA
#   cae a las genéricas. Si faltan -> config incompleta controlada (no rompe el import).
# El standalone actual (sin variables nuevas) sigue leyendo GSHEETS_* igual que antes.
_PENAFLOR_REQUIRE_NS_GSHEETS = os.environ.get("PENAFLOR_REQUIRE_NAMESPACED_GSHEETS", "").strip() == "1"


def _penv(name, default=""):
    """Lee una variable de Sheets priorizando el prefijo PENAFLOR_GSHEETS_.
    - Si existe PENAFLOR_GSHEETS_<name> -> la usa.
    - Si no, y NO estamos en modo estricto -> cae a la genérica GSHEETS_<name> (compat standalone).
    - En modo estricto -> NO mira la genérica (aislamiento total) y devuelve `default`.
    Devuelve el valor .strip()eado. No expone ni loguea el contenido (puede ser un secreto)."""
    v = os.environ.get("PENAFLOR_GSHEETS_" + name, "").strip()
    if v:
        return v
    if _PENAFLOR_REQUIRE_NS_GSHEETS:
        return default
    return os.environ.get("GSHEETS_" + name, "").strip() or default


_GSHEETS_SPREADSHEET_ID = _penv("SPREADSHEET_ID")
_GSHEETS_SHEET_NAME     = _penv("SHEET_NAME", "planificaciones") or "planificaciones"
_GSHEETS_SCOPES         = ["https://www.googleapis.com/auth/spreadsheets"]

# Orden canónico de columnas en la hoja (orden aprobado).
PLAN_SHEET_COLS = [
    "id", "fecha", "vendedor_id", "zona", "dia_visita", "venta_esperada",
    "ccc_tradicional", "ccc_autoservicio", "ccc_onpremise", "once_t",
    "marcas", "clientes_clave", "acciones", "estado",
    "created_at", "updated_at", "editado_por", "comentario_gerencia",
]

def _plan_id(fecha, vendedor_id):
    """ID determinístico del plan: fecha + '_' + vendedor_id. Ej: 2026-06-04_V8"""
    return f"{str(fecha).strip()}_{str(vendedor_id).strip().upper()}"

def _gsheet_cell(v):
    """Normaliza un valor de plan a celda de hoja (None -> '')."""
    return "" if v is None else v

def _gsheets_credentials_info():
    """Devuelve el dict de credenciales del service account o None.
    Lee vía _penv: prioriza PENAFLOR_GSHEETS_CREDENTIALS_JSON; en modo estricto NO usa la genérica."""
    raw = _penv("CREDENTIALS_JSON")
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        try:
            import base64
            return json.loads(base64.b64decode(raw).decode("utf-8"))
        except Exception:
            return None

def gsheets_enabled():
    """True si hay id de hoja y credenciales configuradas."""
    return bool(_GSHEETS_SPREADSHEET_ID) and _gsheets_credentials_info() is not None


def get_gsheets_config():
    """M1 — estado NO sensible de la config de Sheets (para diagnóstico/mount). No expone
    valores: sólo banderas de presencia, el origen efectivo y qué variables faltan (por NOMBRE).
    Útil para que Orbit Home confirme el aislamiento sin leer secretos."""
    strict = _PENAFLOR_REQUIRE_NS_GSHEETS
    sid_ns  = bool(os.environ.get("PENAFLOR_GSHEETS_SPREADSHEET_ID", "").strip())
    sid_leg = bool(os.environ.get("GSHEETS_SPREADSHEET_ID", "").strip())
    cred_ns  = bool(os.environ.get("PENAFLOR_GSHEETS_CREDENTIALS_JSON", "").strip())
    cred_leg = bool(os.environ.get("GSHEETS_CREDENTIALS_JSON", "").strip())
    name_ns = bool(os.environ.get("PENAFLOR_GSHEETS_SHEET_NAME", "").strip())
    # Origen efectivo del spreadsheet_id (la señal más relevante de aislamiento).
    if sid_ns:
        source = "PENAFLOR_GSHEETS"
    elif sid_leg and not strict:
        source = "GSHEETS_LEGACY"
    else:
        source = "missing"
    id_ok   = sid_ns or (sid_leg and not strict)
    cred_ok = cred_ns or (cred_leg and not strict)
    missing = []
    if not id_ok:
        missing.append("PENAFLOR_GSHEETS_SPREADSHEET_ID" if strict else "GSHEETS_SPREADSHEET_ID")
    if not cred_ok:
        missing.append("PENAFLOR_GSHEETS_CREDENTIALS_JSON" if strict else "GSHEETS_CREDENTIALS_JSON")
    if strict and not name_ns:
        missing.append("PENAFLOR_GSHEETS_SHEET_NAME")
    return {
        "configured": gsheets_enabled(),
        "source": source,
        "strict": strict,
        "skip_boot": _PENAFLOR_SKIP_BOOT,
        "missing": missing,
        "spreadsheet_id_present": id_ok,
        "credentials_present": cred_ok,
    }

def _gsheets_get_worksheet():
    """Devuelve (worksheet, None) o (None, mensaje_error). Crea la pestaña/header si falta."""
    if not _GSHEETS_SPREADSHEET_ID:
        return None, "GSHEETS_SPREADSHEET_ID no configurado"
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except Exception as e:
        return None, f"gspread/google-auth no disponible: {e}"
    info = _gsheets_credentials_info()
    if not info:
        return None, "GSHEETS_CREDENTIALS_JSON no configurado o inválido"
    try:
        creds = Credentials.from_service_account_info(info, scopes=_GSHEETS_SCOPES)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(_GSHEETS_SPREADSHEET_ID)
        try:
            ws = sh.worksheet(_GSHEETS_SHEET_NAME)
        except Exception:
            ws = sh.add_worksheet(title=_GSHEETS_SHEET_NAME,
                                  rows=1000, cols=len(PLAN_SHEET_COLS))
            ws.update(range_name="A1", values=[PLAN_SHEET_COLS])
            return ws, None
        if not ws.row_values(1):   # header faltante
            ws.update(range_name="A1", values=[PLAN_SHEET_COLS])
        return ws, None
    except Exception as e:
        return None, f"error abriendo Google Sheet: {e}"

def _gsheets_find_row(values, header, plan_id):
    """Devuelve (num_fila_1based, fila_existente) buscando por columna 'id'. (None, None) si no está."""
    idx = {name: i for i, name in enumerate(header)}
    id_i = idx.get("id")
    if id_i is None:
        return None, None
    for r in range(1, len(values)):
        row = values[r]
        if len(row) > id_i and str(row[id_i]).strip() == str(plan_id).strip():
            return r + 1, row
    return None, None

def gsheets_upsert_plan(plan):
    """Inserta o actualiza una fila por 'id' (= fecha_vendedor_id).
    Devuelve (True, None) o (False, error). Preserva created_at en re-envíos."""
    ws, err = _gsheets_get_worksheet()
    if err:
        return False, err
    try:
        plan = dict(plan)
        plan["id"] = _plan_id(plan.get("fecha"), plan.get("vendedor_id"))
        values = ws.get_all_values()
        header = values[0] if values else list(PLAN_SHEET_COLS)
        row_num, existing = _gsheets_find_row(values, header, plan["id"])
        idx = {name: i for i, name in enumerate(header)}
        # No pisar created_at en re-envíos: conservar el de la hoja si existe.
        if row_num and existing and "created_at" in idx:
            ci = idx["created_at"]
            ex_created = existing[ci] if len(existing) > ci else ""
            if ex_created:
                plan["created_at"] = ex_created
        ordered = [_gsheet_cell(plan.get(c)) for c in header]
        if row_num:
            ws.update(range_name=f"A{row_num}", values=[ordered])
        else:
            ws.append_row(ordered, value_input_option="RAW")
        return True, None
    except Exception as e:
        return False, str(e)

def gsheets_verify_plan(plan_id, expected):
    """Relee la fila por 'id' y verifica que los campos esperados coinciden. True/False."""
    ws, err = _gsheets_get_worksheet()
    if err:
        return False
    try:
        values = ws.get_all_values()
        if not values:
            return False
        header = values[0]
        idx = {name: i for i, name in enumerate(header)}
        _, row = _gsheets_find_row(values, header, plan_id)
        if not row:
            return False
        for k, v in (expected or {}).items():
            if k in idx:
                cell = row[idx[k]] if len(row) > idx[k] else ""
                if str(cell).strip() != str(v).strip():
                    return False
        return True
    except Exception:
        return False

def gsheets_read_all():
    """Lee todas las filas de la hoja como lista de dicts. [] si no hay/no disponible."""
    ws, err = _gsheets_get_worksheet()
    if err:
        return []
    try:
        values = ws.get_all_values()
        if len(values) < 2:
            return []
        header = values[0]
        out = []
        for r in range(1, len(values)):
            row = values[r]
            d = {header[i]: (row[i] if i < len(row) else "") for i in range(len(header))}
            if str(d.get("fecha", "")).strip() and str(d.get("vendedor_id", "")).strip():
                out.append(d)
        return out
    except Exception:
        return []

def hydrate_planificacion_from_sheets():
    """Carga la cache SQLite desde Google Sheets (fuente de verdad). Devuelve nº de filas."""
    rows = gsheets_read_all()
    if not rows:
        return 0
    def _num(v, cast):
        try:
            return cast(v) if str(v).strip() != "" else None
        except Exception:
            return None
    conn = sqlite3.connect(str(DB_PATH))
    inserted = 0
    for row in rows:
        try:
            conn.execute("""
                INSERT OR IGNORE INTO planificacion
                  (fecha, vendedor_id, zona, dia_visita, venta_esperada,
                   ccc_tradicional, ccc_autoservicio, ccc_onpremise,
                   once_t, marcas, clientes_clave, acciones, estado,
                   editado_por, comentario_gerencia, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                str(row.get("fecha", "")).strip(),
                str(row.get("vendedor_id", "")).strip().upper(),
                row.get("zona") or None, row.get("dia_visita") or None,
                _num(row.get("venta_esperada"), float),
                _num(row.get("ccc_tradicional"), int),
                _num(row.get("ccc_autoservicio"), int),
                _num(row.get("ccc_onpremise"), int),
                _num(row.get("once_t"), int),
                row.get("marcas") or None, row.get("clientes_clave") or None,
                row.get("acciones") or None, row.get("estado") or "enviada",
                row.get("editado_por") or None, row.get("comentario_gerencia") or None,
                row.get("created_at") or None, row.get("updated_at") or None,
            ))
            inserted += 1
        except Exception:
            pass
    conn.commit()
    conn.close()
    print(f"[ORBIT] Cache hidratada desde Google Sheets: {inserted} planes")
    return inserted

def restore_planificacion_if_empty():
    """Si planificacion está vacía: restaura desde CSV de backup; si no hay CSV,
    restaura desde Google Sheets (fuente de verdad)."""
    try:
        conn = sqlite3.connect(str(DB_PATH))
        count = conn.execute("SELECT COUNT(*) FROM planificacion").fetchone()[0]
        conn.close()
        if count > 0:
            return
        if not PLAN_CSV_LATEST.exists():
            # No hay CSV local: intentar fuente de verdad (Google Sheets).
            if gsheets_enabled():
                hydrate_planificacion_from_sheets()
            return
        with open(str(PLAN_CSV_LATEST), "r", encoding="utf-8") as f:
            rows = list(_csv.DictReader(f))
        if not rows:
            # CSV existe pero está vacío: intentar fuente de verdad (Google Sheets).
            if gsheets_enabled():
                hydrate_planificacion_from_sheets()
            return
        conn = sqlite3.connect(str(DB_PATH))
        restored = 0
        for row in rows:
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO planificacion
                      (id, fecha, vendedor_id, zona, dia_visita, venta_esperada,
                       ccc_tradicional, ccc_autoservicio, ccc_onpremise,
                       once_t, marcas, clientes_clave, acciones, estado,
                       editado_por, comentario_gerencia, created_at, updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    row.get("id") or None,
                    row.get("fecha"), row.get("vendedor_id"),
                    row.get("zona") or None, row.get("dia_visita") or None,
                    float(row["venta_esperada"]) if row.get("venta_esperada") else None,
                    int(row["ccc_tradicional"]) if row.get("ccc_tradicional") else None,
                    int(row["ccc_autoservicio"]) if row.get("ccc_autoservicio") else None,
                    int(row["ccc_onpremise"]) if row.get("ccc_onpremise") else None,
                    int(row["once_t"]) if row.get("once_t") else None,
                    row.get("marcas") or None, row.get("clientes_clave") or None,
                    row.get("acciones") or None, row.get("estado") or "enviada",
                    row.get("editado_por") or None, row.get("comentario_gerencia") or None,
                    row.get("created_at") or None, row.get("updated_at") or None,
                ))
                restored += 1
            except Exception:
                pass
        conn.commit()
        conn.close()
        print(f"[ORBIT] AUTO-RESTORE: {restored} planes recuperados desde CSV de seguridad")
    except Exception as e:
        print(f"[WARN] restore_planificacion_if_empty: {e}")

def clean_code(v):
    s = str(v).upper().replace("V","").strip()
    return "".join(filter(str.isdigit, s))

def normalizar_vendedor_codigo(valor):
    if valor is None:
        return ""
    s = str(valor).strip().upper()
    if not s or s in ("NONE", "NAN"):
        return ""
    if s.startswith("V"):
        n = s[1:].strip()
    else:
        n = s
    try:
        return f"V{int(float(n))}"
    except Exception:
        return s

_READ_CSV_CACHE = {}

def read_csv(path):
    """Lee un CSV cacheando el parseo por (ruta, mtime). Devuelve una COPIA para que los
    endpoints puedan agregar/transformar columnas sin contaminar el caché. Evita re-parsear
    los datasets en cada request — clave para la velocidad del portal en Render."""
    if not path.exists(): return pd.DataFrame()
    try:
        key = (str(path), os.path.getmtime(path))
    except OSError:
        key = (str(path), 0)
    cached = _READ_CSV_CACHE.get(key)
    if cached is not None:
        return cached.copy()
    df = None
    for enc in ("utf-8-sig", "latin1", "utf-8"):
        try:
            df = pd.read_csv(path, encoding=enc); break
        except Exception:
            continue
    if df is None:
        return pd.DataFrame()
    _READ_CSV_CACHE[key] = df
    # No acumular versiones viejas del mismo archivo (mtime distinto)
    for k in [k for k in _READ_CSV_CACHE if k[0] == str(path) and k != key]:
        _READ_CSV_CACHE.pop(k, None)
    return df.copy()

_VENDEDORES_EXCLUIDOS = {1, 2, 5, 20}
_VENDEDORES_ACTIVOS_PLAN = {"V3","V4","V6","V7","V8","V9","V10"}

# _LEEME_EMPRESA — NO filtrar por la columna `Empresa` en ninguna métrica.
# `P&P LOGISTICA S.R.L` NO es otro distribuidor: es nuestra segunda razón social. En las
# ventas, `Proveedor` es `GRUPO PEÑAFLOR SA` en el 100% de las filas, facture quien facture.
# Medimos SIEMPRE con las dos empresas (confirmado por el usuario 2026-07-13).
# Filtrar `Empresa == 'Empresa'` borraba los clientes facturados vía P&P — en julio 2026,
# 135 de 229 clientes con compra: V6 perdía 30 de sus 34 y V10 35 de sus 40. Rompía el 11T
# (CCC a la mitad), el CCC empresa (Autoservicios daba 2 contra un objetivo de 145),
# innovaciones, acciones, la ruta del vendedor y el cierre.
# La única excepción legítima es un corte donde la razón social ES el dato pedido.

def _parse_num_ar(valor):
    """Parsea número en formato argentino (punto=miles, coma=decimal)."""
    try:
        s = str(valor).strip().replace(" ", "")
        s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return 0.0

def _feriados_config():
    path = CONFIG / "feriados.csv"
    if not path.exists():
        return set()
    try:
        df = pd.read_csv(path)
        col = "fecha" if "fecha" in df.columns else df.columns[0]
        fechas = pd.to_datetime(df[col], errors="coerce").dropna()
        return {d.date() for d in fechas}
    except Exception:
        return set()

def _es_dia_operativo(d):
    # Matinal Penaflor opera lunes a sabado; domingos y feriados no cuentan.
    return d.weekday() != 6 and d not in _feriados_config()

def _siguiente_dia_operativo(d):
    nxt = d + timedelta(days=1)
    while not _es_dia_operativo(nxt):
        nxt += timedelta(days=1)
    return nxt

def _fecha_planificacion_default(now=None):
    """Fecha objetivo para planes enviados por vendedores."""
    now = now or datetime.now(_ARG_TZ)
    hoy = now.date()
    if now.hour >= 12 or not _es_dia_operativo(hoy):
        return _siguiente_dia_operativo(hoy).isoformat()
    return hoy.isoformat()

def _clasificar_segmento(ramo: str, subsegmento: str) -> str:
    """EL SUBSEGMENTO MANDA SOBRE EL RAMO — ver `_clasificar()` en generar_datasets_acum.py,
    esta función es su espejo para el portal y tiene que dar el mismo resultado. El ERP mete
    carnicerías, verdulerías y panaderías bajo Ramo = AWAY FROM HOME; mirando el Ramo primero
    quedaban como On Premise (cobertura 6 botellas en vez de 3). Corregido 2026-07-30."""
    r = str(ramo).upper()
    s = str(subsegmento).upper()
    # "CADENAS REGIONALES" (plural) va sí o sí antes que On Premise: `CADENAS REGIONALES (BAR)`
    # es un formato de supermercado grande, no un bar, y el `(BAR)` matcheaba la clave "BAR".
    auto = ["AUTOSERVICIO","CADENA REGIONAL","CADENAS REGIONALES","SAR","LARGE FORMAT",
            "CASH&CARRY","CASH & CARRY","MAYORISTA","MAYORISTAS","TIENDA DE BEBIDAS"]
    if any(k in f"{r} | {s}" for k in auto):
        return "AUTOSERVICIO"
    # PROXIMITY (estaciones de servicio) = canal propio, decisión del negocio 2026-07-30.
    # Va antes que On Premise: el SubSegmento dice "Estacion de Servicio - AXION".
    if any(k in f"{r} | {s}" for k in ["PROXIMITY","ESTACION DE SERVICIO","ESTACIONES DE SERVICIO"]):
        return "PROXIMITY"
    on = ["ON PREMISE","AWAY FROM HOME","VINOTECA","VINOTECAS","BAR",
          "RESTAURANT","RESTAURANTE",
          "EVENTOS","TEMPORADA","CATERING","ON DIA","ON NOCHE"]
    # "KIOSK" cubre las dos grafías del ERP (KIOSCO y KIOSKO).
    trad = ["TRADITIONAL TRADE","ALMACEN","DESPENSA","KIOSCO","KIOSK","MAXIKIOSCO",
            "FIAMBRERIA","CARNICERIA","GRANJA","PANADERIA","CASA DE PASTAS",
            "VERDULERIA","TRADICIONAL"]
    for texto in (s, r):          # el SubSegmento decide solo; el Ramo es el fallback
        if any(k in texto for k in on):
            return "ON_PREMISE_VTK"
        if any(k in texto for k in trad):
            return "TRADICIONAL"
    return "OTROS"

_VENTAS_PARSED_CACHE = {}

def _ventas_parsed() -> pd.DataFrame:
    """ventas.csv parseado UNA sola vez por mtime (numérico + segmento vectorizado).
    Todos los endpoints filtran sobre este DataFrame en vez de releer/reparsear el CSV
    en cada request — clave para la velocidad del portal en Render (CPU limitada).
    Columnas: cliente_id, vendedor_codigo, importe_neto, fecha, segmento_operativo."""
    path = INPUTS / "ventas.csv"
    if not path.exists():
        return pd.DataFrame()
    try:
        key = os.path.getmtime(path)
    except OSError:
        key = 0
    cached = _VENTAS_PARSED_CACHE.get(key)
    if cached is not None:
        return cached
    df = pd.DataFrame()
    for enc in ("latin1", "utf-8-sig", "utf-8"):
        try:
            # dtype=str: parseo determinístico en Render (la inferencia de coma decimal
            # difiere entre versiones de pandas). El importe se convierte con _parse_num_ar.
            df = pd.read_csv(path, sep=";", encoding=enc, dtype=str, low_memory=False)
            break
        except Exception:
            continue
    req = {"Cliente", "CodVendedor", "ImporteNetoItem", "FechaComprobante", "Ramo", "Subramo"}
    if df.empty or not req.issubset(set(df.columns)):
        return pd.DataFrame()
    df["cliente_id"]      = pd.to_numeric(df["Cliente"], errors="coerce")
    df["vendedor_codigo"] = pd.to_numeric(df["CodVendedor"], errors="coerce")
    df["importe_neto"]    = df["ImporteNetoItem"].apply(_parse_num_ar)
    df["fecha"]           = pd.to_datetime(df["FechaComprobante"], dayfirst=True, errors="coerce")
    df["codigo_art"]      = pd.to_numeric(df["Codigo"], errors="coerce") if "Codigo" in df.columns else pd.NA
    df["cant_base"]       = df["CantBase"].apply(_parse_num_ar) if "CantBase" in df.columns else 0.0
    # Segmento vectorizado: clasificar SÓLO las combinaciones únicas (Ramo, Subramo)
    # y mapear — evita el .apply() fila por fila sobre miles de filas en cada request.
    rm = df["Ramo"].fillna("").astype(str)
    sb = df["Subramo"].fillna("").astype(str)
    seg_map = {par: _clasificar_segmento(par[0], par[1]) for par in set(zip(rm, sb))}
    df["segmento_operativo"] = [seg_map[(a, b)] for a, b in zip(rm, sb)]
    _VENTAS_PARSED_CACHE.clear()
    _VENTAS_PARSED_CACHE[key] = df
    return df

def _cargar_ventas_mes_actual() -> pd.DataFrame:
    """Ventas del mes calendario actual, ImporteNetoItem > 0, sin vendedores excluidos.
    Filtra sobre _ventas_parsed() (cacheado). Columnas: cliente_id, vendedor_codigo, segmento_operativo."""
    df = _ventas_parsed()
    if df.empty:
        return pd.DataFrame()
    hoy = datetime.now()
    mes_inicio = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    d = df[(df["fecha"] >= mes_inicio) & (df["importe_neto"] > 0)
           & (~df["vendedor_codigo"].isin(_VENDEDORES_EXCLUIDOS))]
    d = d.dropna(subset=["cliente_id", "vendedor_codigo"])
    return d[["cliente_id", "vendedor_codigo", "segmento_operativo"]].copy()

def _cargar_ventas_dia(fecha_str: str = None):
    """Ventas de UN día (fecha_str 'YYYY-MM-DD', o el más reciente si None).
    Filtra sobre _ventas_parsed() (cacheado). Excluye V2/V5/V20.
    Devuelve (DataFrame[cliente_id, vendedor_codigo, importe_neto, segmento_operativo], fecha_usada)."""
    df = _ventas_parsed()
    if df.empty:
        return pd.DataFrame(), ""
    df = df[(~df["vendedor_codigo"].isin(_VENDEDORES_EXCLUIDOS)) & (df["importe_neto"] > 0)]
    df = df.dropna(subset=["cliente_id", "vendedor_codigo", "fecha"])
    if df.empty:
        return pd.DataFrame(), ""
    target = pd.to_datetime(fecha_str).date() if fecha_str else df["fecha"].dt.date.max()
    df_dia = df[df["fecha"].dt.date == target]
    fecha_usada = str(target)
    if df_dia.empty:
        return pd.DataFrame(), fecha_usada
    return df_dia[["cliente_id", "vendedor_codigo", "importe_neto", "segmento_operativo"]].copy(), fecha_usada

def _ccc_mes_por_vendedor(ventas_mes: pd.DataFrame) -> dict:
    """
    Calcula CCC Compradores Mes por vendedor y segmento.
    Retorna dict keyed por int(vendedor_codigo):
      {tradicional, autoservicio, onpremise, proximity, total}
    V3 tiene autoservicio=0 y onpremise=0 por regla de negocio (Proximity sí lo trabaja).
    'total' suma los CUATRO canales: Proximity es canal propio desde 2026-07-30 y si no se
    sumara, esos clientes desaparecerían del CCC total del vendedor sin que nada lo avise.
    """
    if ventas_mes.empty:
        return {}

    # Un cliente puede aparecer en múltiples líneas → deduplicate por cliente+vendedor
    buyers = ventas_mes.drop_duplicates(subset=["cliente_id", "vendedor_codigo"]).copy()

    result = {}
    for cod, grp in buyers.groupby("vendedor_codigo"):
        cod_int = int(cod)
        trad  = int((grp["segmento_operativo"] == "TRADICIONAL").sum())
        aas   = int((grp["segmento_operativo"] == "AUTOSERVICIO").sum())
        op    = int((grp["segmento_operativo"] == "ON_PREMISE_VTK").sum())
        prox  = int((grp["segmento_operativo"] == "PROXIMITY").sum())
        if cod_int == 3:  # V3 no trabaja autoservicio ni on premise (regla de negocio)
            aas = 0
            op  = 0
        result[cod_int] = {
            "tradicional": trad,
            "autoservicio": aas,
            "onpremise":   op,
            "proximity":   prox,
            "total":       trad + aas + op + prox,
        }
    return result

def _clientes_por_dia(dia: str) -> pd.DataFrame:
    """Calcula cartera del día desde clientes.xlsx + compra_mes_flag desde ventas.csv."""
    cli_path = INPUTS / "clientes.xlsx"
    if not cli_path.exists():
        return pd.DataFrame()
    try:
        cli = pd.read_excel(cli_path)
    except Exception:
        return pd.DataFrame()

    dia_cap = dia.capitalize()
    cli = cli[cli["DiasVisita"].astype(str).str.strip() == dia_cap]
    cli = cli[~cli["codven"].isin([2, 5, 20])]
    if cli.empty:
        return pd.DataFrame()

    ventas_mes = _cargar_ventas_mes_actual()
    ccc_ids = set(ventas_mes["cliente_id"].dropna().astype(int)) if not ventas_mes.empty else set()

    sub_col = next((c for c in cli.columns if "subseg" in c.lower() or "subramo" in c.lower()), None)
    nombre_col = next(
        (c for c in cli.columns if c.lower().replace("_","").replace(" ","") in ("razonsocial","nombre","cliente")),
        next((c for c in cli.columns if "razon" in c.lower() or "nombre" in c.lower()), None)
    )

    rows = []
    for _, row in cli.iterrows():
        cid_raw = row.get("Codigo")
        vcod_raw = row.get("codven")
        if pd.isna(cid_raw) or pd.isna(vcod_raw):
            continue
        cid = int(cid_raw)
        vcod = int(vcod_raw)
        seg = _clasificar_segmento(str(row.get("Ramo", "")),
                                   str(row.get(sub_col, "") if sub_col else ""))
        # V3 (Nadia) solo trabaja Tradicional almacén/despensa/kiosco
        if vcod == 3:
            _ss = str(row.get(sub_col, "")).upper() if sub_col else ""
            if seg != "TRADICIONAL" or not any(k in _ss for k in ("ALMACEN", "DESPENSA", "KIOSCO")):
                continue
        compra_mes = 1 if cid in ccc_ids else 0
        estado = "SIN_COMPRA_MES" if compra_mes == 0 else "COBERTURA_OK"
        rows.append({
            "cliente_id": cid,
            "cliente_nombre": str(row[nombre_col]) if (nombre_col and pd.notna(row.get(nombre_col))) else str(cid),
            "vendedor_codigo": vcod,
            "vendedor_id": f"V{vcod}",
            "dias_visita": dia_cap,
            "segmento": seg,
            "segmento_operativo": seg,
            "compra_mes_flag": compra_mes,
            "compra_ayer_flag": 0,
            "estado": estado,
            "estado_cliente": estado,
            "prioridad_label": "ALTA" if compra_mes == 0 else "NORMAL",
            "importe_mes": 0.0,
            "botellas_mes": 0,
            "ultima_compra_fecha": None,
            "ultima_compra_importe": None,
        })
    df = pd.DataFrame(rows)

    # Enriquecer con última compra
    try:
        hist = read_csv(BASE / "02_HISTORY" / "historial_ventas_cliente.csv")
        if not hist.empty:
            hist.columns = [c.lstrip("﻿") for c in hist.columns]
            f_col = next((c for c in hist.columns if "fecha" in c.lower()), None)
            i_col = next((c for c in hist.columns if "importe" in c.lower() or "neto" in c.lower()), None)
            id_col = next((c for c in hist.columns if c.lower() in ("cliente", "cliente_id", "cod_cliente")), None)
            if all([f_col, i_col, id_col]):
                hist[id_col] = pd.to_numeric(hist[id_col], errors="coerce")
                hist[i_col] = pd.to_numeric(hist[i_col], errors="coerce")
                ultima = (hist.sort_values(f_col, ascending=False)
                          .drop_duplicates(subset=[id_col])[[id_col, f_col, i_col]]
                          .rename(columns={id_col: "cliente_id", f_col: "ultima_compra_fecha",
                                           i_col: "ultima_compra_importe"}))
                ultima["cliente_id"] = pd.to_numeric(ultima["cliente_id"], errors="coerce")
                df["cliente_id"] = pd.to_numeric(df["cliente_id"], errors="coerce")
                df = df.merge(ultima, on="cliente_id", how="left", suffixes=("", "_h"))
                if "ultima_compra_fecha_h" in df.columns:
                    df["ultima_compra_fecha"] = df["ultima_compra_fecha_h"].fillna(df["ultima_compra_fecha"])
                    df["ultima_compra_importe"] = df["ultima_compra_importe_h"].fillna(df["ultima_compra_importe"])
                    df.drop(columns=["ultima_compra_fecha_h", "ultima_compra_importe_h"], inplace=True, errors="ignore")
    except Exception:
        pass

    records = df.to_dict(orient="records")
    for rec in records:
        for k, v in rec.items():
            if isinstance(v, float) and not math.isfinite(v):
                rec[k] = None
    return pd.DataFrame(records)


def _cargar_feriados():
    import csv
    p = CONFIG / "feriados.csv"
    if not p.exists():
        return set()
    with open(p, encoding="utf-8-sig") as f:
        return {row["fecha"] for row in csv.DictReader(f)}

def contar_dias_habiles(fecha_corte=None):
    if fecha_corte is None:
        fecha_corte = datetime.now()
    feriados = _cargar_feriados()
    inicio = datetime(fecha_corte.year, fecha_corte.month, 1)
    fin_mes = datetime(fecha_corte.year, fecha_corte.month + 1, 1) - timedelta(days=1)
    total, corridos = 0, 0
    feriados_del_mes = []
    d = inicio
    while d <= fin_mes:
        fecha_str = d.strftime("%Y-%m-%d")
        es_domingo = d.weekday() == 6
        es_feriado = fecha_str in feriados
        if es_feriado:
            feriados_del_mes.append(fecha_str)
        if not es_domingo and not es_feriado:
            total += 1
            if d <= fecha_corte:
                corridos += 1
        d += timedelta(days=1)
    print(f"[ORBIT calendario] fecha_corte={fecha_corte.strftime('%Y-%m-%d')} | "
          f"total_comerciales={total} | corridos={corridos} | "
          f"feriados_mes={feriados_del_mes}")
    return {
        "total": total,
        "corridos": corridos,
        "restantes": total - corridos,
        "fecha_corte": fecha_corte.strftime("%Y-%m-%d"),
        "total_dias_comerciales_mes": total,
        "dias_comerciales_corridos": corridos,
        "feriados_detectados_del_mes": feriados_del_mes,
        "fecha_corte_calendario": fecha_corte.strftime("%Y-%m-%d"),
    }

@app.route("/favicon.ico")
def favicon():
    return "", 204

# ====== STATIC ======
@app.route("/")
@app.route("/portal.html")
def portal_html():
    """Sirve portal.html sin ETag ni Last-Modified para evitar cache en móviles."""
    try:
        content = (FRONTEND / "portal.html").read_bytes()
        resp = make_response(content)
        resp.headers["Content-Type"] = "text/html; charset=utf-8"
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp
    except Exception as e:
        return f"Error sirviendo portal: {e}", 500

@app.route("/<path:filename>")
def frontend(filename):
    return send_from_directory(str(FRONTEND), filename)

# ====== AUTH ======
@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json()
    user = (data.get("user") or "").lower().strip()
    u = USERS.get(user)
    if u and u["pwd"] == data.get("pwd",""):
        return jsonify({"ok":True,"rol":u["rol"],"nombre":u["nombre"],"vendedor_id":user.upper()})
    return jsonify({"ok":False,"error":"Credenciales inválidas"}), 401

# ====== HEALTHCHECK (Render / UptimeRobot) ======
@app.route("/api/healthz")
def healthz():
    # commit: SHA que Render inyecta en el deploy (RENDER_GIT_COMMIT). Permite que el
    # cierre (.bat) espere a que la instancia NUEVA esté sirviendo antes de abrir el
    # portal — durante el redeploy Render sigue respondiendo 200 desde la vieja.
    return jsonify({"status": "ok", "service": "orbit-penaflor-pav", "healthcheck": True,
                    "commit": os.environ.get("RENDER_GIT_COMMIT", "")}), 200

# ====== DIAGNÓSTICO ======
@app.route("/api/diagnostico")
def diagnostico():
    vol = read_csv(DATASETS / "mod_volumen_vendedor.csv")
    ccc_df = read_csv(DATASETS / "mod_ccc_segmento.csv")
    t11_df = read_csv(DATASETS / "mod_11_titulares.csv")
    vend = read_csv(CONFIG / "vendedores_activos.csv")

    fuentes = []
    for nombre, path in [("volumen", DATASETS / "mod_volumen_vendedor.csv"),
                          ("ccc", DATASETS / "mod_ccc_segmento.csv"),
                          ("11t", DATASETS / "mod_11_titulares.csv"),
                          ("vendedores", CONFIG / "vendedores_activos.csv")]:
        if path.exists():
            fuentes.append({"archivo": nombre, "path": str(path), "filas": len(read_csv(path)), "estado": "OK"})
        else:
            fuentes.append({"archivo": nombre, "path": str(path), "filas": 0, "estado": "NO ENCONTRADO"})

    vendedores_detectados = []
    if not vend.empty:
        vend_activos = vend[vend["activo"]==1]
        for _, v in vend_activos.iterrows():
            cod = str(v["codigo_vendedor"]).strip()
            vv = vol[vol["vendedor_codigo"].astype(str).apply(clean_code) == clean_code(cod)] if not vol.empty else pd.DataFrame()
            obj = float(vv["objetivo_mes"].sum()) if not vv.empty else 0
            acum = float(vv["acumulado_mes"].sum()) if not vv.empty else 0
            vendedores_detectados.append({"codigo": cod, "nombre": str(v["nombre_vendedor"]), "objetivo": obj, "acumulado": acum})

    # fecha_corte desde ventas.csv con sep=";" explícito.
    # sep=None falla en Linux (mismo patrón que ventas_mes.csv): columnas mal alineadas,
    # filas del día más reciente quedan con ImporteNetoItem=0 y se pierden.
    _fecha_corte_datos = None
    try:
        _sv = INPUTS / "ventas.csv"
        if _sv.exists():
            _dv2 = pd.DataFrame()
            for _enc in ("latin1", "utf-8-sig", "utf-8"):
                try:
                    _dv2 = pd.read_csv(_sv, sep=";", encoding=_enc,
                                       usecols=["FechaComprobante"], low_memory=False)
                    break
                except (UnicodeDecodeError, ValueError):
                    continue
            if not _dv2.empty:
                _ult = pd.to_datetime(_dv2["FechaComprobante"], format="%d/%m/%Y", errors="coerce").max()
                if pd.notna(_ult):
                    _fecha_corte_datos = _ult.to_pydatetime()
    except Exception:
        pass
    if _fecha_corte_datos is None:
        _fecha_corte_datos = datetime.now(_ARG_TZ).replace(tzinfo=None) - timedelta(days=1)
    dias = contar_dias_habiles(fecha_corte=_fecha_corte_datos)
    total_acum = sum(v["acumulado"] for v in vendedores_detectados)

    # Segmentos: denominadores desde cartera real (clientes.xlsx); cubiertos desde mod_ccc_segmento (ayer)
    seg_ids = [
        ("TRADICIONAL",    "Tradicional",          3,  "#5BC23A"),
        ("AUTOSERVICIO",   "Autoservicio",          6,  "#4DA3FF"),
        ("ON_PREMISE_VTK", "On Premise / Vinoteca", 6,  "#9B7BFF"),
        ("PROXIMITY",      "Proximity",             6,  "#F5A623"),
    ]
    cartera_real_total = 0
    cartera_segs_real = {sid: 0 for sid, *_ in seg_ids}
    # Cartera real desde clientes.xlsx CACHEADO (_clientes_maestro, ya excluye V2/V5/V20).
    cli_df = _clientes_maestro()
    if not cli_df.empty:
        try:
            cartera_real_total = len(cli_df)
            ramo_col = next((c for c in cli_df.columns if c.lower() == "ramo"), None)
            sub_col  = next((c for c in cli_df.columns if "subramo" in c.lower() or "subseg" in c.lower()), None)
            if ramo_col:
                rm = cli_df[ramo_col].fillna("").astype(str)
                sb = cli_df[sub_col].fillna("").astype(str) if sub_col else pd.Series([""] * len(cli_df), index=cli_df.index)
                seg_map = {par: _clasificar_segmento(par[0], par[1]) for par in set(zip(rm, sb))}
                _seg = pd.Series([seg_map[(a, b)] for a, b in zip(rm, sb)], index=cli_df.index)
                for sid, *_ in seg_ids:
                    cartera_segs_real[sid] = int((_seg == sid).sum())
        except Exception:
            pass
    cdia_df = read_csv(DATASETS / "clientes_dia.csv")
    segmentos = []
    for sid, snombre, req, color in seg_ids:
        total_cli = cartera_segs_real.get(sid) or \
                    (int((cdia_df["segmento_operativo"] == sid).sum()) if not cdia_df.empty else 0)
        cubiertos = int(
            ccc_df.loc[ccc_df["segmento_operativo"] == sid, "coberturas_logradas"]
            .apply(pd.to_numeric, errors="coerce").sum()
        ) if not ccc_df.empty else 0
        segmentos.append({"id": sid, "nombre": snombre, "req": req,
                          "clientes": total_cli, "cubiertos": cubiertos, "color": color,
                          "cubiertos_label": "ayer"})

    # Titulares11: agrega tiene_flag por marca desde mod_11_titulares
    titulares11 = []
    if not t11_df.empty and "marca_objetivo" in t11_df.columns and "tiene_flag" in t11_df.columns:
        t11_df["tiene_flag"] = pd.to_numeric(t11_df["tiene_flag"], errors="coerce").fillna(0)
        agg = (t11_df.groupby("marca_objetivo", dropna=False)
               .agg(objetivo=("tiene_flag", "count"), cubiertos=("tiene_flag", "sum"))
               .reset_index())
        agg["cubiertos"] = agg["cubiertos"].astype(int)
        for _, row in agg.iterrows():
            titulares11.append({"marca": row["marca_objetivo"],
                                 "objetivo": int(row["objetivo"]),
                                 "cubiertos": row["cubiertos"]})
        titulares11.sort(key=lambda x: -x["cubiertos"])

    botellas_dia = int(pd.to_numeric(ccc_df["botellas_vendidas"], errors="coerce").sum()) if not ccc_df.empty and "botellas_vendidas" in ccc_df.columns else 0
    botellas_mes = None
    try:
        hv_path = BASE / "02_HISTORY" / "historial_ventas_cliente.csv"
        if hv_path.exists():
            hv = read_csv(hv_path)
            if not hv.empty and "cant_base" in hv.columns and "importe_neto" in hv.columns:
                hoy = datetime.now()
                mes_str = hoy.strftime("%Y-%m")
                hv["_fecha"] = hv["fecha_comprobante"].astype(str).str[:7]
                hv["_importe"] = pd.to_numeric(hv["importe_neto"], errors="coerce").fillna(0)
                hv["_cant"] = pd.to_numeric(hv["cant_base"], errors="coerce").fillna(0)
                hv["_vend"] = pd.to_numeric(hv["vendedor_codigo"], errors="coerce").fillna(0).astype(int)
                mes_df = hv[(hv["_fecha"] == mes_str) & (hv["_importe"] > 0) & (~hv["_vend"].isin(_VENDEDORES_EXCLUIDOS))]
                botellas_mes = int(mes_df["_cant"].sum())
    except Exception:
        botellas_mes = None

    # dia_operativo = próxima matinal (siguiente día operativo desde hoy AR)
    _DIAS_AR = {0:'LU', 1:'MA', 2:'MI', 3:'JU', 4:'VI', 5:'SA', 6:'DO'}
    _now_diag = datetime.now(_ARG_TZ)
    dia_op       = _DIAS_AR[_now_diag.weekday()]
    fecha_corte_rt = _now_diag.strftime("%Y-%m-%d")
    _sig_matinal   = _siguiente_dia_operativo(_fecha_corte_datos.date())
    dia_op_matinal = _DIAS_AR[_sig_matinal.weekday()]

    # fecha_datos = cuándo fue generado el último dataset (para transparencia)
    fecha_datos = fecha_corte_rt  # fallback: hoy
    fecha_obj_str = None
    modo_fecha = "SIN_DATOS"
    if not vol.empty:
        try:
            if "fecha_ejecucion" in vol.columns:
                fecha_datos = str(vol["fecha_ejecucion"].iloc[0])  # fecha del último regenerar
            if "fecha_objetivo" in vol.columns:
                fecha_obj_str = str(vol["fecha_objetivo"].iloc[0])
            modo_fecha = "REAL"
        except Exception:
            pass

    return jsonify({
        "modo_datos": "REAL",
        "generado_en": _now_ar(),
        "calendario": dias,
        "fuentes": fuentes,
        "vendedores_detectados": vendedores_detectados,
        "total_acumulado_csv": total_acum,
        "excluidos": ["V2", "V5"],
        "advertencias": [] if fuentes[0]["estado"] == "OK" else ["Faltan archivos de datos"],
        "v3_autoservicio": False,
        "segmentos": segmentos,
        "titulares11": titulares11,
        "botellas_dia": botellas_dia,
        "botellas_mes": botellas_mes,
        "cartera_real_total": cartera_real_total,
        "dia_snapshot": dia_op,
        "fecha_datos": fecha_datos,          # cuándo se regeneraron los datasets (puede ser de ayer)
        "fecha_corte": _fecha_corte_datos.strftime("%Y-%m-%d"),  # última fecha de datos en ventas.csv
        "fecha_objetivo": fecha_obj_str,
        "fecha_matinal": _sig_matinal.isoformat(),  # próxima matinal = siguiente día operativo
        "fecha_planificacion_default": _fecha_planificacion_default(_now_diag),
        "dia_operativo": dia_op_matinal,     # día de la próxima matinal (siguiente día operativo)
        "modo_fecha": modo_fecha,
    })

# ====== DASHBOARD REAL ======
@app.route("/api/dashboard")
def dashboard():
    dia_param = request.args.get("dia", "").strip()

    # Precompute day-specific client counts when a day is requested
    clientes_dia_map = {}  # {cn: {total, sin_compra}}
    if dia_param:
        cli_dia_df = _clientes_por_dia(dia_param)
        if cli_dia_df is not None and not cli_dia_df.empty and "vendedor_codigo" in cli_dia_df.columns:
            cli_dia_df["_cn"] = cli_dia_df["vendedor_codigo"].astype(str).apply(clean_code)
            for vcn, grp in cli_dia_df.groupby("_cn"):
                clientes_dia_map[vcn] = {
                    "total": len(grp),
                    "sin_compra": int((grp["compra_mes_flag"] == 0).sum()),
                }

    vol     = read_csv(DATASETS / "mod_volumen_vendedor.csv")
    vend    = read_csv(CONFIG  / "vendedores_activos.csv")
    ccc_df  = read_csv(DATASETS / "mod_ccc_segmento.csv")   # CCC día (ayer)
    t11_df  = read_csv(DATASETS / "mod_11_titulares.csv")
    cdia_df = read_csv(DATASETS / "clientes_dia.csv")        # para oportunidades (día base)

    # 11T cumplidos por vendedor: desde mod_11t_acum.csv (cobertura acumulada con mínimo).
    # mod_11_titulares.csv (objetivo del día) llega del motor con tiene_flag/botellas en 0,
    # por eso el KPI "11T ✓" daba 0 para todos. mod_11t_acum sí está poblado.
    t11_acum_map = {}
    try:
        _t11a = read_csv(DATASETS / "mod_11t_acum.csv")
        if not _t11a.empty and {"vendedor_codigo", "tiene_flag"}.issubset(_t11a.columns):
            _t11a["_tf"] = pd.to_numeric(_t11a["tiene_flag"], errors="coerce").fillna(0)
            _t11a["_cn"] = _t11a["vendedor_codigo"].astype(str).apply(clean_code)
            for _cnk, _g in _t11a.groupby("_cn"):
                t11_acum_map[_cnk] = {"cumplidos": int(_g["_tf"].sum()), "total": int(len(_g))}
    except Exception:
        pass

    # CCC Compradores Mes — desde ventas.csv del mes actual (no clientes_dia)
    ventas_mes = _cargar_ventas_mes_actual()
    ccc_mes_map = _ccc_mes_por_vendedor(ventas_mes)

    # Corridos: usar última fecha con datos reales en ventas.csv, no datetime.now()
    # Si ventas.csv tiene datos hasta June 1 y hoy es June 2, corridos=1 (no 2)
    _fecha_corte_ventas = None
    try:
        _src_v = INPUTS / "ventas.csv"
        if _src_v.exists():
            _dv = _preparar_df_ventas(_src_v)
            if not _dv.empty and "FechaComprobante" in _dv.columns:
                _ultima = pd.to_datetime(_dv["FechaComprobante"], dayfirst=True, errors="coerce").max()
                if pd.notna(_ultima):
                    _fecha_corte_ventas = _ultima.to_pydatetime()
    except Exception:
        pass
    dias = contar_dias_habiles(fecha_corte=_fecha_corte_ventas)
    _corridos = max(dias["corridos"], 1)
    _total    = dias["total"]

    # Venta ayer live — desde ventas.csv (día más reciente), evita depender del CSV estático
    _ventas_dia_live, _ = _cargar_ventas_dia()
    _venta_ayer_live = {}
    if not _ventas_dia_live.empty:
        for _cv, _grp in _ventas_dia_live.groupby("vendedor_codigo"):
            _venta_ayer_live[int(_cv)] = round(float(_grp["importe_neto"].sum()), 2)

    # Fallback: objetivo/acumulado/avance desde resultado.xlsx para vendedores sin maestro de clientes
    resultado_fallback = {}
    resultado_path = INPUTS / "resultado.xlsx"
    if resultado_path.exists():
        try:
            avance_df = pd.read_excel(resultado_path, sheet_name="Avance")
            for _, r in avance_df.iterrows():
                cn_r = clean_code(str(r.get("VendedorCodigo", "")))
                if cn_r:
                    resultado_fallback[cn_r] = {
                        "objetivo": float(r.get("ValorObjetivo", 0) or 0),
                        "acumulado": float(r.get("Acumulado", 0) or 0),
                        "avance": float(r.get("Avance", 0) or 0),
                    }
        except Exception:
            pass

    if vol.empty:
        return jsonify({"error": "No se encontró mod_volumen_vendedor.csv", "modo_datos": "SIN_DATOS"}), 500

    vend = vend[vend["activo"] == 1]
    result = []
    for _, v in vend.iterrows():
        cod = str(v["codigo_vendedor"]).strip()
        nombre = str(v["nombre_vendedor"]).strip()
        cn = clean_code(cod)

        vv = vol[vol["vendedor_codigo"].astype(str).apply(clean_code) == cn] if not vol.empty else pd.DataFrame()
        cv = ccc_df[ccc_df["vendedor_codigo"].astype(str).apply(clean_code) == cn] if not ccc_df.empty else pd.DataFrame()
        tv = t11_df[t11_df["vendedor_codigo"].astype(str).apply(clean_code) == cn] if not t11_df.empty else pd.DataFrame()

        cod_int = int(cn) if cn.isdigit() else 0

        # resultado.xlsx es la fuente primaria para obj/acum/avance
        # (se actualiza diariamente sin necesidad de regenerar el motor)
        # mod_volumen_vendedor.csv es el fallback si resultado.xlsx no tiene al vendedor
        sin_maestro = False
        if cn in resultado_fallback:
            fb = resultado_fallback[cn]
            obj  = fb["objetivo"]
            acum = fb["acumulado"]
            av   = fb["avance"]
        elif not vv.empty:
            obj  = float(vv["objetivo_mes"].sum())
            acum = float(vv["acumulado_mes"].sum())
            av   = float(vv["avance_pct"].mean())
        else:
            obj = acum = av = 0
            sin_maestro = True

        # venta_ayer live desde ventas.csv; fallback al CSV estático
        venta_ayer = _venta_ayer_live.get(cod_int, float(vv["venta_ayer"].sum()) if not vv.empty else 0)

        # tendencia_pct = Avance de resultado.xlsx (Tendencia/Objetivo) cuando hay fuente;
        # para vendedores sin resultado.xlsx (fallback mod_volumen) se proyecta por días hábiles.
        if cn in resultado_fallback:
            tendencia_pct = round(av, 2)
        else:
            tendencia_pct = round((acum / _corridos) * _total / obj * 100, 2) if obj else 0
        cli_total = int(vv["clientes_planificados"].sum()) if not vv.empty and "clientes_planificados" in vv.columns else 0
        cli_sin = int(vv["clientes_sin_compra_mes"].sum()) if not vv.empty and "clientes_sin_compra_mes" in vv.columns else 0
        # Override with day-specific counts when dia requested (0 if vendor not scheduled that day)
        if dia_param:
            day_data = clientes_dia_map.get(cn, {"total": 0, "sin_compra": 0})
            cli_total = day_data["total"]
            cli_sin   = day_data["sin_compra"]

        # CCC Compradores Mes — desde ventas.csv del mes actual
        cod_int = int(cn) if cn.isdigit() else 0
        ccc_mes = ccc_mes_map.get(cod_int, {"tradicional": 0, "autoservicio": 0, "onpremise": 0,
                                            "proximity": 0, "total": 0})
        ccc_mes_trad = ccc_mes["tradicional"]
        ccc_mes_as   = ccc_mes["autoservicio"]   # ya es 0 para V3 por regla en _ccc_mes_por_vendedor
        ccc_mes_op   = ccc_mes["onpremise"]
        ccc_mes_prox = ccc_mes.get("proximity", 0)   # estaciones de servicio: canal propio, V3 sí lo trabaja

        # CCC DÍA — desde mod_ccc_segmento (clientes con compra ayer)
        def _ccc_dia_seg(df, seg_pattern):
            if df.empty:
                return 0
            mask = df["segmento_operativo"].astype(str).str.upper().str.contains(seg_pattern, na=False)
            return int(df.loc[mask, "clientes_con_compra"].sum())

        ccc_dia_trad = _ccc_dia_seg(cv, "TRADICIONAL")
        ccc_dia_as   = _ccc_dia_seg(cv, "AUTOSERVICIO")
        ccc_dia_op   = _ccc_dia_seg(cv, "ON_PREMISE|VTK")
        ccc_dia_prox = _ccc_dia_seg(cv, "PROXIMITY")

        # 11T cumplidos desde mod_11t_acum (cobertura real); fallback a mod_11_titulares.
        _t11v = t11_acum_map.get(cn)
        if _t11v is not None:
            t11_cumplidos = _t11v["cumplidos"]
            t11_total = _t11v["total"]
        else:
            t11_cumplidos = int(tv["tiene_flag"].sum()) if not tv.empty and "tiene_flag" in tv.columns else 0
            t11_total = len(tv)

        # V3 no trabaja autoservicio ni on premise (ccc_mes ya es 0; refuerzo ccc_dia)
        if cod.upper() == "V3":
            ccc_dia_as = 0
            ccc_dia_op = 0

        # Oportunidades: clientes sin compra del día (V3 excluye AS)
        oportunidades = 0
        if dia_param:
            oportunidades = clientes_dia_map.get(cn, {"sin_compra": 0})["sin_compra"]
        elif not cdia_df.empty and "estado_cliente" in cdia_df.columns and "vendedor_codigo" in cdia_df.columns:
            opv = cdia_df[cdia_df["vendedor_codigo"].astype(str).apply(clean_code) == cn]
            omask = opv["estado_cliente"].astype(str).str.lower().str.contains("sin", na=False)
            if cod.upper() == "V3" and "segmento_operativo" in opv.columns:
                _segu = opv["segmento_operativo"].astype(str).str.upper()
                omask = omask & (_segu != "AUTOSERVICIO") & (~_segu.str.contains("ON_PREMISE|VTK", na=False))
            oportunidades = int(omask.sum())

        result.append({
            "vendedor_id": cod,
            "vendedor_nombre": nombre,
            "sin_maestro": sin_maestro,
            "last_sync": _now_ar(),
            "kpis": {
                "objetivo": obj, "acumulado": acum, "avance_pct": av, "tendencia_pct": tendencia_pct,
                "venta_hoy_total": venta_ayer, "venta_mes_actual": acum,
                "clientes_total": cli_total, "clientes_pendientes": cli_sin,
                # CCC Compradores Mes — fuente: ventas.csv mes actual, ImporteNetoItem > 0
                "ccc_tradicional": ccc_mes_trad, "ccc_autoservicio": ccc_mes_as, "ccc_onpremise": ccc_mes_op,
                "ccc_proximity": ccc_mes_prox,
                "ccc_total": ccc_mes_trad + ccc_mes_as + ccc_mes_op + ccc_mes_prox,
                # CCC Compradores Día — fuente: mod_ccc_segmento (ayer)
                "ccc_dia_tradicional": ccc_dia_trad, "ccc_dia_autoservicio": ccc_dia_as, "ccc_dia_onpremise": ccc_dia_op,
                "ccc_dia_proximity": ccc_dia_prox,
                "ccc_dia_total": ccc_dia_trad + ccc_dia_as + ccc_dia_op + ccc_dia_prox,
                "once_titulares_cumplidos": t11_cumplidos, "once_titulares_total": t11_total,
                "cobertura_pct": cli_total and (100 - (cli_sin/cli_total*100)) or 0,
                "alertas_criticas": 0, "oportunidades": oportunidades,
                "inversion_desc_ars": 0.0, "sin_cargo_ars": 0.0,
                "impacto_alertas_ars": 0.0, "venta_mes_anterior": 0.0,
                "trabaja_autoservicio": cod.upper() != "V3",
                "trabaja_onpremise": cod.upper() != "V3"
            }
        })
    return jsonify(result)

# ====== CLIENTES, ALERTAS ======
@app.route("/api/clientes")
def clientes():
    dia_param = request.args.get("dia", "").strip()
    if dia_param:
        df = _clientes_por_dia(dia_param)
        if df is None or df.empty:
            return jsonify([])
        records = df.to_dict(orient="records")
        for rec in records:
            for k, v in rec.items():
                if isinstance(v, float) and not math.isfinite(v):
                    rec[k] = None
        return jsonify(records)
    df = read_csv(DATASETS / "clientes_dia.csv")
    if df.empty: return jsonify([])
    df.columns = [c.lstrip("﻿") for c in df.columns]
    for col in ["vendedor_codigo", "cliente_id", "botellas_ayer", "botellas_mes",
                "importe_ayer", "importe_mes", "compra_ayer_flag", "compra_mes_flag",
                "ccc_ayer_flag", "ccc_mes_flag", "cobertura_ayer_flag", "cobertura_mes_flag"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df["vendedor_id"] = "V" + df["vendedor_codigo"].astype("Int64").astype(str)
    df["segmento"] = df["segmento_operativo"]
    df["estado"] = df["estado_cliente"]
    df["prioridad_label"] = df["prioridad_comercial"]
    df["impacto_alertas_ars"] = df["importe_mes"].fillna(0)
    df["faltan_11t"] = 11
    df["kernel_accion"] = ""
    # V3 (Nadia) solo Tradicional (no AS / On Premise / Mayorista)
    df = df[~((df["vendedor_id"] == "V3") & (df["segmento"].astype(str).str.upper() != "TRADICIONAL"))]

    # Enriquecer con última compra desde historial
    try:
        hist = read_csv(BASE / "02_HISTORY" / "historial_ventas_cliente.csv")
        if not hist.empty:
            hist.columns = [c.lstrip("﻿") for c in hist.columns]
            fecha_col = next((c for c in hist.columns if "fecha" in c.lower()), None)
            imp_col = next((c for c in hist.columns if "importe" in c.lower() or "neto" in c.lower()), None)
            id_col = next((c for c in hist.columns if c.lower() in ("cliente", "cliente_id", "cod_cliente")), None)
            if all([fecha_col, imp_col, id_col]):
                hist[id_col] = pd.to_numeric(hist[id_col], errors="coerce")
                hist[imp_col] = pd.to_numeric(hist[imp_col], errors="coerce")
                ultima = (hist.sort_values(fecha_col, ascending=False)
                          .drop_duplicates(subset=[id_col])[[id_col, fecha_col, imp_col]]
                          .rename(columns={id_col: "cliente_id", fecha_col: "ultima_compra_fecha", imp_col: "ultima_compra_importe"}))
                ultima["cliente_id"] = pd.to_numeric(ultima["cliente_id"], errors="coerce")
                df["cliente_id"] = pd.to_numeric(df["cliente_id"], errors="coerce")
                df = df.merge(ultima, on="cliente_id", how="left")
    except Exception:
        pass

    records = df.to_dict(orient="records")
    for rec in records:
        for k, v in rec.items():
            if isinstance(v, float) and not math.isfinite(v):
                rec[k] = None
    return jsonify(records)

_CLIENTES_MAESTRO_CACHE = {}

def _clientes_maestro(incluir_deposito=False):
    """Cartera real desde clientes.xlsx, sin vendedores excluidos.

    En el maestro el Depósito (V20 en ventas) figura como codven=1 y por defecto
    queda excluido como el resto de las métricas con objetivo. incluir_deposito=True
    lo conserva SOLO para la pantalla de Clientes de gerencia (buscar/ficha); no
    afecta cobertura, planes AS ni ninguna métrica con objetivo."""
    p = INPUTS / "clientes.xlsx"
    if not p.exists():
        return pd.DataFrame()
    try:
        key = os.path.getmtime(p)
    except OSError:
        key = 0
    df = _CLIENTES_MAESTRO_CACHE.get(key)
    if df is None:
        try:
            df = pd.read_excel(p)
        except Exception:
            return pd.DataFrame()
        df.columns = [str(c).strip() for c in df.columns]
        if "Codigo" not in df.columns:
            return pd.DataFrame()
        df["_cliente_id"] = pd.to_numeric(df["Codigo"], errors="coerce")
        df["_vend"] = pd.to_numeric(df.get("codven"), errors="coerce")
        df = df.dropna(subset=["_cliente_id"]).copy()
        df["_cliente_id"] = df["_cliente_id"].astype(int)
        # Un cliente = una fila. Si el ERP vuelve a exportar un cliente en dos rutas,
        # sin esto queda en las DOS carteras e infla el denominador de todas las
        # tarjetas en silencio (pasó con 10 clientes en V3/V8, corregido 2026-07-30).
        _dups = df[df.duplicated(subset=["_cliente_id"], keep=False)]
        if not _dups.empty:
            print(f"[AVISO] clientes.xlsx: {_dups['_cliente_id'].nunique()} cliente(s) duplicado(s) "
                  f"{sorted(_dups['_cliente_id'].unique().tolist())} — se deja la primera fila. "
                  f"Corregir la cartera en el ERP.")
            df = df.drop_duplicates(subset=["_cliente_id"], keep="first").copy()
        df["_vend_id"] = df["_vend"].apply(lambda x: f"V{int(x)}" if pd.notna(x) else "")
        _CLIENTES_MAESTRO_CACHE.clear()
        _CLIENTES_MAESTRO_CACHE[key] = df
    # Depósito = codven 1; el resto de excluidos (2/5/20) nunca se muestran.
    excluir = _VENDEDORES_EXCLUIDOS - {1} if incluir_deposito else _VENDEDORES_EXCLUIDOS
    return df[~df["_vend"].isin(excluir)].copy()


_CLIENTE_VENTAS_CACHE = {}

def _cliente_ventas_base():
    """Ventas vivas disponibles para ficha cliente. Litros por la misma cascada que Sell Out
    (_litros_por_linea: maestro 04D → PesoKg → nombre del artículo).

    incluir_deposito=True: la pantalla de Clientes de gerencia muestra la cartera CON
    depósito (_clientes_maestro(incluir_deposito=True)); si la base de ventas excluía V20,
    los clientes que sólo compran por venta directa (ej. #786 ANSELMI) salían como
    'Sin compras en el mes vigente' teniendo venta real. La ficha es informativa (sin
    objetivo), así que V20 suma acá; el ficha de vendedor lo vuelve a excluir."""
    paths = [INPUTS / "ventas_acumulada.csv", INPUTS / "ventas.csv"]
    key = tuple((str(p), os.path.getmtime(p) if p.exists() else 0) for p in paths)
    df = _CLIENTE_VENTAS_CACHE.get(key)
    if df is not None:
        return df
    frames = []
    for p in paths:
        if p.exists():
            v = _preparar_df_ventas(p, incluir_deposito=True)
            if not v.empty:
                frames.append(v)
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    df.columns = [str(c).strip() for c in df.columns]
    df["_cli"] = pd.to_numeric(df.get("Cliente"), errors="coerce")
    df["_vend"] = pd.to_numeric(df.get("CodVendedor"), errors="coerce")
    df["_fecha"] = pd.to_datetime(df.get("FechaComprobante"), dayfirst=True, errors="coerce")
    df["_litros"] = _litros_por_linea(df)
    df["_importe"] = pd.to_numeric(df.get("ImporteNetoItem"), errors="coerce").fillna(0)
    # astype(str) sobre una celda vacía del ERP deja el literal "nan": sin esto la ficha
    # mostraba una marca llamada "nan" (Cinzano/Dada Sweet vienen sin Marca en ventas.csv).
    def _txt(col):
        s = df.get(col, pd.Series([""] * len(df), index=df.index)).astype(str).str.strip()
        return s.where(~s.str.lower().isin(["nan", "none", "nat"]), "")
    df["_marca"] = _txt("Marca")
    df["_linea"] = _txt("Linea")
    df["_articulo"] = _txt("Articulo")
    df["_codigo"] = _txt("Codigo")
    df["_botellas"] = pd.to_numeric(df.get("CantBase"), errors="coerce").fillna(0)
    df = df.dropna(subset=["_cli", "_vend", "_fecha"])
    if "NroComprobante" in df.columns and "Codigo" in df.columns:
        df = df.drop_duplicates(subset=["NroComprobante", "Cliente", "Codigo", "CantBase", "ImporteNetoItem"])
    else:
        df = df.drop_duplicates()
    df["_cli"] = df["_cli"].astype(int)
    df["_vend"] = df["_vend"].astype(int)
    _CLIENTE_VENTAS_CACHE.clear()
    _CLIENTE_VENTAS_CACHE[key] = df
    return df


def _cliente_row_to_dict(row):
    vend = int(row.get("_vend")) if pd.notna(row.get("_vend")) else None
    return {
        "cliente_id": int(row.get("_cliente_id")),
        "nombre": str(row.get("Razon_Social", row.get("_cliente_id", ""))).strip(),
        "direccion": str(row.get("Direccion", "")).strip(),
        "localidad": str(row.get("Localidad", "")).strip(),
        "telefono": "" if pd.isna(row.get("Telefono", "")) else str(row.get("Telefono", "")).strip(),
        "vendedor_codigo": vend,
        "vendedor_id": f"V{vend}" if vend is not None else "",
        "vendedor_nombre": str(row.get("Vendedor", "")).strip(),
        "dia_visita": str(row.get("DiasVisita", "")).strip(),
        "frecuencia_visita": str(row.get("Frecuencia", "")).strip(),
        "subcanal": str(row.get("SubSegmento", row.get("Ramo", ""))).strip(),
        "ramo": str(row.get("Ramo", "")).strip(),
    }


@app.route("/api/clientes/buscar")
def clientes_buscar():
    q = (request.args.get("q") or "").strip().upper()
    vend = normalizar_vendedor_codigo(request.args.get("vendedor") or "")
    try:
        limit = min(max(int(request.args.get("limit", 20)), 1), 50)
    except Exception:
        limit = 20
    df = _clientes_maestro(incluir_deposito=True)
    if df.empty:
        return jsonify([])
    if vend:
        df = df[df["_vend_id"] == vend]
    if q:
        nom = df.get("Razon_Social", pd.Series([""] * len(df), index=df.index)).astype(str).str.upper()
        cid = df["_cliente_id"].astype(str)
        loc = df.get("Localidad", pd.Series([""] * len(df), index=df.index)).astype(str).str.upper()
        df = df[nom.str.contains(q, na=False) | cid.str.contains(q, na=False) | loc.str.contains(q, na=False)]
    df = df.sort_values(["Razon_Social", "_cliente_id"], na_position="last").head(limit)
    return jsonify(_to_native([_cliente_row_to_dict(r) for _, r in df.iterrows()]))


@app.route("/api/clientes/<int:cliente_id>/ficha")
def cliente_ficha(cliente_id):
    vend = normalizar_vendedor_codigo(request.args.get("vendedor") or "")
    cli = _clientes_maestro(incluir_deposito=True)
    if cli.empty:
        return jsonify({"error": "clientes.xlsx no disponible"}), 404
    row_df = cli[cli["_cliente_id"] == int(cliente_id)]
    if row_df.empty:
        return jsonify({"error": "cliente no encontrado"}), 404
    row = row_df.iloc[0]
    base = _cliente_row_to_dict(row)
    if vend and base.get("vendedor_id") != vend:
        return jsonify({"error": "cliente fuera de la cartera del vendedor"}), 403

    ventas = _cliente_ventas_base()
    if not ventas.empty and vend:
        # Ficha pedida desde el perfil de un vendedor: V20 (Depósito / venta directa) no
        # entra en lo que ve el vendedor. En gerencia (sin ?vendedor=) sí se muestra.
        ventas = ventas[ventas["_vend"] != 20]
    vc = ventas[ventas["_cli"] == int(cliente_id)].copy() if not ventas.empty else pd.DataFrame()
    if vc.empty:
        base.update({
            "marcas_mes": [],
            "productos_mes": [],
            "ventas_mensuales": [],
            "frecuencia_compra_mensual": 0,
            "fecha_ultima_compra": None,
            "promedio_12m": {"litros": 0, "importe": 0, "meses_con_datos": 0, "meses_ventana": 12},
            "posibilidad_venta": {"litros": 0, "importe": 0, "criterio": "sin ventas disponibles"},
            "fuente": "clientes.xlsx + ventas_acumulada.csv/ventas.csv",
        })
        return jsonify(_to_native(base))

    latest = vc["_fecha"].max()
    periodo_actual = latest.to_period("M")
    inicio_12 = (periodo_actual - 11).to_timestamp()
    vc12 = vc[vc["_fecha"] >= inicio_12].copy()
    vc12["_periodo"] = vc12["_fecha"].dt.to_period("M").astype(str)
    mes_actual = vc12[vc12["_fecha"].dt.to_period("M") == periodo_actual]

    marca_col = "_marca"
    marcas = []
    productos = []
    if not mes_actual.empty:
        mm = (mes_actual.groupby(marca_col, dropna=False)
              .agg(litros=("_litros", "sum"), importe=("_importe", "sum"),
                   botellas=("_botellas", "sum"))
              .reset_index())
        for _, r in mm.sort_values("litros", ascending=False).iterrows():
            nombre = str(r[marca_col]).strip() or "Sin marca"
            marcas.append({
                "marca": nombre,
                "litros": round(float(r["litros"]), 1),
                "importe": round(float(r["importe"]), 0),
                "botellas": round(float(r["botellas"]), 1),
            })
        # Detalle por producto (SKU) del mes: qué compró exactamente dentro de cada marca
        # (ej. marca "Dada" → "DADA LATA TINTO VERANO 4X6X355").
        pm = (mes_actual.groupby(["_marca", "_codigo", "_articulo"], dropna=False)
              .agg(litros=("_litros", "sum"), importe=("_importe", "sum"),
                   botellas=("_botellas", "sum"), compras=("_fecha", lambda s: int(s.dt.date.nunique())))
              .reset_index())
        for _, r in pm.sort_values("importe", ascending=False).iterrows():
            productos.append({
                "marca": str(r["_marca"]).strip() or "Sin marca",
                "codigo": str(r["_codigo"]).strip(),
                "producto": str(r["_articulo"]).strip() or "Sin descripción",
                "botellas": round(float(r["botellas"]), 1),
                "litros": round(float(r["litros"]), 1),
                "importe": round(float(r["importe"]), 0),
                "compras": int(r["compras"]),
            })

    mensual = (vc12.groupby("_periodo")
               .agg(litros=("_litros", "sum"), importe=("_importe", "sum"),
                    compras=("_fecha", lambda s: int(s.dt.date.nunique())))
               .reset_index()
               .sort_values("_periodo"))
    ventas_mensuales = []
    prev_litros = None
    for _, r in mensual.iterrows():
        litros = round(float(r["litros"]), 1)
        if prev_litros is None:
            tendencia = "neutro"
        elif litros > prev_litros:
            tendencia = "verde"
        elif litros < prev_litros:
            tendencia = "rojo"
        else:
            tendencia = "neutro"
        ventas_mensuales.append({
            "periodo": str(r["_periodo"]),
            "litros": litros,
            "importe": round(float(r["importe"]), 0),
            "compras": int(r["compras"]),
            "tendencia_litros": tendencia,
        })
        prev_litros = litros

    meses_con_datos = int(len(mensual))
    promedio_litros = float(mensual["litros"].mean()) if meses_con_datos else 0
    promedio_importe = float(mensual["importe"].mean()) if meses_con_datos else 0
    litros_mes = float(mes_actual["_litros"].sum()) if not mes_actual.empty else 0
    importe_mes = float(mes_actual["_importe"].sum()) if not mes_actual.empty else 0
    frecuencia = float(mensual["compras"].mean()) if meses_con_datos else 0

    base.update({
        "periodo_mes": str(periodo_actual),
        "marcas_mes": marcas,
        "productos_mes": productos,
        "ventas_mensuales": ventas_mensuales,
        "frecuencia_compra_mensual": round(frecuencia, 1),
        "fecha_ultima_compra": latest.strftime("%Y-%m-%d") if pd.notna(latest) else None,
        "venta_mes": {"litros": round(litros_mes, 1), "importe": round(importe_mes, 0)},
        "promedio_12m": {
            "litros": round(promedio_litros, 1),
            "importe": round(promedio_importe, 0),
            "meses_con_datos": meses_con_datos,
            "meses_ventana": 12,
        },
        "posibilidad_venta": {
            "litros": round(max(0, promedio_litros - litros_mes), 1),
            "importe": round(max(0, promedio_importe - importe_mes), 0),
            "criterio": "promedio mensual de los meses disponibles dentro de la ventana de 12 meses",
        },
        "fuente": "clientes.xlsx + ventas_acumulada.csv/ventas.csv",
    })
    return jsonify(_to_native(base))

def _v3_clientes_tradicional():
    """Set de cliente_id de V3 que son Tradicional almacén/despensa/kiosco (su único canal).
    Devuelve None si no hay maestro (no filtrar)."""
    m = _clientes_maestro()
    if m is None or m.empty:
        return None
    sub_col = next((c for c in m.columns if "subseg" in c.lower() or "subramo" in c.lower()), None)
    v3 = m[m["_vend"] == 3]
    out = set()
    for _, r in v3.iterrows():
        seg = _clasificar_segmento(str(r.get("Ramo", "")), str(r.get(sub_col, "") if sub_col else ""))
        sub = str(r.get(sub_col, "")).upper() if sub_col else ""
        if seg == "TRADICIONAL" and any(k in sub for k in ("ALMACEN", "DESPENSA", "KIOSCO")):
            out.add(int(r["_cliente_id"]))
    return out


def _alerta_clave(a):
    """Clave estable de una alerta, para poder descartarla.
    Incluye el mes (YYYY-MM) para que el descarte NO se herede al mes siguiente, y la fecha
    del pedido: si el mismo cliente vuelve a excederse otro día, es una alerta nueva y se
    muestra igual.
    En 'descuento' se agrega la magnitud (% aplicado, neto, cantidad) porque un mismo
    artículo puede tener DOS líneas el mismo día con importes distintos: son infracciones
    distintas y descartar una no debe tapar la otra. Líneas idénticas comparten clave, que
    es lo correcto: son indistinguibles.
    En 'tope' NO se incluye la magnitud a propósito: es un acumulado del mes y si sumara
    cajas volvería a aparecer todos los días — justo lo que se quiere evitar. Una por
    cliente + acción + mes."""
    mes = pd.Timestamp.today().strftime("%Y-%m")
    partes = [mes, str(a.get("tipo") or ""), str(a.get("vendedor_id") or ""),
              str(a.get("cliente_id") or ""), str(a.get("articulo") or ""),
              str(a.get("fecha_pedido") or "")]
    if str(a.get("tipo") or "") == "descuento":
        partes.append(f"{a.get('descuento_aplicado_pct')}/{a.get('importe_neto')}/{a.get('cant_base')}")
    return "|".join(partes)


def _alertas_descartadas():
    """Set de claves descartadas. Si la tabla no existe todavía, no filtra nada."""
    try:
        conn = sqlite3.connect(str(DB_PATH), timeout=10, check_same_thread=False)
        try:
            return {r[0] for r in conn.execute("SELECT clave FROM alerta_descartada")}
        finally:
            conn.close()
    except Exception as e:
        print(f"[WARN] _alertas_descartadas: {e}")
        return set()


@app.route("/api/alertas")
def alertas():
    # Alertas en vivo desde el catálogo de acciones del mes (acciones_comerciales_<mes>_penaflor.csv):
    #  - descuento: % aplicado supera el tramo máximo de la acción (Plan AS / 11T contemplados).
    #  - tope: cajas/mes por cliente superan el tope mensual de la acción (combinable entre marcas).
    data = _alertas_descuento_mes() + _alertas_tope_cajas_mes()
    # V3 (Nadia) solo Tradicional almacén/despensa/kiosco: descartar alertas de clientes de otros canales
    v3_ok = _v3_clientes_tradicional()
    if v3_ok is not None:
        def _keep(a):
            es_v3 = str(a.get("vendedor_id") or "").upper() == "V3" or str(a.get("vendedor_codigo")) == "3"
            if not es_v3:
                return True
            try:
                return int(a.get("cliente_id")) in v3_ok
            except (TypeError, ValueError):
                return True
        data = [a for a in data if _keep(a)]
    # Descartadas por gerencia: se ocultan para todos (gerencia y vendedor consumen esta ruta).
    descartadas = _alertas_descartadas()
    out = []
    for a in data:
        a["clave_descarte"] = _alerta_clave(a)
        if a["clave_descarte"] not in descartadas:
            out.append(a)
    return jsonify(out)


@app.route("/api/alertas/descartar", methods=["POST"])
def alertas_descartar():
    """Descarta alertas por clave (`clave_descarte` de /api/alertas) para que dejen de
    mostrarse. No borra ninguna venta: sólo registra que ya fueron revisadas.
    Body: {"claves": [...], "autor": "Gerencia", "resumenes": {clave: texto}}"""
    d = request.get_json(silent=True) or {}
    claves = [str(k).strip() for k in (d.get("claves") or []) if str(k).strip()]
    if not claves:
        return jsonify({"error": "falta 'claves'"}), 400
    autor = str(d.get("autor", "Gerencia")).strip() or "Gerencia"
    resumenes = d.get("resumenes") or {}
    ts = _now_ar()
    conn = sqlite3.connect(str(DB_PATH), timeout=10, check_same_thread=False)
    try:
        conn.executemany(
            """INSERT INTO alerta_descartada(clave, autor, resumen, descartada_at)
               VALUES(?,?,?,?)
               ON CONFLICT(clave) DO UPDATE SET autor=excluded.autor,
                   resumen=excluded.resumen, descartada_at=excluded.descartada_at""",
            [(k, autor, str(resumenes.get(k, ""))[:300], ts) for k in claves])
        conn.commit()
        total = conn.execute("SELECT COUNT(*) FROM alerta_descartada").fetchone()[0]
    finally:
        conn.close()
    return jsonify({"ok": True, "descartadas": len(claves), "total_descartadas": total,
                    "autor": autor, "descartada_at": ts})


# ====== SEGUIMIENTO GERENCIAL DE ALERTAS ======
@app.route("/api/alertas/seguimiento", methods=["GET", "POST"])
def alertas_seguimiento():
    """Nota gerencial por alerta (si fue vista y hablada con el vendedor).
    Clave estable = 'vendedor_id|cliente_id|articulo'. Persiste en orbit.db (disco persistente)."""
    conn = sqlite3.connect(str(DB_PATH), timeout=10, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        if request.method == "POST":
            d = request.get_json(silent=True) or {}
            clave = str(d.get("clave", "")).strip()
            if not clave:
                return jsonify({"error": "falta 'clave'"}), 400
            mensaje = str(d.get("mensaje", "")).strip()
            autor = str(d.get("autor", "Gerencia")).strip() or "Gerencia"
            ts = _now_ar()
            if mensaje:
                conn.execute(
                    """INSERT INTO alerta_seguimiento(clave, mensaje, autor, updated_at)
                       VALUES(?,?,?,?)
                       ON CONFLICT(clave) DO UPDATE SET mensaje=excluded.mensaje,
                           autor=excluded.autor, updated_at=excluded.updated_at""",
                    (clave, mensaje, autor, ts))
            else:
                conn.execute("DELETE FROM alerta_seguimiento WHERE clave=?", (clave,))
            conn.commit()
            return jsonify({"ok": True, "clave": clave, "mensaje": mensaje,
                            "autor": autor, "updated_at": ts})
        # GET: devuelve todas las notas {clave: {mensaje, autor, updated_at}}
        rows = conn.execute("SELECT clave, mensaje, autor, updated_at FROM alerta_seguimiento").fetchall()
        return jsonify({r["clave"]: {"mensaje": r["mensaje"], "autor": r["autor"],
                                     "updated_at": r["updated_at"]} for r in rows})
    finally:
        conn.close()


# ====== DETALLE VENDEDOR ======
@app.route("/api/vendedor/<vid>")
def vendedor_detalle(vid):
    vid_norm = normalizar_vendedor_codigo(vid)
    cn = clean_code(vid_norm)

    vend = read_csv(CONFIG / "vendedores_activos.csv")
    nombre = vid_norm
    if not vend.empty:
        activos = vend[vend["activo"] == 1]
        fila = activos[activos["codigo_vendedor"].astype(str).apply(clean_code) == cn]
        if fila.empty:
            return jsonify({"error": f"Vendedor {vid} no encontrado o inactivo"}), 404
        nombre = str(fila.iloc[0]["nombre_vendedor"])

    # Regla de negocio: V3 no trabaja autoservicio ni on premise
    trabaja_as = (vid_norm != "V3")
    trabaja_op = (vid_norm != "V3")

    # Clientes / venta del día — desde el motor (mod_volumen_vendedor.csv)
    vol = read_csv(DATASETS / "mod_volumen_vendedor.csv")
    vv = vol[vol["vendedor_codigo"].astype(str).apply(clean_code) == cn] if not vol.empty else pd.DataFrame()
    venta_hoy = float(vv["venta_ayer"].sum())        if not vv.empty else 0
    cli_total = int(vv["clientes_planificados"].sum()) if not vv.empty and "clientes_planificados" in vv.columns else 0
    cli_sin   = int(vv["clientes_sin_compra_mes"].sum()) if not vv.empty and "clientes_sin_compra_mes" in vv.columns else 0

    # objetivo/acumulado/avance — fuente primaria resultado.xlsx (igual que /api/dashboard);
    # se actualiza a diario sin regenerar el motor. Fallback a mod_volumen_vendedor.csv.
    obj = acum = av = 0.0
    _usa_resultado = False
    _resultado_path = INPUTS / "resultado.xlsx"
    if _resultado_path.exists():
        try:
            _adf = pd.read_excel(_resultado_path, sheet_name="Avance")
            _row = _adf[_adf["VendedorCodigo"].astype(str).apply(clean_code) == cn]
            if not _row.empty:
                _r = _row.iloc[0]
                obj  = float(_r.get("ValorObjetivo", 0) or 0)
                acum = float(_r.get("Acumulado", 0) or 0)
                av   = float(_r.get("Avance", 0) or 0)   # Avance = Tendencia/Objetivo*100 (regla Peñaflor)
                _usa_resultado = True
        except Exception:
            pass
    if not _usa_resultado:
        obj  = float(vv["objetivo_mes"].sum())  if not vv.empty else 0
        acum = float(vv["acumulado_mes"].sum()) if not vv.empty else 0
        av   = float(vv["avance_pct"].mean())   if not vv.empty else 0

    # tendencia_pct = Avance de resultado.xlsx (Tendencia/Objetivo); sin recálculo por días para
    # no divergir del dashboard. Fallback: proyección por días hábiles si no hay resultado.xlsx.
    if _usa_resultado:
        tendencia_pct_vd = round(av, 2)
    else:
        dias_vd          = contar_dias_habiles()
        tendencia_pct_vd = round((acum / max(dias_vd["corridos"], 1)) * dias_vd["total"] / obj * 100, 2) if obj else 0

    # CCC Compradores Mes — desde ventas.csv del mes actual
    ventas_mes_vd = _cargar_ventas_mes_actual()
    ccc_mes_map_vd = _ccc_mes_por_vendedor(ventas_mes_vd)
    cod_int = int(cn) if cn.isdigit() else 0
    ccc_mes_vd = ccc_mes_map_vd.get(cod_int, {"tradicional": 0, "autoservicio": 0, "onpremise": 0,
                                              "proximity": 0, "total": 0})
    ccc_trad = ccc_mes_vd["tradicional"]
    ccc_as   = ccc_mes_vd["autoservicio"]   # ya es 0 para V3
    ccc_op   = ccc_mes_vd["onpremise"]
    ccc_prox = ccc_mes_vd.get("proximity", 0)   # estaciones de servicio: V3 sí lo trabaja

    # CCC Compradores Día — desde mod_ccc_segmento (ayer)
    ccc_df = read_csv(DATASETS / "mod_ccc_segmento.csv")
    cv = ccc_df[ccc_df["vendedor_codigo"].astype(str).apply(clean_code) == cn] if not ccc_df.empty else pd.DataFrame()

    def _ccc_dia(df, pat):
        if df.empty:
            return 0
        return int(df.loc[df["segmento_operativo"].astype(str).str.upper().str.contains(pat, na=False), "clientes_con_compra"].sum())

    ccc_dia_trad = _ccc_dia(cv, "TRADICIONAL")
    ccc_dia_as   = _ccc_dia(cv, "AUTOSERVICIO")
    ccc_dia_op   = _ccc_dia(cv, "ON_PREMISE|VTK")
    ccc_dia_prox = _ccc_dia(cv, "PROXIMITY")

    # V3 no trabaja autoservicio ni on premise (ccc_as/ccc_op ya son 0; refuerzo ccc_dia)
    if vid_norm == "V3":
        ccc_dia_as = 0
        ccc_dia_op = 0

    # 11 Titulares por vendedor — cantidad de clientes a los que logró vender cada marca:
    # cubiertos_dia = clientes de la ZONA DEL DÍA; cubiertos = total de todas las zonas.
    # Fuente: mod_11t_acum.csv (tiene_flag poblado por cliente). mod_11_titulares.csv viene en 0.
    dia_req = request.args.get("dia", "").strip()
    if not dia_req:
        _DIAS_AR = {0: "LU", 1: "MA", 2: "MI", 3: "JU", 4: "VI", 5: "SA", 6: "DO"}
        dia_req = _DIAS_AR[datetime.now(_ARG_TZ).weekday()]
    dia_ids = set()
    try:
        cd = _clientes_por_dia(dia_req)
        if not cd.empty and "cliente_id" in cd.columns:
            dia_ids = set(pd.to_numeric(cd["cliente_id"], errors="coerce").dropna().astype(int))
    except Exception:
        pass

    titulares11 = []
    t11a = read_csv(DATASETS / "mod_11t_acum.csv")
    if not t11a.empty and "marca_objetivo" in t11a.columns and "tiene_flag" in t11a.columns:
        tv = t11a[t11a["vendedor_codigo"].astype(str).apply(clean_code) == cn].copy()
        if not tv.empty:
            tv["tiene_flag"] = pd.to_numeric(tv["tiene_flag"], errors="coerce").fillna(0)
            tv["cliente_id"] = pd.to_numeric(tv["cliente_id"], errors="coerce")
            for marca, grp in tv.groupby("marca_objetivo", dropna=False):
                cli_si = set(grp[grp["tiene_flag"] == 1]["cliente_id"].dropna().astype(int))
                titulares11.append({"marca": marca,
                                    "cubiertos_dia": len(cli_si & dia_ids),
                                    "cubiertos": len(cli_si),
                                    "dia_zona": dia_req})
            titulares11.sort(key=lambda x: -x["cubiertos"])

    once_t_cumplidos = sum(1 for t in titulares11 if t["cubiertos"] > 0)
    once_t_total     = len(titulares11)

    return jsonify({
        "vendedor_id":       vid_norm,
        "vendedor_nombre":   nombre,
        "trabaja_autoservicio": trabaja_as,
        "trabaja_onpremise": trabaja_op,
        "objetivo":          obj,
        "acumulado":         acum,
        "avance_pct":        round(av, 2),
        "tendencia_pct":    tendencia_pct_vd,
        "venta_hoy":         venta_hoy,
        "clientes_total":    cli_total,
        "clientes_pendientes": cli_sin,
        # CCC Compradores Mes — fuente: ventas.csv mes actual, ImporteNetoItem > 0
        "ccc_tradicional":   ccc_trad,
        "ccc_autoservicio":  ccc_as,
        "ccc_onpremise":     ccc_op,
        "ccc_proximity":     ccc_prox,
        "ccc_total":         ccc_trad + ccc_as + ccc_op + ccc_prox,
        # CCC Compradores Día — fuente: mod_ccc_segmento (ayer)
        "ccc_dia_tradicional": ccc_dia_trad,
        "ccc_dia_autoservicio": ccc_dia_as,
        "ccc_dia_onpremise": ccc_dia_op,
        "ccc_dia_proximity": ccc_dia_prox,
        "ccc_dia_total":     ccc_dia_trad + ccc_dia_as + ccc_dia_op + ccc_dia_prox,
        "once_t_cumplidos":  once_t_cumplidos,
        "once_t_total":      once_t_total,
        "titulares11":       titulares11,
        "modo_datos":        "REAL",
        "generado_en": _now_ar(),
    })

@app.route("/api/planificacion", methods=["GET","POST"])
def planificacion():
    fecha_q = request.args.get("fecha")
    vid_q   = request.args.get("vendedor_id")
    limit_q = (request.args.get("limit") or "").strip().lower()
    conn = sqlite3.connect(str(DB_PATH), timeout=10, check_same_thread=False)
    conn.row_factory = sqlite3.Row

    if request.method == "POST":
        d = request.get_json() or {}
        _ip = request.remote_addr or "?"
        _ts = _now_ar()
        _log_line = f"{_ts} | IP={_ip} | payload={d}\n"
        try:
            with open(str(BASE / "99_LOGS_ORBIT" / "planificacion_post.log"), "a", encoding="utf-8") as _lf:
                _lf.write(_log_line)
        except Exception:
            pass
        print(f"[PLAN POST] {_log_line.strip()}")

        # Normalizar y validar vendedor_id
        vid_raw = normalizar_vendedor_codigo(d.get("vendedor_id",""))
        if vid_raw not in _VENDEDORES_ACTIVOS_PLAN:
            conn.close()
            return jsonify({"ok": False, "error": f"Vendedor {vid_raw} no autorizado"}), 400

        # Regla: V3 no trabaja autoservicio ni on premise
        ccc_as = 0 if vid_raw == "V3" else int(d.get("ccc_autoservicio") or 0)
        ccc_op = 0 if vid_raw == "V3" else int(d.get("ccc_onpremise") or 0)
        fecha_raw = str(d.get("fecha") or "").strip().lower()
        fecha  = _fecha_planificacion_default() if fecha_raw in ("", "auto", "default") else d.get("fecha")
        _ts    = _now_ar()  # hora Argentina para ambos timestamps

        # FAIL-CLOSED: Google Sheets es la fuente de verdad. Si no guarda y verifica,
        # no se toca SQLite y se devuelve 503. SQLite queda solo como cache.
        plan = {
            "fecha": fecha, "vendedor_id": vid_raw,
            "zona": d.get("zona"), "dia_visita": d.get("dia_visita"),
            "venta_esperada": float(d.get("venta_esperada") or 0),
            "ccc_tradicional": int(d.get("ccc_tradicional") or 0),
            "ccc_autoservicio": ccc_as,
            "ccc_onpremise": ccc_op,
            "once_t": int(d.get("once_t") or 0),
            "marcas": d.get("marcas"), "clientes_clave": d.get("clientes_clave"),
            "acciones": d.get("acciones"), "estado": "enviada",
            "created_at": _ts, "updated_at": _ts,
        }
        ok_w, gerr = gsheets_upsert_plan(plan)
        if not ok_w:
            conn.close()
            return jsonify({"ok": False,
                            "error": f"No se pudo guardar en Google Sheets: {gerr}"}), 503
        if not gsheets_verify_plan(_plan_id(fecha, vid_raw), {"updated_at": _ts}):
            conn.close()
            return jsonify({"ok": False,
                            "error": "Guardado no verificado en Google Sheets"}), 503

        conn.execute("""
            INSERT INTO planificacion
              (fecha, vendedor_id, zona, dia_visita, venta_esperada,
               ccc_tradicional, ccc_autoservicio, ccc_onpremise,
               once_t, marcas, clientes_clave, acciones, estado, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,'enviada',?,?)
            ON CONFLICT(fecha, vendedor_id) DO UPDATE SET
              zona=excluded.zona, dia_visita=excluded.dia_visita,
              venta_esperada=excluded.venta_esperada,
              ccc_tradicional=excluded.ccc_tradicional,
              ccc_autoservicio=excluded.ccc_autoservicio,
              ccc_onpremise=excluded.ccc_onpremise,
              once_t=excluded.once_t, marcas=excluded.marcas,
              clientes_clave=excluded.clientes_clave, acciones=excluded.acciones,
              estado='enviada', updated_at=excluded.updated_at""",
            (fecha, vid_raw, d.get("zona"), d.get("dia_visita"),
             float(d.get("venta_esperada") or 0),
             int(d.get("ccc_tradicional") or 0), ccc_as,
             ccc_op, int(d.get("once_t") or 0),
             d.get("marcas"), d.get("clientes_clave"), d.get("acciones"),
             _ts, _ts))
        conn.commit(); conn.close()
        export_planificacion_csv()
        return jsonify({"ok": True, "vendedor_id": vid_raw, "fecha": fecha, "hora_envio": _ts})

    # GET — lee SQLite (cache) si tiene datos; si está vacío, hidrata desde Google Sheets.
    total = conn.execute("SELECT COUNT(*) FROM planificacion").fetchone()[0]
    if total == 0 and gsheets_enabled():
        conn.close()
        hydrate_planificacion_from_sheets()
        conn = sqlite3.connect(str(DB_PATH), timeout=10, check_same_thread=False)
        conn.row_factory = sqlite3.Row

    # filtros opcionales por fecha y/o vendedor_id
    if fecha_q and vid_q:
        rows = conn.execute(
            "SELECT * FROM planificacion WHERE fecha=? AND vendedor_id=? ORDER BY updated_at DESC",
            (fecha_q, vid_q.upper())).fetchall()
    elif fecha_q:
        rows = conn.execute(
            "SELECT * FROM planificacion WHERE fecha=? ORDER BY vendedor_id",
            (fecha_q,)).fetchall()
    elif vid_q:
        rows = conn.execute(
            "SELECT * FROM planificacion WHERE vendedor_id=? ORDER BY fecha DESC LIMIT 30",
            (vid_q.upper(),)).fetchall()
    else:
        if limit_q == "all":
            rows = conn.execute(
                "SELECT * FROM planificacion ORDER BY fecha DESC, vendedor_id").fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM planificacion ORDER BY fecha DESC, vendedor_id LIMIT 500").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/planificacion/<int:plan_id>", methods=["PATCH"])
def planificacion_patch(plan_id):
    d = request.get_json() or {}
    _ESTADOS = {"enviada", "modificada", "aprobada"}

    estado = d.get("estado")
    if estado and estado not in _ESTADOS:
        return jsonify({"ok": False, "error": f"Estado '{estado}' inválido. Válidos: {sorted(_ESTADOS)}"}), 400

    conn = sqlite3.connect(str(DB_PATH), timeout=10, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM planificacion WHERE id=?", (plan_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": f"Plan {plan_id} no encontrado"}), 404

    row_d = dict(row)
    vid_row = (row_d.get("vendedor_id") or "").upper()
    fields, vals = [], []
    # row_merge: estado del plan tras aplicar el patch (payload para Google Sheets).
    row_merge = dict(row_d)

    for f in ["zona","dia_visita","venta_esperada","ccc_tradicional",
              "once_t","marcas","clientes_clave","acciones"]:
        if f in d:
            fields.append(f"{f}=?"); vals.append(d[f]); row_merge[f] = d[f]

    if "ccc_autoservicio" in d:
        val_as = 0 if vid_row == "V3" else int(d.get("ccc_autoservicio") or 0)
        fields.append("ccc_autoservicio=?"); vals.append(val_as); row_merge["ccc_autoservicio"] = val_as

    # V3 no trabaja on premise → su CCC On Premise planificado siempre 0
    if "ccc_onpremise" in d:
        val_op = 0 if vid_row == "V3" else int(d.get("ccc_onpremise") or 0)
        fields.append("ccc_onpremise=?"); vals.append(val_op); row_merge["ccc_onpremise"] = val_op

    if estado:
        fields.append("estado=?"); vals.append(estado); row_merge["estado"] = estado
    if "editado_por" in d:
        fields.append("editado_por=?"); vals.append(d["editado_por"]); row_merge["editado_por"] = d["editado_por"]
    if "comentario_gerencia" in d:
        fields.append("comentario_gerencia=?"); vals.append(d["comentario_gerencia"]); row_merge["comentario_gerencia"] = d["comentario_gerencia"]

    new_ts = _now_ar()
    fields.append("updated_at=?")
    vals.append(new_ts)
    row_merge["updated_at"] = new_ts

    # FAIL-CLOSED: escribir y verificar en Google Sheets antes de tocar SQLite (cache).
    pid = _plan_id(row_merge.get("fecha"), row_merge.get("vendedor_id"))
    ok_w, gerr = gsheets_upsert_plan(row_merge)
    if not ok_w:
        conn.close()
        return jsonify({"ok": False,
                        "error": f"No se pudo guardar en Google Sheets: {gerr}"}), 503
    if not gsheets_verify_plan(pid, {"updated_at": new_ts}):
        conn.close()
        return jsonify({"ok": False,
                        "error": "Cambio no verificado en Google Sheets"}), 503

    conn.execute(f"UPDATE planificacion SET {', '.join(fields)} WHERE id=?", vals + [plan_id])
    conn.commit()
    updated = dict(conn.execute("SELECT * FROM planificacion WHERE id=?", (plan_id,)).fetchone())
    conn.close()
    export_planificacion_csv()
    return jsonify({"ok": True, "plan": updated})

@app.route("/api/mensajes", methods=["GET","POST"])
def mensajes():
    vid = request.args.get("vendedor_id")
    conn = sqlite3.connect(str(DB_PATH)); conn.row_factory = sqlite3.Row
    if request.method == "POST":
        d = request.get_json()
        conn.execute("INSERT INTO mensajes(vendedor_id,mensaje,created_at) VALUES(?,?,?)", (d.get("vendedor_id"), d.get("mensaje"), _now_ar()))
        conn.commit(); conn.close()
        return jsonify({"ok":True})
    if vid:
        rows = conn.execute("SELECT * FROM mensajes WHERE vendedor_id=? ORDER BY created_at DESC", (vid,)).fetchall()
        conn.execute("UPDATE mensajes SET leido=1 WHERE vendedor_id=? AND leido=0", (vid,))
    else:
        rows = conn.execute("SELECT * FROM mensajes ORDER BY created_at DESC").fetchall()
    conn.commit(); conn.close()
    return jsonify([dict(r) for r in rows])

# ====== GASTOS POR ACCION COMERCIAL ======
@app.route("/api/gastos_accion")
def gastos_accion():
    df = read_csv(DATASETS / "mod_gastos_accion.csv")
    if df.empty:
        return jsonify({"modo_datos": "SIN_DATOS", "detalle": [], "resumen": {}, "top_acciones": [], "top_vendedores": []})

    for col in ["gasto_real_total", "gasto_teorico_total", "exceso_pesos_total", "exceso_pct_promedio",
                "clientes_afectados", "lineas_alertadas", "vendedor_codigo"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    resumen = {
        "filas_con_exceso":      int(len(df)),
        "gasto_real_total":      round(float(df["gasto_real_total"].sum()), 2),
        "gasto_teorico_total":   round(float(df["gasto_teorico_total"].sum()), 2),
        "exceso_pesos_total":    round(float(df["exceso_pesos_total"].sum()), 2),
        "exceso_pct_promedio":   round(float(df["exceso_pct_promedio"].mean()), 2),
        "vendedores_alertados":  int(df["vendedor_codigo"].nunique()),
        "clientes_afectados_total": int(df["clientes_afectados"].sum()),
        "acciones_csv":          int(df["es_regla_csv"].astype(str).str.upper().eq("TRUE").sum()),
        "acciones_fallback":     int(df["es_regla_csv"].astype(str).str.upper().ne("TRUE").sum()),
    }

    top_acc = (
        df.groupby(["accion_id", "canal", "categoria"], dropna=False)
        .agg(
            gasto_real_total   =("gasto_real_total",    "sum"),
            gasto_teorico_total=("gasto_teorico_total", "sum"),
            exceso_pesos_total =("exceso_pesos_total",  "sum"),
            clientes_afectados =("clientes_afectados",  "sum"),
            lineas_alertadas   =("lineas_alertadas",    "sum"),
            vendedores         =("vendedor_codigo",     "nunique"),
        )
        .reset_index()
        .sort_values("exceso_pesos_total", ascending=False)
        .head(5)
    )
    top_acciones = [
        {
            "accion_id":          str(r["accion_id"]),
            "canal":              str(r["canal"]),
            "categoria":          str(r["categoria"]),
            "gasto_real_total":   round(float(r["gasto_real_total"]), 2),
            "gasto_teorico_total":round(float(r["gasto_teorico_total"]), 2),
            "exceso_pesos_total": round(float(r["exceso_pesos_total"]), 2),
            "clientes_afectados": int(r["clientes_afectados"]),
            "lineas_alertadas":   int(r["lineas_alertadas"]),
            "vendedores":         int(r["vendedores"]),
        }
        for _, r in top_acc.iterrows()
    ]

    top_vend = (
        df.groupby(["vendedor_codigo", "vendedor_nombre"], dropna=False)
        .agg(
            gasto_real_total   =("gasto_real_total",   "sum"),
            gasto_teorico_total=("gasto_teorico_total","sum"),
            exceso_pesos_total =("exceso_pesos_total", "sum"),
            acciones_con_exceso=("accion_id",          "count"),
        )
        .reset_index()
        .sort_values("exceso_pesos_total", ascending=False)
        .head(5)
    )
    top_vendedores = [
        {
            "vendedor_codigo":    normalizar_vendedor_codigo(r["vendedor_codigo"]),
            "vendedor_nombre":    str(r["vendedor_nombre"]),
            "gasto_real_total":   round(float(r["gasto_real_total"]), 2),
            "gasto_teorico_total":round(float(r["gasto_teorico_total"]), 2),
            "exceso_pesos_total": round(float(r["exceso_pesos_total"]), 2),
            "acciones_con_exceso":int(r["acciones_con_exceso"]),
        }
        for _, r in top_vend.iterrows()
    ]

    detalle = df.replace({float("nan"): None}).to_dict(orient="records")

    return jsonify({
        "modo_datos":    "REAL",
        "generado_en": _now_ar(),
        "resumen":       resumen,
        "top_acciones":  top_acciones,
        "top_vendedores":top_vendedores,
        "detalle":       detalle,
    })

# ====== ORBIT DATA para el frontend ======
@app.route("/api/orbit-data")
def orbit_data():
    p = APP_DATA / "orbit_portal_data.json"
    if not p.exists():
        return jsonify({"error": "orbit_portal_data.json no encontrado. Ejecute tools/orbit_truth_audit.py primero."}), 503
    with open(p, "r", encoding="utf-8") as f:
        data = json.load(f)
    return jsonify(data)

# ====== MATINAL RESUMEN: plan vs real ======
def _real_dia_resultado(fecha_objetivo=None):
    """Real del día por vendedor = Acumulado(fecha) − Acumulado(día anterior del MISMO mes),
    desde 02_HISTORY/acumulado_resultado_historico.csv (snapshots diarios de resultado.xlsx).
    Es la forma correcta del 'real del día' (diferencia de acumulado), robusta a la FechaComprobante.
    Devuelve (dict {vendedor_codigo:int -> real_$}, fecha_hoy, fecha_ayer)."""
    p = BASE / "02_HISTORY" / "acumulado_resultado_historico.csv"
    if not p.exists():
        return {}, None, None
    df = read_csv(p)
    if df.empty or "fecha" not in df.columns:
        return {}, None, None
    df["fecha"] = df["fecha"].astype(str)
    df["vc"] = pd.to_numeric(df["vendedor_codigo"], errors="coerce")
    df["ac"] = pd.to_numeric(df["acumulado"], errors="coerce").fillna(0.0)
    df = df.dropna(subset=["vc"])
    fechas = sorted(df["fecha"].dropna().unique())
    if not fechas:
        return {}, None, None
    cand = [f for f in fechas if (not fecha_objetivo or f <= fecha_objetivo)]
    f_hoy = cand[-1] if cand else fechas[-1]
    prev = [f for f in fechas if f < f_hoy and f[:7] == f_hoy[:7]]   # mismo mes (evita borde de mes)
    f_ayer = prev[-1] if prev else None
    hoy  = dict(zip(df[df["fecha"] == f_hoy]["vc"].astype(int), df[df["fecha"] == f_hoy]["ac"]))
    ayer = (dict(zip(df[df["fecha"] == f_ayer]["vc"].astype(int), df[df["fecha"] == f_ayer]["ac"]))
            if f_ayer else {})
    real = {}
    for vc, a in hoy.items():
        d = float(a) - float(ayer.get(vc, 0.0))
        real[int(vc)] = round(d if d > 0 else 0.0, 2)   # negativo (reset) -> 0
    return real, f_hoy, f_ayer


def _snapshot_matinal_resumen():
    """Fallback de LECTURA para Plan vs Real cuando NO hay fuentes vivas (modo embebido en Orbit
    Home: sin SQLite poblado, sin 02_HISTORY, sin 01_INPUTS/ventas.csv). Arma el resumen desde el
    ÚLTIMO SNAPSHOT COMPLETO ya generado (04_DATASETS_ORBIT/mod_volumen_vendedor.csv), sin usar la
    fecha del reloj y sin recomputar reglas comerciales (sólo lee valores YA calculados por el
    cierre). Devuelve el MISMO contrato que matinal_resumen, o None si no hay snapshot.

    Fecha efectiva: metadata real del dataset (fecha_ejecucion = último cierre completo). Plan: de
    Google Sheets (fuente de verdad) hidratando la cache si hace falta, para esa MISMA fecha. Real:
    real_resultado/venta_ayer (venta del día) + clientes_compra_ayer (CCC total del día) del snapshot.
    El CCC real POR SEGMENTO del día no está en los datasets vendorizados (mod_ccc_segmento es del
    MES): en modo embebido queda en 0 (no se inventa); el real principal (venta + CCC total) sí sale.
    """
    vol = read_csv(DATASETS / "mod_volumen_vendedor.csv")
    if vol.empty or "fecha_ejecucion" not in vol.columns:
        return None
    fechas = sorted(str(f).strip() for f in vol["fecha_ejecucion"].dropna().unique() if str(f).strip())
    if not fechas:
        return None
    fecha_efectiva = fechas[-1]                      # último cierre completo (no hardcode, no reloj)
    vol = vol[vol["fecha_ejecucion"].astype(str).str.strip() == fecha_efectiva]

    def _num(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    # Real del día por vendedor (precomputado en el snapshot; no se recalcula).
    real_snap = {}
    for _, row in vol.iterrows():
        cn = clean_code(str(row.get("vendedor_codigo", "")))
        if not cn.isdigit():
            continue
        rr, va = _num(row.get("real_resultado")), _num(row.get("venta_ayer"))
        real_snap[int(cn)] = {"venta": rr if rr > 0 else va,
                              "ccc_total": int(_num(row.get("clientes_compra_ayer")))}

    # Plan de la MISMA fecha efectiva: Google Sheets (fuente de verdad), cache SQLite si ya está.
    planes = {}
    try:
        conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
        conn.row_factory = sqlite3.Row
        n = conn.execute("SELECT COUNT(*) FROM planificacion").fetchone()[0]
        if n == 0 and gsheets_enabled():
            conn.close()
            try:
                hydrate_planificacion_from_sheets()
            except Exception:
                pass
            conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
            conn.row_factory = sqlite3.Row
        planes = {r["vendedor_id"]: dict(r) for r in conn.execute(
            "SELECT * FROM planificacion WHERE fecha=?", (fecha_efectiva,)).fetchall()}
        conn.close()
    except Exception:
        planes = {}

    vend = read_csv(CONFIG / "vendedores_activos.csv")
    resultado, tiene_real_any = [], False
    if not vend.empty:
        for _, v in vend[vend["activo"] == 1].iterrows():
            cod = str(v["codigo_vendedor"]).strip()
            cn = clean_code(cod)
            cod_int = int(cn) if cn.isdigit() else 0
            r = real_snap.get(cod_int, {})
            v_tiene_real = cod_int in real_snap
            if v_tiene_real:
                tiene_real_any = True
            real_venta = float(r.get("venta") or 0)
            plan = planes.get(cod, {})
            plan_venta = float(plan.get("venta_esperada") or 0)
            pct = round(real_venta / plan_venta * 100, 1) if plan_venta else None
            resultado.append({
                "vendedor_id": cod, "vendedor_nombre": str(v["nombre_vendedor"]).strip(),
                "fecha_plan": fecha_efectiva, "fecha_real": fecha_efectiva,
                "plan_venta": plan_venta, "real_ayer": real_venta,
                "delta": round(real_venta - plan_venta, 2), "pct_cumplimiento": pct,
                "plan_ccc_trad": int(plan.get("ccc_tradicional") or 0),
                "plan_ccc_as":   int(plan.get("ccc_autoservicio") or 0),
                "plan_ccc_op":   int(plan.get("ccc_onpremise") or 0),
                "plan_once_t":   int(plan.get("once_t") or 0),
                "real_ccc_trad": 0, "real_ccc_as": 0, "real_ccc_op": 0,   # por-segmento no vendorizado
                "real_ccc_total": int(r.get("ccc_total") or 0),
                "plan_acciones": plan.get("acciones") or "",
                "plan_estado": plan.get("estado") or "sin_plan",
                "plan_id": plan.get("id"),
                "tiene_plan": bool(plan),
                "tiene_real": v_tiene_real,
            })
    if not resultado:
        return None
    return {
        "fecha_plan": fecha_efectiva, "fecha_real": fecha_efectiva,
        "tiene_real": tiene_real_any, "modo": "cierre",
        "fuente_real": "snapshot mod_volumen_vendedor.csv (último cierre completo)",
        "fecha_real_ayer": None, "fecha_efectiva": fecha_efectiva, "origen": "snapshot",
        "generado_en": _now_ar(), "resumen": resultado,
    }


@app.route("/api/matinal/resumen")
def matinal_resumen():
    """
    Compara plan vs real. El PLAN ancla la fecha (no ventas.csv).
    - fecha_plan = parámetro ?fecha, o la fecha más reciente con planes en orbit.db.
    - real = ventas de esa misma fecha en ventas.csv. Si no existe todavía → tiene_real=False.

    Flujo:
      Mañana matinal → muestra planes aprobados, real = '–' (ventas aún no actualizadas).
      Después del .bat → ventas.csv se actualiza → real aparece solo, sin tocar nada más.
      Planes del día siguiente → tienen otra fecha, no pisan estos.
    """
    fecha_param = request.args.get("fecha")
    modo = (request.args.get("modo") or "cierre").strip().lower()

    # PASO 1: Determinar fecha_plan desde orbit.db
    conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
    conn.row_factory = sqlite3.Row
    if fecha_param:
        fecha_plan = fecha_param
    else:
        today_ar = _now_ar()[:10]
        # Ventana acotada: evita que un plan viejo de prueba ancle Plan vs Real indefinidamente.
        cutoff = (datetime.now(_ARG_TZ) - timedelta(days=10)).strftime("%Y-%m-%d")
        if modo in ("actual", "ultimo", "plan"):
            row = conn.execute(
                "SELECT fecha FROM planificacion WHERE fecha >= ? ORDER BY fecha DESC LIMIT 1",
                (cutoff,)
            ).fetchone()
        else:
            # modo "cierre": anclar en el ÚLTIMO día con cierre hecho (último snapshot
            # disponible), no en `fecha < hoy`. El cierre del día agrega el snapshot de
            # hoy a acumulado_resultado_historico.csv; recién entonces Plan vs Real debe
            # mostrar plan(hoy) vs real(hoy), y mantenerlo hasta el próximo cierre.
            _, last_snap, _ = _real_dia_resultado()
            anchor = last_snap or (datetime.now(_ARG_TZ) - timedelta(days=1)).strftime("%Y-%m-%d")
            row = conn.execute(
                "SELECT fecha FROM planificacion WHERE fecha <= ? AND fecha >= ? ORDER BY fecha DESC LIMIT 1",
                (anchor, cutoff)
            ).fetchone()
            if not row:
                row = conn.execute(
                    "SELECT fecha FROM planificacion WHERE fecha >= ? ORDER BY fecha DESC LIMIT 1",
                    (cutoff,)
                ).fetchone()
        fecha_plan = row["fecha"] if row else today_ar

    planes = {row["vendedor_id"]: dict(row)
              for row in conn.execute(
                  "SELECT * FROM planificacion WHERE fecha=?", (fecha_plan,)).fetchall()}
    conn.close()

    # PASO 2: Buscar real para esa misma fecha en ventas.csv (CCC) + real $ por diferencia de acumulado
    ventas_dia, fecha_real = _cargar_ventas_dia(fecha_plan)
    tiene_real = not ventas_dia.empty  # False = aún no se corrió el .bat con ese día
    # Real del día ($) = Acumulado(hoy) − Acumulado(ayer) desde resultado.xlsx (snapshots).
    real_dia_map, f_real_hoy, f_real_ayer = _real_dia_resultado(fecha_plan)
    usa_resultado = (f_real_hoy == fecha_plan) and (f_real_ayer is not None) and bool(real_dia_map)
    if usa_resultado:
        tiene_real = True
        fecha_real = f_real_hoy

    # PASO 3: Calcular real_map solo si hay datos de ventas
    real_map = {}
    if tiene_real:
        for cod_v, grp in ventas_dia.groupby("vendedor_codigo"):
            cod_int = int(cod_v)
            ccc_t = int(grp[grp["segmento_operativo"] == "TRADICIONAL"]["cliente_id"].nunique())
            ccc_a = int(grp[grp["segmento_operativo"] == "AUTOSERVICIO"]["cliente_id"].nunique())
            ccc_o = int(grp[grp["segmento_operativo"] == "ON_PREMISE_VTK"]["cliente_id"].nunique())
            ccc_p = int(grp[grp["segmento_operativo"] == "PROXIMITY"]["cliente_id"].nunique())
            if cod_int == 3:           # V3 no trabaja autoservicio ni on premise (Proximity sí)
                ccc_a = 0
                ccc_o = 0
            real_map[cod_int] = {
                "venta":     round(float(grp["importe_neto"].sum()), 2),
                "ccc_trad":  ccc_t,
                "ccc_as":    ccc_a,
                "ccc_op":    ccc_o,
                "ccc_prox":  ccc_p,
                "ccc_total": int(grp["cliente_id"].nunique()),
            }

    vend = read_csv(CONFIG / "vendedores_activos.csv")
    resultado = []
    if not vend.empty:
        for _, v in vend[vend["activo"] == 1].iterrows():
            cod     = str(v["codigo_vendedor"]).strip()
            cn      = clean_code(cod)
            nombre  = str(v["nombre_vendedor"]).strip()
            cod_int = int(cn) if cn.isdigit() else 0

            real       = real_map.get(cod_int, {})
            # Venta real del día: diferencia de acumulado (resultado.xlsx); fallback a ventas.csv
            if usa_resultado and cod_int in real_dia_map:
                real_venta = float(real_dia_map[cod_int])
                v_tiene_real = True
            else:
                real_venta = float(real.get("venta") or 0)
                v_tiene_real = bool(real)

            plan       = planes.get(cod, {})
            plan_venta = float(plan.get("venta_esperada") or 0)
            delta      = real_venta - plan_venta
            pct        = round(real_venta / plan_venta * 100, 1) if plan_venta else None

            resultado.append({
                "vendedor_id":      cod,
                "vendedor_nombre":  nombre,
                "fecha_plan":       fecha_plan,
                "fecha_real":       fecha_real,
                "plan_venta":       plan_venta,
                "real_ayer":        real_venta,
                "delta":            round(delta, 2),
                "pct_cumplimiento": pct,
                "plan_ccc_trad":    int(plan.get("ccc_tradicional") or 0),
                "plan_ccc_as":      int(plan.get("ccc_autoservicio") or 0),
                "plan_ccc_op":      int(plan.get("ccc_onpremise") or 0),
                "plan_once_t":      int(plan.get("once_t") or 0),
                "real_ccc_trad":    real.get("ccc_trad", 0),
                "real_ccc_as":      real.get("ccc_as", 0),
                "real_ccc_op":      real.get("ccc_op", 0),
                "real_ccc_total":   real.get("ccc_total", 0),
                "plan_acciones":    plan.get("acciones") or "",
                "plan_estado":      plan.get("estado") or "sin_plan",
                "plan_id":          plan.get("id"),
                "tiene_plan":       bool(plan),
                "tiene_real":       v_tiene_real,
            })

    # Fecha efectiva única: si el día vivo NO está completo (falta plan o falta real, p.ej. modo
    # embebido en Orbit Home sin SQLite/02_HISTORY/01_INPUTS, o mañana sin real todavía), usar el
    # ÚLTIMO cierre completo del snapshot generado — nunca la fecha del reloj, sin mezclar fechas.
    # Si el día vivo tiene plan Y real (comportamiento standalone habitual), se mantiene igual.
    live_completo = bool(planes) and bool(tiene_real)
    if (not fecha_param) and (not live_completo):
        snap = _snapshot_matinal_resumen()
        if snap is not None:
            return jsonify(snap)

    return jsonify({
        "fecha_plan":  fecha_plan,
        "fecha_real":  fecha_real or None,
        "tiene_real":  tiene_real,
        "modo":        modo,
        "fuente_real": ("resultado.xlsx (acumulado hoy − ayer)" if usa_resultado else "ventas.csv"),
        "fecha_real_ayer": f_real_ayer if usa_resultado else None,
        "generado_en": _now_ar(),
        "resumen":     resultado,
    })


# ====== GERENCIA: CCC EMPRESA ======
# Canales de la cobertura CCC total empresa (objetivo en 01_INPUTS/objccc.xlsx).
# Autoservicio se identifica por SUBRAMO, no por Ramo: "AUTOSERVICIO TRADICIONAL" tiene
# Ramo=TRADITIONAL TRADE y es el grueso del canal (764 de 826 filas AS). Clasificar por Ramo
# dejaba una cartera AS de sólo 18 clientes contra un objetivo de 145 -> la tarjeta mostraba
# 3.4% de avance, imposible (corregido 2026-07-20). Mismo criterio que `_clasificar()` de
# generar_datasets_acum.py, para que esta tarjeta y la de Cobertura acumulada no se
# contradigan sobre el mismo objetivo.
# MAYORISTA/Cash&Carry queda FUERA de los canales con objetivo (objccc.xlsx no lo abre),
# igual que en mod_cobertura_acum.csv: no se mide contra un objetivo que no existe.
_CCC_CANALES_ORDEN = ["Tradicionales", "Autoservicios", "On Premise", "Vinotecas", "On Premise Noche"]
_CCC_AS_SUBRAMOS = ("AUTOSERVICIO", "CADENA REGIONAL", "CADENAS REGIONALES", "LARGE FORMAT")

def _canal_ccc_empresa(df):
    """Serie con el canal objccc por fila, clasificado por Ramo/Subramo."""
    ramo = df["Ramo"].astype(str).str.upper().str.strip() if "Ramo" in df.columns else pd.Series("", index=df.index)
    sub  = df["Subramo"].astype(str).str.upper().str.strip() if "Subramo" in df.columns else pd.Series("", index=df.index)
    txt  = ramo.str.cat(sub, sep=" ")
    canal = pd.Series("Tradicionales", index=df.index)
    es_as = ramo.isin(("AUTOSERVICIO", "LARGE FORMAT"))
    for clave in _CCC_AS_SUBRAMOS:
        es_as = es_as | sub.str.startswith(clave, na=False)
    canal[es_as] = "Autoservicios"
    canal[txt.str.contains("ON PREMISE", regex=False, na=False)
          | txt.str.contains("AWAY FROM HOME", regex=False, na=False)
          | txt.str.contains("RESTAURANT", regex=False, na=False)] = "On Premise"
    canal[txt.str.contains("VINOTECA", regex=False, na=False)] = "Vinotecas"
    canal[txt.str.contains("NOCHE", regex=False, na=False)] = "On Premise Noche"
    # Mayorista al final: nunca debe quedar contado como Autoservicio ni como Tradicional.
    canal[ramo.str.contains("CASH", regex=False, na=False)
          | ramo.str.startswith("MAYORISTA", na=False)
          | sub.str.startswith("MAYORISTA", na=False)] = "Mayoristas"
    return canal

def _objetivos_ccc_empresa():
    """canal -> objetivo CCC desde 01_INPUTS/objccc.xlsx (columnas Canal / Objetivo CCC).

    Lee la hoja "total" POR NOMBRE: desde 2026-07-20 el archivo tiene 4 hojas (total +
    apertura por vendedor) y tomar la primera por posición se rompería en silencio si
    alguien las reordena. Cae a la primera hoja si no existe "total" (formato viejo,
    hoja única "Hoja1")."""
    p = INPUTS / "objccc.xlsx"
    out = {}
    if not p.exists():
        return out
    try:
        hojas = pd.ExcelFile(p).sheet_names
        hoja = next((s for s in hojas if str(s).strip().lower() == "total"), hojas[0])
        d = pd.read_excel(p, sheet_name=hoja)
        cols = {str(c).strip().lower(): c for c in d.columns}
        c_can = cols.get("canal")
        c_obj = next((cols[k] for k in cols if "objetivo" in k), None)
        if c_can and c_obj:
            for _, r in d.iterrows():
                can = str(r[c_can]).strip()
                try:
                    out[can] = int(float(r[c_obj]))
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        print(f"[AVISO] objccc.xlsx: {e}")
    return out


# Hojas de objccc.xlsx con la apertura por vendedor -> segmento de mod_cobertura_acum.csv.
# La hoja "On premise" ya incluye Vinotecas y On Premise Noche (30+15+11=56), igual que el
# segmento ON_PREMISE del clasificador (_OP_KEYWORDS incluye VINOTECA). MAYORISTA no tiene
# objetivo cargado: se muestra sin objetivo, nunca contra 0.
_OBJCCC_HOJAS = {"autoservicio": "AUTOSERVICIO", "tradicional": "TRADICIONAL", "on premise": "ON_PREMISE"}
_OBJCCC_VEND_CACHE = {"mtime": None, "data": None}

def _objetivos_ccc_vendedor():
    """Objetivos CCC aperturados por vendedor desde 01_INPUTS/objccc.xlsx.

    Las hojas por vendedor no tienen encabezado real (columnas 'Unnamed'), así que NO se
    leen por posición: se busca en cada fila la celda 'V<n>' y se toma el último numérico
    de esa fila como objetivo. La fila 'Total' se guarda aparte como total declarado.

    Devuelve {"por_segmento": {SEG: {cod_vend: obj}}, "declarado": {SEG: total_hoja}}."""
    p = INPUTS / "objccc.xlsx"
    vacio = {"por_segmento": {}, "declarado": {}}
    if not p.exists():
        return vacio
    mt = p.stat().st_mtime
    if _OBJCCC_VEND_CACHE["data"] is not None and _OBJCCC_VEND_CACHE["mtime"] == mt:
        return _OBJCCC_VEND_CACHE["data"]
    por_seg, declarado = {}, {}
    try:
        x = pd.ExcelFile(p)
        hojas = {str(s).strip().lower(): s for s in x.sheet_names}
        for clave, seg in _OBJCCC_HOJAS.items():
            hoja = hojas.get(clave)
            if hoja is None:
                continue
            d = x.parse(hoja, header=None)
            por_seg[seg] = {}
            for _, fila in d.iterrows():
                celdas = [c for c in fila.tolist() if pd.notna(c)]
                if not celdas:
                    continue
                textos = [str(c).strip() for c in celdas]
                nums = [pd.to_numeric(c, errors="coerce") for c in celdas]
                nums = [n for n in nums if pd.notna(n)]
                if not nums:
                    continue
                valor = int(float(nums[-1]))
                cod = next((t[1:] for t in textos
                            if len(t) > 1 and t[0] in "Vv" and t[1:].isdigit()), None)
                if cod is not None:
                    por_seg[seg][int(cod)] = valor
                elif any(t.lower() == "total" for t in textos):
                    declarado[seg] = valor
    except Exception as e:
        print(f"[AVISO] objccc.xlsx (apertura por vendedor): {e}")
        return vacio
    data = {"por_segmento": por_seg, "declarado": declarado}
    _OBJCCC_VEND_CACHE.update({"mtime": mt, "data": data})
    return data

@app.route("/api/gerencia/ccc_empresa")
def gerencia_ccc_empresa():
    """CCC del mes vivo (cobertura) total empresa vs objetivo por canal.
    Real: ventas.csv (mes actual, neto>0, excluye V1/V2/V5/V20), clientes únicos por canal
    (Ramo+Subramo, ver `_canal_ccc_empresa`). NO se filtra por Empresa: P&P Logística es
    nuestra 2da razón social (ver `_LEEME_EMPRESA`). Objetivo: 01_INPUTS/objccc.xlsx."""
    obj_map = _objetivos_ccc_empresa()
    ccc_map = {}
    total_ccc = 0
    fuera_obj = 0        # clientes en canales sin objetivo (Mayoristas): no entran al %
    vpath = INPUTS / "ventas.csv"
    if vpath.exists():
        try:
            v = pd.read_csv(vpath, sep=";", encoding="latin1", low_memory=False)
            v["imp"] = pd.to_numeric(
                v["ImporteNetoItem"].astype(str).str.replace(",", ".", regex=False), errors="coerce")
            v = v[~v["CodVendedor"].isin(_VENDEDORES_EXCLUIDOS)]
            v = v[v["imp"] > 0]
            if not v.empty:
                v["canal"] = _canal_ccc_empresa(v)
                ccc_map = v.groupby("canal")["Cliente"].nunique().to_dict()
                # El total se arma con los canales QUE TIENEN objetivo, para que numerador y
                # denominador midan lo mismo (antes era nunique() global e incluía Mayoristas).
                total_ccc = int(v[v["canal"].isin(_CCC_CANALES_ORDEN)]["Cliente"].nunique())
                fuera_obj = int(v[~v["canal"].isin(_CCC_CANALES_ORDEN)]["Cliente"].nunique())
        except Exception as e:
            print(f"[AVISO] ccc_empresa ventas.csv: {e}")

    canales = []
    total_obj = 0
    for can in _CCC_CANALES_ORDEN:
        ccc = int(ccc_map.get(can, 0))
        obj = int(obj_map.get(can, 0))
        pct = round(ccc / obj * 100, 1) if obj else None
        canales.append({"canal": can, "ccc": ccc, "objetivo": obj, "pct": pct})
        total_obj += obj

    _op = sum(int(ccc_map.get(c, 0)) for c in ("On Premise", "Vinotecas", "On Premise Noche"))
    return jsonify({
        "generado_en": _now_ar(),
        "fuente": "ventas.csv (mes vivo)",
        "canales": canales,
        "empresa": {
            "total":         total_ccc,
            "objetivo_total": total_obj,
            "pct":           round(total_ccc / total_obj * 100, 1) if total_obj else None,
            "fuera_objetivo": fuera_obj,
            # compat con el render previo (Trad / AS / OP)
            "tradicional":   int(ccc_map.get("Tradicionales", 0)),
            "autoservicio":  int(ccc_map.get("Autoservicios", 0)),
            "onpremise":     _op,
        },
    })


# ══════════════════════════════════════════════════════════════════════════════
# SEMANAL — apertura del mes en 4 semanas (pantalla Semanal de gerencia)
# ══════════════════════════════════════════════════════════════════════════════
# Definiciones acordadas con el usuario (2026-07-27):
#   · Semana = bloque de días del mes: S1 = 1-7, S2 = 8-14, S3 = 15-21, S4 = 22-fin.
#     Siempre 4 semanas → los meses se comparan entre sí sin ajustes.
#   · Facturación = suma de ImporteNetoItem de la semana (medida FACTURADO, por
#     FechaComprobante). El % de cada semana suma 100% del mes.
#   · CCC = aporte INCREMENTAL: cada cliente cuenta en la semana en que compró por
#     PRIMERA VEZ en el mes. Las 4 semanas suman 100% del CCC del mes (si se contara
#     el CCC bruto semanal, un cliente que compra 2 semanas contaría 2 veces y el
#     total pasaría el 100%, que no es lo que se quiere planificar).
#   · Canal: se reusa `_canal_ccc_empresa` (misma clasificación que CCC empresa /
#     objccc.xlsx). On Premise agrupa On Premise + Vinotecas + On Premise Noche,
#     igual que el objetivo del Excel (30+15+11).
#   · V20 Depósito queda EXCLUIDO (además de V1/V2/V5): la planificación semanal es
#     de ruta y se mide contra objetivos, donde V20 nunca entra.
#
# Fuente por mes (se resuelve sola, sin tocar código cada mes):
#   1) 01_INPUTS/cierres mes/ventas_mes_MMAAAA.csv  → cierre mensual congelado
#   2) 02_HISTORY/historial_ventas.csv              → fallback para los meses viejos
#      (export estático 2024-03 → 2026-04; es la única fuente con detalle diario
#      anterior a los cierres versionados)
#   3) 01_INPUTS/ventas.csv                         → mes en curso
_SEMANAL_DESDE = "2025-07"
_SEMANAL_MESES_ES = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                     7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
_SEMANAL_KPIS = [
    {"id":"facturacion",      "label":"Facturación",       "tipo":"moneda"},
    {"id":"ccc_tradicional",  "label":"CCC Tradicionales",  "tipo":"clientes"},
    {"id":"ccc_autoservicio", "label":"CCC Autoservicios",  "tipo":"clientes"},
    {"id":"ccc_onpremise",    "label":"CCC On Premise",     "tipo":"clientes"},
]
_SEMANAL_KPI_IDS = [k["id"] for k in _SEMANAL_KPIS]
_SEMANAL_CANALES = {
    "ccc_tradicional":  ("Tradicionales",),
    "ccc_autoservicio": ("Autoservicios",),
    "ccc_onpremise":    ("On Premise", "Vinotecas", "On Premise Noche"),
}
_SEMANAL_COLS = ["Cliente", "FechaComprobante", "CodVendedor", "ImporteNetoItem", "Ramo", "Subramo"]


def _semanal_num(serie):
    """Importe a float tolerando los dos formatos que conviven en las fuentes.
    Si el texto trae coma decimal ('1.234,56' / '15800,82') el punto es separador de
    miles; si no la trae, se parsea tal cual. Nunca al revés: strippear el punto en
    un '15800.82' lo multiplicaría por 100."""
    t = serie.astype(str).str.strip().str.strip('"')
    con_coma = t.str.contains(",", regex=False)
    t = t.where(~con_coma, t.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    return pd.to_numeric(t, errors="coerce").fillna(0.0)


def _semanal_fechas(serie):
    """FechaComprobante a datetime. Los cierres versionados vienen en ISO
    (2026-06-01) y ventas.csv / historial en dd/mm/aaaa: parsear todo con
    dayfirst=True desarma las ISO (mes y día invertidos → meses fantasma)."""
    s = serie.astype(str).str.strip()
    iso = s.str.match(r"^\d{4}-\d{2}-\d{2}")
    fec = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")
    if iso.any():
        fec[iso] = pd.to_datetime(s[iso].str.slice(0, 10), format="%Y-%m-%d", errors="coerce")
    if (~iso).any():
        fec[~iso] = pd.to_datetime(s[~iso], dayfirst=True, errors="coerce")
    return fec


def _leer_ventas_min(path, cols):
    """Lee una fuente de ventas del ERP tomando SÓLO las columnas pedidas.
    Sirve para las 3 fuentes (ventas.csv, cierre versionado, historial): sniff del
    separador (`;` vs `,`) y encoding en cascada. `usecols` es lo que hace esto viable
    en Render: de las 57 columnas del ERP se parsean 4-6 (historial_ventas.csv pesa
    63 MB y leerlo entero es inviable)."""
    if not path.exists():
        return pd.DataFrame()
    try:
        with open(path, "r", encoding="utf-8-sig", errors="ignore") as f:
            cab = f.readline()
    except OSError:
        return pd.DataFrame()
    sep = ";" if cab.count(";") >= cab.count(",") else ","
    df = None
    for enc in ("utf-8-sig", "latin-1", "windows-1252"):
        try:
            df = pd.read_csv(path, sep=sep, encoding=enc, dtype=str, quotechar='"',
                             usecols=cols, low_memory=False)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = [c.strip() for c in df.columns]
    return df


def _semanal_leer(path):
    """Lee una fuente de ventas (cualquiera de las 3) y devuelve el DataFrame mínimo
    normalizado: cliente, fecha, imp, canal, sem, periodo. Ya filtrado a neto>0 y sin
    vendedores excluidos."""
    df = _leer_ventas_min(path, _SEMANAL_COLS)
    if df.empty:
        return pd.DataFrame()
    df["imp"] = _semanal_num(df["ImporteNetoItem"])
    df["cv"]  = pd.to_numeric(df["CodVendedor"], errors="coerce")
    df["fec"] = _semanal_fechas(df["FechaComprobante"])
    df = df[(~df["cv"].isin(_VENDEDORES_EXCLUIDOS)) & (df["imp"] > 0)]
    df = df.dropna(subset=["fec", "Cliente"])
    if df.empty:
        return pd.DataFrame()
    df = df.copy()
    df["canal"]   = _canal_ccc_empresa(df)
    df["sem"]     = (((df["fec"].dt.day - 1) // 7) + 1).clip(upper=4)
    df["periodo"] = df["fec"].dt.strftime("%Y-%m")
    return df[["Cliente", "fec", "imp", "canal", "sem", "periodo"]]


def _semanal_agg(df):
    """{kpi: {total, valores[4], pcts[4]}} para un mes ya filtrado."""
    out = {}
    tot_f = float(df["imp"].sum())
    val_f = [float(df.loc[df["sem"] == s, "imp"].sum()) for s in (1, 2, 3, 4)]
    out["facturacion"] = {
        "total":   round(tot_f, 2),
        "valores": [round(v, 2) for v in val_f],
        "pcts":    [round(v / tot_f * 100, 1) if tot_f else 0.0 for v in val_f],
    }
    for kpi, canales in _SEMANAL_CANALES.items():
        g = df[df["canal"].isin(canales)]
        if g.empty:
            out[kpi] = {"total": 0, "valores": [0, 0, 0, 0], "pcts": [0.0] * 4}
            continue
        # Semana de la PRIMERA compra del mes de cada cliente → aporte incremental
        primera = g.groupby("Cliente")["sem"].min()
        n = int(primera.shape[0])
        val = [int((primera == s).sum()) for s in (1, 2, 3, 4)]
        out[kpi] = {
            "total":   n,
            "valores": val,
            "pcts":    [round(v / n * 100, 1) if n else 0.0 for v in val],
        }
    return out


def _semanal_periodos_cerrados(hoy=None):
    """Lista de 'YYYY-MM' desde _SEMANAL_DESDE hasta el mes anterior al actual."""
    hoy = hoy or datetime.now(_ARG_TZ).date()
    a0, m0 = int(_SEMANAL_DESDE[:4]), int(_SEMANAL_DESDE[5:7])
    a1, m1 = (hoy.year, hoy.month - 1) if hoy.month > 1 else (hoy.year - 1, 12)
    out = []
    a, m = a0, m0
    while (a, m) <= (a1, m1):
        out.append(f"{a:04d}-{m:02d}")
        a, m = (a, m + 1) if m < 12 else (a + 1, 1)
    return out


def _semanal_fuente_de(periodo):
    """(path, etiqueta) del archivo que manda para ese mes. Prioriza el cierre
    versionado; si no está, cae al historial estático."""
    p = INPUTS / "cierres mes" / f"ventas_mes_{periodo[5:7]}{periodo[:4]}.csv"
    if p.exists():
        return p, "cierre mensual"
    return BASE / "02_HISTORY" / "historial_ventas.csv", "historial"


def _semanal_mtime(path):
    try:
        return path.stat().st_mtime
    except OSError:
        return 0


_SEMANAL_HIST_CACHE = {"key": None, "data": None}


def _semanal_historico():
    """Apertura semanal de todos los meses cerrados desde julio 2025.
    Cacheado por (archivo, mtime) de todas las fuentes: el historial se parsea una
    sola vez por proceso y sólo se recalcula si cambia algún archivo."""
    periodos = _semanal_periodos_cerrados()
    fuentes = {per: _semanal_fuente_de(per) for per in periodos}
    key = tuple(sorted({(str(p), _semanal_mtime(p)) for p, _ in fuentes.values()}))
    if _SEMANAL_HIST_CACHE["data"] is not None and _SEMANAL_HIST_CACHE["key"] == key:
        return _SEMANAL_HIST_CACHE["data"]
    cache_df = {}
    meses = []
    for per in periodos:
        path, etiqueta = fuentes[per]
        sp = str(path)
        if sp not in cache_df:
            cache_df[sp] = _semanal_leer(path)
        df = cache_df[sp]
        if df.empty:
            continue
        g = df[df["periodo"] == per]
        if g.empty:
            continue
        a, m = int(per[:4]), int(per[5:7])
        meses.append({
            "periodo": per,
            "label":   f"{_SEMANAL_MESES_ES[m]} {a}",
            "fuente":  etiqueta,
            "kpis":    _semanal_agg(g),
        })
    promedio = {}
    for kpi in _SEMANAL_KPI_IDS:
        filas = [m["kpis"][kpi]["pcts"] for m in meses if m["kpis"][kpi]["total"]]
        promedio[kpi] = ([round(sum(f[i] for f in filas) / len(filas), 1) for i in range(4)]
                         if filas else [0.0] * 4)
    data = {"meses": meses, "promedio": promedio, "meses_promediados":
            {kpi: sum(1 for m in meses if m["kpis"][kpi]["total"]) for kpi in _SEMANAL_KPI_IDS}}
    _SEMANAL_HIST_CACHE.update({"key": key, "data": data})
    return data


_SEMANAL_ACTUAL_CACHE = {"key": None, "data": None}


def _semanal_actual():
    """Apertura semanal del mes en curso (ventas.csv). Parcial por definición:
    las semanas que todavía no ocurrieron valen 0 y se marcan 'pendiente'."""
    from calendar import monthrange
    path = INPUTS / "ventas.csv"
    hoy = datetime.now(_ARG_TZ).date()
    per = f"{hoy.year:04d}-{hoy.month:02d}"
    key = (str(path), _semanal_mtime(path), per, hoy.isoformat())
    if _SEMANAL_ACTUAL_CACHE["data"] is not None and _SEMANAL_ACTUAL_CACHE["key"] == key:
        return _SEMANAL_ACTUAL_CACHE["data"]
    ultimo = monthrange(hoy.year, hoy.month)[1]
    rangos = [(1, 7), (8, 14), (15, 21), (22, ultimo)]
    semanas = []
    for i, (d1, d2) in enumerate(rangos, start=1):
        if hoy.day > d2:
            estado = "cerrada"
        elif hoy.day >= d1:
            estado = "en_curso"
        else:
            estado = "pendiente"
        semanas.append({
            "semana": i, "estado": estado,
            "desde": f"{per}-{d1:02d}", "hasta": f"{per}-{d2:02d}",
            "label": f"{d1}–{d2}",
        })
    df = _semanal_leer(path)
    g = df[df["periodo"] == per] if not df.empty else pd.DataFrame()
    if g.empty:
        kpis = {k: {"total": 0, "valores": [0] * 4, "pcts": [0.0] * 4} for k in _SEMANAL_KPI_IDS}
        ultima_fecha = None
    else:
        kpis = _semanal_agg(g)
        ultima_fecha = str(g["fec"].max().date())
    data = {
        "periodo": per,
        "label":   f"{_SEMANAL_MESES_ES[hoy.month]} {hoy.year}",
        "hoy":     hoy.isoformat(),
        "ultima_fecha_venta": ultima_fecha,
        "semanas": semanas,
        "kpis":    kpis,
    }
    _SEMANAL_ACTUAL_CACHE.update({"key": key, "data": data})
    return data


def _semanal_objetivos():
    """Objetivo del mes por KPI: CCC de objccc.xlsx (mismos canales que CCC empresa)
    y facturación de resultado.xlsx hoja Avance (suma de ValorObjetivo de la ruta).
    Sin fuente devuelve None — nunca 0, que se leería como 'objetivo cero'."""
    obj = {k: None for k in _SEMANAL_KPI_IDS}
    ccc = _objetivos_ccc_empresa()
    if ccc:
        obj["ccc_tradicional"]  = int(ccc.get("Tradicionales", 0)) or None
        obj["ccc_autoservicio"] = int(ccc.get("Autoservicios", 0)) or None
        op = sum(int(ccc.get(c, 0)) for c in _SEMANAL_CANALES["ccc_onpremise"])
        obj["ccc_onpremise"] = op or None
    p = INPUTS / "resultado.xlsx"
    if p.exists():
        try:
            av = pd.read_excel(p, sheet_name="Avance")
            tot = 0.0
            for _, r in av.iterrows():
                cn = clean_code(str(r.get("VendedorCodigo", "")))
                if not cn or int(cn) in _VENDEDORES_EXCLUIDOS:
                    continue
                tot += float(r.get("ValorObjetivo", 0) or 0)
            obj["facturacion"] = round(tot, 2) or None
        except Exception as e:
            print(f"[AVISO] semanal objetivos resultado.xlsx: {e}")
    return obj


def _semanal_plan_leer(periodo):
    """{kpi: [p1,p2,p3,p4]} + metadata de la última edición."""
    plan = {k: [None] * 4 for k in _SEMANAL_KPI_IDS}
    meta = {"updated_at": None, "editado_por": None}
    try:
        conn = sqlite3.connect(str(DB_PATH))
        rows = conn.execute(
            "SELECT kpi, semana, pct, editado_por, updated_at FROM plan_semanal WHERE periodo=?",
            (periodo,)).fetchall()
        conn.close()
    except Exception as e:
        print(f"[AVISO] plan_semanal leer: {e}")
        return plan, meta
    for kpi, sem, pct, autor, ts in rows:
        if kpi in plan and 1 <= int(sem) <= 4:
            plan[kpi][int(sem) - 1] = None if pct is None else round(float(pct), 2)
        if ts and (meta["updated_at"] is None or ts > meta["updated_at"]):
            meta = {"updated_at": ts, "editado_por": autor}
    return plan, meta


def _semanal_plan_export_csv():
    """Respaldo del plan semanal a CSV (mismo criterio que la planificación diaria:
    la tabla vive en el disco persistente, pero el CSV permite recuperarla a mano)."""
    try:
        PLAN_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(str(DB_PATH))
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM plan_semanal ORDER BY periodo, kpi, semana").fetchall()
        conn.close()
        if not rows:
            return
        with open(str(PLAN_BACKUP_DIR / "plan_semanal_latest.csv"), "w", newline="", encoding="utf-8") as f:
            w = _csv.DictWriter(f, fieldnames=rows[0].keys())
            w.writeheader()
            w.writerows([dict(r) for r in rows])
    except Exception as e:
        print(f"[WARN] _semanal_plan_export_csv: {e}")


@app.route("/api/gerencia/semanal")
def gerencia_semanal():
    """Pantalla Semanal: histórico de distribución semanal + plan del mes en curso.
    Real (histórico y mes vivo) = ventas facturadas por FechaComprobante.
    Plan = carga manual de gerencia (tabla plan_semanal)."""
    hist = _semanal_historico()
    actual = _semanal_actual()
    plan, plan_meta = _semanal_plan_leer(actual["periodo"])
    return jsonify({
        "generado_en": _now_ar(),
        "fuente": ("cierres mes/ventas_mes_MMAAAA.csv + 02_HISTORY/historial_ventas.csv "
                   "(meses cerrados) · 01_INPUTS/ventas.csv (mes en curso)"),
        "definicion_semana": "S1 1-7 · S2 8-14 · S3 15-21 · S4 22-fin",
        "definicion_ccc": "aporte incremental: el cliente cuenta en la semana de su primera compra del mes",
        "excluidos": "V1 · V2 · V5 · V20 Depósito",
        "kpis": _SEMANAL_KPIS,
        "historico": hist["meses"],
        "promedio": hist["promedio"],
        "meses_promediados": hist["meses_promediados"],
        "actual": actual,
        "objetivos": _semanal_objetivos(),
        "plan": plan,
        "plan_meta": plan_meta,
    })


@app.route("/api/gerencia/semanal/plan", methods=["POST"])
def gerencia_semanal_plan():
    """Guarda el % planificado por semana y KPI del mes indicado.
    Body: {periodo:'YYYY-MM', autor:'...', plan:{kpi:[p1,p2,p3,p4]}}
    Un valor null/'' borra la celda (queda sin planificar, no en 0)."""
    body = request.get_json(silent=True) or {}
    periodo = str(body.get("periodo", "")).strip()
    _p = periodo.split("-")
    if len(_p) != 2 or not (_p[0].isdigit() and len(_p[0]) == 4
                            and _p[1].isdigit() and len(_p[1]) == 2 and 1 <= int(_p[1]) <= 12):
        return jsonify({"ok": False, "error": "periodo inválido (YYYY-MM)"}), 400
    plan = body.get("plan") or {}
    if not isinstance(plan, dict):
        return jsonify({"ok": False, "error": "plan inválido"}), 400
    autor = str(body.get("autor", "gerencia")).strip()[:40]
    ts = _now_ar()
    filas, borrar = [], []
    for kpi, vals in plan.items():
        if kpi not in _SEMANAL_KPI_IDS or not isinstance(vals, list):
            continue
        for i, v in enumerate(vals[:4], start=1):
            if v is None or v == "":
                borrar.append((periodo, kpi, i))
                continue
            try:
                pct = float(v)
            except (TypeError, ValueError):
                return jsonify({"ok": False, "error": f"valor no numérico en {kpi} S{i}"}), 400
            if pct < 0 or pct > 100:
                return jsonify({"ok": False, "error": f"{kpi} S{i}: el % debe estar entre 0 y 100"}), 400
            filas.append((periodo, kpi, i, round(pct, 2), autor, ts))
    try:
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        if borrar:
            c.executemany("DELETE FROM plan_semanal WHERE periodo=? AND kpi=? AND semana=?", borrar)
        if filas:
            c.executemany("""INSERT INTO plan_semanal(periodo,kpi,semana,pct,editado_por,updated_at)
                             VALUES(?,?,?,?,?,?)
                             ON CONFLICT(periodo,kpi,semana) DO UPDATE SET
                               pct=excluded.pct, editado_por=excluded.editado_por,
                               updated_at=excluded.updated_at""", filas)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[ERROR] plan_semanal guardar: {e}")
        return jsonify({"ok": False, "error": f"no se pudo guardar: {e}"}), 500
    _semanal_plan_export_csv()
    guardado, meta = _semanal_plan_leer(periodo)
    return jsonify({"ok": True, "periodo": periodo, "plan": guardado, "plan_meta": meta})


# ══════════════════════════════════════════════════════════════════════════════
# DÍAS DE STOCK — 11 Titulares · Innovaciones · MPA (tarjeta de la pantalla Semanal)
# ══════════════════════════════════════════════════════════════════════════════
# Días de stock = existencia actual / venta diaria del MES ANTERIOR (cerrado).
#   · Venta: CantBase del mes anterior, MISMA unidad que el stock (unidades/botellas).
#     Se suma con signo: las devoluciones vienen en negativo (TipoDeVenta "Devolución
#     por Rechazo"/"por Canje") y tienen que netear, porque vuelven al depósito.
#   · Divisor = días operativos del mes anterior (lun-sáb sin feriados): el depósito
#     no despacha los domingos, así que un "día de stock" es un día que se vende.
# Universos: 11T (matriz oficial por código), Innovaciones (Innovaciones.xlsx) y
# MPA (01_INPUTS/MPA/MPA.xlsx + el mapeo a código de 09_CONFIG/mpa_codigos.csv).
#
# DOS DEPÓSITOS, DOS TARJETAS (definido por el usuario 2026-07-28). No se suman ni se
# comparten: cada uno tiene su propio archivo de stock y su propia ruta de vendedores,
# así que la venta que descuenta cada stock es SÓLO la de sus vendedores.
#   · Stock PyP  → stock.xlsx           → V3, V4, V6, V8, V10
#   · VSB Cuyo   → stock_VSB_Cuyo.xlsx  → V7, V9
# V20 (Depósito) queda fuera de los dos: no pertenece a ninguna de las dos rutas.
_DIAS_STOCK_CRITICO = 15    # días: rojo
_DIAS_STOCK_ATENCION = 30   # días: ámbar; por encima, verde
_STOCK_BLOQUES = [
    {"id": "pyp", "label": "Stock PyP",  "archivo": "stock.xlsx",
     "vendedores": [3, 4, 6, 8, 10]},
    {"id": "vsb", "label": "VSB Cuyo",   "archivo": "stock_VSB_Cuyo.xlsx",
     "vendedores": [7, 9]},
]


def _innovaciones_codigos_todas():
    """[{codigo, nombre}] de TODAS las innovaciones de Innovaciones.xlsx.
    Se diferencia de `_inov_plan_as_productos()`, que sólo devuelve las marcadas con
    'x' para Planes AS: acá interesa el portfolio de innovación completo."""
    p = INPUTS / "INNOVACIONES" / "Innovaciones.xlsx"
    if not p.exists():
        return []
    out, vistos = [], set()
    try:
        df = pd.read_excel(p, sheet_name=0, header=None, dtype=str)
        for _, fila in df.iterrows():
            for celda in [str(v).strip() for v in fila.tolist() if pd.notna(v)]:
                m = _re.match(r"^0*(\d{4,6})\s*-\s*(.+)$", celda)
                if m:
                    cod = int(m.group(1))
                    if cod not in vistos:
                        vistos.add(cod)
                        out.append({"codigo": cod,
                                    "nombre": _re.sub(r"\s+", " ", m.group(2).replace("\xa0", " ")).strip()})
                    break
    except Exception as e:
        print(f"[WARN] innovaciones (días de stock): {e}")
        return []
    return out


_MPA_CACHE = {"key": None, "data": None}


def _mpa_universo():
    """({codigo: nombre_mpa}, [nombres sin código]) del universo MPA.

    MPA.xlsx trae los productos por NOMBRE COMERCIAL ("Alaris Malbec 0.75L"), que no
    coincide con el `Descripción Art.` del ERP ("TRAPICHE ALARIS MALBEC 6X750"). El
    puente es 09_CONFIG/mpa_codigos.csv, un mapeo REVISADO A MANO (el match
    automático por texto se equivocaba, p.ej. Alma Mora Cabernet → F.LAS MORAS CABSAU).
    Lo que está en MPA.xlsx y no en el mapeo se informa como "sin código asignado":
    nunca se descarta en silencio ni se adivina."""
    px = INPUTS / "MPA" / "MPA.xlsx"
    pc = CONFIG / "mpa_codigos.csv"
    key = (_semanal_mtime(px), _semanal_mtime(pc))
    if _MPA_CACHE["data"] is not None and _MPA_CACHE["key"] == key:
        return _MPA_CACHE["data"]

    def _norm(s):
        return _re.sub(r"[^a-z0-9]+", "", str(s).lower())

    nombres = []
    if px.exists():
        try:
            d = pd.read_excel(px, header=None, dtype=str)
            # Las filas de productos son las que arrancan con el plan (Inicial / Silver).
            # Ambas listan el MISMO universo, así que se unifican por nombre.
            for _, fila in d.iterrows():
                vals = [str(v).strip() for v in fila.tolist() if pd.notna(v)]
                if not vals or _norm(vals[0]) not in ("inicial", "silver"):
                    continue
                for v in vals[1:]:
                    if v and v not in nombres:
                        nombres.append(v)
        except Exception as e:
            print(f"[WARN] MPA.xlsx: {e}")

    mapa = {}
    if pc.exists():
        try:
            m = pd.read_csv(pc, dtype=str, encoding="utf-8-sig")
            for _, r in m.iterrows():
                cod = pd.to_numeric(r.get("codigo"), errors="coerce")
                if pd.isna(cod):
                    continue
                mapa.setdefault(_norm(r.get("mpa_nombre", "")), []).append(int(cod))
        except Exception as e:
            print(f"[WARN] mpa_codigos.csv: {e}")

    cod2nom, sin_codigo = {}, []
    for n in nombres:
        cods = mapa.get(_norm(n))
        if not cods:
            sin_codigo.append(n)
            continue
        for c in cods:
            cod2nom.setdefault(c, n)   # un código puede venir de un SKU y de una agrupación
    data = (cod2nom, sin_codigo)
    _MPA_CACHE.update({"key": key, "data": data})
    return data


_DIAS_STOCK_VENTA_CACHE = {}


def _dias_stock_venta_base(vendedores):
    """({codigo: unidades}, meta) del MES ANTERIOR cerrado, SÓLO de los vendedores
    indicados (la ruta del depósito que se está midiendo).
    Fuente resuelta igual que el histórico semanal: cierre versionado → historial."""
    from calendar import monthrange
    hoy = datetime.now(_ARG_TZ).date()
    a, m = (hoy.year, hoy.month - 1) if hoy.month > 1 else (hoy.year - 1, 12)
    per = f"{a:04d}-{m:02d}"
    path, etiqueta = _semanal_fuente_de(per)
    vend = tuple(sorted(int(v) for v in vendedores))
    key = (per, str(path), _semanal_mtime(path), vend)
    cache = _DIAS_STOCK_VENTA_CACHE.get(key)
    if cache is not None:
        return cache

    ultimo = monthrange(a, m)[1]
    habiles = sum(1 for d in range(1, ultimo + 1)
                  if _es_dia_operativo(datetime(a, m, d).date()))
    unidades = {}
    df = _leer_ventas_min(path, ["Codigo", "CantBase", "FechaComprobante", "CodVendedor"])
    if not df.empty:
        df["cod"] = pd.to_numeric(df["Codigo"], errors="coerce")
        df["cv"]  = pd.to_numeric(df["CodVendedor"], errors="coerce")
        df["cant"] = _semanal_num(df["CantBase"])
        df["fec"] = _semanal_fechas(df["FechaComprobante"])
        df = df.dropna(subset=["cod", "fec"])
        df = df[(df["fec"].dt.strftime("%Y-%m") == per) & (df["cv"].isin(vend))]
        if not df.empty:
            unidades = {int(c): float(v) for c, v in
                        df.groupby("cod")["cant"].sum().items()}
    meta = {
        "periodo": per,
        "label": f"{_SEMANAL_MESES_ES.get(m, m)} {a}",
        "dias_habiles": habiles,
        "fuente": etiqueta,
        "sin_fuente": not bool(unidades),
    }
    data = (unidades, meta)
    if len(_DIAS_STOCK_VENTA_CACHE) > 8:
        _DIAS_STOCK_VENTA_CACHE.clear()
    _DIAS_STOCK_VENTA_CACHE[key] = data
    return data


def _dias_stock_filas(cod2etq, stock_map, unidades, habiles, desc_map):
    """Una fila por código del universo, ordenada por días de stock ascendente
    (primero lo que se queda sin stock).

    Cada fila lleva un `grupo` EXCLUYENTE, para que los contadores del resumen y las
    listas que se muestran coincidan exactamente (si "bajo 30" incluyera a los de
    "bajo 15", el mismo producto aparecería dos veces y los números no cerrarían):
      sin_stock  → sin existencia en el depósito (o el código no figura en el archivo)
      critico    → hay stock, alcanza para menos de _DIAS_STOCK_CRITICO días
      atencion   → hay stock, entre _DIAS_STOCK_CRITICO y _DIAS_STOCK_ATENCION días
      sin_venta  → hay stock pero el mes anterior no se vendió: no hay días que calcular
      ok         → por encima del umbral de atención
    """
    filas = []
    for cod, etq in cod2etq.items():
        st = stock_map.get(cod)
        vendido = unidades.get(cod, 0.0)
        diaria = (vendido / habiles) if habiles else 0.0
        dias = (st["disponible"] / diaria) if (st and diaria > 0) else None
        disponible = int(round(st["disponible"])) if st else None
        if st is None or (disponible or 0) <= 0:
            grupo = "sin_stock"
        elif dias is None:
            grupo = "sin_venta"
        elif dias < _DIAS_STOCK_CRITICO:
            grupo = "critico"
        elif dias < _DIAS_STOCK_ATENCION:
            grupo = "atencion"
        else:
            grupo = "ok"
        filas.append({
            "codigo":      int(cod),
            "etiqueta":    etq,
            "descripcion": (st or {}).get("descripcion") or desc_map.get(str(cod), {}).get("descripcion", ""),
            "en_stock":    st is not None,
            "disponible":  disponible,
            "transito":    int(round(st["transito"])) if st else None,
            "vendido_mes": int(round(vendido)),
            "venta_diaria": round(diaria, 1),
            "dias_stock":  round(dias, 1) if dias is not None else None,
            "grupo":       grupo,
        })
    # None al final: sin venta el mes pasado no es "0 días".
    filas.sort(key=lambda f: (f["dias_stock"] is None, f["dias_stock"] if f["dias_stock"] is not None else 0))
    return filas


def _dias_stock_resumen(filas, habiles):
    """Días de stock del conjunto: existencia total / venta diaria total del universo.
    Los contadores por grupo son excluyentes (ver `_dias_stock_filas`)."""
    disp = sum(f["disponible"] or 0 for f in filas)
    vend = sum(f["vendido_mes"] or 0 for f in filas)
    diaria = (vend / habiles) if habiles else 0.0
    n = lambda g: sum(1 for f in filas if f["grupo"] == g)
    return {
        "productos":   len(filas),
        "en_stock":    sum(1 for f in filas if f["en_stock"]),
        "disponible":  int(disp),
        "vendido_mes": int(vend),
        "venta_diaria": round(diaria, 1),
        "dias_stock":  round(disp / diaria, 1) if diaria > 0 else None,
        "sin_stock":   n("sin_stock"),
        "criticos":    n("critico"),
        "atencion":    n("atencion"),
        "sin_venta":   n("sin_venta"),
        "ok":          n("ok"),
        "en_riesgo":   n("sin_stock") + n("critico") + n("atencion"),
    }


def _dias_stock_bloque(cfg, desc_map, universos_base):
    """Un depósito: su stock, su ruta de vendedores y los 3 universos de producto."""
    stock = _stock_disponible(cfg["archivo"])
    stock_map = {}
    if not stock.empty:
        for _, r in stock.iterrows():
            c = int(r["codigo"])
            prev = stock_map.get(c)
            if prev:   # el mismo código puede venir en varias sedes/sectores
                prev["disponible"] += float(r["disponible"])
                prev["transito"]   += float(r["transito"])
            else:
                stock_map[c] = {"descripcion": str(r["descripcion"]),
                                "disponible": float(r["disponible"]),
                                "transito": float(r["transito"])}

    unidades, meta = _dias_stock_venta_base(cfg["vendedores"])
    habiles = meta["dias_habiles"]

    universos = []
    for uid, label, fuente, cod2etq, sin_codigo in universos_base:
        filas = _dias_stock_filas(cod2etq, stock_map, unidades, habiles, desc_map)
        universos.append({"id": uid, "label": label, "fuente": fuente,
                          "productos": filas,
                          "resumen": _dias_stock_resumen(filas, habiles),
                          "sin_codigo": sin_codigo})

    # Diagnóstico de la fuente de stock: si NINGÚN código del portfolio aparece en el
    # archivo, el export cargado no es el de Peñaflor. Se avisa explícito en vez de
    # dibujar una tabla de ceros que se leería como "no tenemos nada".
    codigos_universo = {f["codigo"] for u in universos for f in u["productos"]}
    match = len(codigos_universo & set(stock_map))
    return {
        "id": cfg["id"], "label": cfg["label"],
        "archivo": cfg["archivo"],
        "vendedores": [f"V{v}" for v in cfg["vendedores"]],
        "base": meta,
        "stock_codigos": len(stock_map),
        "stock_match": match,
        "stock_ok": bool(stock_map) and match > 0,
        "universos": universos,
    }


@app.route("/api/gerencia/dias_stock")
def gerencia_dias_stock():
    """Días de stock de 11 Titulares, Innovaciones y MPA contra la venta del mes
    anterior, por depósito (Stock PyP y VSB Cuyo, cada uno con su ruta)."""
    desc_map = _acc_desc_articulo_map()
    # Los universos de producto son los mismos para los dos depósitos: se arman una
    # sola vez y cada bloque los cruza con SU stock y SU venta.
    universos_base = []
    u11 = _codigos_11t_map()
    if u11:
        universos_base.append(("11t", "11 Titulares", "matriz oficial 11T (Código Art.)",
                               {int(c): str(mk) for c, mk in u11.items()}, []))
    inov = _innovaciones_codigos_todas()
    if inov:
        universos_base.append(("innovaciones", "Innovaciones", "INNOVACIONES/Innovaciones.xlsx",
                               {int(p["codigo"]): p["nombre"] for p in inov}, []))
    cod2nom, sin_codigo = _mpa_universo()
    if cod2nom or sin_codigo:
        universos_base.append(("mpa", "MPA · Plan AASS",
                               "MPA/MPA.xlsx × 09_CONFIG/mpa_codigos.csv",
                               cod2nom, sin_codigo))

    bloques = [_dias_stock_bloque(cfg, desc_map, universos_base) for cfg in _STOCK_BLOQUES]
    return jsonify({
        "generado_en": _now_ar(),
        "fuente": "01_INPUTS/Stock/<depósito>.xlsx + venta del mes anterior (unidades)",
        "umbrales": {"critico": _DIAS_STOCK_CRITICO, "atencion": _DIAS_STOCK_ATENCION},
        "base": bloques[0]["base"] if bloques else {},
        "bloques": bloques,
    })


# ====== 11 TITULARES: MATCH POR CÓDIGO DE ARTÍCULO (matriz oficial AS) ======
# Contrato de datos: 01_INPUTS/11 titulares autoservicio/11_titulares_autoservicios_match_codigos.xlsx
# (hoja DETALLE_SKU_11T_AS). El match por Código Art. exacto es la fuente PRIMARIA para asignar
# cada SKU a su marca titular; el match por texto de Marca queda como FALLBACK (todas las variedades
# de una marca suman a la misma marca — validado con el usuario 2026-07-06).
_COD11T_CACHE = {"mtime": None, "map": None}

def _codigos_11t_map():
    """codigo_articulo (int) -> marca_objetivo, desde la matriz oficial de 11 Titulares.
    Si el archivo no está, devuelve {} y el 11T cae al match por texto de Marca (comportamiento previo)."""
    p = INPUTS / "11 titulares autoservicio" / "11_titulares_autoservicios_match_codigos.xlsx"
    if not p.exists():
        return {}
    mt = p.stat().st_mtime
    if _COD11T_CACHE["map"] is not None and _COD11T_CACHE["mtime"] == mt:
        return _COD11T_CACHE["map"]
    _NORM = {"SMF ICE": "SMIRNOFF ICE"}   # nombre de línea en la matriz -> marca_objetivo canónica
    out = {}
    try:
        d = pd.read_excel(p, sheet_name="DETALLE_SKU_11T_AS")
        for _, r in d.iterrows():
            cod = pd.to_numeric(r.get("codigo_articulo"), errors="coerce")
            marca = str(r.get("linea_comercial_11t", "")).upper().strip()
            marca = _NORM.get(marca, marca)
            if pd.notna(cod) and marca and marca != "NAN":
                out[int(cod)] = marca
    except Exception as e:
        print(f"[AVISO] 11T códigos matriz: {e}")
        return {}
    _COD11T_CACHE["map"], _COD11T_CACHE["mtime"] = out, mt
    return out

def _marca_11t_por_codigo(df):
    """Serie marca_objetivo por match de Codigo exacto contra la matriz (NaN si el código no está)."""
    lk = _codigos_11t_map()
    if not lk or "Codigo" not in df.columns:
        return pd.Series(pd.NA, index=df.index, dtype="object")
    return pd.to_numeric(df["Codigo"], errors="coerce").map(lk)


# ── Superficie de medición del 11T (validado con el usuario 2026-07-06) ──
# El 11T se mide SOLO en Autoservicio + Almacén + Kiosco (antes sumaba todo canal, lo que
# inflaba el CCC con On Premise, Vinotecas, Away From Home, etc.). Criterio por Ramo/Subramo:
#   - Autoservicio : Ramo == AUTOSERVICIO  o  Subramo con "AUTOSERVICIO" (incluye
#                    "Autoservicio Tradicional" — autoservicios chicos; validado 2026-07-06).
#                    NO entra Cash&Carry ni Mayoristas (Subramo Mayoristas).
#   - Almacén      : Subramo Almacén/Despensa  (o Ramo ALMACENES)
#   - Kiosco       : Subramo Kiosco/Maxikiosco
# Queda EXCLUIDO: On Premise, Vinotecas, Away From Home, Mayoristas, Cash&Carry, Fiambrería,
# Carnicería, Panadería y resto de tradicionales sin formato self-service.
def _mask_superficie_11t(df):
    """Máscara booleana de filas cuya superficie mide para el 11T (AS + Almacén + Kiosco)."""
    # Fail-open: si la fuente no trae Ramo ni Subramo no se puede clasificar; se deja pasar
    # todo (no filtrar) en vez de anular el CCC por columnas ausentes.
    if "Ramo" not in df.columns and "Subramo" not in df.columns:
        return pd.Series(True, index=df.index)
    ramo = (df["Ramo"].astype(str).str.upper().str.strip()
            if "Ramo" in df.columns else pd.Series("", index=df.index))
    sub  = (df["Subramo"].astype(str).str.upper()
            if "Subramo" in df.columns else pd.Series("", index=df.index))
    es_as  = (ramo == "AUTOSERVICIO") | sub.str.contains("AUTOSERVICIO", regex=False, na=False)
    es_alm = (sub.str.contains("ALMACEN", regex=False, na=False)
              | sub.str.contains("DESPENSA", regex=False, na=False)
              | ramo.str.contains("ALMACEN", regex=False, na=False))
    es_kio = (sub.str.contains("KIOSCO", regex=False, na=False)
              | sub.str.contains("MAXIKIOSCO", regex=False, na=False))
    return es_as | es_alm | es_kio


# ====== GERENCIA: 11 TITULARES POR MARCA ======
@app.route("/api/gerencia/once_titulares")
def gerencia_once_titulares():
    """11 Titulares: CCC acumulado vs objetivo CCC.
    REGLA:
      - Período = TRIMESTRE calendario en curso (ene-mar / abr-jun / jul-sep / oct-dic);
        en julio arranca de cero. Se filtra por FechaComprobante >= inicio del trimestre.
      - NO se filtra por Empresa (corregido 2026-07-13 contra reporte de Peñaflor). P&P
        Logística es nuestra segunda razón social, no otro distribuidor: Proveedor =
        GRUPO PEÑAFLOR SA en el 100% de las filas. El filtro Empresa=='Empresa' anterior
        borraba 135 de 229 clientes con compra (V6 y V10 perdían el 88% de su cartera) y
        dejaba el CCC en la mitad del que reporta Peñaflor.
      - CCC = clientes únicos con compra válida (neto>0) por marca titular; excluye V2/V5/V20.
    Fuente primaria : ventas_acumulada.csv.
    Fuente fallback : mod_11_titulares.csv  (tiene_flag = 1).
    Fuente objetivo : objetivo 11T.xlsx."""

    # ── Brand alias lookup ──
    _MARCA_LOOKUP = {
        "ALMA MORA": "ALMA MORA", "ALARIS": "ALARIS", "TRAPICHE ALARIS": "ALARIS",
        "DON DAVID": "DON DAVID", "DADA": "DADA", "LOS ARBOLES": "LOS ARBOLES",
        "FINCA LAS MORAS": "FINCA LAS MORAS", "F LAS MORAS": "FINCA LAS MORAS",
        "TRAPICHE RESERVA": "TRAPICHE RESERVA",
        "FOND DE CAVE": "FOND DE CAVE", "FOND CAVE": "FOND DE CAVE",
        "CAZADOR": "CAZADOR", "ANTARES": "ANTARES",
        "GORDON'S FLAVOURS": "GORDON'S FLAVOURS", "GORDONS FLAVOURS": "GORDON'S FLAVOURS",
        "GORDON'S": "GORDON'S FLAVOURS", "GORDONS": "GORDON'S FLAVOURS",
        "GORDON S": "GORDON'S FLAVOURS",
        "SMIRNOFF": "SMIRNOFF FLAVOURS",
        "SMIRNOFF FLAVOURS": "SMIRNOFF FLAVOURS",
        "SMIRNOFF ICE FLAVOURS": "SMIRNOFF ICE",
        "SMIRNOFF ICE": "SMIRNOFF ICE",
        "JW": "JW BLACK", "JW BLACK": "JW BLACK", "JW RED": "JW RED",
        "MASCOTA": "MASCOTA", "LA MASCOTA": "MASCOTA",
        "NC ESPUMANTES": "NC ESPUMANTES", "NAVARRO CORREAS": "NC ESPUMANTES",
        "TRAPICHE MEDALLA": "TRAPICHE MEDALLA", "GRAN MEDALLA": "TRAPICHE MEDALLA",
    }
    _ART_KW = [
        ("SMIRNOFF ICE", "SMIRNOFF ICE"), ("SMF ICE", "SMIRNOFF ICE"),
        ("SMIRNOFF", "SMIRNOFF FLAVOURS"), ("GORDON", "GORDON'S FLAVOURS"),
        ("ANTARES", "ANTARES"), ("CAZADOR", "CAZADOR"),
        ("FOND DE CAVE", "FOND DE CAVE"), ("ALMA MORA", "ALMA MORA"),
        ("LOS ARBOLES", "LOS ARBOLES"), ("DADA", "DADA"),
        ("FINCA LAS MORAS", "FINCA LAS MORAS"), ("F.LAS MORAS", "FINCA LAS MORAS"),
        ("DON DAVID", "DON DAVID"), ("ALARIS", "ALARIS"),
        ("TRAPICHE RESERVA", "TRAPICHE RESERVA"),
        ("JW BLACK", "JW BLACK"), ("JW RED", "JW RED"),
    ]

    ccc_map = {}
    ccc_dep_map = {}   # CCC del Depósito (V20), bloque aparte sin objetivo
    fuente = ""

    # ── Fuente primaria: ventas_acumulada.csv ──
    vac_path = INPUTS / "ventas_acumulada.csv"
    if vac_path.exists():
        try:
            vac = pd.read_csv(vac_path, sep=";", encoding="latin1", low_memory=False)
            vac["ImporteNetoItem"] = pd.to_numeric(
                vac["ImporteNetoItem"].astype(str).str.replace(",", ".", regex=False), errors="coerce")
            # Ruta: excluye V1/V2/V5, neto>0. No se filtra por Empresa (ver docstring).
            # V20 (Depósito) se CONSERVA para su CCC aparte.
            vac = vac[~vac["CodVendedor"].isin(_VENDEDORES_EXCLUIDOS - {20})]
            vac = vac[vac["ImporteNetoItem"] > 0]
            vac = vac[_mask_superficie_11t(vac)]   # 11T mide solo AS + Almacén + Kiosco
            # Período = trimestre calendario en curso (en julio arranca de cero)
            _f = pd.to_datetime(vac.get("FechaComprobante"), dayfirst=True, errors="coerce")
            if _f.notna().any():
                _hoy = datetime.now(_ARG_TZ)
                _ini_trim = pd.Timestamp(_hoy.year, ((_hoy.month - 1) // 3) * 3 + 1, 1)
                vac = vac[_f >= _ini_trim]
            vac["marca_upper"] = vac["Marca"].astype(str).str.upper().str.strip()
            # 1) match por Código Art. exacto (matriz oficial) — fuente primaria
            vac["marca_objetivo"] = _marca_11t_por_codigo(vac)
            # 2) fallback: texto de Marca (para variedades fuera de la matriz)
            _falta = vac["marca_objetivo"].isna()
            vac.loc[_falta, "marca_objetivo"] = vac.loc[_falta, "marca_upper"].map(_MARCA_LOOKUP)
            unresolved = vac["marca_objetivo"].isna()
            if unresolved.any():
                for kw, mo in _ART_KW:
                    still = vac["marca_objetivo"].isna() & unresolved
                    if not still.any():
                        break
                    hits = vac.loc[still, "Articulo"].astype(str).str.upper().str.contains(kw, regex=False, na=False)
                    vac.loc[still & hits, "marca_objetivo"] = mo
            _es_dep = vac["CodVendedor"] == 20
            ccc_map = (vac[~_es_dep & vac["marca_objetivo"].notna()]
                       .groupby("marca_objetivo")["Cliente"]
                       .nunique().to_dict())
            ccc_dep_map = (vac[_es_dep & vac["marca_objetivo"].notna()]
                           .groupby("marca_objetivo")["Cliente"]
                           .nunique().to_dict())
            fuente = "ventas_acumulada.csv"
        except Exception:
            ccc_map = {}

    # ── Fuente fallback: mod_11_titulares.csv (si la primaria no está disponible) ──
    if not ccc_map:
        t11_path = DATASETS / "mod_11_titulares.csv"
        if t11_path.exists():
            try:
                t11 = pd.read_csv(t11_path, sep=",", encoding="utf-8-sig", low_memory=False)
                t11["vendedor_codigo"] = pd.to_numeric(t11["vendedor_codigo"], errors="coerce")
                t11 = t11[~t11["vendedor_codigo"].isin(_VENDEDORES_EXCLUIDOS)]
                t11_ok = t11[t11["tiene_flag"].astype(str) == "1"]
                ccc_map = t11_ok.groupby("marca_objetivo")["cliente_id"].nunique().to_dict()
                fuente = "mod_11_titulares.csv (parcial)"
            except Exception:
                ccc_map = {}

    if not ccc_map:
        return jsonify({"error": "Sin fuente de CCC disponible", "marcas": [], "fuente": "N/A"}), 200

    # ── Objetivos: objetivo 11T.xlsx ──
    _OBJ_ALIAS = {
        "ALMA MORA": "ALMA MORA", "TRAPICHE RESERVA": "TRAPICHE RESERVA",
        "FINCA LAS MORAS": "FINCA LAS MORAS",
        "ALARIS": "ALARIS", "DON DAVID": "DON DAVID", "DADA": "DADA",
        "SIMRNOFF FLAVORS": "SMIRNOFF FLAVOURS", "SMIRNOFF FLAVORS": "SMIRNOFF FLAVOURS",
        "SMIRNOFF FLAVOURS": "SMIRNOFF FLAVOURS",
        "LOS ARBOLES": "LOS ARBOLES", "ANTARES": "ANTARES",
        "SMIRNOFF ICE": "SMIRNOFF ICE", "SMF ICE": "SMIRNOFF ICE",
        "GORDONS FLAVOURS": "GORDON'S FLAVOURS", "GORDONS FLAVORS": "GORDON'S FLAVOURS",
        "GORDON'S FLAVOURS": "GORDON'S FLAVOURS",
    }
    obj_map = {}
    try:
        obj_df = pd.read_excel(INPUTS / "objetivo 11T.xlsx", header=1)
        obj_df = obj_df.dropna(subset=obj_df.columns[1:2])
        for _, row in obj_df.iterrows():
            raw = str(row.iloc[1]).upper().strip()
            marca_key = _OBJ_ALIAS.get(raw, raw)
            try:
                obj_map[marca_key] = int(float(row.iloc[2]))
            except (ValueError, TypeError):
                pass
    except Exception:
        pass

    # ── Resultado: por marca, ordenado por CCC desc ──
    marcas = []
    for marca_obj in sorted(obj_map.keys(), key=lambda x: ccc_map.get(x, 0), reverse=True):
        ccc = int(ccc_map.get(marca_obj, 0))
        obj = obj_map[marca_obj]
        pct = round(ccc / obj * 100, 1) if obj else None
        # Depósito (V20): CCC logrado aparte; no suma al objetivo ni al avance de ruta.
        ccc_dep = int(ccc_dep_map.get(marca_obj, 0))
        marcas.append({"marca": marca_obj, "ccc": ccc, "objetivo_ccc": obj,
                       "pct_objetivo": pct, "ccc_deposito": ccc_dep})

    return jsonify({
        "generado_en": _now_ar(),
        "fuente": fuente,
        "total_marcas": len(marcas),
        "marcas": marcas,
        "ccc_deposito_total": int(sum(ccc_dep_map.values())),
    })


# ====== GERENCIA: 11 TITULARES CCC ZONA DEL DÍA ======
@app.route("/api/gerencia/once_titulares_zona")
def gerencia_once_titulares_zona():
    """11 Titulares: CCC por marca para los clientes de la zona del día.
    Sin objetivos — muestra penetración dentro de la zona.
    Fuente ventas : ventas_acumulada.csv
    Fuente zona   : clientes.xlsx (DiasVisita == dia)"""
    dia_raw = request.args.get("dia", "").strip()
    dia_key = dia_raw.lower()[:2] if dia_raw else ""   # "LU"→"lu", "Ma"→"ma"

    _MARCA_LOOKUP = {
        "ALMA MORA":"ALMA MORA","ALARIS":"ALARIS","TRAPICHE ALARIS":"ALARIS",
        "DON DAVID":"DON DAVID","DADA":"DADA","LOS ARBOLES":"LOS ARBOLES",
        "FINCA LAS MORAS":"FINCA LAS MORAS","F LAS MORAS":"FINCA LAS MORAS",
        "TRAPICHE RESERVA":"TRAPICHE RESERVA",
        "FOND DE CAVE":"FOND DE CAVE","FOND CAVE":"FOND DE CAVE",
        "CAZADOR":"CAZADOR","ANTARES":"ANTARES",
        "GORDON'S FLAVOURS":"GORDON'S FLAVOURS","GORDONS FLAVOURS":"GORDON'S FLAVOURS",
        "GORDON'S":"GORDON'S FLAVOURS","GORDONS":"GORDON'S FLAVOURS","GORDON S":"GORDON'S FLAVOURS",
        "SMIRNOFF":"SMIRNOFF FLAVOURS","SMIRNOFF FLAVOURS":"SMIRNOFF FLAVOURS",
        "SMIRNOFF ICE FLAVOURS":"SMIRNOFF ICE","SMIRNOFF ICE":"SMIRNOFF ICE",
        "JW":"JW BLACK","JW BLACK":"JW BLACK","JW RED":"JW RED",
        "MASCOTA":"MASCOTA","LA MASCOTA":"MASCOTA",
        "NC ESPUMANTES":"NC ESPUMANTES","NAVARRO CORREAS":"NC ESPUMANTES",
        "TRAPICHE MEDALLA":"TRAPICHE MEDALLA","GRAN MEDALLA":"TRAPICHE MEDALLA",
    }
    _ART_KW = [
        ("SMIRNOFF ICE","SMIRNOFF ICE"),("SMF ICE","SMIRNOFF ICE"),
        ("SMIRNOFF","SMIRNOFF FLAVOURS"),("GORDON","GORDON'S FLAVOURS"),
        ("ANTARES","ANTARES"),("CAZADOR","CAZADOR"),
        ("FOND DE CAVE","FOND DE CAVE"),("ALMA MORA","ALMA MORA"),
        ("LOS ARBOLES","LOS ARBOLES"),("DADA","DADA"),
        ("FINCA LAS MORAS","FINCA LAS MORAS"),("F.LAS MORAS","FINCA LAS MORAS"),
        ("DON DAVID","DON DAVID"),("ALARIS","ALARIS"),
        ("TRAPICHE RESERVA","TRAPICHE RESERVA"),
        ("JW BLACK","JW BLACK"),("JW RED","JW RED"),
    ]
    _OBJ_ALIAS = {
        "ALMA MORA":"ALMA MORA","TRAPICHE RESERVA":"TRAPICHE RESERVA",
        "FINCA LAS MORAS":"FINCA LAS MORAS","ALARIS":"ALARIS",
        "DON DAVID":"DON DAVID","DADA":"DADA",
        "SIMRNOFF FLAVORS":"SMIRNOFF FLAVOURS","SMIRNOFF FLAVORS":"SMIRNOFF FLAVOURS",
        "SMIRNOFF FLAVOURS":"SMIRNOFF FLAVOURS","LOS ARBOLES":"LOS ARBOLES",
        "ANTARES":"ANTARES","SMIRNOFF ICE":"SMIRNOFF ICE","SMF ICE":"SMIRNOFF ICE",
        "GORDONS FLAVOURS":"GORDON'S FLAVOURS","GORDONS FLAVORS":"GORDON'S FLAVOURS",
        "GORDON'S FLAVOURS":"GORDON'S FLAVOURS",
    }

    # ── Ventas acumuladas ──
    vac_path = INPUTS / "ventas_acumulada.csv"
    if not vac_path.exists():
        return jsonify({"error": "ventas_acumulada.csv no encontrado", "marcas": [], "dia": dia_raw}), 200
    try:
        vac = pd.read_csv(vac_path, sep=";", encoding="latin1", low_memory=False)
        vac["ImporteNetoItem"] = pd.to_numeric(
            vac["ImporteNetoItem"].astype(str).str.replace(",",".",regex=False), errors="coerce")
        # No se filtra por Empresa — igual criterio que /api/gerencia/once_titulares
        vac = vac[~vac["CodVendedor"].isin(_VENDEDORES_EXCLUIDOS)]
        vac = vac[vac["ImporteNetoItem"] > 0]
        vac = vac[_mask_superficie_11t(vac)]   # 11T mide solo AS + Almacén + Kiosco
        # Período = trimestre calendario en curso (en julio arranca de cero)
        _f = pd.to_datetime(vac.get("FechaComprobante"), dayfirst=True, errors="coerce")
        if _f.notna().any():
            _hoy = datetime.now(_ARG_TZ)
            _ini_trim = pd.Timestamp(_hoy.year, ((_hoy.month - 1) // 3) * 3 + 1, 1)
            vac = vac[_f >= _ini_trim]
    except Exception as e:
        return jsonify({"error": str(e), "marcas": [], "dia": dia_raw}), 200

    # ── Clientes de la zona del día ──
    zona_ids = set()
    cli_total = 0
    try:
        cli_path = INPUTS / "clientes.xlsx"
        if cli_path.exists():
            cli_df = pd.read_excel(cli_path)
            dias_col = next((c for c in cli_df.columns if "diasvisita" in c.lower().replace(" ","")), None)
            cod_col  = next((c for c in cli_df.columns if c.lower() in ("codigo","cod","id")), None)
            vend_col = next((c for c in cli_df.columns if c.lower() == "codven"), None)
            if dias_col and cod_col:
                if vend_col:
                    cli_df = cli_df[~pd.to_numeric(cli_df[vend_col], errors="coerce").isin(_VENDEDORES_EXCLUIDOS)]
                if dia_key:
                    cli_df = cli_df[cli_df[dias_col].astype(str).str.strip().str.lower() == dia_key]
                zona_ids = set(pd.to_numeric(cli_df[cod_col], errors="coerce").dropna().astype(int))
                cli_total = len(zona_ids)
    except Exception:
        pass

    # ── Filtrar ventas a zona ──
    if zona_ids:
        vac = vac[vac["Cliente"].isin(zona_ids)]

    # ── Normalizar marcas ──
    vac["marca_upper"] = vac["Marca"].astype(str).str.upper().str.strip()
    # 1) match por Código Art. exacto (matriz oficial) — fuente primaria
    vac["marca_objetivo"] = _marca_11t_por_codigo(vac)
    # 2) fallback: texto de Marca (para variedades fuera de la matriz)
    _falta = vac["marca_objetivo"].isna()
    vac.loc[_falta, "marca_objetivo"] = vac.loc[_falta, "marca_upper"].map(_MARCA_LOOKUP)
    unresolved = vac["marca_objetivo"].isna()
    if unresolved.any():
        for kw, mo in _ART_KW:
            still = vac["marca_objetivo"].isna() & unresolved
            if not still.any():
                break
            hits = vac.loc[still, "Articulo"].astype(str).str.upper().str.contains(kw, regex=False, na=False)
            vac.loc[still & hits, "marca_objetivo"] = mo

    # CCC por marca en la zona
    ccc_map = (vac[vac["marca_objetivo"].notna()]
               .groupby("marca_objetivo")["Cliente"]
               .nunique().to_dict())

    # Orden oficial desde objetivo 11T.xlsx
    marcas_orden = []
    try:
        obj_df = pd.read_excel(INPUTS / "objetivo 11T.xlsx", header=1)
        obj_df = obj_df.dropna(subset=obj_df.columns[1:2])
        for _, row in obj_df.iterrows():
            raw = str(row.iloc[1]).upper().strip()
            mk = _OBJ_ALIAS.get(raw, raw)
            if mk and mk not in marcas_orden:
                marcas_orden.append(mk)
    except Exception:
        pass
    if not marcas_orden:
        marcas_orden = sorted(ccc_map.keys(), key=lambda x: ccc_map.get(x, 0), reverse=True)

    marcas = [{"marca": m, "ccc": ccc_map.get(m, 0)} for m in marcas_orden]

    return jsonify({
        "generado_en": _now_ar(),
        "dia": dia_raw.upper() if dia_raw else "TODOS",
        "total_clientes_zona": cli_total,
        "fuente": "ventas_acumulada.csv",
        "marcas": marcas,
    })


# ====== GERENCIA: RANKING RECHAZOS ======
@app.route("/api/gerencia/ranking_rechazos")
def gerencia_ranking_rechazos():
    """Ranking de % rechazo por vendedor.
    Fuente: 01_INPUTS/resultado.xlsx hoja Rechazos.
    Filtra: solo Origen == Vendedor, excluye V2/V5/V20."""
    try:
        df = pd.read_excel(INPUTS / "resultado.xlsx", sheet_name="Rechazos")
    except Exception as e:
        return jsonify({"error": str(e), "vendedores": []}), 200

    df.columns = [c.strip() for c in df.columns]

    # Supervisores: filas Origen == Supervisor (rechazo total por supervisor)
    supervisores = []
    if "Origen" in df.columns and "SupervisorNombre" in df.columns:
        sup = df[df["Origen"].astype(str).str.strip().str.lower() == "supervisor"]
        for _, row in sup.iterrows():
            nom_full = str(row.get("SupervisorNombre", "")).strip()
            if not nom_full or nom_full.lower() == "nan":
                continue
            pct = round(float(pd.to_numeric(row.get("PorcRechazo"), errors="coerce") or 0), 1)
            nombre_corto = nom_full.title().split()[-1] if nom_full.split() else nom_full.title()
            supervisores.append({
                "supervisor_nombre": nom_full.title(),
                "nombre":            nombre_corto,   # nombre de pila (ej. Esteban, Raul)
                "rechazo_pct":       pct,
            })
        supervisores.sort(key=lambda x: x["rechazo_pct"], reverse=True)

    # Solo vendedores (excluir filas de supervisores)
    if "Origen" in df.columns:
        df = df[df["Origen"].astype(str).str.strip().str.lower() == "vendedor"]

    df["VendedorCodigo"] = pd.to_numeric(df["VendedorCodigo"], errors="coerce")
    df = df[~df["VendedorCodigo"].isin(_VENDEDORES_EXCLUIDOS)]
    df["PorcRechazo"] = pd.to_numeric(df["PorcRechazo"], errors="coerce").fillna(0)

    resultado = []
    for _, row in df.iterrows():
        cod = int(row["VendedorCodigo"])
        nombre = str(row.get("VendedorNombre", f"V{cod}")).strip()
        pct = round(float(row["PorcRechazo"]), 1)
        resultado.append({
            "vendedor_id":     f"V{cod}",
            "vendedor_nombre": nombre,
            "rechazo_pct":     pct,
        })

    resultado.sort(key=lambda x: x["rechazo_pct"], reverse=True)
    return jsonify({
        "generado_en":  _now_ar(),
        "fuente":       "resultado.xlsx · hoja Rechazos",
        "supervisores": supervisores,
        "vendedores":   resultado,
    })


# ====== GERENCIA: 11T RESUMEN DISTRIBUIDORA ======
@app.route("/api/gerencia/11t_empresa")
def gerencia_11t_empresa():
    """11T distribuidora: por marca, clientes con/sin cada titular (todos los vendedores).
    Fuente: mod_11_titulares.csv  — columnas tiene_flag / falta_flag."""
    df = read_csv(DATASETS / "mod_11_titulares.csv")
    if df.empty:
        return jsonify({"marcas": [], "error": "Sin datos en mod_11_titulares.csv"}), 200

    df.columns = [c.lstrip("﻿").strip() for c in df.columns]
    df["vendedor_codigo"] = pd.to_numeric(df["vendedor_codigo"], errors="coerce")
    df = df[~df["vendedor_codigo"].isin(_VENDEDORES_EXCLUIDOS)]

    for col in ("tiene_flag", "falta_flag"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    vendedores_activos = sorted(df["vendedor_codigo"].dropna().unique())

    result = []
    for marca, grp in df.groupby("marca_objetivo"):
        con   = int(grp["tiene_flag"].sum())
        sin   = int(grp["falta_flag"].sum())
        total = con + sin
        pct   = round(con / total * 100, 1) if total else 0.0

        por_vendedor = {}
        for v in vendedores_activos:
            vg  = grp[grp["vendedor_codigo"] == v]
            vc  = int(vg["tiene_flag"].sum())
            vs  = int(vg["falta_flag"].sum())
            vt  = vc + vs
            por_vendedor[f"V{int(v)}"] = {
                "con": vc, "sin": vs, "total": vt,
                "pct": round(vc / vt * 100, 1) if vt else 0.0,
            }

        result.append({
            "marca":        str(marca),
            "con":          con,
            "sin":          sin,
            "total":        total,
            "pct":          pct,
            "por_vendedor": por_vendedor,
        })

    result.sort(key=lambda x: x["pct"], reverse=True)
    return jsonify({
        "generado_en": _now_ar(),
        "fuente":         "mod_11_titulares.csv",
        "total_marcas":   len(result),
        "vendedores":     [f"V{int(v)}" for v in vendedores_activos],
        "marcas":         result,
    })


# ====== GERENCIA: 11T POR VENDEDOR ======
@app.route("/api/gerencia/11t_vendedor")
def gerencia_11t_vendedor():
    """11T detalle de un vendedor específico.
    Parámetro: vendedor=V3 (o 3). Fuente: mod_11_titulares.csv."""
    import re as _re
    vend_raw = request.args.get("vendedor", "").strip().upper()
    m = _re.search(r"\d+", vend_raw)
    if not m:
        return jsonify({"marcas": [], "error": "Parámetro vendedor inválido"}), 200
    cod = int(m.group())

    df = read_csv(DATASETS / "mod_11_titulares.csv")
    if df.empty:
        return jsonify({"marcas": [], "error": "Sin datos"}), 200

    df.columns = [c.lstrip("﻿").strip() for c in df.columns]
    df["vendedor_codigo"] = pd.to_numeric(df["vendedor_codigo"], errors="coerce")
    df = df[df["vendedor_codigo"] == cod]

    if df.empty:
        return jsonify({"marcas": [], "vendedor": f"V{cod}", "error": "Sin datos para este vendedor"}), 200

    for col in ("tiene_flag", "falta_flag"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    nombre = str(df["vendedor_nombre"].iloc[0]) if "vendedor_nombre" in df.columns else f"V{cod}"

    result = []
    for marca, grp in df.groupby("marca_objetivo"):
        con   = int(grp["tiene_flag"].sum())
        sin   = int(grp["falta_flag"].sum())
        total = con + sin
        pct   = round(con / total * 100, 1) if total else 0.0
        result.append({
            "marca": str(marca), "con": con, "sin": sin, "total": total, "pct": pct
        })

    result.sort(key=lambda x: x["pct"], reverse=True)
    return jsonify({
        "generado_en": _now_ar(),
        "fuente":           "mod_11_titulares.csv",
        "vendedor":         f"V{cod}",
        "vendedor_nombre":  nombre,
        "total_marcas":     len(result),
        "marcas":           result,
    })


# ====== GERENCIA: COBERTURA POR SEGMENTO ======
@app.route("/api/gerencia/cobertura_segmento")
def gerencia_cobertura_segmento():
    """Cobertura acumulada del mes por vendedor × segmento.
    Fuente: clientes_dia.csv (cobertura_mes_flag).
    Excluye V2, V5, V20. V3 sin AUTOSERVICIO."""
    df = read_csv(DATASETS / "clientes_dia.csv")
    if df.empty:
        return jsonify({"generado_en": _now_ar(), "vendedores": []})

    df.columns = [c.lstrip("﻿") for c in df.columns]
    for col in ["vendedor_codigo", "cobertura_mes_flag"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df[~df["vendedor_codigo"].isin(_VENDEDORES_EXCLUIDOS)]
    # V3 no trabaja autoservicio ni on premise
    _segu = df["segmento_operativo"].astype(str).str.upper()
    mask_v3 = (df["vendedor_codigo"] == 3) & (_segu.isin(["AUTOSERVICIO", "ON_PREMISE_VTK"]))
    df = df[~mask_v3]

    if "segmento_operativo" not in df.columns or "cobertura_mes_flag" not in df.columns:
        return jsonify({"error": "Columnas esperadas no encontradas", "columnas": list(df.columns)}), 500

    resultado = []
    for (cod, seg), grp in df.groupby(["vendedor_codigo", "segmento_operativo"]):
        cod_int = int(cod) if pd.notnull(cod) else 0
        total      = len(grp)
        con_cob    = int(grp["cobertura_mes_flag"].fillna(0).sum())
        pct        = round(con_cob / total * 100, 1) if total else 0
        nombre_col = str(grp.iloc[0].get("vendedor_nombre", f"V{cod_int}")) if "vendedor_nombre" in grp.columns else f"V{cod_int}"
        resultado.append({
            "vendedor_id":      f"V{cod_int}",
            "vendedor_nombre":  nombre_col,
            "segmento":         str(seg),
            "total_clientes":   total,
            "con_cobertura":    con_cob,
            "pct_cobertura":    pct,
        })

    resultado.sort(key=lambda x: (x["vendedor_id"], x["segmento"]))
    return jsonify({
        "generado_en": _now_ar(),
        "cobertura": resultado,
    })


# ====== GERENCIA: REAL AYER POR SEGMENTO ======
@app.route("/api/gerencia/real_ayer_segmento")
def gerencia_real_ayer_segmento():
    """Real del día anterior por vendedor × segmento.
    Fuente: mod_ccc_segmento.csv. Vendedores sin venta ayer aparecen con cero.
    V3 sin AUTOSERVICIO."""
    ccc_df = read_csv(DATASETS / "mod_ccc_segmento.csv")
    vend   = read_csv(CONFIG / "vendedores_activos.csv")

    ccc_df.columns = [c.lstrip("﻿") for c in ccc_df.columns] if not ccc_df.empty else []
    for col in ["vendedor_codigo", "clientes_con_compra", "coberturas_logradas", "botellas_vendidas", "venta_neta"]:
        if not ccc_df.empty and col in ccc_df.columns:
            ccc_df[col] = pd.to_numeric(ccc_df[col], errors="coerce")

    # Construir índice: (cod_int, segmento) → datos
    ccc_idx = {}
    if not ccc_df.empty:
        for _, row in ccc_df.iterrows():
            cod_int = int(row["vendedor_codigo"]) if pd.notnull(row.get("vendedor_codigo")) else 0
            seg     = str(row.get("segmento_operativo", "")).upper()
            ccc_idx[(cod_int, seg)] = {
                "clientes_con_compra": int(row.get("clientes_con_compra", 0) or 0),
                "coberturas_logradas": int(row.get("coberturas_logradas", 0) or 0),
                "botellas_vendidas":   int(row.get("botellas_vendidas", 0) or 0),
                "venta_neta":          round(float(row.get("venta_neta", 0) or 0), 2),
            }

    # PROXIMITY (estaciones de servicio) es canal propio desde 2026-07-30: si no está acá,
    # la venta de esos clientes no aparece en ninguna fila y el total del vendedor no cierra.
    SEGMENTOS_POSIBLES = ["TRADICIONAL", "AUTOSERVICIO", "ON_PREMISE_VTK", "PROXIMITY"]
    _cero = {"clientes_con_compra": 0, "coberturas_logradas": 0, "botellas_vendidas": 0, "venta_neta": 0.0}

    resultado = []
    if not vend.empty:
        for _, v in vend[vend["activo"] == 1].iterrows():
            cn      = clean_code(str(v["codigo_vendedor"]))
            cod_int = int(cn) if cn.isdigit() else 0
            if cod_int in _VENDEDORES_EXCLUIDOS:
                continue
            nombre = str(v["nombre_vendedor"]).strip()
            es_v3  = (cod_int == 3)
            segs   = []
            for seg in SEGMENTOS_POSIBLES:
                if es_v3 and seg == "AUTOSERVICIO":
                    continue
                datos = ccc_idx.get((cod_int, seg), _cero).copy()
                datos["segmento"] = seg
                segs.append(datos)
            resultado.append({
                "vendedor_id":      f"V{cn}",
                "vendedor_nombre":  nombre,
                "venta_total_ayer": round(sum(s["venta_neta"] for s in segs), 2),
                "segmentos":        segs,
            })

    return jsonify({
        "generado_en": _now_ar(),
        "vendedores":  resultado,
    })


# ====== GERENCIA: PLANES AUTOSERVICIO ======
@app.route("/api/gerencia/planes_autoservicio")
def gerencia_planes_autoservicio():
    """Acciones en canal AUTOSERVICIOS con gasto real vs teórico.
    ADVERTENCIA: no existe fuente de 'planes formales' ni sin cargos detallados.
    Esta respuesta expone mod_gastos_accion.csv filtrado por canal AUTOSERVICIOS.
    V3 excluida. Excluye V2, V5, V20."""
    df = read_csv(DATASETS / "mod_gastos_accion.csv")
    if df.empty:
        return jsonify({
            "fuente_limitada": True,
            "advertencia": "mod_gastos_accion.csv no encontrado o vacío.",
            "acciones": [],
        })

    df.columns = [c.lstrip("﻿") for c in df.columns]
    for col in ["vendedor_codigo", "clientes_afectados", "lineas_alertadas",
                "gasto_real_total", "gasto_teorico_total", "exceso_pesos_total", "exceso_pct_promedio"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Excluir V3 (no trabaja AS), V2, V5, V20
    excluir = _VENDEDORES_EXCLUIDOS | {3}
    df = df[~df["vendedor_codigo"].isin(excluir)]

    # Filtrar canal AUTOSERVICIOS
    if "canal" in df.columns:
        df = df[df["canal"].astype(str).str.upper().str.contains("AUTOSERVICIO", na=False)]

    acciones = df.where(pd.notnull(df), None).to_dict(orient="records")

    return jsonify({
        "fuente_limitada": True,
        "advertencia": (
            "Sin cargos formales no disponibles: no existe tabla de sin cargos en este sistema. "
            "Se muestran únicamente acciones con canal AUTOSERVICIOS de mod_gastos_accion.csv."
        ),
        "total_acciones": len(acciones),
        "acciones": acciones,
    })


# ====== GERENCIA: INNOVACIONES POR SEGMENTO ======
@app.route("/api/gerencia/innovaciones_segmento")
def gerencia_innovaciones_segmento():
    """Frizze Manxana (14620) y Antares XPA (60020) por segmento, empresa completa.
    Fuente: mod_innovaciones_segmento.csv. Excluye V2, V5, V20 y V3/AUTOSERVICIO."""
    df = read_csv(DATASETS / "mod_innovaciones_segmento.csv")
    if df.empty:
        return jsonify({
            "generado_en": _now_ar(),
            "fuente": "mod_innovaciones_segmento.csv",
            "advertencia": "Dato no disponible",
            "resumen_empresa": [], "por_vendedor": [],
        })
    df.columns = [c.lstrip("﻿") for c in df.columns]
    for col in ["vendedor_codigo", "producto_codigo", "clientes_cartera", "clientes_compraron", "pct_cobertura"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df[~df["vendedor_codigo"].isin(_VENDEDORES_EXCLUIDOS)]
    df = df[~((df["vendedor_codigo"] == 3) & (df["segmento"].astype(str).str.upper() == "AUTOSERVICIO"))].copy()
    fecha_ej = str(df["fecha_ejecucion"].iloc[0]) if "fecha_ejecucion" in df.columns and not df.empty else None
    agg = (df.groupby(["producto_codigo", "producto_nombre", "segmento"], dropna=False)
             .agg(clientes_cartera=("clientes_cartera", "sum"),
                  clientes_compraron=("clientes_compraron", "sum"))
             .reset_index())
    resumen_empresa = []
    for _, row in agg.iterrows():
        cartera = int(row["clientes_cartera"]); compraron = int(row["clientes_compraron"])
        resumen_empresa.append({
            "producto_codigo": int(row["producto_codigo"]), "producto_nombre": str(row["producto_nombre"]),
            "segmento": str(row["segmento"]), "clientes_cartera": cartera,
            "clientes_compraron": compraron,
            "pct_cobertura": round(compraron / cartera, 4) if cartera else 0.0,
        })
    resumen_empresa.sort(key=lambda x: (x["producto_codigo"], x["segmento"]))
    por_vendedor = []
    for cod, grp in df.groupby("vendedor_codigo"):
        cod_int = int(cod); nombre = str(grp["vendedor_nombre"].iloc[0])
        productos = []
        for _, row in grp.iterrows():
            cartera = int(row["clientes_cartera"]); compraron = int(row["clientes_compraron"])
            faltantes_raw = str(row.get("clientes_faltantes", "") or "")
            faltantes = [int(x) for x in faltantes_raw.split("|") if x.strip().isdigit()]
            productos.append({
                "segmento": str(row["segmento"]), "producto_codigo": int(row["producto_codigo"]),
                "producto_nombre": str(row["producto_nombre"]), "clientes_cartera": cartera,
                "clientes_compraron": compraron,
                "pct_cobertura": round(compraron / cartera, 4) if cartera else 0.0,
                "clientes_faltantes": faltantes,
            })
        productos.sort(key=lambda x: (x["segmento"], x["producto_codigo"]))
        por_vendedor.append({"vendedor_id": f"V{cod_int}", "vendedor_nombre": nombre, "productos": productos})
    por_vendedor.sort(key=lambda x: x["vendedor_id"])

    # ── Depósito (V20): CCC de innovación LOGRADO, aparte (sin cartera ni faltantes) ──
    deposito = []
    try:
        cods = df[["producto_codigo", "producto_nombre"]].drop_duplicates()
        cod2nom = {int(c): str(n) for c, n in zip(cods["producto_codigo"], cods["producto_nombre"])
                   if pd.notnull(c)}
        vd = _df_deposito_ventas()
        if cod2nom and not vd.empty:
            vd = vd.copy()
            vd["_cod"] = pd.to_numeric(vd["Codigo"], errors="coerce")
            for pc, nom in sorted(cod2nom.items()):
                sub = vd[vd["_cod"] == pc]
                cc = int(sub["Cliente"].nunique()) if not sub.empty else 0
                if cc > 0:
                    deposito.append({"producto_codigo": pc, "producto_nombre": nom,
                                     "clientes_compraron": cc})
    except Exception:
        deposito = []

    return jsonify({
        "generado_en": _now_ar(),
        "fuente": "mod_innovaciones_segmento.csv",
        "fecha_ejecucion": fecha_ej,
        "resumen_empresa": resumen_empresa,
        "por_vendedor": por_vendedor,
        "deposito": deposito,
    })


# ====== VENDEDOR: INNOVACIONES POR SEGMENTO ======
@app.route("/api/vendedor/<vid>/innovaciones_segmento")
def vendedor_innovaciones_segmento(vid):
    """Cobertura de innovaciones por segmento para un vendedor específico.
    Fuente: mod_innovaciones_segmento.csv. V3 sin AUTOSERVICIO."""
    vid_norm = normalizar_vendedor_codigo(vid)
    cn = clean_code(vid_norm)
    cod_int = int(cn) if cn.isdigit() else 0
    if cod_int in _VENDEDORES_EXCLUIDOS:
        return jsonify({"error": f"Vendedor {vid_norm} excluido"}), 403
    vend = read_csv(CONFIG / "vendedores_activos.csv")
    nombre = vid_norm
    if not vend.empty:
        mask = (vend["activo"] == 1) & (vend["codigo_vendedor"].astype(str).apply(clean_code) == cn)
        fila = vend[mask]
        if fila.empty:
            return jsonify({"error": f"Vendedor {vid_norm} no encontrado o inactivo"}), 404
        nombre = str(fila.iloc[0]["nombre_vendedor"])
    df = read_csv(DATASETS / "mod_innovaciones_segmento.csv")
    if df.empty:
        return jsonify({
            "generado_en": _now_ar(),
            "vendedor_id": vid_norm, "vendedor_nombre": nombre,
            "fuente": "mod_innovaciones_segmento.csv",
            "advertencia": "Dato no disponible", "productos": [],
        })
    df.columns = [c.lstrip("﻿") for c in df.columns]
    for col in ["vendedor_codigo", "producto_codigo", "clientes_cartera", "clientes_compraron"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    dv = df[df["vendedor_codigo"] == cod_int].copy()
    if cod_int == 3:
        dv = dv[dv["segmento"].astype(str).str.upper() != "AUTOSERVICIO"]
    fecha_ej = str(dv["fecha_ejecucion"].iloc[0]) if "fecha_ejecucion" in dv.columns and not dv.empty else None
    productos = []
    for _, row in dv.iterrows():
        cartera = int(row["clientes_cartera"]); compraron = int(row["clientes_compraron"])
        faltantes_raw = str(row.get("clientes_faltantes", "") or "")
        faltantes = [int(x) for x in faltantes_raw.split("|") if x.strip().isdigit()]
        productos.append({
            "segmento": str(row["segmento"]), "producto_codigo": int(row["producto_codigo"]),
            "producto_nombre": str(row["producto_nombre"]), "clientes_cartera": cartera,
            "clientes_compraron": compraron,
            "pct_cobertura": round(compraron / cartera, 4) if cartera else 0.0,
            "clientes_faltantes": faltantes,
        })
    productos.sort(key=lambda x: (x["segmento"], x["producto_codigo"]))
    return jsonify({
        "generado_en": _now_ar(),
        "vendedor_id": vid_norm, "vendedor_nombre": nombre,
        "fuente": "mod_innovaciones_segmento.csv",
        "fecha_ejecucion": fecha_ej, "productos": productos,
    })


# ====== VENDEDOR: PLAN DE ACCIÓN INNOVACIONES ======
@app.route("/api/vendedor/<vid>/plan_innovaciones")
def vendedor_plan_innovaciones(vid):
    """Faltantes de Innovaciones enriquecidos con datos de cliente para plan de acción.
    Fuente: mod_innovaciones_segmento.csv + clientes_dia.csv + clientes_master.csv.
    V3 sin AUTOSERVICIO. Excluye V2, V5, V20."""
    def parse_clientes_faltantes(v):
        if isinstance(v, list):
            result = []
            for x in v:
                try: result.append(int(x))
                except (ValueError, TypeError): pass
            return result
        s = str(v or "").strip()
        if not s or s in ("nan", "None", "[]"):
            return []
        s = s.strip("[]")
        sep = "|" if "|" in s else ","
        result = []
        for x in s.split(sep):
            x = x.strip()
            if x.isdigit():
                result.append(int(x))
        return result

    vid_norm = normalizar_vendedor_codigo(vid)
    cn = clean_code(vid_norm)
    cod_int = int(cn) if cn.isdigit() else 0
    if cod_int in _VENDEDORES_EXCLUIDOS:
        return jsonify({"error": f"Vendedor {vid_norm} excluido"}), 403
    vend = read_csv(CONFIG / "vendedores_activos.csv")
    nombre = vid_norm
    if not vend.empty:
        mask = (vend["activo"] == 1) & (vend["codigo_vendedor"].astype(str).apply(clean_code) == cn)
        fila = vend[mask]
        if fila.empty:
            return jsonify({"error": f"Vendedor {vid_norm} no encontrado o inactivo"}), 404
        nombre = str(fila.iloc[0]["nombre_vendedor"])
    df = read_csv(DATASETS / "mod_innovaciones_segmento.csv")
    if df.empty:
        return jsonify({
            "generado_en": _now_ar(),
            "vendedor_id": vid_norm, "vendedor_nombre": nombre,
            "advertencia": "Dato no disponible", "productos": [],
        })
    df.columns = [str(c).lstrip(chr(65279)).strip() for c in df.columns]
    for col in ["vendedor_codigo", "producto_codigo", "clientes_cartera", "clientes_compraron"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    dv = df[df["vendedor_codigo"] == cod_int].copy()
    if cod_int == 3:
        dv = dv[dv["segmento"].astype(str).str.upper() != "AUTOSERVICIO"]
    if dv.empty:
        return jsonify({
            "generado_en": _now_ar(),
            "vendedor_id": vid_norm, "vendedor_nombre": nombre,
            "advertencia": "Sin datos de innovaciones para este vendedor", "productos": [],
        })
    cli_dia = read_csv(DATASETS / "clientes_dia.csv")
    cli_dia_idx = {}
    if not cli_dia.empty:
        cli_dia["cliente_id"] = pd.to_numeric(cli_dia["cliente_id"], errors="coerce")
        if "vendedor_codigo" in cli_dia.columns:
            cli_dia["vendedor_codigo"] = pd.to_numeric(cli_dia["vendedor_codigo"], errors="coerce")
            cli_dia = cli_dia[cli_dia["vendedor_codigo"] == cod_int]
        for _, r in cli_dia.iterrows():
            cid = int(r["cliente_id"]) if not pd.isna(r["cliente_id"]) else None
            if cid is not None:
                cli_dia_idx[cid] = r
    cli_master = read_csv(BASE / "05_MASTER_DATA" / "clientes_master.csv")
    cli_master_idx = {}
    if not cli_master.empty:
        cli_master.columns = [str(c).lstrip(chr(65279)).strip() for c in cli_master.columns]
        cli_master["cliente_id"] = pd.to_numeric(cli_master["cliente_id"], errors="coerce")
        for _, r in cli_master.iterrows():
            cid = int(r["cliente_id"]) if not pd.isna(r["cliente_id"]) else None
            if cid is not None:
                cli_master_idx[cid] = r
    PRIO_ORD = {"ALTA": 0, "MEDIA": 1, "BAJA": 2}
    productos = []
    for _, row in dv.iterrows():
        cartera = int(row["clientes_cartera"]); compraron = int(row["clientes_compraron"])
        faltantes_ids = parse_clientes_faltantes(row.get("clientes_faltantes", ""))
        plan = []
        for cid in faltantes_ids:
            if cid in cli_dia_idx:
                r = cli_dia_idx[cid]
                plan.append({
                    "cliente_id": cid,
                    "cliente_nombre": str(r.get("cliente_nombre", "") or ""),
                    "segmento": str(r.get("segmento_operativo", r.get("segmento", "")) or ""),
                    "localidad": str(r.get("localidad", "") or ""),
                    "ruta": str(r.get("ruta", "") or ""),
                    "dias_visita": str(r.get("dias_visita", "") or ""),
                    "prioridad": str(r.get("prioridad_comercial", "") or "") or None,
                    "en_zona_hoy": True,
                    "enriquecimiento": "completo",
                })
            elif cid in cli_master_idx:
                r = cli_master_idx[cid]
                plan.append({
                    "cliente_id": cid,
                    "cliente_nombre": str(r.get("cliente_nombre", "") or ""),
                    "segmento": str(r.get("segmento", "") or ""),
                    "localidad": str(r.get("localidad", "") or ""),
                    "ruta": None,
                    "dias_visita": None,
                    "prioridad": None,
                    "en_zona_hoy": False,
                    "enriquecimiento": "parcial",
                })
            else:
                plan.append({
                    "cliente_id": cid,
                    "cliente_nombre": None,
                    "segmento": None, "localidad": None,
                    "ruta": None, "dias_visita": None, "prioridad": None,
                    "en_zona_hoy": False,
                    "enriquecimiento": "sin_datos",
                })
        plan.sort(key=lambda x: (
            0 if x["en_zona_hoy"] else 1,
            PRIO_ORD.get(x.get("prioridad") or "", 9),
            x.get("cliente_nombre") or "",
        ))
        productos.append({
            "segmento": str(row["segmento"]),
            "producto_codigo": int(row["producto_codigo"]),
            "producto_nombre": str(row["producto_nombre"]),
            "clientes_cartera": cartera,
            "clientes_compraron": compraron,
            "pct_cobertura": round(compraron / cartera, 4) if cartera else 0.0,
            "total_faltantes": len(plan),
            "en_zona_hoy": sum(1 for c in plan if c["en_zona_hoy"]),
            "plan": plan,
        })
    productos.sort(key=lambda x: (x["segmento"], x["producto_codigo"]))
    return jsonify({
        "generado_en": _now_ar(),
        "vendedor_id": vid_norm, "vendedor_nombre": nombre,
        "fuente": "mod_innovaciones_segmento.csv + clientes_dia.csv + clientes_master.csv",
        "productos": productos,
    })


# ====== COBERTURA ACUMULADA ======
@app.route("/api/gerencia/cobertura_acum")
def gerencia_cobertura_acum():
    df = read_csv(DATASETS / "mod_cobertura_acum.csv")
    if df.empty:
        return jsonify({"error": "Sin datos"}), 404
    df["vendedor_codigo"] = pd.to_numeric(df["vendedor_codigo"], errors="coerce")
    df = df[~df["vendedor_codigo"].isin(_VENDEDORES_EXCLUIDOS)]
    for c in ["cartera", "cubiertos", "sin_cobertura"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
    df["pct_cobertura"] = pd.to_numeric(df.get("pct_cobertura", 0), errors="coerce").fillna(0)
    fecha = str(df["fecha_calculo"].iloc[0]) if "fecha_calculo" in df.columns else ""
    # Objetivos aperturados por vendedor (objccc.xlsx). MAYORISTA no tiene objetivo cargado:
    # queda visible como informativo y FUERA del total, para no comparar contra 0.
    objs = _objetivos_ccc_vendedor()
    obj_seg = objs.get("por_segmento", {})
    declarado = objs.get("declarado", {})

    por_vendedor = {}
    for _, row in df.iterrows():
        cod = int(row["vendedor_codigo"])
        vid = f"V{cod}"
        if vid not in por_vendedor:
            por_vendedor[vid] = {"vendedor_id": vid, "vendedor_nombre": str(row.get("vendedor_nombre", "")),
                                 "segmentos": [], "cubiertos": 0, "objetivo": 0}
        seg = str(row["segmento"])
        cubiertos = int(row["cubiertos"])
        obj = obj_seg.get(seg, {}).get(cod)
        por_vendedor[vid]["segmentos"].append({
            "segmento": seg,
            "cartera": int(row["cartera"]),
            "cubiertos": cubiertos,
            "sin_cobertura": int(row["sin_cobertura"]),
            "pct_cobertura": round(float(row["pct_cobertura"]), 4),
            "objetivo": obj,
            "pct_objetivo": round(cubiertos / obj * 100, 1) if obj else None,
        })
        if obj is not None:
            por_vendedor[vid]["cubiertos"] += cubiertos
            por_vendedor[vid]["objetivo"] += obj
    for v in por_vendedor.values():
        v["pct_objetivo"] = round(v["cubiertos"] / v["objetivo"] * 100, 1) if v["objetivo"] else None

    # ── Depósito (V20): CCC informativo. No tiene cartera en el maestro ni objetivo propio,
    # pero su CCC SÍ suma al logrado total de la empresa (decisión usuario 2026-07-20). ──
    deposito = None
    try:
        vd = _df_deposito_ventas()
        if not vd.empty:
            deposito = {
                "vendedor_id": "V20",
                "vendedor_nombre": "Depósito (venta directa)",
                "ccc": int(vd["Cliente"].nunique()),
                "clientes": int(vd["Cliente"].nunique()),
                "botellas": round(float(vd["CantBase"].sum()), 1),
            }
    except Exception:
        deposito = None

    # Empresa: sólo los segmentos CON objetivo (AS/TRAD/OP), en numerador y denominador.
    log_ruta = sum(v["cubiertos"] for v in por_vendedor.values())
    log_dep = int(deposito["ccc"]) if deposito else 0
    obj_asignado = sum(v["objetivo"] for v in por_vendedor.values())
    obj_total = sum(declarado.get(s, 0) for s in _OBJCCC_HOJAS.values()) or obj_asignado
    log_total = log_ruta + log_dep

    return jsonify({
        "generado_en": _now_ar(),
        "fecha_calculo": fecha,
        "fuente": "mod_cobertura_acum.csv + objccc.xlsx",
        "por_vendedor": list(por_vendedor.values()),
        "deposito": deposito,
        "empresa": {
            "logrado_ruta":      log_ruta,
            "logrado_deposito":  log_dep,
            "logrado_total":     log_total,
            "objetivo_total":    obj_total,
            # Los objetivos por vendedor suman menos que el total declarado en objccc.xlsx
            # (Tradicional: 809 vs 845). Se exponen ambos para que la diferencia sea visible
            # en la tarjeta y no quede escondida dentro del %.
            "objetivo_asignado": obj_asignado,
            "pct":               round(log_total / obj_total * 100, 1) if obj_total else None,
        },
    })


def _cobertura_faltantes_rows(df, cod=None):
    """Normaliza mod_cobertura_acum_detalle.csv. Si cod, filtra a ese vendedor.
    Devuelve dict {(vid, segmento): [clientes faltantes]}."""
    out = {}
    if df.empty:
        return out
    df.columns = [c.lstrip("﻿") for c in df.columns]
    df["vendedor_codigo"] = pd.to_numeric(df["vendedor_codigo"], errors="coerce")
    df = df[~df["vendedor_codigo"].isin(_VENDEDORES_EXCLUIDOS)]
    if cod is not None:
        df = df[df["vendedor_codigo"] == cod]
    for _, row in df.iterrows():
        vid = f"V{int(row['vendedor_codigo'])}" if pd.notnull(row.get("vendedor_codigo")) else "V0"
        seg = str(row.get("segmento", ""))
        out.setdefault((vid, seg), []).append({
            "cliente_id":     int(row["cliente_id"]) if pd.notnull(row.get("cliente_id")) else None,
            "cliente_nombre": str(row.get("cliente_nombre", "") or ""),
            "localidad":      str(row.get("localidad", "") or ""),
            "botellas":       round(float(row.get("cant_base_acum", 0) or 0), 1),
            "umbral":         int(row.get("umbral", 0) or 0),
        })
    return out


# ====== GERENCIA: FALTANTES DE COBERTURA POR SEGMENTO (drill-down) ======
@app.route("/api/gerencia/cobertura_acum_faltantes")
def gerencia_cobertura_acum_faltantes():
    """Clientes que aún no lograron cobertura (acumulado) por vendedor, para un segmento.
    Fuente: mod_cobertura_acum_detalle.csv. Excluye V2, V5, V20."""
    seg = request.args.get("segmento", "").strip()
    df = read_csv(DATASETS / "mod_cobertura_acum_detalle.csv")
    rows = _cobertura_faltantes_rows(df)
    por_vendedor = {}
    for (vid, segk), clientes in rows.items():
        if seg and segk.upper() != seg.upper():
            continue
        v = por_vendedor.setdefault(vid, {"vendedor_id": vid, "vendedor_nombre": "", "faltantes": []})
        v["faltantes"].extend(clientes)
    # vendedor_nombre desde el primer registro disponible
    if not df.empty:
        df.columns = [c.lstrip("﻿") for c in df.columns]
        for _, r in df.iterrows():
            vid = f"V{int(r['vendedor_codigo'])}" if pd.notnull(r.get("vendedor_codigo")) else None
            if vid in por_vendedor and not por_vendedor[vid]["vendedor_nombre"]:
                por_vendedor[vid]["vendedor_nombre"] = str(r.get("vendedor_nombre", "") or "")
    res = sorted(por_vendedor.values(),
                 key=lambda x: int(x["vendedor_id"][1:]) if x["vendedor_id"][1:].isdigit() else 999)
    return jsonify({
        "generado_en": _now_ar(),
        "segmento": seg,
        "fuente": "mod_cobertura_acum_detalle.csv",
        "por_vendedor": res,
    })


# ====== VENDEDOR: COBERTURA ACUMULADA PROPIA + FALTANTES ======
@app.route("/api/vendedor/<vid>/cobertura_acum")
def vendedor_cobertura_acum(vid):
    """Cobertura acumulada del mes del vendedor por segmento, con clientes faltantes.
    Fuente: mod_cobertura_acum.csv (agregado) + mod_cobertura_acum_detalle.csv (faltantes).
    Solo devuelve datos del propio vendedor (sin exponer a otros)."""
    try:
        cod = int(str(vid).upper().replace("V", ""))
    except ValueError:
        return jsonify({"error": "vendedor inválido"}), 400

    agg = read_csv(DATASETS / "mod_cobertura_acum.csv")
    det = read_csv(DATASETS / "mod_cobertura_acum_detalle.csv")
    falt = _cobertura_faltantes_rows(det, cod=cod)

    fecha = ""
    segmentos = []
    if not agg.empty:
        agg.columns = [c.lstrip("﻿") for c in agg.columns]
        agg["vendedor_codigo"] = pd.to_numeric(agg["vendedor_codigo"], errors="coerce")
        a = agg[agg["vendedor_codigo"] == cod].copy()
        if "fecha_calculo" in a.columns and not a.empty:
            fecha = str(a["fecha_calculo"].iloc[0])
        for _, row in a.sort_values("segmento").iterrows():
            segk = str(row["segmento"])
            segmentos.append({
                "segmento":      segk,
                "cartera":       int(row.get("cartera", 0) or 0),
                "cubiertos":     int(row.get("cubiertos", 0) or 0),
                "sin_cobertura": int(row.get("sin_cobertura", 0) or 0),
                "pct_cobertura": round(float(row.get("pct_cobertura", 0) or 0), 4),
                "faltantes":     falt.get((f"V{cod}", segk), []),
            })
    return jsonify({
        "generado_en": _now_ar(),
        "vendedor_id": f"V{cod}",
        "fecha_calculo": fecha,
        "fuente": "mod_cobertura_acum.csv",
        "segmentos": segmentos,
    })


# ====== 11T ACUMULADO ======
@app.route("/api/gerencia/11t_acum")
def gerencia_11t_acum():
    df = read_csv(DATASETS / "mod_11t_acum.csv")
    if df.empty:
        return jsonify({"error": "Sin datos"}), 404
    df["vendedor_codigo"] = pd.to_numeric(df["vendedor_codigo"], errors="coerce")
    df = df[~df["vendedor_codigo"].isin(_VENDEDORES_EXCLUIDOS)]
    df["tiene_flag"] = pd.to_numeric(df["tiene_flag"], errors="coerce").fillna(0)
    # Resumen por marca (distribuidora total)
    por_marca = (df.groupby("marca_objetivo")
                 .agg(cartera=("cliente_id", "count"), cubiertos=("tiene_flag", "sum"))
                 .reset_index())
    por_marca["cubiertos"] = por_marca["cubiertos"].astype(int)
    por_marca["pct"] = (por_marca["cubiertos"] / por_marca["cartera"].replace(0, 1) * 100).round(1)
    por_marca = por_marca.sort_values("cubiertos", ascending=False)
    # Resumen por vendedor × marca
    por_vend = (df.groupby(["vendedor_codigo", "vendedor_nombre", "marca_objetivo"])
                .agg(cartera=("cliente_id", "count"), cubiertos=("tiene_flag", "sum"))
                .reset_index())
    por_vend["cubiertos"] = por_vend["cubiertos"].astype(int)
    por_vend["vendedor_id"] = "V" + por_vend["vendedor_codigo"].astype(int).astype(str)
    fecha = str(df["fecha_calculo"].iloc[0]) if "fecha_calculo" in df.columns else ""
    return jsonify({
        "generado_en": _now_ar(),
        "fecha_calculo": fecha,
        "fuente": "mod_11t_acum.csv",
        "por_marca": por_marca[["marca_objetivo", "cartera", "cubiertos", "pct"]].to_dict("records"),
        "por_vendedor": por_vend[["vendedor_id", "vendedor_nombre", "marca_objetivo",
                                   "cartera", "cubiertos"]].to_dict("records"),
    })


# ====== INNOVACIONES TOTAL GERENCIA ======
@app.route("/api/gerencia/innovaciones_total")
def gerencia_innovaciones_total():
    """Cobertura de innovaciones. Acepta ?dia=MA para agregar compraron_dia por producto."""
    dia_param = request.args.get("dia", "").strip().capitalize()
    df = read_csv(DATASETS / "mod_innovaciones_segmento.csv")
    if df.empty:
        return jsonify({"error": "Sin datos"}), 404
    df["vendedor_codigo"] = pd.to_numeric(df["vendedor_codigo"], errors="coerce")
    df = df[~df["vendedor_codigo"].isin(_VENDEDORES_EXCLUIDOS)]
    df["clientes_cartera"]  = pd.to_numeric(df["clientes_cartera"],  errors="coerce").fillna(0)
    df["clientes_compraron"] = pd.to_numeric(df["clientes_compraron"], errors="coerce").fillna(0)
    df["producto_codigo"] = pd.to_numeric(df["producto_codigo"], errors="coerce").fillna(0).astype(int)
    # Total por producto × segmento (toda la distribuidora)
    tot = (df.groupby(["producto_codigo", "producto_nombre", "segmento"])
           .agg(cartera_total=("clientes_cartera", "sum"),
                compraron_total=("clientes_compraron", "sum"))
           .reset_index())
    tot["pct_cobertura"] = (tot["compraron_total"] / tot["cartera_total"].replace(0, 1) * 100).round(1)
    # compraron_dia: clientes del día que compraron el producto (ventas.csv MES VIVO × clientes del día)
    compraron_dia_map = {}   # producto_codigo → count
    if dia_param:
        try:
            acum_path = INPUTS / "ventas.csv"
            if acum_path.exists():
                acum = pd.read_csv(acum_path, encoding="latin1", sep=";", engine="python")
                acum["Codigo"]         = pd.to_numeric(acum["Codigo"], errors="coerce")
                acum["Cliente"]        = pd.to_numeric(acum["Cliente"], errors="coerce")
                acum["ImporteNetoItem"] = (acum["ImporteNetoItem"].astype(str)
                                           .str.replace(".", "", regex=False)
                                           .str.replace(",", ".", regex=False))
                acum["ImporteNetoItem"] = pd.to_numeric(acum["ImporteNetoItem"], errors="coerce").fillna(0)
                acum = acum[acum["ImporteNetoItem"] > 0]
                cli_dia = _clientes_por_dia(dia_param)
                if not cli_dia.empty and "cliente_id" in cli_dia.columns:
                    dia_ids = set(cli_dia["cliente_id"].dropna().astype(int))
                    prod_codes = set(df["producto_codigo"].dropna().astype(int).tolist())
                    for cod in prod_codes:
                        compraron_dia_map[int(cod)] = int(
                            acum[(acum["Codigo"] == cod) & (acum["Cliente"].isin(dia_ids))]
                            ["Cliente"].nunique()
                        )
        except Exception as e:
            print(f"[WARN] innovaciones_total ?dia: {e}")
    records = []
    for _, row in tot.iterrows():
        pcod = int(row["producto_codigo"])
        rec = {
            "producto_codigo": pcod,
            "producto_nombre": str(row["producto_nombre"]),
            "segmento": str(row["segmento"]),
            "cartera_total": int(row["cartera_total"]),
            "compraron_total": int(row["compraron_total"]),
            "pct_cobertura": float(row["pct_cobertura"]),
            "compraron_dia": compraron_dia_map.get(pcod, 0),
        }
        records.append(rec)
    # Por vendedor (sin desglose en portal, se mantiene para API externa)
    por_v = df[["vendedor_codigo", "vendedor_nombre", "segmento",
                "producto_codigo", "producto_nombre",
                "clientes_cartera", "clientes_compraron", "pct_cobertura"]].copy()
    por_v["vendedor_id"] = "V" + por_v["vendedor_codigo"].astype(int).astype(str)
    fecha = str(df["fecha_ejecucion"].iloc[0]) if "fecha_ejecucion" in df.columns else ""

    # ── Depósito (V20): CCC de innovación LOGRADO, aparte (sin cartera ni faltantes) ──
    deposito = []
    try:
        cod2nom = {int(c): str(n) for c, n in zip(df["producto_codigo"], df["producto_nombre"])
                   if pd.notnull(c)}
        vd = _df_deposito_ventas()
        if cod2nom and not vd.empty:
            vd = vd.copy()
            vd["_cod"] = pd.to_numeric(vd["Codigo"], errors="coerce")
            for pc, nom in sorted(cod2nom.items()):
                sub = vd[vd["_cod"] == pc]
                cc = int(sub["Cliente"].nunique()) if not sub.empty else 0
                if cc > 0:
                    deposito.append({"producto_codigo": pc, "producto_nombre": nom,
                                     "clientes_compraron": cc})
    except Exception:
        deposito = []

    return jsonify({
        "generado_en": _now_ar(),
        "fecha_ejecucion": fecha,
        "fuente": "mod_innovaciones_segmento.csv",
        "dia": dia_param or None,
        "por_producto": records,
        "deposito": deposito,
        "por_vendedor": por_v[["vendedor_id", "vendedor_nombre", "segmento",
                                "producto_codigo", "producto_nombre",
                                "clientes_cartera", "clientes_compraron"]].to_dict("records"),
    })


# ====== STOCK SIN VENTA EN EL MES — GERENCIA ======
_STOCK_CACHE = {}

def _stock_disponible(archivo: str = "stock.xlsx") -> pd.DataFrame:
    """Stock de depósito desde 01_INPUTS/Stock/<archivo>, cacheado por (archivo, mtime).
    Columnas normalizadas: codigo (int), descripcion, disponible (UniTotalDisponible),
    bultos_total, reserva, transito, proveedor. El disponible es la existencia real
    (Total = Disponible + Reserva; Tránsito es mercadería en camino, informativo).
    El default `stock.xlsx` es el depósito PyP; VSB Cuyo tiene su propio export
    (`stock_VSB_Cuyo.xlsx`) — son dos depósitos distintos, no se suman."""
    p = INPUTS / "Stock" / archivo
    if not p.exists():
        return pd.DataFrame()
    try:
        key = (archivo, os.path.getmtime(p))
    except OSError:
        key = (archivo, 0)
    df = _STOCK_CACHE.get(key)
    if df is not None:
        return df
    try:
        raw = pd.read_excel(p)
    except Exception:
        return pd.DataFrame()
    raw.columns = [str(c).strip() for c in raw.columns]
    if "Codigo" not in raw.columns:
        return pd.DataFrame()
    def _num(col):
        return pd.to_numeric(raw[col], errors="coerce").fillna(0) if col in raw.columns else 0
    df = pd.DataFrame({
        "codigo":       pd.to_numeric(raw["Codigo"], errors="coerce"),
        "descripcion":  raw.get("Descripcion", "").astype(str).str.strip(),
        "disponible":   _num("UniTotalDisponible"),
        "bultos_total": _num("BultosTotal"),
        "reserva":      _num("BultosReserva"),
        "transito":     _num("BultosTransito"),
        "proveedor":    raw.get("NombreProvedor", "").astype(str).str.strip(),
    })
    df = df.dropna(subset=["codigo"]).copy()
    df["codigo"] = df["codigo"].astype(int)
    # Consolidar por código (por si hay más de una sede/sector): sumar existencias
    df = (df.groupby(["codigo", "descripcion", "proveedor"], as_index=False)
            .agg({"disponible": "sum", "bultos_total": "sum",
                  "reserva": "sum", "transito": "sum"}))
    for k in [k for k in _STOCK_CACHE if k[0] == archivo and k != key]:
        _STOCK_CACHE.pop(k, None)
    _STOCK_CACHE[key] = df
    return df


def _stock_sin_venta_payload():
    """Productos con existencia en depósito (disponible > 0) que NO registraron
    ninguna venta en el mes calendario en curso (ventas.csv, todas las empresas).
    Devuelve el dict de salida, o None si no hay fuente de stock.
    Fuente stock: 01_INPUTS/Stock/stock.xlsx · Fuente ventas: 01_INPUTS/ventas.csv."""
    stock = _stock_disponible()
    if stock.empty:
        return None
    con_stock = stock[stock["disponible"] > 0].copy()

    # Códigos con al menos una unidad vendida en el mes en curso (cualquier empresa/vendedor:
    # si el producto se movió por algún canal NO es stock muerto).
    v = _ventas_parsed()
    codigos_vendidos = set()
    if not v.empty and "codigo_art" in v.columns:
        hoy = datetime.now()
        mes_inicio = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        vm = v[(v["fecha"] >= mes_inicio) & (v["cant_base"] > 0)]
        codigos_vendidos = set(vm["codigo_art"].dropna().astype(int).tolist())

    sin_venta = con_stock[~con_stock["codigo"].isin(codigos_vendidos)].copy()
    sin_venta = sin_venta.sort_values("disponible", ascending=False)

    productos = [{
        "codigo":       int(r["codigo"]),
        "descripcion":  str(r["descripcion"]),
        "disponible":   int(round(r["disponible"])),
        "bultos_total": int(round(r["bultos_total"])),
        "reserva":      int(round(r["reserva"])),
        "transito":     int(round(r["transito"])),
    } for _, r in sin_venta.iterrows()]

    meses = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio",
             "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    hoy = datetime.now()
    return {
        "generado_en": _now_ar(),
        "fuente": "01_INPUTS/Stock/stock.xlsx + 01_INPUTS/ventas.csv",
        "mes": f"{meses[hoy.month]} {hoy.year}",
        "total_productos_stock": int(len(con_stock)),
        "total_sin_venta": int(len(productos)),
        "unidades_sin_venta": int(round(sin_venta["disponible"].sum())),
        "productos": productos,
    }


@app.route("/api/gerencia/stock_sin_venta")
def gerencia_stock_sin_venta():
    """JSON del reporte Stock sin Venta (ver _stock_sin_venta_payload)."""
    payload = _stock_sin_venta_payload()
    if payload is None:
        return jsonify({"error": "Stock no disponible (falta 01_INPUTS/Stock/stock.xlsx)"}), 404
    return jsonify(payload)


@app.route("/api/gerencia/stock_sin_venta/export")
def gerencia_stock_sin_venta_export():
    """Descarga Excel (.xlsx) SOLO de los productos con stock y sin ventas en el mes."""
    payload = _stock_sin_venta_payload()
    if payload is None:
        return jsonify({"error": "Stock no disponible (falta 01_INPUTS/Stock/stock.xlsx)"}), 404

    cols = ["codigo", "descripcion", "disponible", "reserva", "transito", "bultos_total"]
    df = pd.DataFrame(payload["productos"], columns=cols).rename(columns={
        "codigo": "Código", "descripcion": "Producto", "disponible": "Disponible",
        "reserva": "Reserva", "transito": "En tránsito", "bultos_total": "Bultos total",
    })

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        df.to_excel(xl, index=False, sheet_name="Stock sin venta")
        ws = xl.sheets["Stock sin venta"]
        for i, col in enumerate(df.columns, start=1):
            if len(df):
                ancho = max(len(str(col)), int(df.iloc[:, i - 1].astype(str).map(len).max()))
            else:
                ancho = len(str(col))
            ws.column_dimensions[chr(64 + i)].width = min(max(ancho + 2, 10), 50)
    buf.seek(0)

    mes = str(payload.get("mes", "")).replace(" ", "_")
    fname = f"stock_sin_venta_{mes}.xlsx" if mes else "stock_sin_venta.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ====== PLANES AS — GERENCIA ======
def _cargar_sincargos_envios():
    """Detalle de envíos de sin cargo por cliente desde mod_sincargos_envios.csv.
    Devuelve {cliente_id: [{producto, fecha, cajas, categoria}, ...]} para la tarjeta
    desplegable del portal (fecha en que se envió cada sin cargo)."""
    df = read_csv(DATASETS / "mod_sincargos_envios.csv")
    out = {}
    if df.empty:
        return out
    df["cliente_id"] = pd.to_numeric(df["cliente_id"], errors="coerce")
    for _, r in df.dropna(subset=["cliente_id"]).iterrows():
        out.setdefault(int(r["cliente_id"]), []).append({
            "producto":  str(r.get("producto", "")),
            "fecha":     str(r.get("fecha", "")),
            "cajas":     int(pd.to_numeric(r.get("cajas", 0), errors="coerce") or 0),
            "categoria": str(r.get("categoria", "")),
        })
    return out


# ── INNOVACIONES SEGUIDAS EN PLANES AS ──
_INOV_PLAN_AS_CACHE = {}

def _inov_plan_as_productos():
    """Innovaciones marcadas con 'x' en la columna 'AASS c/plan' de Innovaciones.xlsx:
    las que se siguen cliente por cliente en la pantalla de Planes AS. La fila sin marca
    (Termidor) queda fuera. Mismo archivo que la pantalla de Innovaciones -> agregar un
    producto ahi lo suma en las dos. Cacheado por mtime."""
    p = INPUTS / "INNOVACIONES" / "Innovaciones.xlsx"
    if not p.exists():
        return []
    try:
        key = os.path.getmtime(p)
    except OSError:
        key = 0
    cached = _INOV_PLAN_AS_CACHE.get(key)
    if cached is not None:
        return cached
    out = []
    try:
        df = pd.read_excel(p, sheet_name=0, header=None, dtype=str)
        for _, fila in df.iterrows():
            celdas = [str(v).strip() for v in fila.tolist() if pd.notna(v)]
            if not any(c.lower() == "x" for c in celdas):
                continue
            for c in celdas:
                m = _re.match(r"^0*(\d{4,6})\s*-\s*(.+)$", c)
                if m:
                    nombre = _re.sub(r"\s+", " ", m.group(2).replace("\xa0", " ")).strip()
                    out.append({"codigo": int(m.group(1)), "nombre": nombre})
                    break
    except Exception as e:
        print(f"[WARN] innovaciones plan AS: {e}")
        out = []
    _INOV_PLAN_AS_CACHE.clear()
    _INOV_PLAN_AS_CACHE[key] = out
    return out


# Cobertura de una innovacion en un cliente de Plan AS = misma regla que la cobertura
# de Autoservicio del resto del sistema (UMBRAL["AUTOSERVICIO"] en generar_datasets_acum):
# minimo 6 unidades (CantBase) del producto en el mes. Los Planes AS son todos AS.
_INOV_PLAN_AS_MIN_UNID = 6

def _inov_plan_as_compras():
    """cliente_id -> {codigo_innovacion: unidades (CantBase) del mes en curso}.
    ventas.csv (mes vivo), ImporteNetoItem > 0, sin filtro de Empresa. El flag de
    'comprado' (cobertura) se decide luego contra _INOV_PLAN_AS_MIN_UNID."""
    prods = {p["codigo"] for p in _inov_plan_as_productos()}
    if not prods:
        return {}
    df = _ventas_parsed()
    if df.empty:
        return {}
    hoy = datetime.now(_ARG_TZ)
    mes_inicio = pd.Timestamp(hoy.year, hoy.month, 1)
    d = df[(df["fecha"] >= mes_inicio) & (df["importe_neto"] > 0)
           & (df["codigo_art"].isin(prods))]
    out = {}
    for cid, cod, u in zip(d["cliente_id"], d["codigo_art"], d["cant_base"]):
        if pd.notna(cid) and pd.notna(cod):
            porcli = out.setdefault(int(cid), {})
            porcli[int(cod)] = porcli.get(int(cod), 0.0) + (float(u) if pd.notna(u) else 0.0)
    return out


def _inov_plan_as_cliente(cid, prods, compras):
    """Lista de innovaciones del plan para un cliente. 'comprado' = alcanzo la cobertura
    de Autoservicio (>= 6 unidades del producto en el mes). 'unidades' viaja para el tooltip."""
    porcli = compras.get(cid, {})
    out = []
    for pr in prods:
        u = round(porcli.get(pr["codigo"], 0.0), 1)
        out.append({"codigo": pr["codigo"], "nombre": pr["nombre"],
                    "unidades": u, "comprado": u >= _INOV_PLAN_AS_MIN_UNID})
    return out


@app.route("/api/gerencia/planes_as")
def gerencia_planes_as():
    df = read_csv(DATASETS / "mod_planes_as.csv")
    if df.empty:
        return jsonify({"error": "Sin datos"}), 404
    envios_map = _cargar_sincargos_envios()
    inov_prods = _inov_plan_as_productos()
    inov_compras = _inov_plan_as_compras()
    _num_cols = ["total_facturado", "dcto_plan", "cant_cajas", "tope", "escala_actual", "escala_max",
                 "sc_alaris", "sc_alma_mora", "sc_frizze", "sc_antares_ipa", "sc_smf_flavours",
                 "sc_total_ganado", "sc_cajas_enviadas_total", "sc_pendiente",
                 "sc_env_alaris", "sc_env_alma_mora", "sc_env_frizze", "sc_env_antares_ipa", "sc_env_smf_flavours",
                 "sc_pend_alaris", "sc_pend_alma_mora", "sc_pend_frizze", "sc_pend_antares_ipa", "sc_pend_smf_flavours",
                 "pt_disponible", "pt_enviado", "pt_pendiente"]
    for c in _num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # Join con clientes.xlsx CACHEADO (_clientes_maestro) para localidad y dia_visita
    _cli_info = {}   # cliente_id → {localidad, dia_visita}
    try:
        cli_xl = _clientes_maestro()
        if not cli_xl.empty and "Codigo" in cli_xl.columns:
            for _, r in cli_xl.iterrows():
                cidx = pd.to_numeric(r.get("Codigo"), errors="coerce")
                if pd.notna(cidx):
                    _cli_info[int(cidx)] = {
                        "localidad":   str(r.get("Localidad", "") or "").strip(),
                        "dia_visita":  str(r.get("DiasVisita", "") or "").strip(),
                        "direccion":   str(r.get("Direccion", "") or "").strip(),
                    }
    except Exception as e:
        print(f"[WARN] planes_as join clientes: {e}")

    def _int(v): return int(v) if pd.notna(v) else 0
    registros = []
    for _, row in df.iterrows():
        cid = int(row["cliente_id"]) if pd.notna(row["cliente_id"]) else None
        ci  = _cli_info.get(cid, {})
        registros.append({
            "cliente_id":      cid,
            "cliente_nombre":  str(row.get("cliente_nombre", "")),
            "direccion":       str(row.get("direccion", "") or ci.get("direccion", "")),
            "localidad":       ci.get("localidad", ""),
            "dia_visita":      ci.get("dia_visita", ""),
            "vendedor_id":     f"V{int(row['vendedor_codigo'])}" if pd.notna(row.get("vendedor_codigo")) else None,
            "vendedor_nombre": str(row.get("vendedor_nombre", "")),
            "plan_as":         str(row.get("plan_as", "")),
            "escala_actual":   _int(row.get("escala_actual", 0)),
            "escala_max":      _int(row.get("escala_max", 0)),
            "total_facturado": round(float(row["total_facturado"]), 2),
            "dcto_plan":       round(float(row["dcto_plan"]), 2),
            "cant_cajas":      _int(row["cant_cajas"]),
            "tope":            _int(row["tope"]),
            "sc_alaris":       _int(row["sc_alaris"]),
            "sc_alma_mora":    _int(row["sc_alma_mora"]),
            "sc_frizze":       _int(row["sc_frizze"]),
            "sc_antares_ipa":  _int(row["sc_antares_ipa"]),
            "sc_smf_flavours": _int(row["sc_smf_flavours"]),
            "sc_total_ganado": _int(row["sc_total_ganado"]),
            "sc_env_alaris":   _int(row.get("sc_env_alaris", 0)),
            "sc_env_alma_mora":_int(row.get("sc_env_alma_mora", 0)),
            "sc_env_frizze":   _int(row.get("sc_env_frizze", 0)),
            "sc_env_antares_ipa": _int(row.get("sc_env_antares_ipa", 0)),
            "sc_env_smf_flavours": _int(row.get("sc_env_smf_flavours", 0)),
            "sc_pend_alaris":  _int(row.get("sc_pend_alaris", 0)),
            "sc_pend_alma_mora": _int(row.get("sc_pend_alma_mora", 0)),
            "sc_pend_frizze":  _int(row.get("sc_pend_frizze", 0)),
            "sc_pend_antares_ipa": _int(row.get("sc_pend_antares_ipa", 0)),
            "sc_pend_smf_flavours": _int(row.get("sc_pend_smf_flavours", 0)),
            "sc_enviadas_total": _int(row["sc_cajas_enviadas_total"]),
            "sc_pendiente":    _int(row["sc_pendiente"]),
            "pf_disponible":   _int(row.get("pf_disponible", 0)),
            "pf_enviado":      _int(row.get("pf_enviado", 0)),
            "pf_estado":       str(row.get("pf_estado", "")),
            "pt_disponible":   _int(row.get("pt_disponible", 0)),
            "pt_enviado":      _int(row.get("pt_enviado", 0)),
            "pt_pendiente":    _int(row.get("pt_pendiente", 0)),
            "pt_estado":       str(row.get("pt_estado", "")),
            "pt_producto":     str(row.get("pt_producto", "") or ""),
            "envios":          envios_map.get(cid, []),
            "innovaciones":    _inov_plan_as_cliente(cid, inov_prods, inov_compras),
        })
    fecha = str(df["fecha_calculo"].iloc[0]) if "fecha_calculo" in df.columns else ""
    return jsonify({
        "generado_en": _now_ar(),
        "fecha_calculo": fecha,
        "fuente": "mod_planes_as.csv",
        "innovaciones_productos": inov_prods,
        "total_clientes": len(registros),
        "clientes": registros,
    })


# ====== ACCIONES VIGENTES (todos los roles) ======
@app.route("/api/acciones_vigentes")
def acciones_vigentes():
    """
    Acciones comerciales del mes con tramos de descuento y marcas.
    Lee el CSV de reglas (fuente de verdad) para devolver info completa al vendedor.
    Excluye PLAN_AS (tiene su propia pestaña).
    """
    # Buscar el CSV de reglas más reciente en 09_CONFIG/
    archivos = sorted(CONFIG.glob("reglas_acciones_*.csv"))
    if not archivos:
        return jsonify([])

    try:
        df = pd.read_csv(str(archivos[-1]), encoding="utf-8-sig")
    except Exception as e:
        print(f"[WARN] acciones_vigentes: {e}")
        return jsonify([])

    if df.empty:
        return jsonify([])

    # Excluir Plan AS: tiene pestaña propia
    if "tipo_accion" in df.columns:
        df = df[df["tipo_accion"] != "PLAN_AS"]

    grupos = {}
    orden  = []

    for _, row in df.iterrows():
        grupo = str(row.get("accion_grupo", "") or "").strip()
        if not grupo:
            continue

        if grupo not in grupos:
            lineas = str(row.get("lineas_segmentos", "") or "").strip()
            grupos[grupo] = {
                "accion_grupo":     grupo,
                "accion_nombre":    str(row.get("accion_nombre",    "") or "").strip(),
                "canal":            str(row.get("canal",            "") or "").strip(),
                "tipo_accion":      str(row.get("tipo_accion",      "") or "").strip(),
                "categoria":        str(row.get("categoria",        "") or "").strip(),
                "lineas_segmentos": lineas,
                "tramos":           [],
            }
            orden.append(grupo)

        # Construir tramo desde condicion_original
        condicion = str(row.get("condicion_original", "") or "").strip()
        if not condicion:
            continue

        dto_raw  = row.get("descuento_pct")
        cant_min = row.get("cantidad_min")
        cant_max = row.get("cantidad_max")
        bonif_e  = row.get("bonif_entrega_cajas")
        unidad   = str(row.get("unidad_condicion", "") or "").strip()

        grupos[grupo]["tramos"].append({
            "condicion":     condicion.replace("->", "→"),
            "descuento_pct": round(float(dto_raw) * 100, 1) if pd.notna(dto_raw) and dto_raw else None,
            "cant_min":      int(cant_min) if pd.notna(cant_min) else None,
            "cant_max":      int(cant_max) if pd.notna(cant_max) else None,
            "bonif_cajas":   int(bonif_e)  if pd.notna(bonif_e)  else None,
            "unidad":        unidad,
        })

    return jsonify([grupos[g] for g in orden])


# ====== ACCIONES COMERCIALES DEL MES (catálogo mensual × ventas) ======
# Fuente oficial: 01_INPUTS/ACCIONES COMERCIALES/<YYYY-MM>/acciones_comerciales_<mes>_<año>_penaflor.csv
# Calcula por acción: inversión real (ImporteItem-ImporteNetoItem en líneas con descuento),
# litros, clientes alcanzados, clientes nuevos (no compraron esas marcas el mes anterior).
# Display desde el catálogo: segmento, tipo (descuento/sin cargo), escala, marcas, topes.
_ACC_CAT_CANON = [
    (("VINOS DEL AÑO", "VINOS DEL ANO", "VDA"), "VDA"),
    (("VINOS DE GUARDA", "VDG"), "VDG"),
    (("VINOS DE MESA", "VDM"), "VDM"),
    (("ESPUMANTE",), "ESPUMANTES"),
    (("SIDRA",), "SIDRA"),
    (("SPIRIT",), "SPIRITS"),
    (("RTD",), "RTD"),
    (("CERVEZA",), "CERVEZA"),
]
_ACC_LINEA_TOK = {"VDA": "VDA", "VDG": "VDG", "VDM": "VDM",
                  "ESPUMANTE": "ESPUMANTES", "ESPUMANTES": "ESPUMANTES",
                  "SIDRA": "SIDRA", "SPIRIT": "SPIRITS", "SPIRITS": "SPIRITS",
                  "CERVEZA": "CERVEZA"}
_ACC_PROD_GENERICOS = {"", "SEGUN MAESTRO PRODUCTOS ACTIVOS", "RESTO SKU", "RESTO",
                       "TODOS", "TODOS_ACTIVOS", "LISTA CERRADA DE INNOVACIONES JUNIO 2026"}
_ACC_11T_CACHE = None
_ACC_INNOV_CACHE = None
_ACC_PLAN_AS_CACHE = None
_ACC_DETALLE_CACHE = {}


def _acc_mes_dir():
    """Carpeta de Acciones Comerciales a usar: el MES EN CURSO (YYYY-MM en hora AR) si
    ya tiene carpeta cargada; si todavía no se subió, cae a la carpeta más reciente
    disponible (fail-safe: la pantalla nunca queda vacía). Así el módulo se actualiza
    solo cada mes al crear 01_INPUTS/ACCIONES COMERCIALES/<YYYY-MM>/, y subir el mes
    siguiente por adelantado NO adelanta el cambio (recién entra al llegar ese mes)."""
    base = INPUTS / "ACCIONES COMERCIALES"
    if not base.exists():
        return None
    meses = sorted(sub.name for sub in base.iterdir()
                   if sub.is_dir() and _re.match(r"^\d{4}-\d{2}$", sub.name))
    if not meses:
        return None
    actual = datetime.now(_ARG_TZ).strftime("%Y-%m")
    if actual in meses:
        elegido = actual
    else:
        # mes en curso aún sin carpeta: usar el mes real más reciente que NO sea futuro
        # (subir el mes siguiente por adelantado nunca adelanta el cambio). Si todas las
        # carpetas fueran futuras (estado degenerado), caer a la más nueva.
        pasados = [m for m in meses if m <= actual]
        elegido = pasados[-1] if pasados else meses[-1]
    return base / elegido


def _acc_detalle_map():
    """Lee detalle_categorias_*.csv del mes en curso (';' UTF-8-BOM) y devuelve
    dict: detalle_click_ref -> {detalle_click_ref, categoria_tarjeta, items:[...]}.
    Cada item = {grupo, marca_o_linea, tipo_detalle, observaciones}.
    Vacío si el mes no tiene detalle (esquema histórico, ej. junio). NO hardcodea marcas:
    todo el detalle sale del CSV. Cacheado por (path, mtime)."""
    import csv as _csvm
    mdir = _acc_mes_dir()
    if mdir is None:
        return {}
    det = None
    for c in sorted(mdir.glob("detalle_categorias*.csv")):
        if "salida" in str(c).lower() or "_backup" in str(c).lower():
            continue
        det = c
        break
    if det is None:
        return {}
    try:
        key = (str(det), os.path.getmtime(det))
    except OSError:
        key = (str(det), 0)
    cached = _ACC_DETALLE_CACHE.get(key)
    if cached is not None:
        return cached
    mapa = {}
    try:
        with open(det, encoding="utf-8-sig", newline="") as f:
            for row in _csvm.DictReader(f, delimiter=";"):
                ref = str(row.get("detalle_click_ref", "")).strip()
                if not ref:
                    continue
                entry = mapa.setdefault(ref, {
                    "detalle_click_ref": ref,
                    "categoria_tarjeta": str(row.get("categoria_tarjeta", "")).strip(),
                    "items": [],
                })
                entry["items"].append({
                    "grupo":         str(row.get("grupo", "")).strip(),
                    "marca_o_linea": str(row.get("marca_o_linea", "")).strip(),
                    "tipo_detalle":  str(row.get("tipo_detalle", "")).strip(),
                    "observaciones": str(row.get("observaciones", "")).strip(),
                })
    except Exception:
        return {}
    _ACC_DETALLE_CACHE.clear()
    _ACC_DETALLE_CACHE[key] = mapa
    return mapa


def _acc_canon_cat(c):
    u = str(c or "").strip().upper()
    for kws, canon in _ACC_CAT_CANON:
        if any(k in u for k in kws):
            return canon
    return u or None


# Orden comercial de segmentos VDA/VDG para listar marcas de menor a mayor gama.
_ACC_SEG_ORDER = {"MEDIO": 1, "MEDIO ALTO": 2, "ALTO": 3, "SUPERIOR": 4}


def _acc_marcas_maestro():
    """Desde el maestro 04D: dict cat_canon -> [{segmento, marca}] (marca = Linea Comercial,
    la 'marca canónica'). Deduplicado y ordenado por gama de segmento. NO hardcodea: sale
    del maestro vigente. Barato (itera dicts ya cacheados de _cargar_maestro_04D)."""
    cod2cat, cod2seg, cod2lxu, cod2linea = _cargar_maestro_04D()
    out, seen = {}, set()
    for cod, cat in cod2cat.items():
        canon = _acc_canon_cat(cat)
        if not canon:
            continue
        marca = str(cod2linea.get(cod, "") or "").strip()
        if not marca or marca.lower() == "nan":
            continue
        seg = str(cod2seg.get(cod, "") or "").strip()
        if seg.lower() == "nan":
            seg = ""
        key = (canon, seg.upper(), marca.upper())
        if key in seen:
            continue
        seen.add(key)
        out.setdefault(canon, []).append({"segmento": seg, "marca": marca})
    for k in out:
        out[k].sort(key=lambda x: (_ACC_SEG_ORDER.get(x["segmento"].upper(), 99),
                                   x["segmento"], x["marca"]))
    return out


_ACC_DESC_ART_CACHE = {}
_ACC_MESES_ES = {1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
                 7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"}


def _acc_desc_articulo_file():
    """Archivo de productos crudo (01_INPUTS/RAW_PRODUCTOS/productos<mes>.xlsx). Prefiere el
    del mes en curso por nombre; si no, el más reciente por mtime.
    Ignora los `_NO_USAR_*` (exports viejos/inflados): si uno quedara con mtime nuevo, sería
    elegido como maestro del mes y cuelga el cierre (ver raw de mayo, 19MB con filas fantasma)."""
    base = INPUTS / "RAW_PRODUCTOS"
    if not base.exists():
        return None
    xls = [p for p in base.glob("*.xlsx")
           if not p.name.startswith("~$") and not p.name.startswith("_NO_USAR_")]
    if not xls:
        return None
    mes = _ACC_MESES_ES.get(datetime.now(_ARG_TZ).month, "")
    delmes = [p for p in xls if mes and mes in p.name.lower()]
    cand = delmes or xls
    return max(cand, key=lambda p: p.stat().st_mtime)


def _acc_desc_articulo_map():
    """Lee el archivo de productos crudo y mapea Código Art. -> {descripcion, linea, categoria}.
    La descripción (columna H 'Descripción Art.') se muestra en la tarjeta de la acción para que
    el vendedor sepa el producto exacto. Header autodetectado. Cacheado por (path, mtime)."""
    p = _acc_desc_articulo_file()
    if p is None:
        return {}
    try:
        key = (str(p), os.path.getmtime(p))
    except OSError:
        key = (str(p), 0)
    cached = _ACC_DESC_ART_CACHE.get(key)
    if cached is not None:
        return cached
    mapa = {}
    try:
        raw = pd.read_excel(p, header=None, dtype=str)
        hdr, cols = None, None
        for i, row in raw.iterrows():
            vals = [str(x).strip() for x in row.tolist()]
            if (any(v.startswith("Código Art") or v.startswith("Codigo Art") for v in vals)
                    and any(v.startswith("Descripci") for v in vals)):
                hdr, cols = i, vals
                break
        if hdr is None:
            return {}
        df = raw.iloc[hdr + 1:].copy()
        df.columns = cols
        ccod = next(c for c in cols if c.startswith("Código Art") or c.startswith("Codigo Art"))
        cdesc = next(c for c in cols if c.startswith("Descripci"))
        clin = next((c for c in cols if c.strip().lower().startswith("linea comercial")), None)
        ccat = next((c for c in cols if c.startswith("Categor")), None)
        for _, r in df.iterrows():
            cod = _re.sub(r"\.0$", "", str(r.get(ccod, "")).strip())
            if not cod or cod.lower() == "nan":
                continue
            desc = str(r.get(cdesc, "") or "").strip()
            mapa[cod] = {
                "descripcion": "" if desc.lower() == "nan" else desc,
                "linea": str(r.get(clin, "") or "").strip() if clin else "",
                "categoria": str(r.get(ccat, "") or "").strip() if ccat else "",
            }
    except Exception:
        return {}
    _ACC_DESC_ART_CACHE.clear()
    _ACC_DESC_ART_CACHE[key] = mapa
    return mapa


def _acc_universo_productos(v_act=None):
    """Universo de productos para el buscador de acciones: SKUs del maestro de productos
    activos (01_INPUTS/RAW_PRODUCTOS/productos<mes>.xlsx) + los códigos vendidos en el mes.

    Cada item trae `_args`, exactamente los cinco argumentos con los que se evalúa el
    predicado de una acción sobre una línea de venta (_acc_product_pred): categoría canónica
    y línea comercial SALEN DEL 04D (igual que _acc_preparar_from_df), no del archivo de
    productos. Si un código no está en el 04D, su categoría queda vacía y no matchea las
    reglas por categoría — que es lo que hoy pasa con su venta real. Se marca en_maestro=False
    en vez de adivinarle la categoría, para que el faltante se vea en lugar de taparse.
    """
    cod2cat, _cod2seg, _cod2lxu, cod2linea = _cargar_maestro_04D()
    desc = _acc_desc_articulo_map()
    art_ventas, marca_ventas = {}, {}
    if v_act is not None and not v_act.empty:
        for cod, art, mar in zip(v_act["_cod"], v_act["_art"], v_act["_marca"]):
            c = str(cod).strip()
            if c and c not in art_ventas:
                art_ventas[c] = str(art or "").strip()
                marca_ventas[c] = str(mar or "").strip()

    def _clean(x):
        s = str(x or "").strip()
        return "" if s.lower() == "nan" else s

    out = []
    for cod in sorted(set(desc) | set(art_ventas)):
        d = desc.get(cod, {})
        linea04 = _clean(cod2linea.get(cod))
        linea = linea04 or _clean(d.get("linea"))
        producto = _clean(d.get("descripcion")) or art_ventas.get(cod, "")
        # marca mostrada = Línea Comercial (marca canónica del maestro). La columna Marca de
        # ventas se guarda aparte (alias): alimenta el predicado y el buscador, pero no la UI.
        alias = marca_ventas.get(cod, "")
        marca = linea or alias
        cat_canon = _acc_canon_cat(cod2cat.get(cod))   # None si el código no está en el 04D
        out.append({
            "codigo": cod,
            "producto": producto or linea or cod,
            "marca": marca,
            "alias": alias if alias.upper() != marca.upper() else "",
            "linea": linea,
            "categoria": _clean(cod2cat.get(cod)) or _clean(d.get("categoria")),
            "cat_canon": cat_canon or "",
            "en_maestro": cod in cod2cat,
            "_args": (cat_canon, linea04.upper(), (producto or "").upper(),
                      (alias or marca).upper(), cod),
        })
    return out


def _acc_item_cats(txt):
    """Categorías canónicas del maestro que menciona un texto de detalle (para expandir
    items FILTRO_MAESTRO/FILTRO_EXCLUSION a marcas reales)."""
    t = _acc_norm(txt)
    cats = set()
    for kw, canon in (("VINOS DEL", "VDA"), ("VDA", "VDA"), ("VINOS DE GUARDA", "VDG"),
                      ("VDG", "VDG"), ("ESPUMANTE", "ESPUMANTES"), ("SIDRA", "SIDRA"),
                      ("SPIRIT", "SPIRITS"), ("CERVEZA", "CERVEZA")):
        if kw in t:
            cats.add(canon)
    return cats


def _acc_enriquecer_grupo(g, cat_to_marcas, cod2seg, cod2linea, cod2cat=None):
    """Copia el grupo de DETALLE_CATEGORIAS agregando a cada item la lista `marcas`
    (con segmento) resuelta del maestro. Reglas por tipo_detalle:
      - FILTRO_MAESTRO / FILTRO_EXCLUSION → todas las marcas de la(s) categoría(s).
      - FAMILIA / MARCA_EXPLICITA → marcas del maestro cuya Linea Comercial matchea el token,
        RESTRINGIDAS a la categoría del grupo si ésta se puede inferir (ej. 'Smirnoff' bajo
        Spirits NO trae Smirnoff Ice, que es RTD).
      - resto (PRODUCTO_EXPLICITO, PRODUCTO_O_FAMILIA, SUBREGLA, EXCLUSION) → literal."""
    cod2cat = cod2cat or {}
    grp_cats = _acc_item_cats(f"{g.get('categoria_tarjeta','')} {g.get('detalle_click_ref','')}")
    items_out = []
    for it in g.get("items", []):
        tipo = _acc_norm(it.get("tipo_detalle"))
        ml = str(it.get("marca_o_linea", "")).strip()
        marcas = []
        if "FILTRO" in tipo or "MAESTRO" in tipo:
            cats = _acc_item_cats(f"{ml} {g.get('categoria_tarjeta','')} {g.get('detalle_click_ref','')}")
            seen = set()
            for c in cats:
                for m in cat_to_marcas.get(c, []):
                    k = (m["segmento"].upper(), m["marca"].upper())
                    if k not in seen:
                        seen.add(k); marcas.append(m)
        elif tipo in ("FAMILIA", "MARCA EXPLICITA"):
            tok = _acc_norm(ml.replace("FAMILIA", "").replace("FLIA", ""))
            seen = set()
            for cod, ln in cod2linea.items():
                lnn = _acc_norm(ln)
                if tok and lnn and (tok in lnn or lnn in tok):
                    if grp_cats and _acc_canon_cat(cod2cat.get(cod)) not in grp_cats:
                        continue  # respeta la categoría del grupo (excluye p.ej. Ice/RTD)
                    seg = str(cod2seg.get(cod, "") or "").strip()
                    if seg.lower() == "nan":
                        seg = ""
                    k = (seg.upper(), _acc_norm(ln))
                    if k not in seen:
                        seen.add(k); marcas.append({"segmento": seg, "marca": str(ln).strip()})
            marcas.sort(key=lambda x: (x["segmento"], x["marca"]))
        item = dict(it)
        item["marcas"] = marcas  # vacío → el front muestra el label literal (marca_o_linea)
        items_out.append(item)
    return {"detalle_click_ref": g.get("detalle_click_ref", ""),
            "categoria_tarjeta": g.get("categoria_tarjeta", ""),
            "items": items_out}


def _acc_catalogo_mes():
    """Auto-detecta el catálogo del MES EN CURSO (fallback: más reciente disponible).
    Devuelve (mes, fuente, lista_reglas). El catálogo es el primer CSV del mes que NO
    sea el detalle_categorias ni salidas/backups."""
    import csv as _csvm
    mdir = _acc_mes_dir()
    if mdir is None:
        return None, None, []
    fuente = None
    for c in sorted(mdir.glob("*.csv")):
        low = str(c).lower()
        if ("salida" in low or "_backup" in low
                or c.name.lower().startswith("detalle_categorias")):
            continue
        fuente = c; break
    if fuente is None:
        return mdir.name, None, []
    try:
        with open(fuente, encoding="utf-8-sig", newline="") as f:
            reglas = list(_csvm.DictReader(f, delimiter=";"))
    except Exception:
        return mdir.name, fuente.name, []
    return mdir.name, fuente.name, reglas


def _acc_seg_canon(seg_text, canal_text):
    t = (str(seg_text or "") + " | " + str(canal_text or "")).upper()
    out = set()
    if "TRADICIONAL" in t or "TRAD" in t or "KIOSCO" in t:
        out.add("TRADICIONAL")
    if "AUTOSERVICIO" in t or _re.search(r"\bAS\b", t):
        out.add("AUTOSERVICIO")
    if "ON PREMISE" in t or "ON_PREMISE" in t or "VTK" in t or "TDB" in t or "VINOTECA" in t:
        out.add("ON_PREMISE_VTK")
    # MAYORISTA es su propio canon: una acción de Petit Mayoristas NO debe caer sobre
    # autoservicios (el cliente mayorista se detecta aparte por Ramo/Subramo, _es_mayorista).
    if "MAYORISTA" in t:
        out.add("MAYORISTA")
    if "PROXIMITY" in t or "ESTACION DE SERVICIO" in t:
        out.add("PROXIMITY")
    if not out:
        # Sin canal declarado = la accion aplica a todos, Proximity incluido.
        out = {"TRADICIONAL", "AUTOSERVICIO", "ON_PREMISE_VTK", "MAYORISTA", "PROXIMITY", "OTROS"}
    return out


def _acc_seg_match(row_seg, row_may, rule_segs):
    """¿La fila (segmento clasificado + flag mayorista) cae bajo los segmentos de la regla?
    - Cliente mayorista → solo si la regla apunta a MAYORISTA.
    - Cliente NO mayorista → si su segmento está en la regla (el token MAYORISTA nunca
      iguala un segmento clasificado, así que no lo alcanza)."""
    if row_may:
        return "MAYORISTA" in rule_segs
    return row_seg in rule_segs


# Subtipos de tradicional para acciones que aplican SOLO a almacén/despensa/kiosco.
# Despensa = Almacén (regla de negocio): se canoniza a "ALMACEN" en todas las estadísticas.
_ACC_SUBSEG_TRAD = {"ALMACEN": "ALMACEN", "ALMACENES": "ALMACEN", "DESPENSA": "ALMACEN",
                    "KIOSCO": "KIOSCO", "MAXIKIOSCO": "KIOSCO"}


def _acc_subseg_filtro(seg_text, canal_text):
    """Sub-filtro dentro del canon TRADICIONAL por Subramo de la venta.
    Si la regla nombra subtipos específicos (almacén/despensa/kiosco) SIN el genérico
    de canal ('tradicional'/'trad'), devuelve el set de tokens permitidos; si no, None
    (la regla aplica a todo el canal tradicional, sin sub-restricción)."""
    t = _acc_norm(str(seg_text or "") + " " + str(canal_text or ""))
    if (_re.search(r"\bTRADICIONAL(?:ES)?\b", t) or _re.search(r"\bTRADITIONAL\b", t)
            or _re.search(r"\bTRAD\b", t)):
        return None
    allowed = {tok for kw, tok in _ACC_SUBSEG_TRAD.items() if kw in t}
    return allowed or None


def _acc_norm(s):
    """Normaliza para matchear marcas: mayúsculas, sin acentos, sin puntuación.
    Los apóstrofes/comillas se ELIMINAN (no se vuelven espacio) para que 'Gordon's'
    y 'Gordons' matcheen igual (el catálogo usa 'Gordons', las ventas 'Gordon's')."""
    import unicodedata
    s = str(s or "").upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = _re.sub(r"[\'’‘´`]", "", s)   # apóstrofes/comillas → nada
    s = _re.sub(r"[^A-Z0-9 ]", " ", s)
    return _re.sub(r"\s+", " ", s).strip()


def _acc_lineas_de_marca(token, all_lineas):
    """Devuelve set de líneas comerciales (NORMALIZADAS) que corresponden a la marca del token."""
    base = _acc_norm(token.replace(" VINO", ""))
    wine_only = token.upper().endswith(" VINO") or token.upper().strip() == "VINO"
    out = set()
    if not base:
        return out
    for ln in all_lineas:
        lnn = _acc_norm(ln)
        if lnn == base or lnn.startswith(base + " "):
            if wine_only and ("SIDRA" in lnn or "CHAMPA" in lnn or "ESPUMA" in lnn):
                continue
            out.add(lnn)
    return out


def _acc_once_titulares_tokens():
    """Marcas 11T vigentes desde objetivo 11T.xlsx, normalizadas para match contra ventas.csv."""
    global _ACC_11T_CACHE
    if _ACC_11T_CACHE is not None:
        return _ACC_11T_CACHE
    toks = []
    p = INPUTS / "objetivo 11T.xlsx"
    if p.exists():
        try:
            df = pd.read_excel(p)
            col = next((c for c in df.columns if "linea" in str(c).lower()), None)
            if col is None and len(df.columns) > 1:
                col = df.columns[1]
            vals = df[col].dropna().tolist() if col is not None else []
            for v in vals:
                n = _acc_norm(v).replace("SIMRNOFF", "SMIRNOFF")
                if n and n != "LINEA COMERCIAL":
                    toks.append(n)
        except Exception:
            toks = []
    if not toks:
        toks = [_acc_norm(x) for x in [
            "Alma Mora", "Trapiche Reserva", "Finca Las Moras", "Alaris", "Don David",
            "Dada", "Smirnoff Flavors", "Los Arboles", "Antares", "Smirnoff Ice",
            "Gordons Flavours",
        ]]
    _ACC_11T_CACHE = set(toks)
    return _ACC_11T_CACHE


def _acc_innovaciones_codigos():
    """Codigos de la lista cerrada de Innovaciones.xlsx."""
    global _ACC_INNOV_CACHE
    if _ACC_INNOV_CACHE is not None:
        return _ACC_INNOV_CACHE
    out = set()
    p = INPUTS / "INNOVACIONES" / "Innovaciones.xlsx"
    if p.exists():
        try:
            df = pd.read_excel(p, sheet_name=0, header=None, dtype=str)
            for val in df.stack().dropna().astype(str):
                for m in _re.finditer(r"(?:0{3,})?(\d{5})\s*-", val):
                    out.add(m.group(1).lstrip("0"))
        except Exception:
            out = set()
    _ACC_INNOV_CACHE = out
    return _ACC_INNOV_CACHE


# Tolerancia al comparar el % aplicado contra el tramo de la acción. Los % que llegan de
# ventas.csv son enteros limpios (1763/1770 líneas del mes caen en entero exacto), así que
# 0.5 pp solo absorbe redondeo; no acerca un tramo a otro (el más cercano es 6 vs 7).
_ACC_PCT_TOL = 0.5


def _acc_tramos_pct(rule):
    """Tramos de descuento de la regla, desde `descuento_pct` ('6|7|8' -> [6.0, 7.0, 8.0]).
    [] en las de bonificación/sin cargo (no tienen % que matchear)."""
    return [float(x) for x in str(rule.get("descuento_pct", "")).replace(",", ".").split("|")
            if x.strip().replace(".", "").isdigit()]


def _acc_mask_usa_accion(df, tramos):
    """¿La línea de venta USÓ esta acción? = el % de descuento aplicado coincide con alguno de
    los tramos de la acción.

    Sin esto, la tarjeta contaba a TODO el que compró el producto del alcance (con descuento
    ajeno o sin descuento) como si hubiera usado la acción: ACJ26-029 mostraba 32 clientes y
    $22.769 cuando sus únicos descuentos eran de 3% y 5% (de la acción de Spirits), y ninguno
    del 10% de la acción. También inflaba la inversión por doble conteo: el mismo descuento
    caía en todas las tarjetas cuyo alcance de producto lo alcanzaba.

    Bonificación / sin cargo (sin tramos): no hay % que comparar → cae a "tiene descuento"."""
    if not len(df):
        return pd.Series(dtype=bool, index=df.index)
    if not tramos:
        return df["_desc"] > 0
    pct = df["_pct"]
    mask = pd.Series(False, index=df.index)
    for t in tramos:
        mask |= (pct - t).abs() <= _ACC_PCT_TOL
    return mask


def _acc_plan_as_flags(rule):
    """(requiere_plan_as, excluye_plan_as) de una regla del catálogo.
    - requiere: la acción es SOLO para autoservicios CON plan AASS (ej. ACJ26-015..018, -025).
    - excluye: la acción corre en todos los segmentos MENOS los AASS con plan (ej. ACJ26-024:
      esos clientes tienen su propio tramo del 12%). Se chequea PRIMERO porque el texto de
      exclusión ("menos AASS con planes" / "sin plan AASS") contiene la misma frase que el
      de requerimiento y si no daría el filtro exactamente al revés."""
    txt = _acc_norm(" ".join(str(rule.get(k, "")) for k in (
        "categoria", "canal_aplica", "segmento_cliente_aplica")))
    excluye = ("MENOS AASS CON PLANES" in txt or "SIN PLAN AASS" in txt
               or "SIN PLANES AASS" in txt or "TODOS SIN PLAN AASS" in txt)
    if excluye:
        return False, True
    return ("PLANES AASS" in txt or "PLAN AASS" in txt), False


def _acc_plan_as_clientes():
    """Clientes con Plan AS vigente desde mod_planes_as.csv."""
    global _ACC_PLAN_AS_CACHE
    if _ACC_PLAN_AS_CACHE is not None:
        return _ACC_PLAN_AS_CACHE
    df = read_csv(DATASETS / "mod_planes_as.csv")
    if df.empty or "cliente_id" not in df.columns:
        _ACC_PLAN_AS_CACHE = set()
    else:
        _ACC_PLAN_AS_CACHE = set(pd.to_numeric(df["cliente_id"], errors="coerce").dropna().astype(int))
    return _ACC_PLAN_AS_CACHE


def _acc_es_botella(cat_canon, articulo):
    """True si el producto es BOTELLA (no lata/RTD). Excluye la categoría RTD (Smirnoff Ice,
    Gordon's Tonic, etc.) y descripciones con marcador de lata. Para acciones acotadas a
    'solo botella' (ej. la caja mixta Smirnoff+Gordon's 15%: 3 botellas de cada uno)."""
    if str(cat_canon or "").upper() == "RTD":
        return False
    a = _acc_norm(articulo)
    if _re.search(r"\bLATA?\b", a):
        return False
    return True


def _acc_product_pred(rule, all_lineas):
    """Devuelve función(cat_canon, linea, articulo, marca, cod=None)->bool para esta regla.
    Si la acción menciona 'botella' (condición/unidad), el match se restringe a botella."""
    regla_txt = _acc_norm(" ".join(str(rule.get(k, "")) for k in (
        "categoria", "subcategoria", "productos_marcas", "lineas_comerciales", "tipo_regla"
    )))
    if "11 TITULARES" in regla_txt:
        marcas_11t = _acc_once_titulares_tokens()
        def pred_11t(cat_canon, linea, articulo, marca, cod=None):
            txt = _acc_norm(f"{linea or ''} {articulo or ''} {marca or ''}")
            return any(tok in txt for tok in marcas_11t)
        return pred_11t
    if "LISTA CERRADA DE INNOVACIONES JUNIO 2026" in regla_txt or "INNOVACIONES LISTADAS" in regla_txt:
        codigos = _acc_innovaciones_codigos()
        def pred_innov(cat_canon, linea, articulo, marca, cod=None):
            return bool(codigos) and cod is not None and str(cod).strip() in codigos
        return pred_innov

    raw = (str(rule.get("productos_marcas", "")) + ";" + str(rule.get("lineas_comerciales", ""))).upper()
    toks = [t.strip() for t in raw.replace(",", ";").split(";") if t.strip()]
    line_cats, brand_lineas = set(), set()
    brand_toks = []
    code_set = set()   # códigos de producto exactos (ej. "35103"): acción dirigida a un SKU
    has_resto = any("RESTO" in t for t in toks)
    for t in toks:
        if t.isdigit():   # token numérico = código de producto exacto
            code_set.add(t)
            continue
        # "TODOS ..." (incl. descripciones como "Todos menos importados premium...") es
        # genérico, NO una marca: si se tratara como marca, su token basura cortaría el
        # match por categoría (line_cats) y la acción no matchearía nada (ej. ACJ26-007).
        if (t in _ACC_PROD_GENERICOS or "MAESTRO" in t or "LISTA CERRADA" in t
                or "RESTO" in t or t.startswith("TODOS")):
            continue
        mapped = _ACC_LINEA_TOK.get(t)   # match exacto de token de línea (no substring)
        if mapped:
            line_cats.add(mapped)
        else:
            brand_toks.append(t.replace(" VINO", "").strip())
            brand_lineas |= _acc_lineas_de_marca(t, all_lineas)

    brand_norm = [b for b in (_acc_norm(x) for x in brand_toks) if b]

    def pred(cat_canon, linea, articulo, marca, cod=None):
        if code_set and cod is not None and str(cod).strip() in code_set:
            return True
        if has_resto and not brand_toks and not line_cats and not code_set:
            return True
        if brand_lineas or brand_norm:
            if linea and brand_lineas and _acc_norm(linea) in brand_lineas:
                return True
            am = _acc_norm(str(articulo or "") + " " + str(marca or ""))
            return any(bt in am for bt in brand_norm)
        if line_cats:
            return cat_canon in line_cats
        return False

    # Acción "solo botella" (ej. caja mixta 3 botellas + 3 botellas): excluye latas/RTD.
    # Si la acción lista códigos explícitos, éstos ya están curados → no se aplica el heurístico.
    solo_botella = (not code_set) and "BOTELLA" in _acc_norm(" ".join(str(rule.get(k, "")) for k in (
        "condicion_compra", "unidad_minimo", "unidad_maximo", "subcategoria")))
    if not solo_botella:
        return pred

    def pred_botella(cat_canon, linea, articulo, marca, cod=None):
        return pred(cat_canon, linea, articulo, marca, cod) and _acc_es_botella(cat_canon, articulo)
    return pred_botella


_ACC_VENTAS_CACHE = {}
def _acc_preparar_from_df(df):
    """Computa las columnas de acciones/alertas (_cli,_vend,_cat,_linea,_seg,_subseg,
    _litros,_desc,_pct,_imp_neto,_cant,_mes,_fcomp,...) a partir de un df crudo de ventas
    (cualquier fuente: ventas.csv viva o ventas_mes versionado del cierre). Sin caché ni I/O."""
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    df.columns = [c.strip() for c in df.columns]
    # No se filtra por Empresa: medimos con las dos razones sociales (ver _LEEME_EMPRESA).
    cod2cat, cod2seg, cod2lxu, cod2linea = _cargar_maestro_04D()
    def _n(s):
        return pd.to_numeric(s.astype(str).str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
                             errors="coerce").fillna(0)
    out = pd.DataFrame()
    out["_cli"]  = pd.to_numeric(df.get("Cliente"), errors="coerce")
    out["_vend"] = pd.to_numeric(df.get("CodVendedor"), errors="coerce")
    out["_cod"]  = df.get("Codigo", "").astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    out["_imp_neto"] = _n(df.get("ImporteNetoItem", pd.Series(["0"] * len(df))))
    out["_imp_item"] = _n(df.get("ImporteItem", pd.Series(["0"] * len(df))))
    out["_cant"] = _n(df.get("CantBase", pd.Series(["0"] * len(df))))
    # Descuento REAL = valorDescuento (por unidad) × CantBase.
    # NO usar ImporteItem-ImporteNetoItem: esa diferencia es IVA (~17.4% en TODAS las líneas), no descuento.
    out["_vd"] = _n(df.get("valorDescuento", pd.Series(["0"] * len(df))))
    out["_desc"] = (out["_vd"] * out["_cant"]).clip(lower=0)
    out["_pct"] = (out["_desc"] / (out["_imp_neto"] + out["_desc"]).replace(0, pd.NA) * 100).fillna(0).round(1)
    out["_lxu"]  = out["_cod"].map(cod2lxu).fillna(0)
    # Litros por la cascada única (maestro → PesoKg → ml del nombre): un SKU que el maestro no
    # trae no puede aportar 0 L a una acción. Misma fuente que sell out y la ficha de cliente.
    out["_litros"] = _litros_por_linea(df)
    out["_cat"]  = out["_cod"].map(cod2cat).map(_acc_canon_cat)
    out["_linea"] = out["_cod"].map(cod2linea).astype(str).str.upper().str.strip()
    out["_art"]  = df.get("Articulo", "").astype(str).str.upper()
    out["_marca"] = df.get("Marca", "").astype(str).str.upper()
    out["_clinom"] = (df["RazonSocial"].astype(str) if "RazonSocial" in df.columns
                      else df.get("Cliente", pd.Series([""] * len(df))).astype(str))
    out["_dir"] = df.get("Direccion", pd.Series([""] * len(df))).astype(str)
    out["_loc"] = df.get("Localidad", pd.Series([""] * len(df))).astype(str)
    out["_vnom"] = (df["Vendedor"].astype(str) if "Vendedor" in df.columns
                    else pd.Series([""] * len(df), index=df.index))
    _subr = df.get("Subramo", pd.Series([""] * len(df)))
    out["_seg"]  = [
        _clasificar_segmento(str(r), str(s))
        for r, s in zip(df.get("Ramo", pd.Series([""] * len(df))), _subr)
    ]
    # Subramo normalizado: sub-filtro de acciones acotadas a almacén/kiosco.
    # Despensa = Almacén (regla de negocio): se canoniza despensa→almacén en todas las estadísticas.
    out["_subseg"] = [_acc_norm(s).replace("DESPENSA", "ALMACEN") for s in _subr]
    # Cliente Petit Mayorista (por Ramo/Subramo): se separa de AUTOSERVICIO para que las
    # acciones de MAYORISTA no caigan sobre autoservicios (y viceversa). El clasificador
    # global mete al mayorista en AUTOSERVICIO; acá se detecta aparte, sin tocarlo.
    _ramo_txt = df.get("Ramo", pd.Series([""] * len(df))).astype(str).str.upper()
    out["_es_mayorista"] = (_ramo_txt + " " + _subr.astype(str).str.upper()).str.contains("MAYORISTA", na=False)
    _fc = pd.to_datetime(df.get("FechaComprobante"), dayfirst=True, errors="coerce")
    out["_mes"]  = _fc.dt.to_period("M")
    out["_fcomp"] = _fc.dt.strftime("%d/%m/%Y").fillna("")
    out["_fcarga"] = (pd.to_datetime(df.get("FechaCarga"), dayfirst=True, errors="coerce")
                      .dt.strftime("%d/%m/%Y").fillna(""))
    out = out[(out["_imp_neto"] > 0) & (~out["_vend"].isin(_VENDEDORES_EXCLUIDOS))]
    return out


def _acc_preparar_ventas(nombre="ventas.csv"):
    """Ventas preparadas para acciones/alertas desde 01_INPUTS/<nombre> (sep=';', mes vivo).
    Por defecto ventas.csv. ventas_acumulada.csv solo para el comparativo de 'clientes nuevos'."""
    p = INPUTS / nombre
    if not p.exists():
        return pd.DataFrame()
    try:
        key = (nombre, os.path.getmtime(p))
    except OSError:
        key = (nombre, 0)
    cached = _ACC_VENTAS_CACHE.get(key)
    if cached is not None:
        return cached
    df = None
    for enc in ("latin1", "utf-8-sig", "windows-1252"):
        try:
            df = pd.read_csv(p, sep=";", encoding=enc, dtype=str, low_memory=False)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    out = _acc_preparar_from_df(df)
    # cachear por (archivo, mtime); purgar mtimes viejos del mismo archivo
    for k in [k for k in _ACC_VENTAS_CACHE if k[0] == nombre]:
        _ACC_VENTAS_CACHE.pop(k, None)
    _ACC_VENTAS_CACHE[key] = out
    return out


def _acc_preparar_ventas_mes_versionado(path):
    """Igual que _acc_preparar_ventas pero sobre el ventas_mes CONGELADO del cierre
    (01_INPUTS/cierres mes/ventas_mes_<MMAAAA>.csv, sep=',', utf-8-sig). Cacheado por path+mtime."""
    try:
        key = ("cierre:" + str(path), os.path.getmtime(path))
    except OSError:
        key = ("cierre:" + str(path), 0)
    cached = _ACC_VENTAS_CACHE.get(key)
    if cached is not None:
        return cached
    df = None
    for enc in ("utf-8-sig", "latin1", "windows-1252"):
        try:
            df = pd.read_csv(path, sep=",", quotechar='"', engine="python", dtype=str, encoding=enc)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    out = _acc_preparar_from_df(df)
    _ACC_VENTAS_CACHE[key] = out
    return out


_ACC_MES_CACHE = {}
_ACC_ON_CACHE = {}

def _acc_on_file(mdir=None):
    """xlsx de acciones ON del mes: 01_INPUTS/ACCIONES COMERCIALES/<mes>/*ON.xlsx
    (autodetecta por sufijo 'ON' antes de la extensión). None si no existe."""
    mdir = mdir or _acc_mes_dir()
    if not mdir or not mdir.exists():
        return None
    cands = sorted(c for c in mdir.glob("*.xlsx")
                   if c.stem.lower().endswith("on") and not c.name.startswith("~$"))
    return cands[-1] if cands else None


def _acc_on_cards():
    """Tarjetas informativas de las acciones ON (On Premise / Vinotecas / Tienda de
    Bebidas / Catering) desde el xlsx ...penaflorON.xlsx del mes. Hoja 01_Acciones =
    cabeceras (mecánica, compra, beneficio, tope, canales); hoja 02_Detalle_Productos =
    productos elegibles por subcanal y LC (para el detalle al click). Son combos de
    incorporación sin cargo: se muestran como catálogo, sin footprint de inversión.
    Cacheado por mtime; [] si no hay archivo ON en el mes."""
    path = _acc_on_file()
    if not path:
        return []
    try:
        key = os.path.getmtime(path)
    except OSError:
        key = 0
    cached = _ACC_ON_CACHE.get(key)
    if cached is not None:
        return cached
    try:
        xls = pd.ExcelFile(path)
        acc_df = pd.read_excel(xls, "01_Acciones")
        det_df = pd.read_excel(xls, "02_Detalle_Productos")
    except Exception as e:
        print(f"[WARN] acc_on_cards: {e}")
        return []

    def _s(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        return _re.sub(r"\s+", " ", str(v).replace("\xa0", " ")).strip()

    # Detalle: action_id -> subcanal -> lc_visual -> [productos]
    det = {}
    for _, r in det_df.iterrows():
        aid = _s(r.get("action_id"))
        if not aid:
            continue
        sub = _s(r.get("subcanal")) or "—"
        lc = _s(r.get("lc_visual")) or "—"
        cod = r.get("codigo_art")
        cod = "" if pd.isna(cod) else str(cod).split(".")[0].strip()
        det.setdefault(aid, {}).setdefault(sub, {}).setdefault(lc, []).append(
            {"codigo": cod, "descripcion": _s(r.get("descripcion_art")),
             "estado": _s(r.get("estado_producto"))})

    cards = []
    for _, r in acc_df.iterrows():
        aid = _s(r.get("action_id"))
        if not aid:
            continue
        canales = [c.strip() for c in _re.split(r"[;,]", _s(r.get("canales"))) if c.strip()]
        grupos, n_prod = [], 0
        for sub, lcs in det.get(aid, {}).items():
            lineas = []
            for lc, prods in lcs.items():
                lineas.append({"lc": lc, "productos": prods})
                n_prod += len(prods)
            grupos.append({"subcanal": sub, "n_productos": sum(len(l["productos"]) for l in lineas),
                           "lineas": lineas})
        try:
            orden = int(float(_s(r.get("orden_ppt"))))
        except (TypeError, ValueError):
            orden = 999
        cards.append({
            "id_accion": _s(r.get("tarjeta_codigo")) or aid,
            "action_id": aid,
            "orden": orden,
            "titulo": _s(r.get("tarjeta_titulo")),
            "categoria": _s(r.get("categoria")),
            "segmento_lc": _s(r.get("segmento_lc")),
            "canales": canales,
            "mecanica": _s(r.get("mecanica")),
            "compra_requerida": _s(r.get("compra_requerida")),
            "beneficio": _s(r.get("beneficio")),
            "tope": _s(r.get("tope_combo_pdv")),
            "validacion": _s(r.get("validacion")),
            "n_productos": n_prod,
            "grupos": grupos,
        })
    cards.sort(key=lambda c: c["orden"])
    _ACC_ON_CACHE.clear()
    _ACC_ON_CACHE[key] = cards
    return cards


def _acc_mes_sig():
    """Firma de invalidación del payload de acciones: mtime de las fuentes reales."""
    sig = []
    for p in (INPUTS / "ventas.csv", INPUTS / "ventas_acumulada.csv",
              DATASETS / "mod_planes_as.csv", CONFIG / "maestro_04D_productos.csv",
              INPUTS / "04D_MAESTRO_PRODUCTOS_PENAFLOR.xlsx"):
        try:
            sig.append((p.name, os.path.getmtime(p) if p.exists() else 0))
        except OSError:
            sig.append((p.name, 0))
    # archivo de productos crudo (descripción de artículo para las tarjetas por código)
    _pf = _acc_desc_articulo_file()
    try:
        sig.append(("raw_productos", os.path.getmtime(_pf) if _pf else 0))
    except OSError:
        sig.append(("raw_productos", 0))
    base = INPUTS / "ACCIONES COMERCIALES"
    try:
        cat = max((os.path.getmtime(c) for sub in base.iterdir() if sub.is_dir()
                   for c in sub.glob("*.csv")), default=0) if base.exists() else 0
    except OSError:
        cat = 0
    sig.append(("acc_cat", cat))
    # Archivo de acciones ON del mes (xlsx): al resubirlo, el payload se refresca solo.
    try:
        onf = _acc_on_file()
        sig.append(("acc_on", os.path.getmtime(onf) if onf else 0))
    except OSError:
        sig.append(("acc_on", 0))
    # Mes en curso en la firma: al cambiar de mes el payload se recalcula solo aunque
    # el proceso de Render lleve semanas vivo y ningún archivo haya cambiado su mtime.
    sig.append(("mes_actual", datetime.now(_ARG_TZ).strftime("%Y-%m")))
    return tuple(sig)


def _acciones_mes_payload(vid_filtro=None):
    """Wrapper cacheado por (firma de fuentes, vendedor). Evita recalcular el payload
    (lectura de ventas + apply por regla, ~18s) en cada login; el cómputo real está en
    _acciones_mes_payload_uncached. Se invalida cuando cambia el mtime de alguna fuente."""
    ckey = (_acc_mes_sig(), vid_filtro)
    cached = _ACC_MES_CACHE.get(ckey)
    if cached is not None:
        return cached
    payload = _acciones_mes_payload_uncached(vid_filtro)
    # purgar firmas viejas; conservar variantes por vendedor de la firma actual
    for k in [k for k in _ACC_MES_CACHE if k[0] != ckey[0]]:
        _ACC_MES_CACHE.pop(k, None)
    _ACC_MES_CACHE[ckey] = payload
    return payload


def _acciones_mes_payload_uncached(vid_filtro=None):
    mes, fuente, reglas = _acc_catalogo_mes()
    if not reglas:
        return {"mes": mes, "fuente": fuente, "acciones": [], "nota": "Sin catálogo de acciones del mes."}
    v_cur = _acc_preparar_ventas("ventas.csv")   # MES VIVO
    if v_cur.empty:
        return {"mes": mes, "fuente": fuente, "acciones": [], "nota": "Sin ventas para calcular."}
    # ventas_acumulada solo aporta el mes ANTERIOR para el comparativo de clientes nuevos
    v_acum = _acc_preparar_ventas("ventas_acumulada.csv")

    # mes objetivo = el del catálogo; mes anterior = el previo
    try:
        per_actual = pd.Period(mes, freq="M")
    except Exception:
        per_actual = v_cur["_mes"].max()
    per_ant = per_actual - 1
    v_act = v_cur[v_cur["_mes"] == per_actual]
    if v_act.empty:
        v_act = v_cur   # ventas.csv ya es el mes vivo
    v_ant = v_acum[v_acum["_mes"] == per_ant] if not v_acum.empty else v_cur.iloc[0:0]
    _lin = [v_cur["_linea"]] + ([v_acum["_linea"]] if not v_acum.empty else [])
    all_lineas = set(l for l in pd.concat(_lin).dropna().unique() if l and l != "NAN")
    plan_as_clientes = _acc_plan_as_clientes()
    detalle_map = _acc_detalle_map()   # {} en meses sin detalle (esquema junio)
    _cat_to_marcas = _acc_marcas_maestro() if detalle_map else {}
    _m_cod2cat, _m_cod2seg, _m_cod2lxu, _m_cod2linea = (_cargar_maestro_04D() if detalle_map
                                                        else ({}, {}, {}, {}))
    _desc_art_map = _acc_desc_articulo_map()   # cod -> {descripcion, linea, categoria}
    # Universo de productos + índice producto -> acciones (buscador de la pantalla).
    universo = _acc_universo_productos(v_act)
    prod_hits = {}

    def _detalle_clientes(df, nuevos=None):
        if df.empty:
            return []
        nuevos = set(nuevos or [])
        rows = []
        for cli, g in df.groupby("_cli", dropna=True):
            try:
                cli_int = int(cli)
            except Exception:
                continue
            vend = pd.to_numeric(g["_vend"], errors="coerce").dropna()
            vend_int = int(vend.iloc[0]) if not vend.empty else None
            fechas = pd.to_datetime(g["_fcomp"], dayfirst=True, errors="coerce")
            ult = fechas.max()
            rows.append({
                "cliente_id": cli_int,
                "cliente_nombre": str(g["_clinom"].iloc[0]),
                "direccion": str(g["_dir"].iloc[0]) if "_dir" in g.columns else "",
                "localidad": str(g["_loc"].iloc[0]) if "_loc" in g.columns else "",
                "vendedor_codigo": vend_int,
                "vendedor_id": f"V{vend_int}" if vend_int is not None else "",
                "vendedor_nombre": str(g["_vnom"].iloc[0]) if "_vnom" in g.columns else "",
                "lineas": int(len(g)),
                "importe_neto": round(float(g["_imp_neto"].sum()), 0),
                "descuento_pesos": round(float(g["_desc"].sum()), 0),
                "litros": round(float(g["_litros"].sum()), 1),
                "cant_base": round(float(g["_cant"].sum()), 1),
                "fecha_ultima": ult.strftime("%d/%m/%Y") if pd.notna(ult) else "",
                "nuevo": cli_int in nuevos,
            })
        rows.sort(key=lambda x: (not x["nuevo"], -x["importe_neto"], x["cliente_nombre"]))
        return rows

    acciones = []
    # Union de líneas de venta que USARON al menos una acción (dedup por índice de fila).
    # El total NO es la suma por acción: una misma línea puede usar varias acciones (canal +
    # Plan AASS + 11 Titulares + Innovaciones) y se contaría 2-4 veces.
    matched_idx = set()
    prev_idx = set()
    for r in reglas:
        vends_raw = str(r.get("vendedores_aplica", "")).upper()
        if "TODOS" in vends_raw:
            vend_set = set(_VENDEDORES_ACTIVOS_PLAN)
        else:
            vend_set = set(_re.findall(r"V\s*\d+", vends_raw))
            vend_set = {"V" + _re.sub(r"\D", "", x) for x in vend_set}
        vend_codes = {int(x[1:]) for x in vend_set if x[1:].isdigit()} & {3, 4, 6, 7, 8, 9, 10}

        seg_set = _acc_seg_canon(r.get("segmento_cliente_aplica"), r.get("canal_aplica"))
        sub_allowed = _acc_subseg_filtro(r.get("segmento_cliente_aplica"), r.get("canal_aplica"))
        pred = _acc_product_pred(r, all_lineas)
        requiere_plan_as, excluye_plan_as = _acc_plan_as_flags(r)
        tramos_pct = _acc_tramos_pct(r)

        # filtro por vendedor (vista vendedor) + regla V3 sin autoservicio
        codes = vend_codes
        seg_use = set(seg_set)
        if vid_filtro is not None:
            vnum = int(_re.sub(r"\D", "", vid_filtro) or 0)
            if vnum not in vend_codes:
                continue  # esta acción no aplica a este vendedor
            codes = {vnum}
            if vnum == 3:
                # V3 (Nadia) solo trabaja Tradicional almacén/despensa/kiosco:
                # descarta AS / On Premise / Mayorista. Si la acción no aplica a
                # tradicional, no se le muestra; la footprint se restringe a almacén/kiosco.
                seg_use &= {"TRADICIONAL"}
                if not seg_use:
                    continue
                if sub_allowed is None:
                    sub_allowed = {"ALMACEN", "KIOSCO"}

        wants_may = "MAYORISTA" in seg_use

        # Índice producto -> acciones: MISMO predicado que la footprint, evaluado contra el
        # universo de productos (no contra las ventas), para poder responder "¿en qué acciones
        # entra este producto?" aunque todavía no se haya vendido en el mes. En la vista
        # vendedor corre después del filtro por vendedor: sólo indexa las acciones que ve él.
        _rid = str(r.get("id_accion", "")).strip()
        for u in universo:
            if pred(*u["_args"]):
                prod_hits.setdefault(u["codigo"], []).append(_rid)

        def _match(df, sub_allowed=sub_allowed, wants_may=wants_may):
            if df.empty:
                return df
            # Segmento: clientes NO mayoristas matchean por su segmento clasificado;
            # los mayoristas SOLO si la regla apunta a MAYORISTA (evita el cruce con AS).
            es_may = df["_es_mayorista"] if "_es_mayorista" in df.columns else False
            seg_ok = (~es_may & df["_seg"].isin(seg_use)) | (es_may & wants_may)
            m = df["_vend"].isin(codes) & seg_ok
            if sub_allowed is not None:
                # el sub-filtro almacén/kiosco SOLO restringe el canon TRADICIONAL;
                # Autoservicio / On Premise no se filtran por subramo (acción multicanal).
                is_trad = df["_seg"].astype(str).str.upper().eq("TRADICIONAL")
                sub_ok = df["_subseg"].apply(lambda s: any(tok in s for tok in sub_allowed))
                m = m & (~is_trad | sub_ok)
            if not m.any():
                return df.iloc[0:0]
            sub = df[m]
            # pred() depende SOLO de (_cat,_linea,_art,_marca,_cod): se evalúa una vez por
            # combinación única y se mapea a cada fila. Evita el apply fila-por-fila (~5s en
            # la vista gerencia, que en Render 0.5 vCPU supera el timeout del worker y daba
            # 500). Resultado idéntico, mucho más rápido.
            keys = list(zip(sub["_cat"], sub["_linea"], sub["_art"], sub["_marca"], sub["_cod"]))
            predcache = {}
            keep_vals = []
            for k in keys:
                if k not in predcache:
                    predcache[k] = bool(pred(*k))
                keep_vals.append(predcache[k])
            keep = pd.Series(keep_vals, index=sub.index, dtype=bool)
            sub = sub[keep]
            if requiere_plan_as:
                sub = sub[pd.to_numeric(sub["_cli"], errors="coerce").isin(plan_as_clientes)]
            elif excluye_plan_as:
                sub = sub[~pd.to_numeric(sub["_cli"], errors="coerce").isin(plan_as_clientes)]
            # La acción se mide por USO real: el % aplicado tiene que ser el de la acción.
            # Comprar el producto del alcance sin ese descuento NO es haber usado la acción.
            sub = sub[_acc_mask_usa_accion(sub, tramos_pct)]
            return sub

        # Footprint = ventas que USARON la accion (alcance + el % de descuento de la accion).
        # Clientes, inversion y litros salen todos de aca: miden la accion, no la categoria.
        cur = _match(v_act)
        matched_idx |= set(cur.index)
        cur_desc = cur[cur["_desc"] > 0]
        clientes_act = set(cur["_cli"].dropna().astype(int))
        prev = _match(v_ant)
        prev_idx |= set(prev.index)
        clientes_ant = set(prev["_cli"].dropna().astype(int))
        nuevos = clientes_act - clientes_ant
        clientes_desc = set(cur_desc["_cli"].dropna().astype(int))

        tipo_raw = str(r.get("tipo_regla", "")).upper()
        tipo = "Sin cargo" if ("SIN_CARGO" in tipo_raw or "BONIFIC" in tipo_raw) else "Descuento"
        detalle = _detalle_clientes(cur, nuevos)
        detalle_nuevos = [d for d in detalle if d["nuevo"]]

        # Columnas nuevas del esquema julio (ausentes en junio → cadenas vacías / None).
        # detalle_click_ref admite varias referencias separadas por "|": se resuelven todas.
        cat_tarjeta = str(r.get("categoria_tarjeta", "")).strip()
        refs = [x.strip() for x in str(r.get("detalle_click_ref", "")).split("|") if x.strip()]
        detalle_grupos = [_acc_enriquecer_grupo(detalle_map[ref], _cat_to_marcas,
                                                _m_cod2seg, _m_cod2linea, _m_cod2cat)
                          for ref in refs if ref in detalle_map]
        mostrar_click = (_acc_norm(r.get("mostrar_detalle_click", "")) == "SI") and bool(detalle_grupos)
        try:
            orden_visual = int(float(str(r.get("orden_visual", "")).strip()))
        except (TypeError, ValueError):
            orden_visual = None

        # Acciones por código (SKU curados): resolver cada código a su DESCRIPCIÓN de artículo
        # (01_INPUTS/RAW_PRODUCTOS/productos<mes>.xlsx, col. H) para que el vendedor sepa el
        # producto exacto. Fallback a la Linea Comercial del 04D. Si no está en ninguno, se
        # muestra el CÓDIGO igual.
        codigos_raw = [t.strip() for t in _re.split(r"[;,]", str(r.get("productos_marcas", "")))
                       if t.strip().isdigit()]
        codigos_detalle, _res_names, _res_seen, _pend = [], [], set(), []
        for cod in codigos_raw:
            d = _desc_art_map.get(cod)
            lin04 = _m_cod2linea.get(cod)
            desc = (d or {}).get("descripcion", "")
            linea = (d or {}).get("linea", "") or (str(lin04).strip() if lin04 and str(lin04).strip().lower() != "nan" else "")
            cat = (d or {}).get("categoria", "") or (_acc_canon_cat(_m_cod2cat.get(cod)) or "")
            encontrado = bool(desc or linea)
            producto = desc or linea  # lo que ve el vendedor: descripción del artículo
            codigos_detalle.append({"codigo": cod, "producto": producto, "descripcion": desc,
                                    "linea": linea, "categoria": cat, "encontrado": encontrado})
            if encontrado:
                lbl = linea or desc
                if lbl and lbl not in _res_seen:
                    _res_seen.add(lbl); _res_names.append(lbl)
            else:
                _pend.append(cod)
        if codigos_detalle:
            marcas_display = "; ".join(_res_names)
            if _pend:
                marcas_display += ("; " if marcas_display else "") + "cód. " + ", ".join(_pend)
        else:
            marcas_display = str(r.get("productos_marcas", "")).strip()

        acciones.append({
            "id_accion":     str(r.get("id_accion", "")).strip(),
            "tipo":          tipo,
            "tipo_regla":    str(r.get("tipo_regla", "")).strip(),
            # bloque del catálogo: agrupa las tarjetas por origen. Las acciones dadas de alta
            # después del drop del mes viajan en su propio bloque (07_NUEVAS_*) y el portal
            # las renderiza en una sección aparte, con diseño propio.
            "bloque":        str(r.get("bloque_pptx", "")).strip(),
            "nueva":         str(r.get("bloque_pptx", "")).strip().upper().startswith("07_NUEVAS"),
            "minimo":        str(r.get("minimo", "")).strip(),
            "unidad_minimo": str(r.get("unidad_minimo", "")).strip(),
            "subcategoria":  str(r.get("subcategoria", "")).strip(),
            "segmento":      str(r.get("segmento_cliente_aplica", "")).strip(),
            "canal":         str(r.get("canal_aplica", "")).strip(),
            "vendedores":    sorted(vend_set),
            "marcas":        marcas_display,
            "codigos":       codigos_detalle,
            "escala":        str(r.get("condicion_compra", "")).strip(),
            "descuento_pct": str(r.get("descuento_pct", "")).strip(),
            "tope":          str(r.get("tope", "")).strip(),
            "observaciones": str(r.get("observaciones", "")).strip(),
            # columnas nuevas esquema julio (compat: vacías/None en junio)
            "categoria_tarjeta":     cat_tarjeta,
            "mostrar_detalle_click": mostrar_click,
            "detalle_click_ref":     refs,
            "detalle_categorias":    detalle_grupos,
            "orden_visual":          orden_visual,
            # computado
            "inversion_pesos":   round(float(cur_desc["_desc"].sum()), 0),
            "litros":            round(float(cur["_litros"].sum()), 1),
            "importe_neto":       round(float(cur["_imp_neto"].sum()), 0),
            "clientes_alcanzados": int(len(clientes_act)),
            "clientes_nuevos":   int(len(nuevos)),
            "clientes_con_descuento": int(len(clientes_desc)),
            "clientes_detalle": detalle,
            "clientes_nuevos_detalle": detalle_nuevos,
            "nota_calculo": ("clientes/litros/inversion desde ventas.csv, SOLO lineas que usaron la accion "
                             "(% aplicado = tramo de la accion); inversion desde valorDescuento x CantBase"),
        })

    # Orden de tarjetas por orden_visual ASC (respeta el orden del PPTX julio). Sort
    # estable: en meses sin la columna (junio) todos caen en el mismo bucket y se
    # conserva el orden original del catálogo.
    acciones.sort(key=lambda a: (0, a["orden_visual"]) if isinstance(a.get("orden_visual"), int) else (1, 0))

    # Totales reales = sobre la UNION de líneas (sin doble conteo entre acciones).
    uni = v_act.loc[v_act.index.isin(matched_idx)] if matched_idx else v_act.iloc[0:0]
    uni_desc = uni[uni["_desc"] > 0]
    uni_prev = v_ant.loc[v_ant.index.isin(prev_idx)] if prev_idx else v_ant.iloc[0:0]
    cli_act_union = set(uni["_cli"].dropna().astype(int))
    cli_ant_union = set(uni_prev["_cli"].dropna().astype(int))
    totales = {
        "litros":                 round(float(uni["_litros"].sum()), 1),
        "importe_neto":           round(float(uni["_imp_neto"].sum()), 0),
        "inversion_pesos":        round(float(uni_desc["_desc"].sum()), 0),
        "clientes_alcanzados":    int(len(cli_act_union)),
        "clientes_nuevos":        int(len(cli_act_union - cli_ant_union)),
        "clientes_con_descuento": int(uni_desc["_cli"].dropna().astype(int).nunique()),
    }
    # Catálogo de productos del buscador: cada SKU con las acciones en las que entra.
    # Se incluyen también los SKU sin acción (respuesta honesta a "este producto no está
    # en ninguna acción del mes") y los que faltan en el maestro 04D (en_maestro=False).
    productos = [{
        "codigo":     u["codigo"],
        "producto":   u["producto"],
        "marca":      u["marca"],
        "alias":      u["alias"],
        "linea":      u["linea"],
        "categoria":  u["categoria"],
        "cat_canon":  u["cat_canon"],
        "en_maestro": u["en_maestro"],
        "acciones":   prod_hits.get(u["codigo"], []),
    } for u in universo]
    productos.sort(key=lambda p: (p["marca"].upper(), p["producto"].upper()))

    return {"mes": mes, "fuente": fuente, "periodo": str(per_actual),
            "generado_en": _now_ar(), "acciones": acciones, "totales": totales,
            "acciones_on": _acc_on_cards(), "productos": productos}


def _alertas_descuento_mes():
    """Alertas de descuento del MES EN CURSO desde el catálogo de acciones (no mayo).
    Línea con descuento = alerta si el % aplicado supera el tramo MÁS ALTO de la acción
    del catálogo que aplica (vendedor+segmento+marca). Sin acción aplicable → máximo 0."""
    mes, fuente, reglas = _acc_catalogo_mes()
    v = _acc_preparar_ventas("ventas.csv")   # MES VIVO: evita arrastrar mayo de ventas_acumulada
    if v.empty or not reglas:
        return []
    try:
        per = pd.Period(mes, freq="M")
    except Exception:
        per = v["_mes"].max()
    cur = v[(v["_mes"] == per) & (v["_desc"] > 0)].copy()
    if cur.empty:
        return []
    all_lineas = set(l for l in v["_linea"].dropna().unique() if l and l != "NAN")

    parsed = []
    for r in reglas:
        vends_raw = str(r.get("vendedores_aplica", "")).upper()
        if "TODOS" in vends_raw:
            codes = {3, 4, 6, 7, 8, 9, 10}
        else:
            codes = {int(_re.sub(r"\D", "", x)) for x in _re.findall(r"V\s*\d+", vends_raw)} & {3, 4, 6, 7, 8, 9, 10}
        seg = _acc_seg_canon(r.get("segmento_cliente_aplica"), r.get("canal_aplica"))
        sub_allowed = _acc_subseg_filtro(r.get("segmento_cliente_aplica"), r.get("canal_aplica"))
        pred = _acc_product_pred(r, all_lineas)
        tipo = str(r.get("tipo_regla", "")).upper()
        tramos = [float(x) for x in str(r.get("descuento_pct", "")).replace(",", ".").split("|")
                  if x.strip().replace(".", "").isdigit()]
        maxpct = 100.0 if ("BONIFIC" in tipo or "SIN_CARGO" in tipo) else (max(tramos) if tramos else 0.0)
        requiere_plan_as, excluye_plan_as = _acc_plan_as_flags(r)
        parsed.append((str(r.get("id_accion", "")).strip(), codes, seg, sub_allowed, pred, maxpct,
                       requiere_plan_as, excluye_plan_as))

    # Clientes Plan AS: tienen 10% de descuento en factura SIEMPRE → piso permitido = 10%.
    pas = read_csv(DATASETS / "mod_planes_as.csv")
    pas_ids = (set(pd.to_numeric(pas["cliente_id"], errors="coerce").dropna().astype(int))
               if not pas.empty and "cliente_id" in pas.columns else set())

    alerts = []
    for _, row in cur.iterrows():
        desc = float(row["_desc"])
        apl = round(float(row["_pct"]), 1)   # % descuento real (valorDescuento), no IVA
        if apl <= 0:
            continue
        vend = row["_vend"]; seg_v = row["_seg"]
        row_may = bool(row["_es_mayorista"]) if "_es_mayorista" in row else False
        try: cli_int = int(row["_cli"])
        except Exception: cli_int = None
        allowed, fuente_id = 0.0, None
        for rid, codes, seg, sub_allowed, pred, maxpct, requiere_plan_as, excluye_plan_as in parsed:
            # el sub-filtro almacén/kiosco SOLO restringe el canon TRADICIONAL
            if (sub_allowed is not None and str(row["_seg"]).upper() == "TRADICIONAL"
                    and not any(tok in row["_subseg"] for tok in sub_allowed)):
                continue
            # Acción de Planes AASS: su tope solo aplica a clientes DENTRO del plan; para
            # el resto no autoriza (si no, taparía sobre-descuentos en autoservicios sin plan).
            if requiere_plan_as and (cli_int is None or cli_int not in pas_ids):
                continue
            # Acción que excluye a los AASS con plan: no autoriza descuento en esos clientes
            # (tienen su propio tramo, con otro tope).
            if excluye_plan_as and cli_int is not None and cli_int in pas_ids:
                continue
            if (vend in codes) and _acc_seg_match(seg_v, row_may, seg) and pred(row["_cat"], row["_linea"], row["_art"], row["_marca"], row["_cod"]):
                if maxpct > allowed:
                    allowed, fuente_id = maxpct, rid
        # Piso 10% para clientes Plan AS (descuento de factura)
        if cli_int is not None and cli_int in pas_ids and allowed < 10.0:
            allowed = 10.0
            if not fuente_id:
                fuente_id = "Plan AS (10% factura)"
        exceso = round(apl - allowed, 1)
        if exceso <= 0:   # sin tolerancia: cualquier descuento que supere el permitido alerta
            continue
        try: cod_int = int(vend)
        except Exception: cod_int = 0
        fr = fuente_id or "sin acción aplicable"
        alerts.append({
            "vendedor_codigo": cod_int, "vendedor_nombre": str(row["_vnom"]),
            "vendedor_id": "V" + str(cod_int), "cliente_id": cli_int,
            "cliente_nombre": str(row["_clinom"]), "articulo": str(row["_art"]), "marca": str(row["_marca"]),
            "fecha_pedido": str(row.get("_fcomp", "")), "fecha_carga": str(row.get("_fcarga", "")),
            "cant_base": round(float(row["_cant"]), 1), "cajas_eq": round(float(row["_cant"]), 1),
            "descuento_aplicado_pct": apl, "descuento_maximo_pct": allowed, "exceso_pct": exceso,
            "fuente_regla": fr, "importe_neto": round(float(row["_imp_neto"]), 0),
            "valor_descuento": round(desc, 0),
            "prioridad": "alta", "tipo": "descuento", "titulo": str(row["_clinom"]),
            "detalle": f"{row['_art']} — descuento aplicado: {apl}% / máximo: {allowed}%",
            "accion": f"Revisar descuento ({fr})",
            "impacto_alertas_ars": round(desc, 0),
        })
    alerts.sort(key=lambda a: a["exceso_pct"], reverse=True)
    return alerts


def _acc_botellas_por_caja(articulo):
    """Botellas por caja desde el formato del artículo (ej. 'ALMA MORA MALBEC 6X750' -> 6).
    Default 6 (formato estándar de estas líneas). CantBase viene en botellas (lxu=0.75 L)."""
    m = _re.search(r"(\d+)\s*[xX]\s*\d+", str(articulo or ""))
    if m:
        n = int(m.group(1))
        if 1 <= n <= 48:
            return n
    return 6


def _alertas_tope_cajas_mes():
    """Alerta por exceso del TOPE MENSUAL DE CAJAS por cliente (combinable entre marcas).
    Aplica a acciones del catálogo con tope mensual en cajas: `maximo` numérico +
    `unidad_maximo` con 'caja' y 'mes' (ej. ACJ26-017: maximo=2, unidad_maximo='cajas en el mes').
    Caja = botellas/caja del artículo (6X750 -> 6). Footprint = líneas que matchean la acción
    (vendedor + segmento + sub-segmento + marca) con descuento real (>0); se suman las cajas
    por cliente en el mes y se alerta si superan el tope."""
    mes, fuente, reglas = _acc_catalogo_mes()
    v = _acc_preparar_ventas("ventas.csv")   # MES VIVO
    if v.empty or not reglas:
        return []
    try:
        per = pd.Period(mes, freq="M")
    except Exception:
        per = v["_mes"].max()
    cur = v[(v["_mes"] == per) & (v["_desc"] > 0)].copy()
    if cur.empty:
        return []
    all_lineas = set(l for l in v["_linea"].dropna().unique() if l and l != "NAN")
    cur["_cajas"] = [float(c) / _acc_botellas_por_caja(a) for c, a in zip(cur["_cant"], cur["_art"])]

    alerts = []
    for r in reglas:
        try:
            tope_cajas = float(str(r.get("maximo", "")).replace(",", ".").strip())
        except (ValueError, TypeError):
            continue
        um = _acc_norm(r.get("unidad_maximo"))
        if tope_cajas <= 0 or "CAJA" not in um or "MES" not in um:
            continue   # solo acciones con tope mensual en cajas
        rid = str(r.get("id_accion", "")).strip()
        vends_raw = str(r.get("vendedores_aplica", "")).upper()
        if "TODOS" in vends_raw:
            codes = {3, 4, 6, 7, 8, 9, 10}
        else:
            codes = {int(_re.sub(r"\D", "", x)) for x in _re.findall(r"V\s*\d+", vends_raw)} & {3, 4, 6, 7, 8, 9, 10}
        seg = _acc_seg_canon(r.get("segmento_cliente_aplica"), r.get("canal_aplica"))
        sub_allowed = _acc_subseg_filtro(r.get("segmento_cliente_aplica"), r.get("canal_aplica"))
        pred = _acc_product_pred(r, all_lineas)

        m = cur["_vend"].isin(codes) & cur["_seg"].isin(seg)
        if sub_allowed is not None:
            is_trad = cur["_seg"].astype(str).str.upper().eq("TRADICIONAL")
            sub_ok = cur["_subseg"].apply(lambda s: any(tok in s for tok in sub_allowed))
            m = m & (~is_trad | sub_ok)
        sub = cur[m]
        if sub.empty:
            continue
        keep = sub.apply(lambda x: pred(x["_cat"], x["_linea"], x["_art"], x["_marca"], x["_cod"]), axis=1)
        sub = sub[keep]
        if sub.empty:
            continue

        for cli, grp in sub.groupby("_cli"):
            cajas = round(float(grp["_cajas"].sum()), 2)
            if cajas <= tope_cajas + 1e-9:
                continue
            exceso = round(cajas - tope_cajas, 2)
            try: cli_int = int(cli)
            except Exception: cli_int = None
            try: cod_int = int(grp["_vend"].iloc[0])
            except Exception: cod_int = 0
            desc_total = round(float(grp["_desc"].sum()), 0)
            marcas_txt = ", ".join(sorted({str(x) for x in grp["_marca"].dropna() if str(x).strip()}))
            cnom = str(grp["_clinom"].iloc[0]); vnom = str(grp["_vnom"].iloc[0])
            fcarga = next((f for f in grp["_fcarga"][::-1] if str(f).strip()), "")
            tope_txt = (f"{tope_cajas:g}").rstrip()
            alerts.append({
                "vendedor_codigo": cod_int, "vendedor_nombre": vnom,
                "vendedor_id": "V" + str(cod_int), "cliente_id": cli_int,
                "cliente_nombre": cnom, "articulo": f"TOPE {rid}", "marca": marcas_txt,
                "fecha_pedido": "", "fecha_carga": str(fcarga),
                "cajas_mes": cajas, "cajas_tope": tope_cajas, "exceso_cajas": exceso,
                "fuente_regla": rid, "valor_descuento": desc_total,
                "prioridad": "alta", "tipo": "tope", "titulo": cnom,
                "detalle": (f"Tope {tope_txt} cajas/mes superado en {marcas_txt or 'marcas de la acción'}: "
                            f"{cajas:g} cajas con descuento (acción {rid}). Exceso {exceso:g} cajas."),
                "accion": f"Revisar tope de cajas ({rid})",
                "impacto_alertas_ars": desc_total,
            })
    alerts.sort(key=lambda a: a["exceso_cajas"], reverse=True)
    return alerts


@app.route("/api/gerencia/acciones_mes")
def gerencia_acciones_mes():
    return jsonify(_acciones_mes_payload(None))


@app.route("/api/vendedor/<vid>/acciones_mes")
def vendedor_acciones_mes(vid):
    vid_norm = normalizar_vendedor_codigo(vid)
    if vid_norm in ("V2", "V5", "V20"):
        return jsonify({"error": "Vendedor no autorizado"}), 403
    return jsonify(_acciones_mes_payload(vid_norm))


# ====== PLANES AS — VENDEDOR ======
@app.route("/api/vendedor/<vid>/planes_as")
def vendedor_planes_as(vid):
    vid_norm = normalizar_vendedor_codigo(vid)
    if vid_norm in ("V2", "V5", "V20"):
        return jsonify({"error": "Vendedor no autorizado"}), 403
    try:
        cod = int(vid_norm.replace("V", ""))
    except Exception:
        return jsonify({"error": "ID inválido"}), 400
    df = read_csv(DATASETS / "mod_planes_as.csv")
    if df.empty:
        return jsonify({"clientes": []}), 200
    envios_map = _cargar_sincargos_envios()
    inov_prods = _inov_plan_as_productos()
    inov_compras = _inov_plan_as_compras()
    df["vendedor_codigo"] = pd.to_numeric(df["vendedor_codigo"], errors="coerce")
    df = df[df["vendedor_codigo"] == cod]
    _num = ["total_facturado", "dcto_plan", "cant_cajas", "tope", "escala_actual", "escala_max",
            "sc_alaris", "sc_alma_mora", "sc_frizze", "sc_antares_ipa", "sc_smf_flavours",
            "sc_total_ganado", "sc_cajas_enviadas_total", "sc_pendiente",
            "sc_env_alaris", "sc_env_alma_mora", "sc_env_frizze", "sc_env_antares_ipa", "sc_env_smf_flavours",
            "sc_pend_alaris", "sc_pend_alma_mora", "sc_pend_frizze", "sc_pend_antares_ipa", "sc_pend_smf_flavours",
            "pt_disponible", "pt_enviado", "pt_pendiente"]
    for c in _num:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    def _i(row, k): return int(row[k]) if k in row.index and pd.notna(row[k]) else 0
    registros = []
    for _, row in df.iterrows():
        _cid = int(row["cliente_id"]) if pd.notna(row["cliente_id"]) else None
        registros.append({
            "cliente_id":      _cid,
            "cliente_nombre":  str(row.get("cliente_nombre", "")),
            "direccion":       str(row.get("direccion", "")),
            "plan_as":         str(row.get("plan_as", "")),
            "innovaciones":    _inov_plan_as_cliente(_cid, inov_prods, inov_compras),
            "total_facturado": round(float(row["total_facturado"]), 2),
            "dcto_plan":       round(float(row["dcto_plan"]), 2),
            "cant_cajas":      _i(row, "cant_cajas"),
            "tope":            _i(row, "tope"),
            "escala_actual":   _i(row, "escala_actual"),
            "escala_max":      _i(row, "escala_max"),
            "sc_alaris":       _i(row, "sc_alaris"),
            "sc_alma_mora":    _i(row, "sc_alma_mora"),
            "sc_frizze":       _i(row, "sc_frizze"),
            "sc_antares_ipa":  _i(row, "sc_antares_ipa"),
            "sc_smf_flavours": _i(row, "sc_smf_flavours"),
            "sc_total_ganado": _i(row, "sc_total_ganado"),
            "sc_enviadas_total": _i(row, "sc_cajas_enviadas_total"),
            "sc_pendiente":    _i(row, "sc_pendiente"),
            # por producto: enviados y pendientes
            "sc_env_alaris":       _i(row, "sc_env_alaris"),
            "sc_env_alma_mora":    _i(row, "sc_env_alma_mora"),
            "sc_env_frizze":       _i(row, "sc_env_frizze"),
            "sc_env_antares_ipa":  _i(row, "sc_env_antares_ipa"),
            "sc_env_smf_flavours": _i(row, "sc_env_smf_flavours"),
            "sc_pend_alaris":      _i(row, "sc_pend_alaris"),
            "sc_pend_alma_mora":   _i(row, "sc_pend_alma_mora"),
            "sc_pend_frizze":      _i(row, "sc_pend_frizze"),
            "sc_pend_antares_ipa": _i(row, "sc_pend_antares_ipa"),
            "sc_pend_smf_flavours":_i(row, "sc_pend_smf_flavours"),
            "pf_disponible":       _i(row, "pf_disponible"),
            "pf_enviado":          _i(row, "pf_enviado"),
            "pf_estado":           str(row.get("pf_estado", "")),
            "pt_disponible":       _i(row, "pt_disponible"),
            "pt_enviado":          _i(row, "pt_enviado"),
            "pt_pendiente":        _i(row, "pt_pendiente"),
            "pt_estado":           str(row.get("pt_estado", "")),
            "pt_producto":         str(row.get("pt_producto", "") or ""),
            "envios":              envios_map.get(int(row["cliente_id"]) if pd.notna(row["cliente_id"]) else -1, []),
        })
    return jsonify({
        "generado_en": _now_ar(),
        "vendedor_id": vid_norm,
        "fuente": "mod_planes_as.csv",
        "innovaciones_productos": inov_prods,
        "total_clientes": len(registros),
        "clientes": registros,
    })


# ====== SELLOUT POR CATEGORÍA ======
# REGLA FIJA: Sellout usa ventas.csv (período comercial actual).
# NO cambiar a ventas_acumulada.csv. Validado por usuario 2026-05-23.
@app.route("/api/gerencia/sellout_categoria")
def gerencia_sellout_categoria():
    df = read_csv(DATASETS / "mod_sellout_categoria.csv")
    if df.empty:
        return jsonify({"error": "Sin datos"}), 404
    df.columns = [c.lstrip("﻿") for c in df.columns]
    for c in ["litros", "cajas", "importe", "clientes"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    fecha = str(df["fecha_calculo"].iloc[0]) if "fecha_calculo" in df.columns and len(df) else ""

    por_cat = {}
    for cat, grp in df.groupby("Categoria"):
        segs = []
        for _, row in grp.iterrows():
            segs.append({
                "segmento":  str(row.get("Segmento", "")),
                "litros":    round(float(row["litros"]), 1),
                "cajas":     int(row["cajas"]),
                "importe":   int(row["importe"]),
                "clientes":  int(row["clientes"]),
            })
        segs.sort(key=lambda x: x["litros"], reverse=True)
        por_cat[cat] = {
            "categoria":    cat,
            "litros_total": round(float(grp["litros"].sum()), 1),
            "cajas_total":  int(grp["cajas"].sum()),
            "clientes":     int(grp["clientes"].sum()),
            "por_segmento": segs,
        }
    categorias = sorted(por_cat.values(), key=lambda x: x["litros_total"], reverse=True)
    return jsonify({
        "generado_en": _now_ar(),
        "fuente":        "mod_sellout_categoria.csv",
        "fecha_calculo": fecha,
        "por_categoria": categorias,
    })


# ====== SELLOUT EN LITROS CON OBJETIVOS ======
import re as _re

def _infer_litros_por_nombre(nombre: str) -> float:
    """Fallback: infiere litros/unidad desde el nombre (ej: 6X750 → 0.75)."""
    matches = _re.findall(r'[X\s](\d{3,4})\b', str(nombre).upper())
    if matches:
        ml = int(matches[-1])
        if 100 <= ml <= 9999:
            return ml / 1000.0
    return 0.0


_MAESTRO_04D_CACHE = {}
_MAESTRO_MES_CACHE = {}


def _maestro_mes_productos():
    """Maestro de productos del MES (01_INPUTS/RAW_PRODUCTOS/productos<mes>.xlsx), en el mismo
    formato que el 04D: (cod2cat, cod2seg, cod2lxu, cod2linea).

    Existe porque el 04D quedó CONGELADO: tiene 258 códigos y le faltan SKU vigentes que sí se
    venden (Alaris D.Cosecha, Dada Sweet Red, Los Arboles Rosado, Smirnoff BC...). Sin categoría,
    línea comercial ni litros/caja, esas líneas de venta no entran en las reglas por categoría de
    las acciones ni en sell out, y aportan 0 L. Este export es el MISMO maestro, actualizado cada
    mes (mismo vocabulario de Categoría/Segmento), así que sirve para completar los huecos.
    Cacheado por (path, mtime). Vacío si no hay archivo del mes."""
    p = _acc_desc_articulo_file()
    if p is None:
        return {}, {}, {}, {}
    try:
        key = (str(p), os.path.getmtime(p))
    except OSError:
        key = (str(p), 0)
    cached = _MAESTRO_MES_CACHE.get(key)
    if cached is not None:
        return cached
    cod2cat, cod2seg, cod2lxu, cod2linea = {}, {}, {}, {}
    try:
        raw = pd.read_excel(p, header=None, dtype=str)
        hdr, cols = None, None
        for i, row in raw.iterrows():
            vals = [str(x).strip() for x in row.tolist()]
            if (any(v.startswith("Código Art") or v.startswith("Codigo Art") for v in vals)
                    and any(v.startswith("Descripci") for v in vals)):
                hdr, cols = i, vals
                break
        if hdr is None:
            return cod2cat, cod2seg, cod2lxu, cod2linea
        df = raw.iloc[hdr + 1:].copy()
        df.columns = cols
        c_cod = next(c for c in cols if c.startswith("Código Art") or c.startswith("Codigo Art"))
        c_cat = next((c for c in cols if c.startswith("Categor")), None)
        c_seg = next((c for c in cols if c.strip().lower() == "segmento"), None)
        c_lin = next((c for c in cols if c.strip().lower().startswith("linea comercial")), None)
        c_lxc = next((c for c in cols if "lts" in c.lower() and "caja" in c.lower()), None)
        c_uxc = next((c for c in cols if "unidad" in c.lower() and "caja" in c.lower()), None)
        df["_cod"] = (df[c_cod].astype(str).str.strip().str.upper()
                      .str.replace(r"\.0$", "", regex=True))
        df = df[df["_cod"].str.len() > 0]
        df = df[~df["_cod"].str.lower().isin(["nan", "none"])]
        lxc = pd.to_numeric(df[c_lxc], errors="coerce").fillna(0) if c_lxc else 0
        uxc = pd.to_numeric(df[c_uxc], errors="coerce").fillna(0) if c_uxc else 0
        df["_lxu"] = (lxc / uxc).where(uxc > 0, 0.0) if (c_lxc and c_uxc) else 0.0
        cod2lxu = df.set_index("_cod")["_lxu"].to_dict()
        if c_cat:
            cod2cat = df.set_index("_cod")[c_cat].to_dict()
        if c_seg:
            cod2seg = df.set_index("_cod")[c_seg].to_dict()
        if c_lin:
            cod2linea = df.set_index("_cod")[c_lin].to_dict()
    except Exception as e:
        print(f"[WARN] maestro del mes no se pudo leer ({p.name}): {e}")
        return {}, {}, {}, {}
    _MAESTRO_MES_CACHE.clear()
    _MAESTRO_MES_CACHE[key] = (cod2cat, cod2seg, cod2lxu, cod2linea)
    return _MAESTRO_MES_CACHE[key]


def _cargar_maestro_04D():
    """Wrapper cacheado por mtime del maestro 04D (CSV preferido, xlsx fallback) COMPLETADO con
    el maestro del mes. El cómputo real está en _cargar_maestro_04D_uncached; cachear evita
    reconstruir los 4 dicts en cada request (lo usan acciones, dashboard, sellout, alertas)."""
    csv_path  = CONFIG / "maestro_04D_productos.csv"
    xlsx_path = INPUTS / "04D_MAESTRO_PRODUCTOS_PENAFLOR.xlsx"
    src = csv_path if csv_path.exists() else (xlsx_path if xlsx_path.exists() else None)
    mes_src = _acc_desc_articulo_file()   # el archivo del mes también invalida la caché
    def _mt(p):
        try:
            return os.path.getmtime(p) if p else 0
        except OSError:
            return 0
    key = (str(src), _mt(src), str(mes_src), _mt(mes_src)) if src else None
    if key is not None and key in _MAESTRO_04D_CACHE:
        return _MAESTRO_04D_CACHE[key]
    result = _cargar_maestro_04D_uncached()
    if key is not None:
        _MAESTRO_04D_CACHE.clear()
        _MAESTRO_04D_CACHE[key] = result
    return result


def _cargar_maestro_04D_uncached():
    """
    Carga maestro de productos 04D. Prefiere CSV liviano (09_CONFIG/maestro_04D_productos.csv)
    sobre el xlsx original (19MB con imágenes, tarda ~40s). Fallback al xlsx si el CSV no existe.
    Devuelve cuatro dicts keyed por Código Art. (str, upper, sin .0):
      cod2cat   → Categoria
      cod2seg   → Segmento  (Nacional/Importados spirits; Alto/Medio Alto/Superior/Medio VDA)
      cod2lxu   → litros por unidad (Lts x caja / UxC)
      cod2linea → Linea Comercial (marca canónica)
    """
    cod2cat, cod2seg, cod2lxu, cod2linea = {}, {}, {}, {}
    csv_path  = CONFIG / "maestro_04D_productos.csv"
    xlsx_path = INPUTS / "04D_MAESTRO_PRODUCTOS_PENAFLOR.xlsx"

    try:
        if csv_path.exists():
            df = pd.read_csv(csv_path, dtype=str, encoding="utf-8-sig")
            df.columns = [c.strip() for c in df.columns]
            df["_cod"] = df["Codigo"].astype(str).str.strip().str.upper().str.replace(r"\.0$", "", regex=True)
            lxc = pd.to_numeric(df.get("Lts x caja", pd.Series(dtype=float)), errors="coerce").fillna(0)
            uxc = pd.to_numeric(df.get("UxC", pd.Series(dtype=float)), errors="coerce").fillna(0)
            df["_lxu"] = (lxc / uxc).where(uxc > 0, 0.0)
        elif xlsx_path.exists():
            df = pd.read_excel(xlsx_path, header=3)
            df.columns = [c.strip() for c in df.columns]
            col_cod = next((c for c in df.columns if "dig" in c.lower() or ("c" in c.lower() and "art" in c.lower())), None)
            if not col_cod:
                return cod2cat, cod2seg, cod2lxu, cod2linea
            df["_cod"] = df[col_cod].astype(str).str.strip().str.upper().str.replace(r"\.0$", "", regex=True)
            lxc = pd.to_numeric(df["Lts x caja"] if "Lts x caja" in df.columns else pd.Series(dtype=float), errors="coerce").fillna(0)
            uxc = pd.to_numeric(df["UxC"] if "UxC" in df.columns else pd.Series(dtype=float), errors="coerce").fillna(0)
            df["_lxu"] = (lxc / uxc).where(uxc > 0, 0.0)
            if "Linea Comercial" not in df.columns:
                col_linea = next((c for c in df.columns if "linea" in c.lower() and "comercial" in c.lower()), None)
                if col_linea:
                    df = df.rename(columns={col_linea: "Linea Comercial"})
        else:
            return cod2cat, cod2seg, cod2lxu, cod2linea

        cod2cat   = df.set_index("_cod")["Categoria"].to_dict()
        cod2seg   = df.set_index("_cod")["Segmento"].to_dict()
        cod2lxu   = df.set_index("_cod")["_lxu"].to_dict()
        if "Linea Comercial" in df.columns:
            cod2linea = df.set_index("_cod")["Linea Comercial"].to_dict()
    except Exception:
        pass

    # ── Completar con el maestro del mes (ver _maestro_mes_productos).
    # El 04D MANDA donde tiene dato: el mes sólo agrega los códigos que faltan y rellena los
    # campos vacíos. Sin esto, un SKU vigente que el 04D no trae se vende "sin categoría y sin
    # litros": no suma en las acciones por categoría ni en sell out. Un dato que existe no puede
    # perderse porque el maestro quedó viejo.
    m_cat, m_seg, m_lxu, m_linea = _maestro_mes_productos()
    def _vacio(v):
        s = str(v if v is not None else "").strip()
        return s == "" or s.lower() == "nan"
    for cod in m_cat or m_lxu:
        if _vacio(cod2cat.get(cod)):
            cod2cat[cod] = m_cat.get(cod, "")
        if _vacio(cod2seg.get(cod)):
            cod2seg[cod] = m_seg.get(cod, "")
        if _vacio(cod2linea.get(cod)):
            cod2linea[cod] = m_linea.get(cod, "")
        try:
            actual = float(cod2lxu.get(cod) or 0)
        except (TypeError, ValueError):
            actual = 0.0
        if actual <= 0:
            cod2lxu[cod] = float(m_lxu.get(cod, 0) or 0)

    return cod2cat, cod2seg, cod2lxu, cod2linea


# Mapeo Categoria 04D → bucket sell out
_SO_CAT_MAP = {
    "vinos del año": "VINOS DEL AÑO",
    "vinos del a\xf1o": "VINOS DEL AÑO",
    "vinos de mesa": "VINOS DEL AÑO",
    "vinos de guarda": "VINOS DE GUARDA",
    "espumantes": "CHAMPAÑA",
    "sidra": "CHAMPAÑA",
    "cerveza artesanal": "CERVEZA ARTESANAL",
    "rtd (s)": "RTD",
    "rtd": "RTD",
    "whisky": "SPIRITS",
    "whisky (maltas)": "SPIRITS",
    "gin": "SPIRITS",
    "ron": "SPIRITS",
    "vodka": "SPIRITS",
    "licores": "SPIRITS",
    "bourbon": "SPIRITS",
    # Vermouth (Cinzano 90105/90106/90110) es categoria PROPIA de sell out, no Spirits:
    # su venta no suma al objetivo de Spirits (definicion del negocio, 16/07/2026).
    "vermouth": "VERMOUTH",
}

# Mapeo Segmento 04D → tier sell out
_SO_SEG_VDA = {"Alto": "Alto", "Medio Alto": "Medio Alto", "Superior": "Superior", "Medio": "Medio", "Vinos de Mesa": "Medio"}
_SO_SEG_SPIRITS = {"Nacional": "Nacionales", "Importados": "Importados"}


# Normalización de la categoría del OBJSELLOUT.xlsx → categoría de la tarjeta.
# 'rtd (s)' no es categoría aparte: es un subgrupo de RTD (igual que en _SO_CAT_MAP).
_OBJ_CAT_NORM = {"RTD (S)": "RTD"}

def _cargar_objetivos_sellout() -> dict:
    """Lee 01_INPUTS/OBJSELLOUT.xlsx → {CATEGORIA_UPPER: {"total": litros, "subs": {grupo_pbp: litros}}}.
    Fuente única de objetivos de sell out. El archivo abre el objetivo por 'Grupo PBP'
    (subgrupo de precio/origen) dentro de cada categoría + una fila 'Total' por categoría.
    Columnas: categoria | Grupo PBP | objetivo en litros. Devuelve {} si falta o falla.
    Clave = categoría en mayúsculas (coincide con los buckets de _sellout_desde_ventas);
    subs por nombre de Grupo PBP (Alto/Medio Alto/Superior/Medio, Nacionales/Importados, ...)."""
    path = INPUTS / "OBJSELLOUT.xlsx"
    out = {}
    if not path.exists():
        return out
    try:
        df = pd.read_excel(path, header=1)
        df.columns = [str(c).strip().lower() for c in df.columns]
        cat_col = next((c for c in df.columns if "categor" in c), df.columns[0])
        grp_col = next((c for c in df.columns if "pbp" in c or "grupo" in c), None)
        obj_col = next((c for c in df.columns if "objetivo" in c or "litro" in c), df.columns[-1])
        rows = {}      # card_cat -> [(cat_orig, grupo_pbp, val)]
        totals = {}    # card_cat -> val (fila 'Total')
        for _, r in df.iterrows():
            cat_orig = str(r[cat_col]).strip()
            if not cat_orig or cat_orig.lower() in ("total", "nan"):
                continue
            val = pd.to_numeric(r[obj_col], errors="coerce")
            if pd.isna(val):
                continue
            val = int(round(float(val)))
            card = _OBJ_CAT_NORM.get(cat_orig.upper(), cat_orig.upper())
            grupo = str(r[grp_col]).strip() if grp_col is not None else ""
            if grupo.lower() == "total":
                totals[card] = val
            elif grupo and grupo.lower() != "nan":
                rows.setdefault(card, []).append((cat_orig, grupo, val))
        for card, rs in rows.items():
            grupos = [g for _, g, _ in rs]
            # Si dos filas comparten Grupo PBP (caso RTD: 'rtd' y 'rtd (s)' ambos PBP 'RTD'),
            # se etiqueta el subgrupo por el nombre de categoría para no pisarse.
            use_cat = len(set(grupos)) < len(grupos)
            subs = {}
            for cat_orig, grupo, val in rs:
                label = cat_orig.upper() if use_cat else grupo
                subs[label] = subs.get(label, 0) + val
            total = totals.get(card)
            if total is None:
                total = sum(subs.values()) if subs else None
            out[card] = {"total": total, "subs": subs}
        for card, t in totals.items():           # categorías que sólo tenían fila 'Total'
            out.setdefault(card, {"total": t, "subs": {}})
    except Exception:
        pass
    return out


def _litros_por_linea(df: pd.DataFrame) -> pd.Series:
    """Litros de cada línea de venta, en cascada (fuente única del criterio de litros).

    1) CantBase × (Lts x caja / UxC) del maestro 04D  → primaria
    2) PesoKg del ERP (ya viene en litros)            → si el SKU no está en el maestro
    3) ml inferidos del nombre × CantBase (6X750→0,75) → último recurso

    Regla: un reporte no muestra 0 L porque falte el cálculo. Si falta, se calcula.
    Solo queda en 0 si el artículo no da litros por ninguna de las tres vías.
    """
    _, _, cod2lxu, _ = _cargar_maestro_04D()
    cod = (df["Codigo"].astype(str).str.strip().str.upper().str.replace(r"\.0$", "", regex=True)
           if "Codigo" in df.columns else pd.Series("", index=df.index))
    cant = pd.to_numeric(df.get("CantBase"), errors="coerce").fillna(0)
    litros = cant * cod.map(cod2lxu).fillna(0)
    peso = pd.to_numeric(df.get("PesoKg"), errors="coerce").fillna(0)
    litros = litros.where(litros > 0, peso)
    falta = litros <= 0
    if falta.any() and "Articulo" in df.columns:
        inferido = df["Articulo"].apply(_infer_litros_por_nombre) * cant
        litros = litros.where(~falta, inferido)
    return litros


def _sellout_desde_ventas(df_raw: pd.DataFrame) -> list:
    """
    Calcula sell out en litros por categoría desde un DataFrame de ventas ya filtrado
    (excl V2/V5/V20, importe > 0). Fuente de clasificación: 04D_MAESTRO_PRODUCTOS_PENAFLOR.xlsx.

    Litros: cascada de _litros_por_linea (maestro 04D → PesoKg → nombre del artículo).

    Spirits NO están en maestro 04D (son Diageo/P&P) → clasificar por keyword en Artículo:
      Nacional  = SMIRNOFF, GORDON, WHITE HORSE, J&B
      Importados = resto de spirits

    Devuelve lista de dicts con estructura:
      [{categoria, litros, objetivo, alcance_pct, clientes,
        subcategorias: [{nombre, litros, objetivo, alcance_pct, clientes,
                         marcas: [{marca, litros}]}]}]
    """
    # Objetivos por categoría: fuente única OBJSELLOUT.xlsx (no hardcode).
    obj_file = _cargar_objetivos_sellout()
    # Estructura de subcategorías (solo nombres). OBJSELLOUT.xlsx trae objetivo
    # SOLO a nivel categoría → las subcategorías muestran litros SIN objetivo.
    SUBS = {
        "VINOS DEL AÑO":     ["Alto", "Medio Alto", "Superior", "Medio"],
        "VINOS DE GUARDA":   [],
        "SPIRITS":           ["Nacionales", "Importados"],
        "RTD":               ["RTD", "RTD (S)"],
        "CHAMPAÑA":          [],
        "CERVEZA ARTESANAL": [],
        # Categoria nueva (16/07/2026): todavia sin objetivo en OBJSELLOUT.xlsx, asi que la
        # tarjeta la muestra con litros logrados y sin avance (objetivo/alcance = None).
        "VERMOUTH":          [],
    }
    _NAC_KW = ("SMIRNOFF", "GORDON", "WHITE HORSE", "J&B", "JYB")

    df = df_raw.copy()

    # ── Código artículo normalizado (strip + quitar ".0" de floats: "14605.0" → "14605")
    df["_cod"] = df["Codigo"].astype(str).str.strip().str.upper() if "Codigo" in df.columns else ""
    df["_cod"] = df["_cod"].str.replace(r"\.0$", "", regex=True)

    # ── Cargar maestro 04D (categoría, segmento, litros/unidad, linea comercial)
    cod2cat_04d, cod2seg_04d, cod2lxu_04d, cod2linea_04d = _cargar_maestro_04D()

    # ── Litros: cascada única (maestro 04D → PesoKg → nombre del artículo)
    df["litros"] = _litros_por_linea(df)

    # ── Categoría: maestro 04D → fallback Rubro (where() para evitar loc parcial en pandas 3.x)
    cat_maestro = df["_cod"].map(cod2cat_04d).astype(object)
    cat_maestro_norm = cat_maestro.str.strip().str.lower().map(_SO_CAT_MAP) if cat_maestro.notna().any() else cat_maestro
    cat_rubro = df["Rubro"].astype(str).str.strip().str.lower().map(_SO_CAT_MAP) if "Rubro" in df.columns else pd.Series(None, index=df.index, dtype=object)
    df["_cat"] = cat_maestro_norm.where(cat_maestro_norm.notna(), cat_rubro)
    # Categoría CRUDA del maestro (sin colapsar) para abrir RTD vs RTD (S)
    df["_cat_raw"] = cat_maestro.astype(str).str.strip().str.upper()

    # ── Segmento VDA: maestro 04D
    df["_seg"] = df["_cod"].map(cod2seg_04d).astype(str).str.strip()

    # ── Linea Comercial (marca): maestro 04D → Linea de ventas.csv → Marca de ventas.csv
    # Usar where() para evitar asignación parcial sobre columna float64 (pandas 3.x)
    linea_maestro = df["_cod"].map(cod2linea_04d)
    # Fallback chain: maestro Linea Comercial > Linea ERP > Marca ERP
    if "Linea" in df.columns:
        fallback = df["Linea"].astype(str).str.strip()
    elif "Marca" in df.columns:
        fallback = df["Marca"].astype(str).str.strip()
    else:
        fallback = pd.Series("", index=df.index)
    df["_linea"] = linea_maestro.where(linea_maestro.notna(), fallback)
    df["_linea"] = df["_linea"].fillna("").astype(str).str.strip()

    resultado = []
    for cat, sub_names in SUBS.items():
        grp = df[df["_cat"] == cat]
        litros   = round(float(grp["litros"].sum()), 1)
        clientes = int(grp["Cliente"].nunique()) if "Cliente" in grp.columns else 0
        obj_cat   = obj_file.get(cat, {})
        obj_total = obj_cat.get("total")            # None si la categoría no está en OBJSELLOUT.xlsx
        obj_subs  = obj_cat.get("subs", {})         # objetivo por Grupo PBP
        alcance   = round(litros / obj_total * 100, 1) if obj_total else None
        subs = []

        def _obj_sub(nombre):
            """Objetivo del subgrupo por nombre de Grupo PBP (match case-insensitive)."""
            if nombre in obj_subs:
                return obj_subs[nombre]
            return next((v for k, v in obj_subs.items() if k.lower() == nombre.lower()), None)

        if cat == "VINOS DEL AÑO":
            for sn in sub_names:
                # Segmento del maestro 04D: Alto / Medio Alto / Superior / Medio
                sg = grp[grp["_seg"].map(_SO_SEG_VDA) == sn]
                sl = round(float(sg["litros"].sum()), 1)
                sc = int(sg["Cliente"].nunique()) if "Cliente" in sg.columns else 0
                marcas = _marcas_de_grupo(sg)
                osub = _obj_sub(sn)
                subs.append({"nombre": sn, "litros": sl, "objetivo": osub,
                             "alcance_pct": round(sl / osub * 100, 1) if osub else None,
                             "clientes": sc, "marcas": marcas})

        elif cat == "SPIRITS":
            # Spirits NO están en maestro 04D → keyword por nombre de artículo
            _art = grp["Articulo"].astype(str).str.upper() if "Articulo" in grp.columns else pd.Series("", index=grp.index)
            mask_nac = _art.str.contains("|".join(_NAC_KW), na=False)
            for sn, mask_s in [("Nacionales", mask_nac), ("Importados", ~mask_nac)]:
                sg = grp[mask_s]
                sl = round(float(sg["litros"].sum()), 1)
                sc = int(sg["Cliente"].nunique()) if "Cliente" in sg.columns else 0
                marcas = _marcas_de_grupo(sg)
                osub = _obj_sub(sn)
                subs.append({"nombre": sn, "litros": sl, "objetivo": osub,
                             "alcance_pct": round(sl / osub * 100, 1) if osub else None,
                             "clientes": sc, "marcas": marcas})

        elif cat == "RTD":
            # RTD se abre en RTD (regular) y RTD (S), por la categoría cruda del maestro 04D.
            is_s = grp["_cat_raw"].astype(str).str.upper().str.replace(" ", "", regex=False) == "RTD(S)"
            for sn, mask_s in [("RTD", ~is_s), ("RTD (S)", is_s)]:
                sg = grp[mask_s]
                sl = round(float(sg["litros"].sum()), 1)
                sc = int(sg["Cliente"].nunique()) if "Cliente" in sg.columns else 0
                marcas = _marcas_de_grupo(sg)
                osub = _obj_sub(sn)
                subs.append({"nombre": sn, "litros": sl, "objetivo": osub,
                             "alcance_pct": round(sl / osub * 100, 1) if osub else None,
                             "clientes": sc, "marcas": marcas})

        resultado.append({
            "categoria": cat, "litros": litros, "objetivo": obj_total,
            "alcance_pct": alcance, "clientes": clientes, "subcategorias": subs,
            "marcas": _marcas_de_grupo(grp),
        })
    return resultado


def _marcas_de_grupo(sg: pd.DataFrame) -> list:
    """Devuelve lista [{marca, litros, varietales:[{nombre, litros}]}] ordenada desc.
    Marca   = _linea (Linea Comercial del maestro 04D); fallback Marca.
    Varietales = desglose por Articulo (SKU) dentro de la marca, ordenado desc."""
    col = "_linea" if "_linea" in sg.columns else ("Marca" if "Marca" in sg.columns else None)
    if col is None:
        return []
    art_col = "Articulo" if "Articulo" in sg.columns else None
    out = []
    for mk, sub in sg.groupby(col):
        mv = round(float(sub["litros"].sum()), 1)
        if mv <= 0 or not str(mk).strip():
            continue
        varietales = []
        if art_col:
            for av, lv in (sub.groupby(art_col)["litros"].sum()
                           .sort_values(ascending=False).items()):
                if lv > 0 and str(av).strip():
                    varietales.append({"nombre": str(av).strip(), "litros": round(float(lv), 1)})
        out.append({"marca": str(mk), "litros": mv, "varietales": varietales})
    out.sort(key=lambda x: x["litros"], reverse=True)
    return out


def _preparar_df_ventas(src_path, incluir_deposito=False) -> pd.DataFrame:
    """Lee ventas.csv, parsea columnas numéricas, excluye V2/V5 (+V20 salvo incluir_deposito),
    filtra importe > 0.
    incluir_deposito=True conserva V20 (Depósito / venta directa) para separarlo aguas abajo.
    sep=";" explícito: sep=None/engine=python sniffea mal el separador en Linux (Render)
    → columnas desalineadas, ImporteNetoItem=0, se pierden filas (mismo patrón que diagnóstico)."""
    df = None
    for enc in ("utf-8-sig", "latin-1", "windows-1252"):
        try:
            # dtype=str: no dejar que pandas infiera. En Render la inferencia de columnas
            # con coma decimal ("15800,82") difiere y casi todo quedaba en 0 (mismo patrón
            # que _leer_ventas_mes_csv, que sí funciona en Render).
            df = pd.read_csv(src_path, sep=";", encoding=enc, dtype=str, low_memory=False)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    if df is None:
        return pd.DataFrame()
    df.columns = [c.strip() for c in df.columns]
    for col in ("PesoKg", "CantBase", "ImporteNetoItem", "CodVendedor"):
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = (df[col].astype(str)
                              .str.strip()
                              .str.strip('"')
                              .str.replace(",", ".", regex=False)
                              .pipe(pd.to_numeric, errors="coerce")
                              .fillna(0))
    _excl = {2, 5} if incluir_deposito else {2, 5, 20}
    df = df[~df["CodVendedor"].isin(_excl) & (df["ImporteNetoItem"] > 0)].copy()
    return df


_DEPOSITO_VENTAS_CACHE = {}
def _df_deposito_ventas() -> pd.DataFrame:
    """ventas.csv (mes vivo) SOLO Depósito V20 (venta directa), neto>0.
    Alimenta los bloques 'depósito' aparte de gerencia (innovaciones, cobertura).
    NO filtra por Empresa: el depósito factura parte de su venta directa vía
    P&P Logística, pero es la misma entidad física V20 (igual criterio que el sell out
    y la conciliación con el proveedor).
    Cacheado por mtime de ventas.csv: lo invocan 2 endpoints por carga de gerencia y
    en Render (0.5 vCPU) releer el CSV de 3MB cada vez pesa."""
    src = INPUTS / "ventas.csv"
    if not src.exists():
        return pd.DataFrame()
    try:
        key = src.stat().st_mtime
    except OSError:
        key = 0
    cached = _DEPOSITO_VENTAS_CACHE.get("df")
    if cached is not None and _DEPOSITO_VENTAS_CACHE.get("key") == key:
        return cached
    vd = _preparar_df_ventas(src, incluir_deposito=True)
    vd = vd if vd.empty else vd[vd["CodVendedor"] == 20].copy()
    _DEPOSITO_VENTAS_CACHE.clear()
    _DEPOSITO_VENTAS_CACHE.update({"key": key, "df": vd})
    return vd


def _leer_ventas_mes_csv(src_path, incluir_deposito=False) -> pd.DataFrame:
    """Lee ventas_mes.csv robusto para Windows (CRLF) y Linux (LF).
    - sep=',' + quotechar='"' + engine='python' + dtype=str evita que el motor C
      de pandas deje comillas residuales en campos como "6620,94" al leer en Linux.
    - strip('"') elimina cualquier comilla residual antes de la conversión numérica.
    incluir_deposito=True conserva V20 (Depósito) para separarlo aguas abajo (sell out cierre)."""
    df = None
    for enc in ("utf-8-sig", "latin-1", "windows-1252"):
        try:
            df = pd.read_csv(src_path, sep=",", quotechar='"',
                             engine="python", dtype=str, encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    if df is None:
        return pd.DataFrame()
    df.columns = [c.strip() for c in df.columns]
    for col in ("PesoKg", "CantBase", "ImporteNetoItem", "CodVendedor"):
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = (df[col].astype(str)
                               .str.strip()
                               .str.strip('"')
                               .str.replace(",", ".", regex=False)
                               .pipe(pd.to_numeric, errors="coerce")
                               .fillna(0))
    _excl = {2, 5} if incluir_deposito else {2, 5, 20}
    df = df[~df["CodVendedor"].isin(_excl) & (df["ImporteNetoItem"] > 0)].copy()
    return df


def _sellout_con_deposito(df_full) -> dict:
    """Dado un df de ventas con V20 conservado, separa ruta (6 cat. con objetivo) y
    Depósito V20 (mismas cat., SIN objetivo/avance). Devuelve el dict de sell out con
    `categorias`, `deposito`, `total_ruta`, `total_deposito`, `total_general`.
    Usado por el sell out vivo y por el del cierre de mes (paridad)."""
    df_ruta = df_full[df_full["CodVendedor"] != 20]
    df_dep  = df_full[df_full["CodVendedor"] == 20]
    categorias = _sellout_desde_ventas(df_ruta)
    deposito   = _sellout_desde_ventas(df_dep) if not df_dep.empty else []
    for c in deposito:
        c["objetivo"], c["alcance_pct"] = None, None
        for s in c.get("subcategorias", []):
            s["objetivo"], s["alcance_pct"] = None, None
    total_ruta     = round(sum(float(c["litros"]) for c in categorias), 1)
    total_deposito = round(sum(float(c["litros"]) for c in deposito), 1)
    return {
        "categorias":     categorias,
        "deposito":       deposito,
        "total_ruta":     total_ruta,
        "total_deposito": total_deposito,
        "total_general":  round(total_ruta + total_deposito, 1),
    }


@app.route("/api/gerencia/sellout_litros")
def gerencia_sellout_litros():
    """Sellout en litros vs objetivos. Fuente: ventas.csv × maestro_04D_productos.csv.
    El objetivo de Sell Out es de la EMPRESA (independiente del vendedor): se agrupa TODA
    la venta —ruta (7 activos) + V20 Depósito (venta directa)— y se mide contra el objetivo.
    Por eso se incluye V20 en cada categoría; el total concilia con el reporte del proveedor."""
    src = INPUTS / "ventas.csv"
    if not src.exists():
        return jsonify({"error": "ventas.csv no encontrado en 01_INPUTS"}), 404
    df = _preparar_df_ventas(src, incluir_deposito=True)  # incluye V20: objetivo de empresa
    if df.empty:
        return jsonify({"error": "No se pudo leer ventas.csv"}), 500
    categorias = _sellout_desde_ventas(df)               # agrupa ruta + V20 vs objetivo
    so = {
        "categorias":      categorias,
        "total_litros":    round(sum(float(c["litros"]) for c in categorias), 1),
        "incluye_deposito": True,
        "generado_en":     _now_ar(),
        "fuente":          "ventas.csv (incl. V20 Depósito) + maestro_04D_productos.csv",
    }
    return jsonify(so)


# ====== ACCIONES COMERCIALES RANKING ======
@app.route("/api/gerencia/acciones_ranking")
def gerencia_acciones_ranking():
    """
    Detalle de acciones comerciales del mes, enriquecido con análisis comparativo
    vs mes anterior (clientes nuevos en categoría, delta litros, costo de activación).
    Lee mod_acciones_ranking.csv (detalle) + mod_acciones_analisis.csv (análisis).
    """
    df = read_csv(DATASETS / "mod_acciones_ranking.csv")
    if df.empty:
        return jsonify({"error": "Sin datos"}), 404
    df.columns = [c.strip() for c in df.columns]

    num_cols_det = ["litros_vendidos", "cajas_vendidas", "inversion_pesos",
                    "importe_neto", "clientes_afectados"]
    for c in num_cols_det:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    fecha = str(df["fecha_calculo"].iloc[0]) if "fecha_calculo" in df.columns and len(df) else ""

    # Cargar análisis histórico si existe
    df_an = read_csv(DATASETS / "mod_acciones_analisis.csv")
    analisis_map = {}
    if not df_an.empty:
        df_an.columns = [c.strip() for c in df_an.columns]
        num_cols_an = ["clientes_mes_actual", "clientes_cat_mes_ant", "clientes_nuevos_cat",
                       "clientes_retorno", "pct_clientes_nuevos", "litros_mes_actual",
                       "litros_mes_anterior", "delta_litros_pct", "inversion_pesos",
                       "costo_activacion"]
        for c in num_cols_an:
            if c in df_an.columns:
                df_an[c] = pd.to_numeric(df_an[c], errors="coerce").fillna(0)
        for _, row in df_an.iterrows():
            g = str(row.get("accion_grupo", "")).strip()
            if g:
                analisis_map[g] = {
                    "clientes_nuevos_cat":   int(row.get("clientes_nuevos_cat", 0)),
                    "clientes_retorno":      int(row.get("clientes_retorno", 0)),
                    "clientes_cat_mes_ant":  int(row.get("clientes_cat_mes_ant", 0)),
                    "pct_clientes_nuevos":   round(float(row.get("pct_clientes_nuevos", 0)), 1),
                    "litros_mes_actual":     round(float(row.get("litros_mes_actual", 0)), 1),
                    "litros_mes_anterior":   round(float(row.get("litros_mes_anterior", 0)), 1),
                    "delta_litros_pct":      round(float(row.get("delta_litros_pct", 0)), 1),
                    "costo_activacion":      int(row.get("costo_activacion", 0)),
                    "mes_anterior":          str(row.get("mes_anterior", "")),
                }

    acciones = []
    for _, row in df.iterrows():
        grupo  = str(row.get("accion_grupo", "")).strip()
        nombre = str(row.get("accion_nombre", row.get("canal", ""))).strip()
        an     = analisis_map.get(grupo, {})
        # Cap delta_litros a ±9999% para evitar valores sin sentido (ej: Termidor vs 0 litros)
        delta_l = an.get("delta_litros_pct", 0)
        if abs(delta_l) > 9999:
            delta_l = None
        acciones.append({
            "accion_grupo":         grupo,
            "accion_nombre":        nombre,
            "tipo_accion":          str(row.get("tipo_accion", "")),
            "canal":                str(row.get("canal", "")),
            "categoria":            str(row.get("categoria", "")),
            "descuento_display":    str(row.get("descuento_display", "")),
            "litros_vendidos":      round(float(row["litros_vendidos"]), 1),
            "cajas_vendidas":       int(row["cajas_vendidas"]),
            "inversion_pesos":      int(row["inversion_pesos"]),
            "importe_neto":         int(row["importe_neto"]),
            "clientes_afectados":   int(row["clientes_afectados"]),
            # análisis comparativo
            "clientes_nuevos_cat":  an.get("clientes_nuevos_cat"),
            "clientes_retorno":     an.get("clientes_retorno"),
            "clientes_cat_mes_ant": an.get("clientes_cat_mes_ant"),
            "pct_clientes_nuevos":  an.get("pct_clientes_nuevos"),
            "litros_mes_anterior":  an.get("litros_mes_anterior"),
            "delta_litros_pct":     delta_l,
            "costo_activacion":     an.get("costo_activacion"),
            "mes_anterior":         an.get("mes_anterior", ""),
        })
    acciones.sort(key=lambda x: x["inversion_pesos"], reverse=True)
    return jsonify({
        "generado_en": _now_ar(),
        "fuente":        "mod_acciones_ranking.csv + mod_acciones_analisis.csv",
        "fecha_calculo": fecha,
        "acciones":      acciones,
    })


# ====== RUTA DEL DÍA (vendedor) ======
# 11 Titulares (marcas) — fuente ventas.csv (mes vivo). Mismos para Trad y Autoservicio.
_RUTA_ONCE = ["ALMA MORA", "TRAPICHE RESERVA", "FINCA LAS MORAS", "ALARIS", "DON DAVID",
              "DADA", "SMIRNOFF FLAVOURS", "LOS ARBOLES", "ANTARES", "SMIRNOFF ICE", "GORDON'S FLAVOURS"]
_RUTA_MARCA = {
    "ALMA MORA": "ALMA MORA", "ALARIS": "ALARIS", "TRAPICHE ALARIS": "ALARIS",
    "DON DAVID": "DON DAVID", "DADA": "DADA", "LOS ARBOLES": "LOS ARBOLES",
    "FINCA LAS MORAS": "FINCA LAS MORAS", "F LAS MORAS": "FINCA LAS MORAS",
    "TRAPICHE RESERVA": "TRAPICHE RESERVA", "ANTARES": "ANTARES",
    "GORDON'S FLAVOURS": "GORDON'S FLAVOURS", "GORDONS FLAVOURS": "GORDON'S FLAVOURS",
    "GORDON'S": "GORDON'S FLAVOURS", "GORDONS": "GORDON'S FLAVOURS",
    "SMIRNOFF": "SMIRNOFF FLAVOURS", "SMIRNOFF FLAVOURS": "SMIRNOFF FLAVOURS",
    "SMIRNOFF ICE FLAVOURS": "SMIRNOFF ICE", "SMIRNOFF ICE": "SMIRNOFF ICE",
}
_RUTA_ART = [
    ("SMIRNOFF ICE", "SMIRNOFF ICE"), ("SMF ICE", "SMIRNOFF ICE"),
    ("SMIRNOFF", "SMIRNOFF FLAVOURS"), ("GORDON", "GORDON'S FLAVOURS"),
    ("ANTARES", "ANTARES"), ("ALMA MORA", "ALMA MORA"), ("LOS ARBOLES", "LOS ARBOLES"),
    ("DADA", "DADA"), ("FINCA LAS MORAS", "FINCA LAS MORAS"), ("F.LAS MORAS", "FINCA LAS MORAS"),
    ("DON DAVID", "DON DAVID"), ("ALARIS", "ALARIS"), ("TRAPICHE RESERVA", "TRAPICHE RESERVA"),
]


def _ruta_titular(marca, articulo):
    m = str(marca).upper().strip()
    if m in _RUTA_MARCA:
        return _RUTA_MARCA[m]
    a = str(articulo).upper()
    for kw, t in _RUTA_ART:
        if kw in a:
            return t
    return None


@app.route("/api/vendedor/<vid>/ruta")
def vendedor_ruta(vid):
    """Ruta del día del vendedor. Cartera de clientes.xlsx (DiasVisita = día), con compra/sin
    compra y 11 Titulares faltantes calculados desde ventas.csv (mes vivo). Solo los 11 titulares."""
    vid_norm = normalizar_vendedor_codigo(vid)
    cn = clean_code(vid_norm)
    out = {"vendedor_id": vid_norm, "dia": "", "clientes": [], "total": 0,
           "con_compra": 0, "sin_compra": 0}
    cli_path = INPUTS / "clientes.xlsx"
    if not cli_path.exists():
        return jsonify(out), 200
    dia_req = request.args.get("dia", "").strip()
    if not dia_req:
        _DIAS_AR = {0: "LU", 1: "MA", 2: "MI", 3: "JU", 4: "VI", 5: "SA", 6: "DO"}
        dia_req = _DIAS_AR[datetime.now(_ARG_TZ).weekday()]
    out["dia"] = dia_req
    try:
        cli = pd.read_excel(cli_path)
    except Exception:
        return jsonify(out), 200
    cli.columns = [str(c).strip() for c in cli.columns]
    if not {"DiasVisita", "Codigo", "codven"}.issubset(cli.columns):
        return jsonify(out), 200
    cli = cli[cli["DiasVisita"].astype(str).str.strip().str.upper() == dia_req.upper()]
    cli = cli[cli["codven"].astype(str).apply(clean_code) == cn]
    if cli.empty:
        return jsonify(out), 200

    # Catálogo de innovaciones por segmento (fuente oficial: mod_innovaciones_segmento.csv)
    inov_seg = read_csv(DATASETS / "mod_innovaciones_segmento.csv")
    inov_by_seg = {}   # segmento -> {producto_codigo: producto_nombre}
    if not inov_seg.empty:
        inov_seg.columns = [c.lstrip("﻿") for c in inov_seg.columns]
        for _, ir in inov_seg.iterrows():
            cod_i = pd.to_numeric(ir.get("producto_codigo"), errors="coerce")
            if pd.isna(cod_i):
                continue
            segn = str(ir.get("segmento", "")).upper()
            inov_by_seg.setdefault(segn, {})[int(cod_i)] = str(ir.get("producto_nombre", "") or int(cod_i))
    inov_codes = set().union(*[set(d) for d in inov_by_seg.values()]) if inov_by_seg else set()

    # ventas.csv (mes vivo): clientes con compra + titulares e innovaciones compradas por cliente
    bought_clients, tit_cli, inov_cli = set(), {}, {}
    vpath = INPUTS / "ventas.csv"
    if vpath.exists():
        try:
            v = pd.read_csv(vpath, sep=";", encoding="latin1", engine="python")
            v["imp"] = pd.to_numeric(v["ImporteNetoItem"].astype(str).str.replace(",", ".", regex=False), errors="coerce")
            # No se filtra por Empresa (ver _LEEME_EMPRESA). Sí se excluyen los vendedores
            # que no son de ruta (V1/V2/V5/V20).
            v = v[(v["imp"] > 0) & (~v["CodVendedor"].isin(_VENDEDORES_EXCLUIDOS))]
            v["cli"] = pd.to_numeric(v["Cliente"], errors="coerce")
            bought_clients = set(v["cli"].dropna().astype(int))
            v["_tit"] = [_ruta_titular(m, a) for m, a in zip(v.get("Marca", ""), v.get("Articulo", ""))]
            for cid, grp in v[v["_tit"].notna()].dropna(subset=["cli"]).groupby("cli"):
                tit_cli[int(cid)] = set(grp["_tit"])
            if inov_codes:
                v["_cod"] = pd.to_numeric(v.get("Codigo"), errors="coerce")
                for cid, grp in v[v["_cod"].isin(inov_codes)].dropna(subset=["cli"]).groupby("cli"):
                    inov_cli[int(cid)] = set(grp["_cod"].dropna().astype(int))
        except Exception:
            pass

    sub_col = next((c for c in cli.columns if "subseg" in c.lower() or "subramo" in c.lower()), None)
    nombre_col = next((c for c in cli.columns if c.lower().replace("_", "") in ("razonsocial", "nombre")), None)
    once_set = set(_RUTA_ONCE)
    clientes = []
    for _, r in cli.iterrows():
        if pd.isna(r.get("Codigo")):
            continue
        cid = int(r["Codigo"])
        seg = _clasificar_segmento(str(r.get("Ramo", "")), str(r.get(sub_col, "") if sub_col else ""))
        # V3 (Nadia): solo Tradicional almacén/despensa/kiosco. Quedan fuera AS, On Premise
        # y también los tradicionales que NO son almacén/despensa/kiosco (fiambrería, resto).
        if vid_norm == "V3":
            subseg = str(r.get(sub_col, "")).upper() if sub_col else ""
            es_alm_kio = any(k in subseg for k in ("ALMACEN", "DESPENSA", "KIOSCO"))
            if seg.upper() != "TRADICIONAL" or not es_alm_kio:
                continue
        comp = cid in bought_clients
        tb = tit_cli.get(cid, set()) & once_set
        comp_11 = [t for t in _RUTA_ONCE if t in tb]
        faltan = [t for t in _RUTA_ONCE if t not in tb]
        # Innovaciones aplicables al segmento del cliente (V3 sin AUTOSERVICIO)
        seg_u = seg.upper()
        cat_inov = {} if (vid_norm == "V3" and seg_u == "AUTOSERVICIO") else inov_by_seg.get(seg_u, {})
        bought_inov = inov_cli.get(cid, set())
        inov_comp = sorted({cat_inov[c] for c in cat_inov if c in bought_inov})
        inov_falt = sorted({cat_inov[c] for c in cat_inov if c not in bought_inov})
        orden_val = pd.to_numeric(r.get("Orden"), errors="coerce")
        clientes.append({
            "cliente_id": cid,
            "cliente_nombre": str(r.get(nombre_col, "") or cid) if nombre_col else str(cid),
            "vendedor_id": vid_norm,
            "dia_visita": dia_req,
            "segmento": seg,
            "orden": int(orden_val) if pd.notnull(orden_val) else None,
            "estado": "COBERTURA_OK" if comp else "SIN_COMPRA_MES",
            "compra_mes_flag": 1 if comp else 0,
            "once_t_comprados": len(tb),
            "once_t_total": len(_RUTA_ONCE),
            "titulares_comprados": comp_11,
            "faltan_11t": len(faltan),
            "titulares_faltantes": faltan,
            "inov_comprados": inov_comp,
            "inov_faltantes": inov_falt,
            "inov_comprados_n": len(inov_comp),
            "inov_total": len(cat_inov),
        })
    # Orden de visita (columna Orden de clientes.xlsx). Orden<=0 = sin asignar → al final.
    def _orden_key(c):
        o = c["orden"]
        tiene = o is not None and o > 0
        return (not tiene, o if tiene else 0, c["cliente_nombre"])
    clientes.sort(key=_orden_key)
    out["clientes"] = clientes
    out["total"] = len(clientes)
    out["con_compra"] = sum(1 for c in clientes if c["compra_mes_flag"])
    out["sin_compra"] = sum(1 for c in clientes if not c["compra_mes_flag"])
    return jsonify(out), 200


# ====== OPORTUNIDADES DEL DÍA: innovaciones ======
@app.route("/api/vendedor/<vid>/oportunidades_innovacion")
def vendedor_oportunidades_innovacion(vid):
    """3 clientes de la ZONA DE HOY que compraron el mes pasado y este mes, pero NUNCA
    innovaciones (ni el mes anterior ni el actual). Fuente: ventas_acumulada.csv (mayo+junio)."""
    import random as _random
    vid_norm = normalizar_vendedor_codigo(vid)
    cn = clean_code(vid_norm)
    vacio = {"vendedor_nombre": vid_norm, "clientes": [], "innovaciones": [], "texto": ""}
    if vid_norm in ("V2", "V5", "V20"):
        return jsonify(vacio), 200

    # nombre del vendedor
    vendedor_nombre = vid_norm
    vend = read_csv(CONFIG / "vendedores_activos.csv")
    if not vend.empty and "codigo_vendedor" in vend.columns:
        fila = vend[vend["codigo_vendedor"].astype(str).apply(clean_code) == cn]
        if not fila.empty:
            col_n = next((c for c in ("nombre", "nombre_vendedor", "vendedor") if c in fila.columns), None)
            if col_n:
                vendedor_nombre = str(fila.iloc[0][col_n])
    vacio["vendedor_nombre"] = vendedor_nombre

    inv = read_csv(DATASETS / "mod_innovaciones_segmento.csv")
    vac_path = INPUTS / "ventas_acumulada.csv"
    if inv.empty or not vac_path.exists():
        return jsonify(vacio), 200
    innov_codes = set(pd.to_numeric(inv["producto_codigo"], errors="coerce").dropna().astype(int))
    innov_names = sorted(set(str(x) for x in inv["producto_nombre"].dropna() if str(x).strip()))

    try:
        vac = pd.read_csv(vac_path, sep=";", encoding="latin1", low_memory=False)
    except Exception:
        return jsonify(vacio), 200
    vac["imp"] = pd.to_numeric(vac["ImporteNetoItem"].astype(str).str.replace(",", ".", regex=False), errors="coerce")
    # No se filtra por Empresa (ver _LEEME_EMPRESA)
    vac = vac[(vac["imp"] > 0) & (vac["CodVendedor"].astype(str).apply(clean_code) == cn)]
    if vac.empty:
        return jsonify(vacio), 200
    vac["cod"] = pd.to_numeric(vac["Codigo"], errors="coerce")
    vac["cli"] = pd.to_numeric(vac["Cliente"], errors="coerce")
    vac["mes"] = pd.to_datetime(vac["FechaComprobante"], dayfirst=True, errors="coerce").dt.to_period("M")
    per_act = vac["mes"].max()
    per_ant = per_act - 1
    cur, prev = vac[vac["mes"] == per_act], vac[vac["mes"] == per_ant]
    prev_any = set(prev["cli"].dropna().astype(int))
    prev_inov = set(prev[prev["cod"].isin(innov_codes)]["cli"].dropna().astype(int))
    cur_any = set(cur["cli"].dropna().astype(int))
    cur_inov = set(cur[cur["cod"].isin(innov_codes)]["cli"].dropna().astype(int))
    # compró ambos meses, pero sin innovaciones en ninguno
    cand = (prev_any - prev_inov) & (cur_any - cur_inov)

    # clientes de la zona de HOY
    dia_req = request.args.get("dia", "").strip()
    if not dia_req:
        _DIAS_AR = {0: "LU", 1: "MA", 2: "MI", 3: "JU", 4: "VI", 5: "SA", 6: "DO"}
        dia_req = _DIAS_AR[datetime.now(_ARG_TZ).weekday()]
    try:
        cd = _clientes_por_dia(dia_req)
        if not cd.empty and "cliente_id" in cd.columns:
            dia_ids = set(pd.to_numeric(cd["cliente_id"], errors="coerce").dropna().astype(int))
            cand = cand & dia_ids
    except Exception:
        pass

    innov3 = _random.sample(innov_names, min(3, len(innov_names))) if innov_names else []
    if not cand:
        return jsonify({"vendedor_nombre": vendedor_nombre, "dia": dia_req,
                        "clientes": [], "innovaciones": innov3, "texto": ""}), 200

    # top 3 por volumen ($ acumulado mayo+junio)
    vol = vac[vac["cli"].isin(cand)].groupby("cli")["imp"].sum().sort_values(ascending=False)
    nm = vac[vac["cli"].isin(cand)].dropna(subset=["cli"]).drop_duplicates("cli")
    nombres = dict(zip(nm["cli"].astype(int), nm["RazonSocial"].astype(str))) if "RazonSocial" in vac.columns else {}
    top3 = [int(c) for c in vol.index[:3]]
    clientes = [{"cliente_id": cid, "cliente_nombre": nombres.get(cid, str(cid)),
                 "importe": round(float(vol.get(cid, 0)), 0)} for cid in top3]

    nombres_cli = ", ".join(c["cliente_nombre"] for c in clientes)
    marcas_txt = ", ".join(innov3)
    texto = (f"Hoy {vendedor_nombre}, andá a venderles innovaciones a estos {len(clientes)} clientes: "
             f"{nombres_cli}. El mes pasado hicieron compra, al igual que este mes, pero todavía "
             f"no compraron estas marcas: {marcas_txt}.")
    return jsonify({"vendedor_nombre": vendedor_nombre, "dia": dia_req,
                    "clientes": clientes, "innovaciones": innov3, "texto": texto}), 200


# ====== INCENTIVO CLUB FARO ======
# TODA la definición del incentivo se lee de la HOJA 01_INPUTS/incentivo_club_faro*.xlsx (Hoja1):
# categorías, segmento, umbral, códigos de SKU, tope por cliente, período (meses), objetivos por
# vendedor, premios y mapping de supervisores. Cambia cada bimestre → el usuario SOLO edita el Excel,
# no hay que tocar código (_faro_config parsea la hoja tal como se escribe hoy: grilla de objetivos +
# reglas en texto libre). Avance/no-compradores: ventas_acumulada.csv filtrado a los meses del período.
# Match de producto por CÓDIGO de SKU. Segmento por Ramo+Subramo de la venta (no el maestro).
_FARO_SUP_MAP_DEFAULT = {"Esteban": [3, 4, 6, 8, 10], "Raul": [7, 9]}   # fallback si la hoja no lo trae
_FARO_MESES_ES = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
                  "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
                  "noviembre": 11, "diciembre": 12}


def _faro_xlsx_path():
    cands = sorted(INPUTS.glob("incentivo_club_faro*.xlsx"))
    return cands[0] if cands else None


def _faro_norm(s):
    import unicodedata as _ud
    s = str(s or "")
    s = _ud.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.lower().strip()


def _faro_slug(s):
    import re as _re
    return _re.sub(r"[^a-z0-9]+", "_", _faro_norm(s)).strip("_")


_FARO_CFG_CACHE = {}
def _faro_config():
    """Lee toda la definición del incentivo desde el xlsx (cacheada por mtime).
    Devuelve dict con: cats (orden), cat_nombre/seg/umbral/skus/cap, objetivos{cod:{cat}},
    premios{cat}, meses (tuple), periodo, sup_map. None si no se puede leer lo esencial."""
    import re as _re, datetime as _dt
    p = _faro_xlsx_path()
    if not p:
        return None
    try:
        mt = p.stat().st_mtime
    except OSError:
        mt = 0
    cached = _FARO_CFG_CACHE.get(mt)
    if cached is not None:
        return cached
    try:
        raw = pd.read_excel(p, sheet_name=0, header=None, dtype=str)
    except Exception as e:
        print(f"[AVISO] FARO no se pudo leer {p.name}: {e}")
        return None
    nrows, ncols = raw.shape

    def cell(r, c):
        if r < 0 or r >= nrows or c < 0 or c >= ncols:
            return ""
        v = raw.iat[r, c]
        return "" if pd.isna(v) else str(v).strip()

    # Fila de encabezado de la grilla (col0 = "vendedores")
    hr = next((r for r in range(nrows) if _faro_norm(cell(r, 0)).startswith("vendedor")), None)
    if hr is None:
        return None
    # Categorías (columnas con nombre a la derecha de col0)
    cats, cat_nombre, cat_col = [], {}, {}
    for c in range(1, ncols):
        nom = cell(hr, c)
        if nom:
            k = _faro_slug(nom)
            if k and k not in cat_nombre:
                cats.append(k); cat_nombre[k] = nom; cat_col[k] = c
    if not cats:
        return None
    # Banda de segmento (fila hr-1, con forward-fill por celdas combinadas)
    seg_row, last = {}, ""
    for c in range(1, ncols):
        t = _faro_norm(cell(hr - 1, c))
        if t:
            last = t
        seg_row[c] = last
    def _seg_de(txt):
        if "autoservicio" in txt:
            return "AUTOSERVICIO"
        if any(k in txt for k in ("kiosco", "almacen", "tradicional", "despensa")):
            return "TRADICIONAL"
        return "TRADICIONAL"
    cat_seg = {k: _seg_de(seg_row.get(cat_col[k], "")) for k in cats}

    # Objetivos por vendedor (filas con col0 numérico)
    objetivos = {}
    for r in range(hr + 1, nrows):
        c0 = cell(r, 0)
        if not c0.isdigit():
            continue
        cod = int(c0)
        objetivos[cod] = {}
        for k in cats:
            v = cell(r, cat_col[k])
            try:
                objetivos[cod][k] = int(float(v)) if v else 0
            except ValueError:
                objetivos[cod][k] = 0
    if not objetivos:
        return None

    # Reglas en texto libre (col0): SKUs, umbral y tope por categoría
    text_rows = [_faro_norm(cell(r, 0)) for r in range(nrows)]
    cat_skus, cat_umbral, cat_cap = {}, {}, {}
    for k in cats:
        nom_n = _faro_norm(cat_nombre[k])
        linea = next((t for t in text_rows if nom_n and nom_n in t), "")
        cat_skus[k] = set(_re.findall(r"\b(\d{4,})\b", linea))     # códigos de SKU (>=4 dígitos)
        m = _re.search(r"(\d+)\s*(?:botell|lata)", linea)          # umbral explícito o default por segmento
        cat_umbral[k] = int(m.group(1)) if m else (6 if cat_seg[k] == "AUTOSERVICIO" else 3)
        mc = _re.search(r"(\d+)\s*maxim|maxim\w*\s*(\d+)", linea)  # tope de coberturas por cliente
        cat_cap[k] = int(next(g for g in mc.groups() if g)) if mc else None

    # Premios (fila con "MILLAS") → categoría por coincidencia de tokens del nombre
    premios = {}
    milla_txt = " ".join(str(v) for v in raw.values.ravel() if pd.notna(v))
    for etq, num in _re.findall(r"([A-Za-zÁÉÍÓÚÑ +&']+?)\s*\(\s*([\d.]+)\s*MILLAS\)", milla_txt, _re.IGNORECASE):
        e = set(_faro_norm(etq).split())
        millas = int(float(num.replace(".", "")))
        best, score = None, 0
        for k in cats:
            sc = len(set(_faro_norm(cat_nombre[k]).split()) & e)
            if sc > score:
                best, score = k, sc
        if best:
            premios[best] = millas

    # Período (meses) desde el título; fallback al bimestre en curso
    meses = tuple(v for m, v in _FARO_MESES_ES.items()
                  if _re.search(r"\b" + m + r"\b", _faro_norm(cell(0, 0))))
    meses = tuple(sorted(set(meses)))
    if not meses:
        mm = _dt.date.today().month
        st = mm - ((mm - 1) % 2)
        meses = (st, st + 1)
    _inv = {}
    for m, v in _FARO_MESES_ES.items():
        _inv.setdefault(v, m)
    periodo = "-".join(_inv[v] for v in meses if v in _inv) or "período"

    # Supervisores (línea "Esteban ... Raul ...")
    sup_map = {}
    linea_sup = next((t for t in text_rows if "esteban" in t and "raul" in t), "")
    if linea_sup:
        idx = linea_sup.find("raul")
        est = [int(x) for x in _re.findall(r"\d+", linea_sup[:idx])]
        rau = [int(x) for x in _re.findall(r"\d+", linea_sup[idx:])]
        if est:
            sup_map["Esteban"] = est
        if rau:
            sup_map["Raul"] = rau
    if not sup_map:
        sup_map = dict(_FARO_SUP_MAP_DEFAULT)

    cfg = dict(cats=cats, cat_nombre=cat_nombre, cat_seg=cat_seg, cat_umbral=cat_umbral,
               cat_skus=cat_skus, cat_cap=cat_cap, objetivos=objetivos, premios=premios,
               meses=meses, periodo=periodo, sup_map=sup_map,
               fuente=(p.name if p else None))
    _FARO_CFG_CACHE.clear()
    _FARO_CFG_CACHE[mt] = cfg
    return cfg


_FARO_VENTAS_CACHE = {}
def _faro_ventas(cfg):
    """ventas_acumulada.csv del bimestre del incentivo (cfg['meses']) preparada para FARO,
    cacheada por (mtime ventas, mtime xlsx). Categoría por código de SKU (cfg['cat_skus'])."""
    p = INPUTS / "ventas_acumulada.csv"
    if not p.exists() or not cfg:
        return pd.DataFrame()
    xp = _faro_xlsx_path()
    try:
        key = (os.path.getmtime(p), os.path.getmtime(xp) if xp else 0)
    except OSError:
        key = (0, 0)
    df = _FARO_VENTAS_CACHE.get(key)
    if df is not None:
        return df
    df = pd.read_csv(p, sep=";", encoding="latin1", low_memory=False)
    df.columns = [c.strip() for c in df.columns]
    df["_fecha"] = pd.to_datetime(df.get("FechaComprobante"), dayfirst=True, errors="coerce")
    df["_neto"] = pd.to_numeric(df["ImporteNetoItem"].astype(str).str.replace(",", ".", regex=False), errors="coerce").fillna(0)
    df["_cant"] = pd.to_numeric(df["CantBase"].astype(str).str.replace(",", ".", regex=False), errors="coerce").fillna(0)
    df["_vend"] = pd.to_numeric(df["CodVendedor"], errors="coerce")
    df["_cli"]  = pd.to_numeric(df["Cliente"], errors="coerce")
    df = df[(df["_neto"] > 0) & (~df["_vend"].isin(_VENDEDORES_EXCLUIDOS))].copy()
    df = df[df["_fecha"].dt.month.isin(cfg["meses"])].copy()   # bimestre del incentivo (de la hoja)
    df["_seg"] = [_clasificar_segmento(str(r), str(s))
                  for r, s in zip(df.get("Ramo", ""), df.get("Subramo", ""))]
    au = df["Articulo"].astype(str).str.upper()
    cod = df.get("Codigo", pd.Series([""] * len(df), index=df.index)).astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    df["_cod"] = cod   # código de SKU normalizado (match de categoría FARO por código)
    # Categoría FARO por CÓDIGO de SKU (matriz cfg['cat_skus'], leída de la hoja)
    cat = pd.Series([None] * len(df), index=df.index, dtype=object)
    for _c, _skus in cfg["cat_skus"].items():
        if _skus:
            cat[cod.isin(_skus)] = _c
    df["_cat"] = cat
    df["_art"] = au    # nombre de artículo (para drill-down de SKU)
    df["_clinom"] = (df["RazonSocial"].astype(str) if "RazonSocial" in df.columns else df["Cliente"].astype(str))
    df["_loc"] = (df["Localidad"].astype(str) if "Localidad" in df.columns else pd.Series([""] * len(df), index=df.index))
    _FARO_VENTAS_CACHE.clear()
    _FARO_VENTAS_CACHE[key] = df
    return df


def _faro_detalle_vendedor(df, cod, cfg):
    """Por categoría (de cfg): {logrado, clientes_cubiertos, compradores, no_compradores} para un
    vendedor. Cobertura POR SKU: cada SKU con ≥ umbral botellas del PDV suma 1; se suman por cliente;
    cfg['cat_cap'] limita coberturas/cliente. No-compradores = clientes del canal a los que el
    vendedor vendió en el bimestre y NO lograron ninguna cobertura."""
    out = {}
    dv = df[df["_vend"] == cod] if not df.empty else df
    for cat in cfg["cats"]:
        seg = cfg["cat_seg"][cat]
        um  = cfg["cat_umbral"][cat]
        cap = cfg["cat_cap"].get(cat)
        canal = dv[dv["_seg"] == seg]
        canal_ids = set(canal["_cli"].dropna().astype(int))
        marca = canal[canal["_cat"] == cat]
        # Cobertura POR SKU: un SKU cuenta 1 si el PDV compró ≥ umbral botellas de ESE SKU en el bimestre.
        per_sku = marca.groupby(["_cli", "_cod"])["_cant"].sum().reset_index()
        qual = per_sku[per_sku["_cant"] >= um]
        cob_cli = qual.groupby("_cli")["_cod"].nunique()        # coberturas (SKUs que califican) por cliente
        if cap:
            cob_cli = cob_cli.clip(upper=cap)                   # tope de coberturas por cliente
        cob_map = cob_cli.to_dict()
        cubiertos = set(int(c) for c in cob_cli.index)          # clientes con ≥1 cobertura
        logrado = int(cob_cli.sum())
        bot_cli = marca.groupby("_cli")["_cant"].sum().to_dict()  # botellas totales de la categoría (display)
        meta = canal.groupby("_cli").agg(nom=("_clinom", "first"), loc=("_loc", "first"))
        def _cli_row(cid):
            return {
                "cliente":        int(cid),
                "razon_social":   str(meta["nom"].get(cid, "")).strip()[:45],
                "localidad":      str(meta["loc"].get(cid, "")).strip()[:25],
                "botellas_marca": round(float(bot_cli.get(cid, 0)), 1),
                "peso":           int(cob_map.get(cid, 1)),      # coberturas aportadas por el cliente
            }
        # Clientes CON cobertura lograda (drill-down de gerencia/vendedor)
        compradores = sorted((_cli_row(c) for c in cubiertos),
                             key=lambda x: (-x["peso"], -x["botellas_marca"], x["cliente"]))
        # No-compradores: clientes del canal del vendedor que no lograron ninguna cobertura
        no_comp = sorted((_cli_row(c) for c in (canal_ids - cubiertos)),
                         key=lambda x: (-x["botellas_marca"], x["cliente"]))
        out[cat] = {
            "logrado": logrado,
            "clientes_cubiertos": int(len(cubiertos)),
            "compradores": compradores,
            "no_compradores": no_comp,
        }
    return out


def _faro_nombres_vendedores():
    vend = read_csv(CONFIG / "vendedores_activos.csv")
    m = {}
    if not vend.empty:
        for _, v in vend.iterrows():
            try:
                m[int(clean_code(str(v["codigo_vendedor"])))] = str(v["nombre_vendedor"]).strip()
            except (ValueError, TypeError):
                pass
    return m


@app.route("/api/gerencia/incentivo_faro")
def gerencia_incentivo_faro():
    """Incentivo Club FARO — objetivo vs logrado por vendedor y supervisor (gerencia).
    Toda la definición (categorías/SKUs/umbrales/período/objetivos/premios) sale del xlsx (_faro_config)."""
    cfg = _faro_config()
    if not cfg:
        return jsonify({"error": "No se pudo leer incentivo_club_faro.xlsx", "vendedores": [],
                        "supervisores": [], "categorias": {}, "categorias_orden": []}), 200
    cats = cfg["cats"]; obj = cfg["objetivos"]; premios = cfg["premios"]
    df = _faro_ventas(cfg)
    nombres = _faro_nombres_vendedores()
    logr = {cod: _faro_detalle_vendedor(df, cod, cfg) for cod in obj}

    def _cat_block(o_dict, l_dict):
        b = {"_millas_alcanzadas": 0, "_millas_posibles": 0}
        for cat in cats:
            o = o_dict.get(cat, 0)
            l = l_dict.get(cat, {}).get("logrado", 0)
            millas = premios.get(cat, 0)
            alcanzado = bool(o) and l >= o
            if o:   # solo categorías en juego (con objetivo asignado) cuentan para millas posibles
                b["_millas_posibles"] += millas
                if alcanzado:
                    b["_millas_alcanzadas"] += millas
            b[cat] = {"objetivo": o, "logrado": l, "pct": round(l / o * 100, 1) if o else None,
                      "premio_millas": millas, "alcanzado": alcanzado,
                      "clientes_cubiertos": l_dict.get(cat, {}).get("clientes_cubiertos", 0),
                      "compradores": l_dict.get(cat, {}).get("compradores", [])}
        return b

    def _finalize(row, b):
        row["millas_alcanzadas"] = b.pop("_millas_alcanzadas")
        row["millas_posibles"]   = b.pop("_millas_posibles")
        row.update(b)
        return row

    vendedores = []
    for cod in sorted(obj):
        row = {"codigo": cod, "nombre": nombres.get(cod, f"V{cod}"), "tipo": "vendedor"}
        vendedores.append(_finalize(row, _cat_block(obj[cod], logr.get(cod, {}))))

    supervisores = []
    for nom, vends in cfg["sup_map"].items():
        o_sum = {cat: sum(obj.get(v, {}).get(cat, 0) for v in vends) for cat in cats}
        l_sum = {cat: {"logrado": sum(logr.get(v, {}).get(cat, {}).get("logrado", 0) for v in vends)} for cat in cats}
        row = {"nombre": nom, "tipo": "supervisor", "vendedores": [f"V{v}" for v in vends]}
        supervisores.append(_finalize(row, _cat_block(o_sum, l_sum)))

    return jsonify(_to_native({
        "vendedores": vendedores, "supervisores": supervisores,
        "categorias": cfg["cat_nombre"],
        "categorias_orden": list(cats),
        "categorias_meta": {c: {"nombre": cfg["cat_nombre"][c],
                                "segmento": "Tradicional" if cfg["cat_seg"][c] == "TRADICIONAL" else "Autoservicio",
                                "umbral": cfg["cat_umbral"][c], "premio_millas": premios.get(c, 0)}
                            for c in cats},
        "premios": premios, "fuente": "ventas_acumulada.csv",
        "periodo": cfg["periodo"], "objetivo_fuente": cfg.get("fuente"),
    }))


@app.route("/api/vendedor/<vid>/incentivo_faro")
def vendedor_incentivo_faro(vid):
    """Incentivo Club FARO del vendedor: objetivo vs logrado + clientes no compradores por categoría."""
    import re as _re
    cod = int(_re.sub(r"\D", "", vid) or 0)
    cfg = _faro_config()
    if not cfg:
        return jsonify({"error": "No se pudo leer incentivo_club_faro.xlsx",
                        "vendedor": vid, "codigo": cod, "categorias": []}), 200
    cats = cfg["cats"]; premios = cfg["premios"]
    obj = cfg["objetivos"].get(cod, {c: 0 for c in cats})
    df = _faro_ventas(cfg)
    det = _faro_detalle_vendedor(df, cod, cfg)
    es_v3 = (cod == 3)
    categorias = []
    millas_alcanzadas = 0
    millas_posibles = 0
    for cat in cats:
        # V3 no trabaja Autoservicio → solo categorías de canal tradicional
        if es_v3 and cfg["cat_seg"][cat] != "TRADICIONAL":
            continue
        o = obj.get(cat, 0)
        l = det.get(cat, {}).get("logrado", 0)
        millas = premios.get(cat, 0)
        alcanzado = bool(o) and l >= o
        if o:   # solo categorías con objetivo asignado cuentan para millas posibles
            millas_posibles += millas
            if alcanzado:
                millas_alcanzadas += millas
        categorias.append({
            "cat": cat, "nombre": cfg["cat_nombre"][cat],
            "segmento": "Tradicional" if cfg["cat_seg"][cat] == "TRADICIONAL" else "Autoservicio",
            "umbral": cfg["cat_umbral"][cat],
            "objetivo": o, "logrado": l, "pct": round(l / o * 100, 1) if o else None,
            "premio_millas": millas, "alcanzado": alcanzado,
            "clientes_cubiertos": det.get(cat, {}).get("clientes_cubiertos", 0),
            "compradores": det.get(cat, {}).get("compradores", []),
            "no_compradores": det.get(cat, {}).get("no_compradores", []),
        })
    return jsonify(_to_native({
        "vendedor": vid, "codigo": cod, "categorias": categorias,
        "millas_alcanzadas": millas_alcanzadas, "millas_posibles": millas_posibles,
        "premios": premios, "fuente": "ventas_acumulada.csv", "periodo": cfg["periodo"],
    }))


# ====== PLAN FRIZZE (On Premise Noche · 3+1 misma variedad) ======
_PLAN_FRIZZE_CFG_CACHE = {}
_PF_IMG = {"BLUE": "/frizze_blue.jpg", "BUBBLE": "/frizze_bubble.jpg"}

def _plan_frizze_config():
    """Definición del Plan Frizze desde 01_INPUTS/PLAN FRIZZE/planfrizze.xlsx (fuente
    única, editable). Parsea clientes activos, productos (código+variedad) y mecánica del
    texto del Excel — sin hardcodear códigos ni clientes. Devuelve None si falta la fuente."""
    import re as _re, unicodedata as _ud
    path = INPUTS / "PLAN FRIZZE" / "planfrizze.xlsx"
    if not path.exists():
        return None
    try:
        key = os.path.getmtime(path)
    except OSError:
        key = 0
    cached = _PLAN_FRIZZE_CFG_CACHE.get(key)
    if cached is not None:
        return cached
    try:
        raw = pd.read_excel(path, header=None)
    except Exception as e:
        print(f"[WARN] plan_frizze config: {e}")
        return None

    def _norm(s):
        return _ud.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()

    lineas = [str(x).strip() for row in raw.values.tolist() for x in row if pd.notna(x)]
    nombre = lineas[0] if lineas else "Plan Frizze"
    clientes, productos = [], []
    for ln in lineas:
        n = _norm(ln)
        if "clientes activos" in n:
            clientes = [int(x) for x in _re.findall(r"\d+", ln.split(":", 1)[-1])]
        if "productos que entran" in n or "codigos productos" in n:
            for cod, lab in _re.findall(r"\((\d+)\)\s*([^,]+)", ln):
                lab = lab.strip()
                nl = _norm(lab)
                var = "BLUE" if "blue" in nl else ("BUBBLE" if "bubble" in nl else lab.upper())
                productos.append({"codigo": str(cod), "label": lab.title(),
                                  "variedad": var, "img": _PF_IMG.get(var, "")})
    if not clientes or not productos:
        return None
    cfg = {"nombre": nombre, "mecanica": "3+1 misma variedad", "tope_combos": 3,
           "clientes": clientes, "productos": productos, "fuente": path.name}
    _PLAN_FRIZZE_CFG_CACHE.clear()
    _PLAN_FRIZZE_CFG_CACHE[key] = cfg
    return cfg


def _plan_frizze_clientes():
    """Arma las tarjetas de cliente del Plan Frizze desde datos reales:
    ficha ← clientes.xlsx · ventas ($/litros) y sin cargos ← ventas.csv (mes vivo) ·
    litros/caja ← maestro 04D. Alerta = sin cargo de una variedad sin compra de esa
    misma variedad (el 3+1 debe ser de la misma variedad)."""
    cfg = _plan_frizze_config()
    if not cfg:
        return None
    prod_por_cod = {p["codigo"]: p for p in cfg["productos"]}
    cli_ids = cfg["clientes"]
    _, _, cod2lxu, _ = _cargar_maestro_04D()

    # Ventas vivas del plan: sólo los códigos del plan y los clientes del plan.
    v = _ventas_parsed()
    if not v.empty:
        cod_norm = v["Codigo"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
        vv = v[v["cliente_id"].isin(cli_ids) & cod_norm.isin(prod_por_cod)].copy()
        vv["_cod"] = cod_norm[vv.index]
        vv["_cant"] = pd.to_numeric(vv["CantBase"], errors="coerce").fillna(0)
    else:
        vv = pd.DataFrame()

    # Maestro de clientes (ficha). 1443 no está en el maestro → "Dato no disponible".
    cli_master = _clientes_maestro()
    minfo = {}
    if not cli_master.empty:
        for _, r in cli_master.iterrows():
            minfo[int(r["_cliente_id"])] = r

    def _sc_line(r):   # sin cargo = 100% descuento → ImporteNetoItem == 0
        return float(pd.to_numeric(r.get("importe_neto"), errors="coerce") or 0) <= 0

    clientes = []
    for cid in cli_ids:
        sub = vv[vv["cliente_id"] == cid] if not vv.empty else pd.DataFrame()
        m = minfo.get(cid)

        # Vendedor: maestro; si no está, el más frecuente en ventas del cliente.
        if m is not None:
            vend_id = m.get("_vend_id") or ""
            vend_nom = str(m.get("Vendedor", "") or "")
        elif not sub.empty:
            vc = pd.to_numeric(sub["CodVendedor"], errors="coerce").dropna()
            vend_id = f"V{int(vc.mode().iloc[0])}" if not vc.empty else ""
            vend_nom = str(sub["Vendedor"].mode().iloc[0]) if "Vendedor" in sub and not sub["Vendedor"].dropna().empty else ""
        else:
            vend_id, vend_nom = "", ""

        marcas = []
        alerta_vars = []
        sc_total = 0
        for p in cfg["productos"]:
            cod = p["codigo"]
            lxu = float(cod2lxu.get(cod, 0) or 0)
            pl = sub[sub["_cod"] == cod] if not sub.empty else pd.DataFrame()
            fact = pl[pl["importe_neto"] > 0] if not pl.empty else pd.DataFrame()
            scl = pl[[_sc_line(r) for _, r in pl.iterrows()]] if not pl.empty else pd.DataFrame()
            cajas_fact = float(fact["_cant"].sum()) if not fact.empty else 0.0
            cajas_sc = float(scl["_cant"].sum()) if not scl.empty else 0.0
            sc_total += cajas_sc
            # Envíos de sin cargo con fecha (FechaComprobante) para el desplegable.
            envios = []
            if not scl.empty:
                tmp = scl.assign(_f=scl["fecha"].dt.strftime("%Y-%m-%d"))
                for f, g in tmp.dropna(subset=["_f"]).groupby("_f"):
                    envios.append({"fecha": f, "unidades": int(g["_cant"].sum())})
                envios.sort(key=lambda e: e["fecha"])
            # Alerta: sin cargo de esta variedad sin compra de la misma variedad.
            if cajas_sc > 0 and cajas_fact <= 0:
                alerta_vars.append(p["variedad"])
            marcas.append({
                "codigo": cod, "label": p["label"], "variedad": p["variedad"], "img": p["img"],
                "litros": round(cajas_fact * lxu, 1), "importe": round(float(fact["importe_neto"].sum()) if not fact.empty else 0.0, 0),
                "unidades_facturadas": int(cajas_fact), "sin_cargo_unidades": int(cajas_sc),
                "envios": envios,
            })

        alerta = bool(alerta_vars)
        alerta_msg = ("" if not alerta else
                      "Sin cargo de " + " y ".join(v.title() for v in alerta_vars) +
                      " sin compra de esa variedad — el 3+1 debe ser de la misma variedad.")
        clientes.append({
            "cliente_id": cid,
            "nombre":     (str(m.get("Razon_Social", "")) if m is not None else
                           (str(sub["RazonSocial"].mode().iloc[0]) if not sub.empty and "RazonSocial" in sub and not sub["RazonSocial"].dropna().empty else "Dato no disponible")),
            "direccion":  (str(m.get("Direccion", "")) if m is not None else "Dato no disponible"),
            "localidad":  (str(m.get("Localidad", "")) if m is not None else "Dato no disponible"),
            "sub_canal":  (str(m.get("SubSegmento", "")) if m is not None else "Dato no disponible"),
            "vendedor_id": vend_id, "vendedor_nombre": vend_nom,
            "en_maestro": m is not None,
            "marcas": marcas,
            "sin_cargo_total_unidades": int(sc_total),
            "alerta": alerta, "alerta_msg": alerta_msg,
        })
    return {"plan": cfg, "clientes": clientes,
            "fuente": "planfrizze.xlsx + ventas.csv + clientes.xlsx + maestro 04D"}


@app.route("/api/gerencia/plan_frizze")
def gerencia_plan_frizze():
    data = _plan_frizze_clientes()
    if data is None:
        return jsonify({"error": "Plan Frizze no disponible: falta 01_INPUTS/PLAN FRIZZE/planfrizze.xlsx o no se pudo parsear."}), 404
    return jsonify(_to_native({"generado_en": _now_ar(), **data}))


@app.route("/api/vendedor/<vid>/plan_frizze")
def vendedor_plan_frizze(vid):
    """Sólo los clientes del plan que pertenecen a este vendedor."""
    data = _plan_frizze_clientes()
    if data is None:
        return jsonify({"error": "Plan Frizze no disponible"}), 404
    vid_u = str(vid).upper()
    data = {**data, "clientes": [c for c in data["clientes"] if (c.get("vendedor_id") or "").upper() == vid_u]}
    return jsonify(_to_native({"generado_en": _now_ar(), "vendedor": vid_u, **data}))


# ====== INCENTIVO DADA (cobertura Dada Tinto Verano · autoservicios) ======
_DADA_COB_MIN = 6   # Cobertura Autoservicio = mínimo 6 botellas (CLAUDE.md)

def _incentivo_dada_objetivo():
    """Objetivo desde 01_INPUTS/DADAVERANOOBJ.xlsx (texto de la regla, editable).
    Extrae objetivo de clientes y código de producto del texto; fallback a defaults."""
    import re as _re
    path = INPUTS / "DADAVERANOOBJ.xlsx"
    obj = {"objetivo_clientes": 38, "codigo": "74884", "superficie": "Autoservicios",
           "regla": "", "fuente": path.name}
    if not path.exists():
        return obj
    try:
        raw = pd.read_excel(path, header=None)
        txt = " ".join(str(x) for row in raw.values.tolist() for x in row if pd.notna(x)).strip()
        obj["regla"] = txt
        m = _re.search(r"objetivo\s+(\d+)\s+cliente", txt, _re.I)
        if m:
            obj["objetivo_clientes"] = int(m.group(1))
        m = _re.search(r"producto\s+(\d+)", txt, _re.I)
        if m:
            obj["codigo"] = m.group(1)
    except Exception as e:
        print(f"[WARN] incentivo_dada objetivo: {e}")
    return obj


def _incentivo_dada():
    """Cobertura del producto Dada Tinto Verano en autoservicios (CCC de producto).
    Fuente ventas: 01_INPUTS/dadatinto.csv (ya filtrado al código objetivo).
    Objetivo:      DADAVERANOOBJ.xlsx.
    Cliente cubierto = autoservicio + compra válida (ImporteNeto>0) + ≥6 botellas
    del código objetivo. Excluye V1/V2/V5/V20 (sin ruta con objetivo). Devuelve None
    si falta la fuente de ventas."""
    import unicodedata as _ud
    path = INPUTS / "dadatinto.csv"
    if not path.exists():
        return None
    obj = _incentivo_dada_objetivo()

    df = pd.DataFrame()
    for enc in ("latin1", "utf-8-sig", "utf-8"):
        try:
            df = pd.read_csv(path, sep=";", encoding=enc, dtype=str, low_memory=False)
            break
        except Exception:
            continue
    req = {"Cliente", "CodVendedor", "ImporteNetoItem", "Codigo", "Ramo", "Subramo", "CantBase"}
    if df.empty or not req.issubset(set(df.columns)):
        return None

    def _norm(s):
        return _ud.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()

    cod_norm = df["Codigo"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    df = df[cod_norm == str(obj["codigo"])].copy()
    df["_cli"]  = pd.to_numeric(df["Cliente"], errors="coerce")
    df["_vend"] = pd.to_numeric(df["CodVendedor"], errors="coerce")
    df["_imp"]  = df["ImporteNetoItem"].apply(_parse_num_ar)
    df["_cant"] = pd.to_numeric(df["CantBase"], errors="coerce").fillna(0)
    df["_fec"]  = pd.to_datetime(df.get("FechaComprobante"), dayfirst=True, errors="coerce")
    df["_aut"]  = (df["Subramo"].map(_norm).str.contains("autoservicio")
                   | df["Ramo"].map(_norm).str.contains("autoservicio"))

    # Universo objetivo: autoservicio, compra válida, vendedor con ruta (excl. V1/V2/V5/V20).
    val = df[(df["_aut"]) & (df["_imp"] > 0) & (~df["_vend"].isin(_VENDEDORES_EXCLUIDOS))].copy()

    articulo = str(df["Articulo"].dropna().mode().iloc[0]) if "Articulo" in df and not df["Articulo"].dropna().empty else ""

    clientes = []
    if not val.empty:
        for cid, g in val.groupby("_cli"):
            botellas = int(g["_cant"].sum())
            vc = g["_vend"].dropna()
            vend_id = f"V{int(vc.mode().iloc[0])}" if not vc.empty else ""
            vend_nom = str(g["Vendedor"].mode().iloc[0]) if "Vendedor" in g and not g["Vendedor"].dropna().empty else ""
            # Pedidos del cliente: una línea por fecha de comprobante con las botellas de ese día.
            pedidos = []
            gf = g.assign(_f=g["_fec"].dt.strftime("%Y-%m-%d"))
            for f, gg in gf.dropna(subset=["_f"]).groupby("_f"):
                pedidos.append({"fecha": f, "botellas": int(gg["_cant"].sum())})
            pedidos.sort(key=lambda e: e["fecha"])
            clientes.append({
                "cliente_id": int(cid),
                "nombre":    str(g["RazonSocial"].mode().iloc[0]) if "RazonSocial" in g and not g["RazonSocial"].dropna().empty else "Dato no disponible",
                "localidad": str(g["Localidad"].mode().iloc[0]) if "Localidad" in g and not g["Localidad"].dropna().empty else "",
                "vendedor_id": vend_id, "vendedor_nombre": vend_nom,
                "botellas": botellas,
                "cubierto": botellas >= _DADA_COB_MIN,
                "pedidos": pedidos,
            })
    clientes.sort(key=lambda c: (not c["cubierto"], c["vendedor_id"], -c["botellas"]))

    logrado = sum(1 for c in clientes if c["cubierto"])
    objetivo = int(obj["objetivo_clientes"]) or 0
    avance = round(logrado / objetivo * 100, 1) if objetivo else 0.0

    por_vend = {}
    for c in clientes:
        if not c["cubierto"]:
            continue
        k = c["vendedor_id"] or "—"
        d = por_vend.setdefault(k, {"vendedor_id": c["vendedor_id"], "vendedor_nombre": c["vendedor_nombre"], "clientes": 0, "botellas": 0})
        d["clientes"] += 1
        d["botellas"] += c["botellas"]
    por_vendedor = sorted(por_vend.values(), key=lambda d: -d["clientes"])

    return {
        "producto": {"codigo": obj["codigo"], "articulo": articulo, "imagen": "/dadatinto.png"},
        "objetivo": objetivo, "superficie": obj["superficie"],
        "logrado": logrado, "faltan": max(objetivo - logrado, 0), "avance_pct": avance,
        "min_botellas": _DADA_COB_MIN,
        "por_vendedor": por_vendedor, "clientes": clientes,
        "regla": obj["regla"],
        "fuente": "dadatinto.csv + DADAVERANOOBJ.xlsx",
    }


@app.route("/api/gerencia/incentivo_dada")
def gerencia_incentivo_dada():
    data = _incentivo_dada()
    if data is None:
        return jsonify({"error": "Incentivo Dada no disponible: falta 01_INPUTS/dadatinto.csv o columnas requeridas."}), 404
    return jsonify(_to_native({"generado_en": _now_ar(), **data}))


# ====== INCENTIVO ALMA MORA MALBEC LOW (cobertura 74887 · autoservicios) ======
# Mecánica (definida por el usuario 2026-08-03): cliente con compra del código 74887,
# medido SOLO sobre autoservicios. El objetivo es el 22% de la cartera de autoservicios
# (no hay planilla de objetivo: se calcula sobre clientes.xlsx, que es la cartera viva).
_ALMAMORA_CODIGO   = "74887"
_ALMAMORA_PCT_OBJ  = 0.22
_ALMAMORA_COB_MIN  = 6   # Cobertura Autoservicio = mínimo 6 botellas (CLAUDE.md)
_ALMAMORA_IMG      = "/almamora_low.png"   # KV exportado de 01_INPUTS/incentivo/*.pdf
_ALMAMORA_CACHE = {}


def _cartera_autoservicios():
    """(cantidad, dict cliente_id -> datos) de la cartera Autoservicio del maestro.
    Clasifica con _clasificar_segmento (EL SUBSEGMENTO MANDA SOBRE EL RAMO) — mismo
    criterio y mismo denominador que la tarjeta de cobertura del dashboard."""
    cli = _clientes_maestro()
    if cli.empty:
        return 0, {}
    ramo_col = next((c for c in cli.columns if c.lower() == "ramo"), None)
    sub_col  = next((c for c in cli.columns if "subramo" in c.lower() or "subseg" in c.lower()), None)
    if not ramo_col:
        return 0, {}
    rm = cli[ramo_col].fillna("").astype(str)
    sb = cli[sub_col].fillna("").astype(str) if sub_col else pd.Series([""] * len(cli), index=cli.index)
    seg_map = {par: _clasificar_segmento(par[0], par[1]) for par in set(zip(rm, sb))}
    seg = pd.Series([seg_map[(a, b)] for a, b in zip(rm, sb)], index=cli.index)
    aut = cli[seg == "AUTOSERVICIO"]
    detalle = {}
    for _, r in aut.iterrows():
        detalle[int(r["_cliente_id"])] = {
            "nombre":     str(r.get("Razon_Social", "") or "").strip() or "Dato no disponible",
            "localidad":  str(r.get("Localidad", "") or "").strip(),
            "segmento":   (str(r.get(sub_col, "") or "").strip() if sub_col else "") or str(r.get(ramo_col, "") or "").strip(),
            "vendedor_id":     str(r.get("_vend_id", "") or ""),
            "vendedor_nombre": str(r.get("Vendedor", "") or "").strip(),
        }
    return len(aut), detalle


def _incentivo_almamora():
    """Cobertura del código 74887 (Alma Mora Malbec Dulce Low) en autoservicios.

    Fuente ventas: 01_INPUTS/ventas_acumulada.csv (acumulado vivo, sin filtro de fecha —
    mismo criterio que Incentivo FARO y 11T). Cliente logrado = autoservicio de la cartera
    con compra NETA > 0 del código (los rechazos restan) Y al menos 6 botellas: cobertura
    de Autoservicio son 6 botellas (CLAUDE.md), igual que el Incentivo Dada. Los que
    compraron pero no llegan a 6 salen aparte como `parciales` (es la lista accionable).
    Excluye V1/V2/V5/V20. Objetivo = 22% de la cartera de autoservicios.
    Devuelve None si falta la fuente."""
    path = INPUTS / "ventas_acumulada.csv"
    if not path.exists():
        return None
    try:
        key = (os.path.getmtime(path), os.path.getmtime(INPUTS / "clientes.xlsx") if (INPUTS / "clientes.xlsx").exists() else 0)
    except OSError:
        key = (0, 0)
    cached = _ALMAMORA_CACHE.get(key)
    if cached is not None:
        return cached

    df = pd.DataFrame()
    for enc in ("latin1", "utf-8-sig", "utf-8"):
        try:
            df = pd.read_csv(path, sep=";", encoding=enc, dtype=str, low_memory=False)
            break
        except Exception:
            continue
    req = {"Cliente", "CodVendedor", "ImporteNetoItem", "Codigo", "CantBase", "FechaComprobante"}
    if df.empty or not req.issubset(set(df.columns)):
        return None

    cod = df["Codigo"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    df = df[cod == _ALMAMORA_CODIGO].copy()
    df["_cli"]  = pd.to_numeric(df["Cliente"], errors="coerce")
    df["_vend"] = pd.to_numeric(df["CodVendedor"], errors="coerce")
    df["_imp"]  = df["ImporteNetoItem"].apply(_parse_num_ar)
    df["_cant"] = pd.to_numeric(df["CantBase"], errors="coerce").fillna(0)
    df["_fec"]  = pd.to_datetime(df["FechaComprobante"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["_cli"])
    df = df[~df["_vend"].isin(_VENDEDORES_EXCLUIDOS)].copy()

    articulo = ""
    if "Articulo" in df.columns and not df["Articulo"].dropna().empty:
        articulo = str(df["Articulo"].dropna().mode().iloc[0])

    cartera_as, cartera_det = _cartera_autoservicios()

    logrados, parciales, fuera_as = [], [], 0
    for cid, g in df.groupby("_cli"):
        cid = int(cid)
        neto = float(g["_imp"].sum())
        if neto <= 0:                      # sin cargo / rechazado: no es CCC
            continue
        m = cartera_det.get(cid)
        if m is None:                      # compró el producto pero no es autoservicio
            fuera_as += 1
            continue
        pos = g[g["_imp"] > 0]
        fechas = sorted({f.strftime("%Y-%m-%d") for f in pos["_fec"].dropna()})
        vc = pos["_vend"].dropna()
        # La venta le cuenta a quien la facturó; si no hay, al dueño de la cartera.
        vend_id = f"V{int(vc.mode().iloc[0])}" if not vc.empty else m["vendedor_id"]
        vend_nom = m["vendedor_nombre"]
        if "Vendedor" in pos.columns and not pos["Vendedor"].dropna().empty:
            vend_nom = str(pos["Vendedor"].mode().iloc[0]).strip()
        botellas = int(pos["_cant"].sum())
        reg = {
            "cliente_id": cid,
            "nombre":     m["nombre"],
            "localidad":  m["localidad"],
            "segmento":   m["segmento"] or "Autoservicio",
            "vendedor_id": vend_id, "vendedor_nombre": vend_nom,
            "botellas":   botellas,
            "importe":    round(neto, 2),
            "fecha":      fechas[0] if fechas else "",
            "fechas":     fechas,
        }
        # Cobertura AS = 6 botellas. El que compró menos NO cuenta, pero se muestra:
        # es el cliente más barato de convertir (le faltan pocas botellas).
        if botellas >= _ALMAMORA_COB_MIN:
            logrados.append(reg)
        else:
            reg["faltan_botellas"] = _ALMAMORA_COB_MIN - botellas
            parciales.append(reg)
    parciales.sort(key=lambda c: (c["faltan_botellas"], c["vendedor_id"], c["nombre"]))

    por_vend = {}
    for c in logrados:
        k = c["vendedor_id"] or "—"
        d = por_vend.setdefault(k, {"vendedor_id": c["vendedor_id"], "vendedor_nombre": c["vendedor_nombre"],
                                    "clientes": 0, "botellas": 0, "detalle": []})
        d["clientes"] += 1
        d["botellas"] += c["botellas"]
        d["detalle"].append(c)
    for d in por_vend.values():
        d["detalle"].sort(key=lambda c: (c["fecha"], c["nombre"]))
    por_vendedor = sorted(por_vend.values(), key=lambda d: (-d["clientes"], d["vendedor_id"]))

    objetivo = int(cartera_as * _ALMAMORA_PCT_OBJ + 0.5)
    logrado  = len(logrados)
    avance   = round(logrado / objetivo * 100, 1) if objetivo else 0.0
    cob_pct  = round(logrado / cartera_as * 100, 1) if cartera_as else 0.0

    fechas_src = df["_fec"].dropna()
    periodo = {"desde": fechas_src.min().strftime("%Y-%m-%d") if not fechas_src.empty else "",
               "hasta": fechas_src.max().strftime("%Y-%m-%d") if not fechas_src.empty else ""}

    data = {
        "producto": {"codigo": _ALMAMORA_CODIGO,
                     "articulo": articulo or "ALMA MORA MALBEC DULCE LOW 6X750",
                     "imagen": _ALMAMORA_IMG},
        "cartera_as": cartera_as,
        "pct_objetivo": round(_ALMAMORA_PCT_OBJ * 100, 1),
        "objetivo": objetivo,
        "logrado": logrado,
        "faltan": max(objetivo - logrado, 0),
        "avance_pct": avance,
        "cobertura_pct": cob_pct,
        "min_botellas": _ALMAMORA_COB_MIN,
        "fuera_as": fuera_as,
        "parciales": parciales,
        "por_vendedor": por_vendedor,
        "clientes": sorted(logrados, key=lambda c: (c["vendedor_id"], c["nombre"])),
        "periodo": periodo,
        "fuente": "ventas_acumulada.csv + clientes.xlsx",
    }
    _ALMAMORA_CACHE.clear()
    _ALMAMORA_CACHE[key] = data
    return data


@app.route("/api/gerencia/incentivo_almamora")
def gerencia_incentivo_almamora():
    data = _incentivo_almamora()
    if data is None:
        return jsonify({"error": "Incentivo Alma Mora no disponible: falta 01_INPUTS/ventas_acumulada.csv o columnas requeridas."}), 404
    return jsonify(_to_native({"generado_en": _now_ar(), **data}))


# ====== PLAN COBERTURA (On Premise B&C · restaurantes y bares con carta) ======
# Plan de Grupo Peñaflor: incrementar cobertura en clientes categoría B y C del canal
# On Premise (restaurantes / bares con carta de bebida). Se mide por CCC únicos del
# canal, de JULIO a DICIEMBRE 2026. Fuente del padrón relevado: el xlsx de la carpeta
# 01_INPUTS/Plan cobertura (el PDF del mismo directorio es el resumen de la mecánica).
PLAN_COB_DIR = INPUTS / "Plan cobertura"
# Meta del plan a diciembre 2026: 60 clientes nuevos dados de alta (los del padrón
# relevado + las altas que aparecieron fuera de ese listado, contadas por cliente).
PLAN_COB_OBJETIVO_ALTAS = 60
_PLAN_COB_CACHE = {}
_PLAN_COB_VENTAS_CACHE = {}


def _plan_cob_norm(s):
    """Texto comparable: sin acentos, mayúsculas, espacios colapsados."""
    import unicodedata as _ud
    t = _ud.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper()
    return " ".join(t.split())


def _plan_cob_padron_path():
    """Primer .xlsx de la carpeta del plan (ignora los temporales ~$ de Excel)."""
    if not PLAN_COB_DIR.exists():
        return None
    xs = [p for p in sorted(PLAN_COB_DIR.glob("*.xlsx")) if not p.name.startswith("~$")]
    return xs[0] if xs else None


def _plan_cob_padron():
    """Padrón relevado del plan. El xlsx trae dos filas de encabezado (grupos de columnas):
    la buena es la segunda. Devuelve (DataFrame normalizado, nombre del archivo).

    NO se usa pd.read_excel: el export viene inflado (la hoja declara ~1.048.000 filas
    para 205 reales) y abrirlo así tarda ~15 s, que es toda la pantalla. Con openpyxl en
    read_only y corte por filas vacías tarda milisegundos. Mismo patrón que
    `producto activos.xlsx` (ver CHANGELOG 2026-06)."""
    path = _plan_cob_padron_path()
    if path is None:
        return pd.DataFrame(), ""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        filas, vacias = [], 0
        for r in ws.iter_rows(values_only=True):
            if all(v is None or str(v).strip() == "" for v in r):
                vacias += 1
                if filas and vacias >= 25:
                    break
                continue
            vacias = 0
            filas.append(r)
        wb.close()
    except Exception as e:
        print(f"[WARN] plan_cobertura padrón: {e}")
        return pd.DataFrame(), path.name
    if len(filas) < 3:
        return pd.DataFrame(), path.name
    hdr = []
    for i, x in enumerate(filas[1]):
        n = str(x).strip() if x is not None else ""
        hdr.append(n if n and n.lower() != "none" else f"_c{i}")
    df = pd.DataFrame(filas[2:], columns=hdr)
    df.columns = [str(c).strip() for c in df.columns]

    def col(*claves):
        for c in df.columns:
            cn = _plan_cob_norm(c)
            if any(k in cn for k in claves):
                return c
        return None

    c_id   = col("ID PUNTO DE VENTA", "ID PDV")
    c_nom  = col("NOMBRE LOCAL")
    c_dir  = col("DIRECCION")
    c_part = col("PARTIDO")
    c_loc  = col("LOCALIDAD")
    c_seg  = col("SEGMENTO")
    c_tipo = col("TIPO")
    c_at   = col("LO ATIENDE")
    c_cod  = col("COD. CLIENTE", "CODIGO CLIENTE", "COD CLIENTE")
    c_obs  = col("OBSERVACIONES")
    if c_nom is None or c_loc is None:
        return pd.DataFrame(), path.name

    out = pd.DataFrame({
        "pdv_id":     df[c_id].astype(str).str.strip() if c_id else "",
        "nombre":     df[c_nom].astype(str).str.strip(),
        "direccion":  df[c_dir].fillna("").astype(str).str.strip() if c_dir else "",
        "partido":    df[c_part].fillna("").astype(str).str.strip() if c_part else "",
        "localidad":  df[c_loc].fillna("").astype(str).str.strip(),
        "segmento":   df[c_seg].fillna("").astype(str).str.strip().str.title() if c_seg else "",
        "tipo":       df[c_tipo].fillna("").astype(str).str.strip() if c_tipo else "",
        "observaciones": df[c_obs].fillna("").astype(str).str.strip() if c_obs else "",
    })
    out["cliente_id"] = pd.to_numeric(df[c_cod], errors="coerce") if c_cod else pd.NA
    # "Sí / si / SI" y "No / no" vienen sin criterio de mayúsculas ni acento en el relevamiento.
    at = df[c_at].fillna("").map(_plan_cob_norm) if c_at else pd.Series([""] * len(df))
    out["atiende_raw"] = df[c_at].fillna("").astype(str).str.strip() if c_at else ""

    def _estado(v):
        if v == "SI":
            return "ATENDIDO"
        if v.startswith("POTENCIAL"):
            return "POTENCIAL"
        if v == "NO":
            return "NO_ATENDIDO"
        return "SIN_RELEVAR"
    out["estado"] = at.map(_estado)
    out["_loc"]  = out["localidad"].map(_plan_cob_norm)
    out["_part"] = out["partido"].map(_plan_cob_norm)
    return out, path.name


def _plan_cob_altas_fuera():
    """Altas del plan que NO figuraban en el padrón relevado.

    El negocio las carga a mano en una hoja aparte del MISMO xlsx (la que tiene "fuera"
    en el nombre): son PDV que se dieron de alta durante el plan y que el relevamiento
    original no tenía. Columnas: código de cliente, nombre y número de vendedor.
    Devuelve DataFrame con cliente_id / nombre / vend_hoja (vacío si no está la hoja).

    Mismo cuidado que el padrón: openpyxl read_only, nunca pd.read_excel (el export
    viene inflado, ver `_plan_cob_padron`)."""
    path = _plan_cob_padron_path()
    if path is None:
        return pd.DataFrame()
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        hoja = next((n for n in wb.sheetnames if "FUERA" in _plan_cob_norm(n)), None)
        if hoja is None:
            wb.close()
            return pd.DataFrame()
        filas, vacias = [], 0
        for r in wb[hoja].iter_rows(values_only=True):
            if all(v is None or str(v).strip() == "" for v in r):
                vacias += 1
                if filas and vacias >= 25:
                    break
                continue
            vacias = 0
            filas.append(r)
        wb.close()
    except Exception as e:
        print(f"[WARN] plan_cobertura altas fuera del listado: {e}")
        return pd.DataFrame()
    if len(filas) < 2:
        return pd.DataFrame()

    # El encabezado no arranca en la columna A (la hoja tiene una columna vacía al
    # principio): se ubican las columnas por su título, no por posición.
    hdr = filas[0]
    i_cod = i_nom = i_ven = None
    for i, x in enumerate(hdr):
        n = _plan_cob_norm(x)
        if i_cod is None and n.startswith("COD"):
            i_cod = i
        elif i_nom is None and n.startswith("NOMBRE"):
            i_nom = i
        elif i_ven is None and n.startswith("VENDEDOR"):
            i_ven = i
    if i_cod is None:
        return pd.DataFrame()

    out = []
    for r in filas[1:]:
        cod = pd.to_numeric(r[i_cod] if i_cod < len(r) else None, errors="coerce")
        if pd.isna(cod):
            continue
        ven = r[i_ven] if (i_ven is not None and i_ven < len(r)) else None
        ven = pd.to_numeric(ven, errors="coerce")
        out.append({
            "cliente_id": int(cod),
            "nombre": str(r[i_nom]).strip() if (i_nom is not None and i_nom < len(r)
                                                and r[i_nom] is not None) else "",
            "vend_hoja": f"V{int(ven)}" if pd.notna(ven) else "",
        })
    return pd.DataFrame(out)


def _plan_cob_clave(row):
    """Clave estable de un PDV del padrón, para colgarle el mensaje de seguimiento.
    El 'ID PUNTO DE VENTA' del relevamiento es único (205 de 205 filas), así que el
    mensaje sobrevive a que se reordene o se recargue el xlsx. Si algún día viniera
    vacío, cae a nombre + localidad normalizados."""
    pid = str(row.get("pdv_id", "") or "").strip()
    if pid and pid.lower() not in ("nan", "none"):
        return "PDV:" + pid
    return "NOM:" + _plan_cob_norm(row.get("nombre")) + "|" + _plan_cob_norm(row.get("localidad"))


def _plan_cob_vendedor_por_zona(padron):
    """Vendedor sugerido para un PDV que NO está dado de alta: el que más clientes tiene
    en esa localidad según el maestro. Si la localidad no tiene ni un cliente nuestro,
    cae al vendedor dominante del PARTIDO (mismo departamento, según el propio padrón).
    Devuelve (dict localidad -> asignación, dict partido -> asignación).

    V3 queda FUERA del cálculo: no trabaja On Premise (sólo tradicional almacén /
    despensa / kiosco), así que un bar del plan no puede ser suyo. Sin esto, 28 PDV
    caían en la cartera de V3 y no le llegaban a ningún vendedor."""
    cli = _clientes_maestro()
    por_loc = {}
    if not cli.empty and "Localidad" in cli.columns:
        cli = cli[cli["_vend_id"].astype(str).str.upper() != "V3"].copy()
        cli["_loc"] = cli["Localidad"].fillna("").map(_plan_cob_norm)
        nombres = {}
        for _, r in cli.iterrows():
            vid = str(r.get("_vend_id", "") or "")
            if vid and vid not in nombres:
                nombres[vid] = str(r.get("Vendedor", "") or "").strip()
        for loc, g in cli.groupby("_loc"):
            if not loc:
                continue
            vc = g["_vend_id"].replace("", pd.NA).dropna().value_counts()
            if vc.empty:
                continue
            vid = vc.index[0]
            por_loc[loc] = {"vendedor_id": vid, "vendedor_nombre": nombres.get(vid, ""),
                            "clientes_zona": int(vc.iloc[0]), "clientes_total": int(vc.sum()),
                            "fuente": "localidad"}
    # Partido: se arma con las localidades del padrón que sí tienen dueño.
    por_part = {}
    if not padron.empty:
        for part, g in padron.groupby("_part"):
            if not part:
                continue
            conteo = {}
            for loc in g["_loc"].unique():
                a = por_loc.get(loc)
                if a:
                    conteo[a["vendedor_id"]] = conteo.get(a["vendedor_id"], 0) + a["clientes_zona"]
            if conteo:
                vid = max(conteo, key=conteo.get)
                por_part[part] = {"vendedor_id": vid,
                                  "vendedor_nombre": por_loc_nombre(por_loc, vid),
                                  "clientes_zona": conteo[vid], "clientes_total": sum(conteo.values()),
                                  "fuente": "partido"}
    return por_loc, por_part


def por_loc_nombre(por_loc, vid):
    for a in por_loc.values():
        if a["vendedor_id"] == vid and a["vendedor_nombre"]:
            return a["vendedor_nombre"]
    return ""


def _plan_cob_ventas(ids):
    """Historial de compras COMPLETO de los clientes del plan, encadenando las cuatro
    fuentes que cubren la línea de tiempo sin huecos:
      02_HISTORY/historial_ventas.csv          2024-03 → 2026-05 (archivo histórico congelado)
      02_HISTORY/historial_ventas_cliente.csv  2026-05 → último cierre
      01_INPUTS/ventas_acumulada.csv           trimestre vivo
      01_INPUTS/ventas.csv                     día vivo
    Se filtra a los clientes del plan apenas se lee (son un puñado) y se deduplica por
    fecha+cliente+artículo+cantidad+importe, que es donde se solapan las fuentes.
    Sólo líneas con importe > 0: una devolución o un sin cargo no activan un cliente."""
    ids = {int(i) for i in ids}
    if not ids:
        return pd.DataFrame()
    paths = [BASE / "02_HISTORY" / "historial_ventas.csv",
             BASE / "02_HISTORY" / "historial_ventas_cliente.csv",
             INPUTS / "ventas_acumulada.csv", INPUTS / "ventas.csv"]
    key = (tuple(sorted(ids)),
           tuple((p.name, os.path.getmtime(p) if p.exists() else 0) for p in paths))
    cached = _PLAN_COB_VENTAS_CACHE.get(key)
    if cached is not None:
        return cached

    frames = []
    for p in paths:
        if not p.exists():
            continue
        norm_cols = (p.name == "historial_ventas_cliente.csv")
        # utf-8-sig PRIMERO: historial_ventas_cliente.csv se escribe con BOM y leído como
        # latin1 la primera columna queda "ï»¿fecha_comprobante" → el chequeo de columnas
        # fallaba y la fuente se descartaba en silencio (clientes activos con 0 compras).
        encs = ("utf-8-sig", "latin1", "utf-8") if norm_cols else ("latin1", "utf-8-sig", "utf-8")
        # Sólo las columnas que se usan: historial_ventas.csv son 129k filas x 59 columnas
        # y leerlo entero se llevaba casi todo el tiempo de la pantalla.
        cols_erp = ["Cliente", "FechaComprobante", "Codigo", "Articulo", "Marca",
                    "CantBase", "ImporteNetoItem", "CodVendedor", "Descuento", "valorDescuento"]
        df = None
        for enc in encs:
            try:
                df = pd.read_csv(p, sep=("," if norm_cols else ";"), encoding=enc,
                                 dtype=str, low_memory=False, on_bad_lines="skip",
                                 usecols=(None if norm_cols else
                                          (lambda c: str(c).strip() in cols_erp)))
                break
            except Exception:
                continue
        if df is None or df.empty:
            continue
        df.columns = [str(c).strip().lstrip("﻿") for c in df.columns]
        if norm_cols:
            ren = {"fecha_comprobante": "_f", "cliente_id": "_cli", "articulo": "_art",
                   "marca": "_marca", "cant_base": "_cant", "importe_neto": "_imp",
                   "vendedor_codigo": "_vend", "descuento_pct": "_pct"}
            if not set(ren).issubset(df.columns):
                continue
            d = df.rename(columns=ren)[list(ren.values())].copy()
            d["_cod"] = ""
            d["_f"] = pd.to_datetime(d["_f"], errors="coerce")
            for c in ("_cant", "_imp", "_pct"):
                d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0)
        else:
            req = {"Cliente", "FechaComprobante", "Articulo", "CantBase", "ImporteNetoItem"}
            if not req.issubset(set(df.columns)):
                continue
            d = pd.DataFrame({
                "_f":    pd.to_datetime(df["FechaComprobante"], dayfirst=True, errors="coerce"),
                "_cli":  df["Cliente"],
                "_art":  df["Articulo"].fillna("").astype(str).str.strip(),
                "_marca": df["Marca"].fillna("").astype(str).str.strip() if "Marca" in df.columns else "",
                "_cod":  (df["Codigo"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
                          if "Codigo" in df.columns else ""),
                "_cant": df["CantBase"].apply(_parse_num_ar),
                "_imp":  df["ImporteNetoItem"].apply(_parse_num_ar),
                "_vend": df["CodVendedor"] if "CodVendedor" in df.columns else None,
                # "Descuento" viene como texto tipo "20,00%"
                "_pct":  (df["Descuento"].astype(str).str.replace("%", "", regex=False).apply(_parse_num_ar)
                          if "Descuento" in df.columns else 0.0),
            })
        d["_cli"] = pd.to_numeric(d["_cli"], errors="coerce")
        d = d[d["_cli"].isin(ids)]
        if d.empty:
            continue
        d["_vend"] = pd.to_numeric(d.get("_vend"), errors="coerce")
        frames.append(d.dropna(subset=["_f"]))

    if not frames:
        out = pd.DataFrame()
    else:
        out = pd.concat(frames, ignore_index=True)
        # Se conservan las líneas de importe 0: son los sin cargo de los combos del plan
        # (1+2, 5+1, 2+1), que el negocio necesita ver. Los rechazos (cantidad negativa)
        # sí se descartan. La activación y las recompras se miden sólo con importe > 0.
        out = out[(out["_imp"] > 0) | ((out["_imp"] <= 0) & (out["_cant"] > 0))].copy()
        out["_cli"] = out["_cli"].astype(int)
        out["_artn"] = out["_art"].map(_plan_cob_norm)
        out["_fd"] = out["_f"].dt.strftime("%Y-%m-%d")
        out = out.drop_duplicates(subset=["_fd", "_cli", "_artn", "_cant", "_imp"])
        out["_mes"] = out["_f"].dt.strftime("%Y-%m")
    _PLAN_COB_VENTAS_CACHE.clear()
    _PLAN_COB_VENTAS_CACHE[key] = out
    return out


def _plan_cob_acciones_por_cliente():
    """cliente_id -> acciones comerciales del catálogo VIGENTE que el cliente efectivamente
    usó. Se invierte el payload de Acciones Comerciales en vez de re-detectar acá: esa
    pantalla ya aplica 'acción = uso' (el % aplicado tiene que ser el tramo de la acción)
    y no queremos dos criterios distintos conviviendo."""
    try:
        pay = _acciones_mes_payload(None) or {}
    except Exception as e:
        print(f"[WARN] plan_cobertura acciones: {e}")
        return {}, ""
    mapa = {}
    for a in pay.get("acciones", []):
        etiqueta = " · ".join(x for x in [str(a.get("tipo", "")).strip(),
                                          str(a.get("marcas", "")).strip()] if x)
        for c in a.get("clientes_detalle", []):
            cid = c.get("cliente_id")
            if cid is None:
                continue
            mapa.setdefault(int(cid), []).append({
                "id_accion": a.get("id_accion", ""),
                "accion": etiqueta or a.get("id_accion", ""),
                "escala": a.get("escala", ""),
                "descuento_pct": a.get("descuento_pct", ""),
                "litros": c.get("litros", 0),
                "cant_base": c.get("cant_base", 0),
                "fecha_ultima": c.get("fecha_ultima", ""),
            })
    return mapa, pay.get("mes", "")


def _plan_cobertura():
    """Seguimiento del Plan Cobertura sobre el padrón relevado.

    Tres grupos, que son los que pidió el negocio:
      capturados  → PDV con código de cliente cargado (dados de alta). Se les mide
                    activación (fecha de la PRIMERA compra), recompras por mes,
                    artículos comprados y acciones comerciales usadas.
      potenciales → relevados como 'Potencial': todavía no son clientes.
      no_atendidos→ relevados como 'No' (más los que quedaron sin relevar).
    A los dos últimos, que no tienen código, se les sugiere vendedor por zona.
    Devuelve None si falta el padrón."""
    # La clave del caché se arma ANTES de parsear nada: si no, cada request volvía a
    # abrir el xlsx y el padrón es lo más caro de la pantalla.
    padron_p = _plan_cob_padron_path()
    if padron_p is None:
        return None
    fuentes = [padron_p, INPUTS / "clientes.xlsx", INPUTS / "ventas.csv",
               INPUTS / "ventas_acumulada.csv",
               BASE / "02_HISTORY" / "historial_ventas_cliente.csv"]
    key = tuple((p.name, os.path.getmtime(p) if p.exists() else 0) for p in fuentes)
    cached = _PLAN_COB_CACHE.get(key)
    if cached is not None:
        return cached

    padron, fuente = _plan_cob_padron()
    if padron.empty:
        return None

    por_loc, por_part = _plan_cob_vendedor_por_zona(padron)
    maestro = _clientes_maestro()
    m_idx = {}
    if not maestro.empty:
        for _, r in maestro.iterrows():
            m_idx[int(r["_cliente_id"])] = r

    def _sugerido(row):
        a = por_loc.get(row["_loc"]) or por_part.get(row["_part"])
        if not a:
            return {"vendedor_id": "", "vendedor_nombre": "", "asignacion": "sin_zona",
                    "asignacion_detalle": "No tenemos clientes en esa zona"}
        det = (f'{a["clientes_zona"]} de {a["clientes_total"]} clientes de {row["localidad"]}'
               if a["fuente"] == "localidad"
               else f'Sin clientes en {row["localidad"]}; dominante en el partido {row["partido"]}')
        return {"vendedor_id": a["vendedor_id"], "vendedor_nombre": a["vendedor_nombre"],
                "asignacion": a["fuente"], "asignacion_detalle": det}

    def _base(row):
        return {
            "clave": _plan_cob_clave(row),
            "pdv_id": row["pdv_id"], "nombre": row["nombre"], "direccion": row["direccion"],
            "localidad": row["localidad"], "partido": row["partido"],
            "segmento_plan": row["segmento"], "tipo": row["tipo"],
            "observaciones": row["observaciones"],
        }

    con_codigo = padron[padron["cliente_id"].notna()].copy()
    fuera_df = _plan_cob_altas_fuera()
    ids = {int(x) for x in con_codigo["cliente_id"].tolist()}
    if not fuera_df.empty:
        ids |= {int(x) for x in fuera_df["cliente_id"].tolist()}
    ventas = _plan_cob_ventas(ids)
    acc_map, acc_mes = _plan_cob_acciones_por_cliente()

    def _lineas_sc(df_sc):
        if df_sc is None or df_sc.empty:
            return []
        out = [{"fecha": r["_f"].strftime("%Y-%m-%d"), "articulo": r["_art"],
                "botellas": int(r["_cant"])} for _, r in df_sc.iterrows()]
        out.sort(key=lambda x: x["fecha"], reverse=True)
        return out

    def _lineas_desc(df_c):
        """Compras con descuento aplicado, agrupadas por artículo y %. Es el dato duro
        de la factura: qué se compró y con qué descuento. No se le pone nombre de
        acción salvo que el catálogo vigente la reconozca (campo `acciones`)."""
        if df_c is None or df_c.empty or "_pct" not in df_c.columns:
            return []
        dd = df_c[df_c["_pct"] > 0]
        if dd.empty:
            return []
        filas = []
        for (art, pct), ga in dd.groupby(["_art", "_pct"]):
            filas.append({"articulo": art, "pct": round(float(pct), 1),
                          "botellas": int(ga["_cant"].sum()),
                          "ultima": ga["_f"].max().strftime("%Y-%m-%d")})
        filas.sort(key=lambda x: (-x["pct"], -x["botellas"]))
        return filas

    def _ficha(d, cid):
        """Completa la ficha de un cliente dado de alta: datos del maestro + qué compró
        (activación = primera compra, recompras mes a mes, artículos, descuentos y sin
        cargos). Es la misma medición para los capturados del padrón y para las altas
        fuera del listado, por eso está acá y no dentro del loop."""
        m = m_idx.get(cid)
        todo = ventas[ventas["_cli"] == cid] if not ventas.empty else pd.DataFrame()
        # Compras = importe > 0 (es lo que activa y lo que cuenta como recompra).
        # Sin cargo = lo que entregamos por los combos del plan, se muestra aparte.
        g  = todo[todo["_imp"] > 0] if not todo.empty else todo
        sc = todo[todo["_imp"] <= 0] if not todo.empty else todo
        d["cliente_id"] = cid
        d["en_maestro"] = m is not None
        d["razon_social"] = str(m.get("Razon_Social", "")).strip() if m is not None else ""
        d["subcanal"] = (str(m.get("SubSegmento", "") or m.get("Ramo", "")).strip()
                         if m is not None else "")
        d["sin_cargos"] = _lineas_sc(sc)
        d["descuentos"] = _lineas_desc(g)
        d["acciones"] = acc_map.get(cid, [])

        if g.empty:
            d.update({"activacion": "", "ultima_compra": "", "meses_con_compra": 0,
                      "recompras": 0, "recompro": False, "botellas": 0, "importe": 0.0,
                      "estado_compra": "sin_compras", "meses": [], "articulos": []})
            return d, m

        act = g["_f"].min()
        ult = g["_f"].max()
        mes_act = act.strftime("%Y-%m")
        meses = []
        for mes, gm in g.groupby("_mes"):
            meses.append({"mes": mes, "botellas": int(gm["_cant"].sum()),
                          "importe": round(float(gm["_imp"].sum()), 0),
                          "lineas": int(len(gm))})
        meses.sort(key=lambda x: x["mes"])
        recompras = sum(1 for x in meses if x["mes"] > mes_act)

        articulos = []
        for art, ga in g.groupby("_art"):
            articulos.append({
                "articulo": art,
                "marca": str(ga["_marca"].replace("", pd.NA).dropna().mode().iloc[0])
                         if not ga["_marca"].replace("", pd.NA).dropna().empty else "",
                "botellas": int(ga["_cant"].sum()),
                "importe": round(float(ga["_imp"].sum()), 0),
                "primera": ga["_f"].min().strftime("%Y-%m-%d"),
                "ultima":  ga["_f"].max().strftime("%Y-%m-%d"),
                "veces": int(ga["_f"].dt.strftime("%Y-%m-%d").nunique()),
            })
        articulos.sort(key=lambda a: -a["botellas"])

        d.update({
            "activacion": act.strftime("%Y-%m-%d"),
            "ultima_compra": ult.strftime("%Y-%m-%d"),
            "meses_con_compra": len(meses),
            "recompras": recompras,
            "recompro": recompras > 0,
            "botellas": int(g["_cant"].sum()),
            "importe": round(float(g["_imp"].sum()), 0),
            "estado_compra": "con_recompra" if recompras > 0 else "solo_activacion",
            "meses": meses,
            "articulos": articulos,
        })
        return d, m

    capturados = []
    for _, row in con_codigo.iterrows():
        d, m = _ficha(_base(row), int(row["cliente_id"]))
        if m is not None:
            d["vendedor_id"] = str(m.get("_vend_id", "") or "")
            d["vendedor_nombre"] = str(m.get("Vendedor", "") or "").strip()
            d["asignacion"] = "maestro"
            d["asignacion_detalle"] = "Cartera asignada en el maestro"
        else:
            d.update(_sugerido(row))
            d["asignacion_detalle"] = "El código no figura en el maestro · " + d["asignacion_detalle"]
        capturados.append(d)
    capturados.sort(key=lambda c: (c["estado_compra"] == "sin_compras", -c["recompras"], c["nombre"]))

    # Altas FUERA del listado: PDV que se dieron de alta durante el plan y que el
    # relevamiento original no tenía. Vienen de la hoja aparte del padrón, sin ID de
    # punto de venta ni localidad: lo que falta se completa con el maestro. El vendedor
    # sale del maestro (cartera real) y, si el cliente todavía no está ahí, del número
    # que cargaron en la hoja.
    ids_padron = {c["cliente_id"] for c in capturados}
    altas_fuera = []
    for _, row in (fuera_df.iterrows() if not fuera_df.empty else []):
        cid = int(row["cliente_id"])
        d, m = _ficha({
            "clave": f"CLI:{cid}", "pdv_id": "", "nombre": str(row["nombre"] or "").strip(),
            "direccion": "", "localidad": "", "partido": "",
            "segmento_plan": "", "tipo": "", "observaciones": "",
        }, cid)
        if m is not None:
            d["nombre"] = d["nombre"] or d["razon_social"]
            d["direccion"] = str(m.get("Direccion", "") or "").strip()
            d["localidad"] = str(m.get("Localidad", "") or "").strip()
            d["vendedor_id"] = str(m.get("_vend_id", "") or "")
            d["vendedor_nombre"] = str(m.get("Vendedor", "") or "").strip()
            d["asignacion"] = "maestro"
            d["asignacion_detalle"] = "Cartera asignada en el maestro"
            if row["vend_hoja"] and row["vend_hoja"] != d["vendedor_id"]:
                d["asignacion_detalle"] += (f' · la hoja de altas dice {row["vend_hoja"]}')
        else:
            d["vendedor_id"] = row["vend_hoja"]
            d["vendedor_nombre"] = ""
            d["asignacion"] = "planilla"
            d["asignacion_detalle"] = ("El código no figura en el maestro · vendedor "
                                       "declarado en la hoja de altas")
        # El mismo cliente puede estar además en el padrón (lo cargaron en los dos
        # lados): se muestra igual, pero el total de altas lo cuenta una sola vez.
        d["ya_en_padron"] = cid in ids_padron
        altas_fuera.append(d)
    altas_fuera.sort(key=lambda c: (c["estado_compra"] == "sin_compras", -c["recompras"], c["nombre"]))

    # Atendidos que NO tienen el código cargado en el padrón: no se pueden seguir hasta
    # que alguien complete la columna. Se listan aparte para que el pendiente se vea.
    sin_codigo = [{**_base(r), **_sugerido(r), "atiende": r["atiende_raw"]}
                  for _, r in padron[(padron["estado"] == "ATENDIDO")
                                     & (padron["cliente_id"].isna())].iterrows()]
    sin_codigo.sort(key=lambda c: c["nombre"])

    potenciales = [{**_base(r), **_sugerido(r)}
                   for _, r in padron[padron["estado"] == "POTENCIAL"].iterrows()]
    potenciales.sort(key=lambda c: (c["vendedor_id"] or "ZZ", c["localidad"], c["nombre"]))

    no_at = padron[padron["estado"].isin(["NO_ATENDIDO", "SIN_RELEVAR"])]
    no_atendidos = [{**_base(r), **_sugerido(r), "relevado": r["estado"] == "NO_ATENDIDO"}
                    for _, r in no_at.iterrows()]
    no_atendidos.sort(key=lambda c: (not c["relevado"], c["vendedor_id"] or "ZZ",
                                     c["localidad"], c["nombre"]))

    comprando = [c for c in capturados if c["estado_compra"] != "sin_compras"]
    # Altas del plan = clientes dados de alta, del listado relevado + fuera del listado.
    # Se cuenta por CLIENTE, no por fila: el padrón repite algún PDV con el mismo código
    # y algún cliente está en las dos hojas.
    altas_ids = {c["cliente_id"] for c in capturados} | {c["cliente_id"] for c in altas_fuera}
    data = {
        "plan": {
            "objetivo": "Incrementar la cobertura en clientes de categoría B y C",
            "destinatarios": "Restaurantes o bares con carta de bebida (On Premise)",
            "medicion": "CCC únicos del canal On Premise segmento B&C, de julio a diciembre 2026",
            "objetivo_altas": PLAN_COB_OBJETIVO_ALTAS,
            "objetivo_altas_texto": f"{PLAN_COB_OBJETIVO_ALTAS} clientes nuevos dados de alta a diciembre 2026",
            "acciones": [
                "Incorporación: 1 + 2 cajas por línea comercial (Premium 1 + 1). Tope 1 combo por PDV por LC, máximo 4 LC.",
                "Alma Mora debe estar en todas las cartas. 5 + 1 botellas con tope de 2 combos por PDV hasta diciembre, sólo para los CCC capturados del plan.",
                "Recompra: 2 + 1 en botella para incorporar Cinzano. 1 combo por PDV.",
                "Premium: 5 + 1 en La Mascota para clientes nuevos que hayan incorporado Alma Mora y La Mascota. Tope 1 por PDV.",
                "1 + 1 a los 6 meses para clientes que hayan comprado al menos 1 LC en los últimos 6 meses.",
                "Para el reconocimiento hay que bajar el sin cargo en el campo 'Cantidad sin cargo UM' del cubo de sell out.",
            ],
        },
        "resumen": {
            "padron": int(len(padron)),
            "capturados": len(capturados),
            # El padrón repite algún PDV con el mismo código (el relevamiento marca
            # "repetido"): el CCC se cuenta por cliente, no por fila.
            "capturados_clientes": len({c["cliente_id"] for c in capturados}),
            "comprando": len(comprando),
            # Recompra sobre TODAS las altas (padrón + fuera del listado), por cliente:
            # es el mismo universo que el KPI de altas de arriba. Si contara sólo los
            # capturados, quedaría un número sobre otro denominador al lado del otro.
            "con_recompra": len({c["cliente_id"] for c in capturados + altas_fuera
                                 if c["recompro"]}),
            "sin_codigo": len(sin_codigo),
            "potenciales": len(potenciales),
            "no_atendidos": len(no_atendidos),
            "sin_relevar": sum(1 for c in no_atendidos if not c["relevado"]),
            "altas_fuera": len(altas_fuera),
            "altas_fuera_nuevas": sum(1 for c in altas_fuera if not c["ya_en_padron"]),
            "altas_total": len(altas_ids),
            "altas_objetivo": PLAN_COB_OBJETIVO_ALTAS,
            "altas_faltan": max(0, PLAN_COB_OBJETIVO_ALTAS - len(altas_ids)),
            "altas_pct": round(len(altas_ids) * 100.0 / PLAN_COB_OBJETIVO_ALTAS, 1),
            "altas_comprando": len({c["cliente_id"] for c in capturados + altas_fuera
                                    if c["estado_compra"] != "sin_compras"}),
        },
        "capturados": capturados,
        "altas_fuera": altas_fuera,
        "atendidos_sin_codigo": sin_codigo,
        "potenciales": potenciales,
        "no_atendidos": no_atendidos,
        "acciones_mes": acc_mes,
        "fuente": f"{fuente} + clientes.xlsx + historial de ventas (02_HISTORY + ventas_acumulada + ventas)",
    }
    _PLAN_COB_CACHE.clear()
    _PLAN_COB_CACHE[key] = data
    return data


@app.route("/api/gerencia/plan_cobertura")
def gerencia_plan_cobertura():
    data = _plan_cobertura()
    if data is None:
        return jsonify({"error": "Plan Cobertura no disponible: falta el padrón en 01_INPUTS/Plan cobertura/*.xlsx."}), 404
    return jsonify(_to_native({"generado_en": _now_ar(), **data}))


def _plan_cob_notas_map():
    """clave del PDV -> {mensaje, autor, updated_at}. Los mensajes viven en orbit.db
    (no en el padrón) y se editan a mano en cualquier momento: se leen siempre frescos,
    fuera del caché del plan."""
    try:
        conn = sqlite3.connect(str(DB_PATH), timeout=10, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(
                "SELECT clave, mensaje, autor, updated_at FROM plan_cob_nota").fetchall()
        finally:
            conn.close()
    except Exception as e:
        print(f"[WARN] plan_cobertura notas: {e}")
        return {}
    return {r["clave"]: {"mensaje": r["mensaje"], "autor": r["autor"],
                         "updated_at": r["updated_at"]} for r in rows}


@app.route("/api/vendedor/<vid>/plan_cobertura")
def vendedor_plan_cobertura(vid):
    """Plan Cobertura del vendedor: los PDV del padrón que le tocan a él.

      capturados            → los de SU cartera según el maestro.
      altas fuera del
      listado               → las de SU cartera, de la hoja aparte del padrón.
      potenciales /
      no atendidos /
      atendidos sin código  → los de SUS localidades, con el mismo criterio de zona que
                              usa gerencia (localidad → vendedor dominante del partido).

    El mensaje que dejó gerencia por PDV viaja en cada fila (sólo lectura del lado del
    vendedor). V3 no trabaja On Premise (sólo tradicional almacén/despensa/kiosco), así
    que el plan no le aplica: devuelve listas vacías con `no_aplica`."""
    data = _plan_cobertura()
    if data is None:
        return jsonify({"error": "Plan Cobertura no disponible: falta el padrón en 01_INPUTS/Plan cobertura/*.xlsx."}), 404
    vid_u = str(vid).upper()
    vacio = {"capturados": [], "altas_fuera": [], "atendidos_sin_codigo": [],
             "potenciales": [], "no_atendidos": []}
    if vid_u == "V3":
        return jsonify(_to_native({
            "generado_en": _now_ar(), "vendedor": vid_u, "no_aplica": True,
            "motivo": "V3 no trabaja On Premise: el Plan Cobertura no le aplica.",
            "plan": data["plan"], "resumen": {k: 0 for k in
                ("padron", "capturados", "capturados_clientes", "comprando", "con_recompra",
                 "sin_codigo", "potenciales", "no_atendidos", "sin_relevar",
                 "altas_fuera", "altas_total", "altas_comprando")},
            **vacio, "fuente": data.get("fuente", "")}))

    notas = _plan_cob_notas_map()

    def _mios(lst):
        # Copia por fila: las listas del payload de gerencia están cacheadas en memoria,
        # no se les puede colgar el mensaje encima.
        out = []
        for c in lst:
            if str(c.get("vendedor_id") or "").upper() != vid_u:
                continue
            n = notas.get(c.get("clave", "")) or {}
            out.append({**c, "mensaje": n.get("mensaje", ""),
                        "mensaje_autor": n.get("autor", ""),
                        "mensaje_fecha": n.get("updated_at", "")})
        return out

    cap = _mios(data["capturados"])
    afuera = _mios(data.get("altas_fuera", []))
    sin_cod = _mios(data["atendidos_sin_codigo"])
    pot = _mios(data["potenciales"])
    noat = _mios(data["no_atendidos"])
    # Altas del vendedor = sus capturados + sus altas fuera del listado, por CLIENTE.
    # El objetivo de 60 es del plan entero (no está repartido por vendedor): viaja en
    # `plan.objetivo_altas` como referencia, no se compara contra el número de él.
    altas_ids = {c["cliente_id"] for c in cap} | {c["cliente_id"] for c in afuera}
    resumen = {
        "padron": len(cap) + len(afuera) + len(sin_cod) + len(pot) + len(noat),
        "capturados": len(cap),
        "capturados_clientes": len({c["cliente_id"] for c in cap}),
        "comprando": sum(1 for c in cap if c["estado_compra"] != "sin_compras"),
        "con_recompra": len({c["cliente_id"] for c in cap + afuera if c["recompro"]}),
        "sin_codigo": len(sin_cod),
        "potenciales": len(pot),
        "no_atendidos": len(noat),
        "sin_relevar": sum(1 for c in noat if not c["relevado"]),
        "altas_fuera": len(afuera),
        "altas_total": len(altas_ids),
        "altas_comprando": len({c["cliente_id"] for c in cap + afuera
                                if c["estado_compra"] != "sin_compras"}),
    }
    return jsonify(_to_native({
        "generado_en": _now_ar(), "vendedor": vid_u, "no_aplica": False,
        "plan": data["plan"], "resumen": resumen,
        "capturados": cap, "altas_fuera": afuera, "atendidos_sin_codigo": sin_cod,
        "potenciales": pot, "no_atendidos": noat,
        "acciones_mes": data.get("acciones_mes", ""), "fuente": data.get("fuente", "")}))


@app.route("/api/gerencia/plan_cobertura/notas", methods=["GET", "POST"])
def gerencia_plan_cobertura_notas():
    """Mensaje libre por punto de venta del Plan Cobertura (lo que hay que hacer con ese
    PDV, qué se habló, quién lo va a visitar). Va aparte del payload del plan porque ése
    se cachea por mtime de los archivos y el mensaje se edita a mano en cualquier momento.
    GET  -> {clave: {mensaje, autor, updated_at}}
    POST -> {"clave": ..., "mensaje": ..., "autor": ...}; mensaje vacío BORRA la nota."""
    conn = sqlite3.connect(str(DB_PATH), timeout=10, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        if request.method == "POST":
            d = request.get_json(silent=True) or {}
            clave = str(d.get("clave", "")).strip()
            if not clave:
                return jsonify({"error": "falta 'clave'"}), 400
            mensaje = str(d.get("mensaje", "")).strip()[:1000]
            autor = str(d.get("autor", "Gerencia")).strip() or "Gerencia"
            ts = _now_ar()
            if mensaje:
                conn.execute(
                    """INSERT INTO plan_cob_nota(clave, mensaje, autor, updated_at)
                       VALUES(?,?,?,?)
                       ON CONFLICT(clave) DO UPDATE SET mensaje=excluded.mensaje,
                           autor=excluded.autor, updated_at=excluded.updated_at""",
                    (clave, mensaje, autor, ts))
            else:
                conn.execute("DELETE FROM plan_cob_nota WHERE clave=?", (clave,))
            conn.commit()
            return jsonify({"ok": True, "clave": clave, "mensaje": mensaje,
                            "autor": autor, "updated_at": ts})
        rows = conn.execute("SELECT clave, mensaje, autor, updated_at FROM plan_cob_nota").fetchall()
        return jsonify({r["clave"]: {"mensaje": r["mensaje"], "autor": r["autor"],
                                     "updated_at": r["updated_at"]} for r in rows})
    finally:
        conn.close()
# (el mismo SELECT, sin request, está en _plan_cob_notas_map() para la vista del vendedor)


# ====== CIERRE DE MES ======
@app.route("/api/gerencia/cierre_mes")
def gerencia_cierre_mes():
    """
    Resumen de cierre mensual por vendedor y empresa.
    ?mes=YYYY-MM  (default: mes anterior al actual)
    Fuente $:   resultado.xlsx hoja Avance
    Fuente CCC: ventas_acumulada.csv filtrado al mes
    """
    from calendar import monthrange

    MESES_ES = {
        1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
        7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"
    }

    hoy = datetime.now()
    mes_param = request.args.get("mes", "")
    if mes_param:
        try:
            año, mes = map(int, mes_param.split("-"))
        except Exception:
            return jsonify({"error": "Formato inválido. Use YYYY-MM"}), 400
    else:
        if hoy.month == 1:
            año, mes = hoy.year - 1, 12
        else:
            año, mes = hoy.year, hoy.month - 1

    ultimo_dia = monthrange(año, mes)[1]
    fecha_cierre = datetime(año, mes, ultimo_dia, 23, 59, 59)
    nombre_mes = f"{MESES_ES.get(mes, str(mes))} {año}"

    # Calendario del mes cerrado
    cal = contar_dias_habiles(fecha_corte=fecha_cierre)

    # Objetivos y acumulado del MES CERRADO desde resultado_mes.xlsx (fuente primaria).
    # resultado_mes.xlsx = acumulado congelado del mes cerrado (Acumulado == Tendencia).
    # resultado.xlsx (vivo, mes en curso) solo se usa como fallback si no existe el cierre.
    obj_por_vend = {}
    fuente_objetivos = None
    resultado_mes_path = INPUTS / "resultado_mes.xlsx"
    resultado_path     = INPUTS / "resultado.xlsx"
    fuente_path = resultado_mes_path if resultado_mes_path.exists() else resultado_path
    if fuente_path.exists():
        try:
            avance_df = pd.read_excel(fuente_path, sheet_name="Avance")
            for _, r in avance_df.iterrows():
                cn = clean_code(str(r.get("VendedorCodigo", "")))
                if not cn or int(cn) in _VENDEDORES_EXCLUIDOS:
                    continue
                obj_por_vend[cn] = {
                    "nombre": str(r.get("VendedorNombre", "")).strip().title(),
                    "objetivo": float(r.get("ValorObjetivo", 0) or 0),
                    "acumulado": float(r.get("Acumulado", 0) or 0),
                }
            fuente_objetivos = fuente_path.name
        except Exception:
            pass

    # CCC desde ventas_acumulada.csv filtrado al mes cerrado
    ccc_por_vend = {}
    vac_path = INPUTS / "ventas_acumulada.csv"
    if vac_path.exists():
        try:
            vac = pd.read_csv(vac_path, sep=";", encoding="latin1")
            vac["fecha"]    = pd.to_datetime(vac["FechaComprobante"], dayfirst=True, errors="coerce")
            vac["importe"]  = vac["ImporteNetoItem"].apply(_parse_num_ar)
            vac["vend_cod"] = pd.to_numeric(vac["CodVendedor"], errors="coerce")
            mes_ini = datetime(año, mes, 1)
            # No se filtra por Empresa (ver _LEEME_EMPRESA)
            vac = vac[
                (vac["fecha"] >= mes_ini) & (vac["fecha"] <= fecha_cierre) &
                (vac["importe"] > 0) & (~vac["vend_cod"].isin(_VENDEDORES_EXCLUIDOS))
            ].copy()
            vac["segmento"] = vac.apply(
                lambda r: _clasificar_segmento(str(r.get("Ramo", "")), str(r.get("Subramo", ""))), axis=1
            )
            for (vend_cod, seg), grp in vac.groupby(["vend_cod", "segmento"]):
                cn = clean_code(str(int(vend_cod)))
                if cn not in ccc_por_vend:
                    ccc_por_vend[cn] = {"TRADICIONAL": 0, "AUTOSERVICIO": 0, "ON_PREMISE_VTK": 0,
                                        "PROXIMITY": 0, "OTROS": 0}
                ccc_por_vend[cn][seg] = int(grp["Cliente"].nunique())
        except Exception:
            pass

    # Vendedores activos
    vend_df = read_csv(CONFIG / "vendedores_activos.csv")
    vend_activos = {}
    if not vend_df.empty:
        for _, v in vend_df[vend_df["activo"] == 1].iterrows():
            cn = clean_code(str(v["codigo_vendedor"]))
            if int(cn) not in _VENDEDORES_EXCLUIDOS:
                vend_activos[cn] = str(v["nombre_vendedor"]).strip()

    vendedores = []
    for cn, nombre_config in vend_activos.items():
        ob  = obj_por_vend.get(cn, {})
        ccc = ccc_por_vend.get(cn, {})

        objetivo   = ob.get("objetivo", 0)
        acumulado  = ob.get("acumulado", 0)
        avance_pct = round(acumulado / objetivo * 100, 2) if objetivo else 0
        nombre     = ob.get("nombre") or nombre_config

        ccc_trad  = ccc.get("TRADICIONAL", 0)
        ccc_auto  = 0 if cn == "3" else ccc.get("AUTOSERVICIO", 0)
        ccc_op    = ccc.get("ON_PREMISE_VTK", 0)
        ccc_total = ccc_trad + ccc_auto + ccc_op + ccc.get("PROXIMITY", 0) + ccc.get("OTROS", 0)

        vendedores.append({
            "codigo":           cn,
            "nombre":           nombre,
            "objetivo":         objetivo,
            "acumulado":        acumulado,
            "avance_pct":       avance_pct,
            "ccc_total":        ccc_total,
            "ccc_tradicional":  ccc_trad,
            "ccc_autoservicio": ccc_auto,
            "ccc_onpremise":    ccc_op,
        })

    vendedores.sort(key=lambda x: x["avance_pct"], reverse=True)

    empresa = {
        "objetivo":         sum(v["objetivo"]         for v in vendedores),
        "acumulado":        sum(v["acumulado"]         for v in vendedores),
        "ccc_total":        sum(v["ccc_total"]         for v in vendedores),
        "ccc_tradicional":  sum(v["ccc_tradicional"]  for v in vendedores),
        "ccc_autoservicio": sum(v["ccc_autoservicio"] for v in vendedores),
        "ccc_onpremise":    sum(v["ccc_onpremise"]    for v in vendedores),
    }
    empresa["avance_pct"] = round(
        empresa["acumulado"] / empresa["objetivo"] * 100, 2
    ) if empresa["objetivo"] else 0

    # ── 11 Titulares (totales empresa, sin apertura por vendedor) ──
    # ── 11 Titulares: CCC por marca vs objetivo CCC (mismo cálculo que /api/gerencia/once_titulares) ──
    # Fuente CCC: ventas_acumulada filtrado al mes cerrado
    # Fuente objetivo: objetivo 11T.xlsx
    once_titulares = {"marcas": [], "empresa": {}}
    try:
        _MARCA_LKP = {
            "ALMA MORA": "ALMA MORA", "ALARIS": "ALARIS", "TRAPICHE ALARIS": "ALARIS",
            "DON DAVID": "DON DAVID", "DADA": "DADA", "LOS ARBOLES": "LOS ARBOLES",
            "FINCA LAS MORAS": "FINCA LAS MORAS", "F LAS MORAS": "FINCA LAS MORAS",
            "TRAPICHE RESERVA": "TRAPICHE RESERVA",
            "FOND DE CAVE": "FOND DE CAVE", "FOND CAVE": "FOND DE CAVE",
            "CAZADOR": "CAZADOR", "ANTARES": "ANTARES",
            "GORDON'S FLAVOURS": "GORDON'S FLAVOURS", "GORDONS FLAVOURS": "GORDON'S FLAVOURS",
            "GORDONS": "GORDON'S FLAVOURS", "GORDON'S": "GORDON'S FLAVOURS",
            "SMIRNOFF": "SMIRNOFF FLAVOURS", "SMIRNOFF FLAVOURS": "SMIRNOFF FLAVOURS",
            "SMIRNOFF ICE": "SMIRNOFF ICE",
            "JW BLACK": "JW BLACK", "JW RED": "JW RED",
            "MASCOTA": "MASCOTA", "NC ESPUMANTES": "NC ESPUMANTES",
            "TRAPICHE MEDALLA": "TRAPICHE MEDALLA",
        }
        _ART_KW_11T = [
            ("SMIRNOFF ICE", "SMIRNOFF ICE"), ("SMF ICE", "SMIRNOFF ICE"),
            ("SMIRNOFF", "SMIRNOFF FLAVOURS"), ("GORDON", "GORDON'S FLAVOURS"),
            ("ANTARES", "ANTARES"), ("CAZADOR", "CAZADOR"),
            ("FOND DE CAVE", "FOND DE CAVE"), ("ALMA MORA", "ALMA MORA"),
            ("LOS ARBOLES", "LOS ARBOLES"), ("DADA", "DADA"),
            ("FINCA LAS MORAS", "FINCA LAS MORAS"), ("DON DAVID", "DON DAVID"),
            ("ALARIS", "ALARIS"), ("TRAPICHE RESERVA", "TRAPICHE RESERVA"),
            ("JW BLACK", "JW BLACK"), ("JW RED", "JW RED"),
        ]
        _OBJ_ALIAS_11T = {
            "ALMA MORA": "ALMA MORA", "TRAPICHE RESERVA": "TRAPICHE RESERVA",
            "FINCA LAS MORAS": "FINCA LAS MORAS", "FINCA LAS MORAS": "FINCA LAS MORAS",
            "ALARIS": "ALARIS", "DON DAVID": "DON DAVID", "DADA": "DADA",
            "SIMRNOFF FLAVORS": "SMIRNOFF FLAVOURS", "SMIRNOFF FLAVORS": "SMIRNOFF FLAVOURS",
            "SMIRNOFF FLAVOURS": "SMIRNOFF FLAVOURS",
            "LOS ARBOLES": "LOS ARBOLES", "ANTARES": "ANTARES",
            "SMIRNOFF ICE": "SMIRNOFF ICE", "SMF ICE": "SMIRNOFF ICE",
            "GORDONS FLAVOURS": "GORDON'S FLAVOURS", "GORDONS FLAVORS": "GORDON'S FLAVOURS",
            "GORDON'S FLAVOURS": "GORDON'S FLAVOURS",
        }

        # Objetivos desde objetivo 11T.xlsx
        obj_map_11t = {}
        try:
            obj_df = pd.read_excel(INPUTS / "objetivo 11T.xlsx", header=1)
            obj_df = obj_df.dropna(subset=obj_df.columns[1:2])
            for _, row in obj_df.iterrows():
                raw = str(row.iloc[1]).upper().strip()
                mk = _OBJ_ALIAS_11T.get(raw, raw)
                try:
                    obj_map_11t[mk] = int(float(row.iloc[2]))
                except (ValueError, TypeError):
                    pass
        except Exception:
            pass

        # CCC por marca desde ventas_acumulada COMPLETO (sin filtro de fecha)
        # — mismo criterio que el dashboard /api/gerencia/once_titulares
        ccc_map_11t = {}
        vac_path_11t = INPUTS / "ventas_acumulada.csv"
        if vac_path_11t.exists():
            try:
                vac11 = pd.read_csv(vac_path_11t, sep=";", encoding="latin1", low_memory=False)
                vac11["importe"] = pd.to_numeric(
                    vac11["ImporteNetoItem"].astype(str).str.replace(",", ".", regex=False), errors="coerce")
                vac11 = vac11[
                    (~vac11["CodVendedor"].isin(_VENDEDORES_EXCLUIDOS)) &
                    (vac11["importe"] > 0)
                ]
                vac11 = vac11[_mask_superficie_11t(vac11)]   # 11T mide solo AS + Almacén + Kiosco
                # 1) match por Código Art. exacto (matriz oficial); 2) fallback texto de Marca
                vac11["marca_objetivo"] = _marca_11t_por_codigo(vac11)
                _falta11 = vac11["marca_objetivo"].isna()
                vac11.loc[_falta11, "marca_objetivo"] = (
                    vac11.loc[_falta11, "Marca"].astype(str).str.upper().str.strip().map(_MARCA_LKP))
                unresolved = vac11["marca_objetivo"].isna()
                if unresolved.any():
                    for kw, mo in _ART_KW_11T:
                        still = vac11["marca_objetivo"].isna()
                        if not still.any():
                            break
                        hits = vac11.loc[still, "Articulo"].astype(str).str.upper().str.contains(kw, regex=False, na=False)
                        vac11.loc[still & hits, "marca_objetivo"] = mo
                ccc_map_11t = (vac11[vac11["marca_objetivo"].notna()]
                               .groupby("marca_objetivo")["Cliente"].nunique().to_dict())
            except Exception:
                pass

        # Construir resultado — solo marcas con objetivo definido
        marcas_11t = []
        total_ccc = 0
        total_obj = 0
        for mk in sorted(obj_map_11t.keys(), key=lambda x: ccc_map_11t.get(x, 0), reverse=True):
            ccc_v = ccc_map_11t.get(mk, 0)
            obj_v = obj_map_11t[mk]
            pct   = round(ccc_v / obj_v * 100, 1) if obj_v else None
            marcas_11t.append({"marca": mk, "ccc": ccc_v, "objetivo": obj_v, "pct": pct})
            total_ccc += ccc_v
            total_obj += obj_v

        once_titulares["marcas"] = marcas_11t
        once_titulares["empresa"] = {
            "ccc_total": total_ccc,
            "objetivo_total": total_obj,
            "pct": round(total_ccc / total_obj * 100, 1) if total_obj else 0,
            "marcas_sobre_objetivo": sum(1 for m in marcas_11t if (m["pct"] or 0) >= 100),
            "total_marcas": len(marcas_11t),
        }
    except Exception:
        pass

    # ── Innovaciones (denominador = CCC Mes: clientes con compra en el mes) ──
    innovaciones = {"resumen": {}, "por_producto": []}
    try:
        innov = read_csv(DATASETS / "mod_innovaciones_segmento.csv")
        if not innov.empty:
            # CCC Mes: clientes únicos con compra en el mes cerrado (fuente ventas_acumulada sin filtro empresa)
            ccc_mes_total = 0
            vac_path2 = INPUTS / "ventas_acumulada.csv"
            if vac_path2.exists():
                try:
                    _v2 = pd.read_csv(vac_path2, sep=";", encoding="latin1",
                                      usecols=["FechaComprobante", "ImporteNetoItem", "CodVendedor", "Cliente"])
                    _v2["fecha"]   = pd.to_datetime(_v2["FechaComprobante"], dayfirst=True, errors="coerce")
                    _v2["importe"] = _v2["ImporteNetoItem"].apply(_parse_num_ar)
                    _v2["vend_c"]  = pd.to_numeric(_v2["CodVendedor"], errors="coerce")
                    _v2 = _v2[
                        (_v2["fecha"] >= datetime(año, mes, 1)) & (_v2["fecha"] <= fecha_cierre) &
                        (_v2["importe"] > 0) & (~_v2["vend_c"].isin(_VENDEDORES_EXCLUIDOS))
                    ]
                    ccc_mes_total = int(_v2["Cliente"].nunique())
                except Exception:
                    pass

            denom = ccc_mes_total if ccc_mes_total > 0 else None
            pp = (innov.groupby(["producto_codigo", "producto_nombre"])
                  .agg(compraron=("clientes_compraron", "sum"))
                  .reset_index())
            if denom:
                pp["pct"] = (pp["compraron"] / denom * 100).round(1)
            else:
                pp["pct"] = 0.0
            innovaciones["resumen"] = {
                "productos": int(pp["producto_codigo"].nunique()),
                "compraron_total": int(pp["compraron"].sum()),
                "ccc_mes": ccc_mes_total,
                "pct_promedio": round(pp["pct"].mean(), 1) if denom else 0,
            }
            innovaciones["por_producto"] = (
                pp.sort_values("pct", ascending=False).head(20).to_dict("records")
            )
    except Exception:
        pass

    # ── Sell Out cierre: ventas_mes.csv + maestro 04D. Misma lógica que auditoría.
    # Sin fallback a ventas.csv. Si ventas_mes.csv no existe → disponible: False.
    sellout = {"categorias": [], "fuente": "ventas_mes.csv"}
    so_src = INPUTS / "ventas_mes.csv"
    if not so_src.exists():
        sellout["disponible"] = False
        sellout["error"] = "ventas_mes.csv no encontrado en 01_INPUTS"
    else:
        # Leer ventas_mes.csv con lector específico (sep=',' explícito, no sep=None).
        # incluir_deposito=True: conserva V20 para separar ruta vs Depósito (paridad con el vivo).
        so_df = _leer_ventas_mes_csv(so_src, incluir_deposito=True)
        sellout["filas_ventas_mes"] = len(so_df)
        if so_df.empty:
            sellout["error"] = "ventas_mes.csv sin filas válidas (importe>0, excl V2/V5)"
        else:
            # Mismo cruce ventas × maestro 04D que generó auditoria_sellout_cierre_mes.csv,
            # más el bloque Depósito V20 aparte (sin objetivo).
            try:
                sellout.update(_sellout_con_deposito(so_df))
            except Exception as _e:
                sellout["error"] = str(_e)

    # ── Planes AS ──
    planes_as = {"resumen": {}, "por_plan": []}
    try:
        pa = read_csv(DATASETS / "mod_planes_as.csv")
        if not pa.empty:
            planes_as["resumen"] = {
                "clientes": int(len(pa)),
                "facturado": round(float(pa["total_facturado"].sum()), 0),
                "sc_ganado": int(pa["sc_total_ganado"].sum()),
                "sc_enviado": int(pa["sc_cajas_enviadas_total"].sum()),
            }
            pp = pa.groupby("plan_as").agg(
                clientes=("cliente_id", "count"),
                facturado=("total_facturado", "sum"),
                sc_ganado=("sc_total_ganado", "sum"),
            ).reset_index()
            planes_as["por_plan"] = pp.to_dict("records")
    except Exception:
        pass

    # ── Acciones Comerciales ──
    acciones = {"resumen": {}, "detalle": []}
    try:
        ac = read_csv(DATASETS / "mod_acciones_ranking.csv")
        if not ac.empty:
            acciones["resumen"] = {
                "total_acciones": int(len(ac)),
                "inversion_total": round(float(ac["inversion_pesos"].sum()), 0),
                "clientes_afectados": int(ac["clientes_afectados"].sum()),
                "importe_neto": round(float(ac["importe_neto"].sum()), 0),
            }
            top = ac.sort_values("inversion_pesos", ascending=False).head(10)
            acciones["detalle"] = [
                {
                    "nombre": r["accion_nombre"],
                    "tipo": r.get("tipo_accion", ""),
                    "canal": r.get("canal", ""),
                    "categoria": r.get("categoria", ""),
                    "descuento": r.get("descuento_display", ""),
                    "litros": round(float(r.get("litros_vendidos", 0)), 1),
                    "inversion": round(float(r["inversion_pesos"]), 0),
                    "importe_neto": round(float(r.get("importe_neto", 0)), 0),
                    "clientes": int(r.get("clientes_afectados", 0)),
                }
                for _, r in top.iterrows()
            ]
    except Exception:
        pass

    return jsonify({
        "mes":              f"{año:04d}-{mes:02d}",
        "nombre_mes":       nombre_mes,
        "calendario":       cal,
        "empresa":          empresa,
        "vendedores":       vendedores,
        "once_titulares":   once_titulares,
        "innovaciones":     innovaciones,
        "sellout":          sellout,
        "planes_as":        planes_as,
        "acciones":         acciones,
        "fuente_objetivos": fuente_objetivos or "resultado.xlsx",
        "fuente_ccc":       "ventas_acumulada.csv",
    })


# ====== CIERRE VERSIONADO POR CARPETA (01_INPUTS/cierres mes/) ======
# Cada cierre es autocontenido en 3 archivos versionados por _MMAAAA:
#   resultado_mes_<MMAAAA>.xlsx   → objetivo/acumulado/avance por vendedor
#   ventas_mes_<MMAAAA>.csv       → CCC, 11T (y luego sell-out/planes/acciones/innov)
#   objetivo 11T_<MMAAAA>.xlsx    → objetivo del 11T por marca
# Catálogos compartidos (maestro 04D, escala, acciones, innovaciones) NO van acá.
CIERRES_MES_DIR = INPUTS / "cierres mes"

_MARCA_LKP_CIERRE = {
    "ALMA MORA": "ALMA MORA", "ALARIS": "ALARIS", "TRAPICHE ALARIS": "ALARIS",
    "DON DAVID": "DON DAVID", "DADA": "DADA", "LOS ARBOLES": "LOS ARBOLES",
    "FINCA LAS MORAS": "FINCA LAS MORAS", "F LAS MORAS": "FINCA LAS MORAS",
    "TRAPICHE RESERVA": "TRAPICHE RESERVA",
    "FOND DE CAVE": "FOND DE CAVE", "FOND CAVE": "FOND DE CAVE",
    "CAZADOR": "CAZADOR", "ANTARES": "ANTARES",
    "GORDON'S FLAVOURS": "GORDON'S FLAVOURS", "GORDONS FLAVOURS": "GORDON'S FLAVOURS",
    "GORDONS": "GORDON'S FLAVOURS", "GORDON'S": "GORDON'S FLAVOURS",
    "SMIRNOFF": "SMIRNOFF FLAVOURS", "SMIRNOFF FLAVOURS": "SMIRNOFF FLAVOURS",
    "SMIRNOFF ICE": "SMIRNOFF ICE",
    "JW BLACK": "JW BLACK", "JW RED": "JW RED",
    "MASCOTA": "MASCOTA", "NC ESPUMANTES": "NC ESPUMANTES",
    "TRAPICHE MEDALLA": "TRAPICHE MEDALLA",
}
_ART_KW_11T_CIERRE = [
    ("SMIRNOFF ICE", "SMIRNOFF ICE"), ("SMF ICE", "SMIRNOFF ICE"),
    ("SMIRNOFF", "SMIRNOFF FLAVOURS"), ("GORDON", "GORDON'S FLAVOURS"),
    ("ANTARES", "ANTARES"), ("CAZADOR", "CAZADOR"),
    ("FOND DE CAVE", "FOND DE CAVE"), ("ALMA MORA", "ALMA MORA"),
    ("LOS ARBOLES", "LOS ARBOLES"), ("DADA", "DADA"),
    ("FINCA LAS MORAS", "FINCA LAS MORAS"), ("DON DAVID", "DON DAVID"),
    ("ALARIS", "ALARIS"), ("TRAPICHE RESERVA", "TRAPICHE RESERVA"),
    ("JW BLACK", "JW BLACK"), ("JW RED", "JW RED"),
]
_OBJ_ALIAS_11T_CIERRE = {
    "ALMA MORA": "ALMA MORA", "TRAPICHE RESERVA": "TRAPICHE RESERVA",
    "FINCA LAS MORAS": "FINCA LAS MORAS",
    "ALARIS": "ALARIS", "DON DAVID": "DON DAVID", "DADA": "DADA",
    "SIMRNOFF FLAVORS": "SMIRNOFF FLAVOURS", "SMIRNOFF FLAVORS": "SMIRNOFF FLAVOURS",
    "SMIRNOFF FLAVOURS": "SMIRNOFF FLAVOURS",
    "LOS ARBOLES": "LOS ARBOLES", "ANTARES": "ANTARES",
    "SMIRNOFF ICE": "SMIRNOFF ICE", "SMF ICE": "SMIRNOFF ICE",
    "GORDONS FLAVOURS": "GORDON'S FLAVOURS", "GORDONS FLAVORS": "GORDON'S FLAVOURS",
    "GORDON'S FLAVOURS": "GORDON'S FLAVOURS",
}


_VENTAS_MES_CACHE = {}
def _leer_ventas_mes_cacheado(path, incluir_deposito=False):
    """Lee ventas_mes versionado con caché (archivos del cierre = inmutables, clave path+mtime).
    Evita releer el CSV (3MB, parser python) varias veces por request → era la causa de los ~31s.
    incluir_deposito entra en la clave de caché: la variante con V20 (sell out) no contamina
    la variante sin V20 que usan CCC/once_titulares/ranking del cierre."""
    try:
        key = (str(path), os.path.getmtime(path), incluir_deposito)
    except OSError:
        key = (str(path), 0, incluir_deposito)
    df = _VENTAS_MES_CACHE.get(key)
    if df is None:
        df = _leer_ventas_mes_csv(path, incluir_deposito=incluir_deposito)
        _VENTAS_MES_CACHE[key] = df
    return df


def _to_native(o):
    """Convierte recursivamente numpy/pandas → tipos nativos de Python. El JSON provider de
    Flask no serializa numpy.int64/float64 (causaba HTTP 500 en Render aunque local funcionara)."""
    import numpy as _np
    if isinstance(o, dict):
        return {k: _to_native(v) for k, v in o.items()}
    if isinstance(o, (list, tuple)):
        return [_to_native(x) for x in o]
    if isinstance(o, _np.generic):
        return o.item()
    return o


def _cierre_archivos_mes(periodo):
    """periodo 'AAAA-MM' → dict con los 3 archivos del cierre versionado, o None si faltan."""
    try:
        y, m = periodo.split("-")
        mmaaaa = f"{int(m):02d}{int(y)}"
    except Exception:
        return None
    res   = CIERRES_MES_DIR / f"resultado_mes_{mmaaaa}.xlsx"
    vmes  = CIERRES_MES_DIR / f"ventas_mes_{mmaaaa}.csv"
    obj11 = CIERRES_MES_DIR / f"objetivo 11T_{mmaaaa}.xlsx"
    # 11T es bimestral → su fuente es ventas_acumulada_<MMAAAA>.csv (2 meses). Opcional:
    # si no existe, _cierre_once_titulares cae a ventas_mes (1 mes).
    vacum = CIERRES_MES_DIR / f"ventas_acumulada_{mmaaaa}.csv"
    if res.exists() and vmes.exists() and obj11.exists():
        return {"resultado": res, "ventas_mes": vmes, "objetivo_11t": obj11, "mmaaaa": mmaaaa,
                "ventas_acumulada": vacum if vacum.exists() else None}
    return None


def _cierre_ccc_por_vend_segmento(vmes_df):
    """CCC (clientes únicos, neto>0) por vendedor × segmento desde ventas_mes del cierre.
    No filtra por Empresa (ver _LEEME_EMPRESA). V2/V5/V20 ya excluidos por _leer_ventas_mes_csv."""
    out = {}
    if vmes_df.empty:
        return out
    df = vmes_df.copy()
    df["_cli"]  = pd.to_numeric(df["Cliente"], errors="coerce")
    df["_vend"] = pd.to_numeric(df["CodVendedor"], errors="coerce")
    df["_seg"]  = [
        _clasificar_segmento(str(r), str(s))
        for r, s in zip(df.get("Ramo", pd.Series([""] * len(df))),
                        df.get("Subramo", pd.Series([""] * len(df))))
    ]
    df = df.dropna(subset=["_cli", "_vend"])
    for (vend, seg), grp in df.groupby([df["_vend"].astype(int), "_seg"]):
        cn = clean_code(str(int(vend)))
        out.setdefault(cn, {"TRADICIONAL": 0, "AUTOSERVICIO": 0, "ON_PREMISE_VTK": 0,
                            "PROXIMITY": 0, "OTROS": 0})
        out[cn][seg] = int(grp["_cli"].nunique())
    return out


def _cierre_objetivos_avance(files):
    """objetivos_avance del cierre desde resultado_mes_<MMAAAA>.xlsx (obj/acum) + ventas_mes (CCC)."""
    obj_por_vend = {}
    av = pd.read_excel(files["resultado"], sheet_name="Avance")
    for _, r in av.iterrows():
        cn = clean_code(str(r.get("VendedorCodigo", "")))
        if not cn or int(cn) in _VENDEDORES_EXCLUIDOS:
            continue
        obj_por_vend[cn] = {
            "nombre":    str(r.get("VendedorNombre", "")).strip().title(),
            "objetivo":  float(r.get("ValorObjetivo", 0) or 0),
            "acumulado": float(r.get("Acumulado", 0) or 0),
        }

    ccc_por_vend = _cierre_ccc_por_vend_segmento(_leer_ventas_mes_cacheado(files["ventas_mes"]))

    vend_df = read_csv(CONFIG / "vendedores_activos.csv")
    vend_activos = {}
    if not vend_df.empty:
        for _, v in vend_df[vend_df["activo"] == 1].iterrows():
            cn = clean_code(str(v["codigo_vendedor"]))
            if int(cn) not in _VENDEDORES_EXCLUIDOS:
                vend_activos[cn] = str(v["nombre_vendedor"]).strip()

    vendedores = []
    for cn, nombre_config in vend_activos.items():
        ob  = obj_por_vend.get(cn, {})
        ccc = ccc_por_vend.get(cn, {})
        objetivo   = ob.get("objetivo", 0)
        acumulado  = ob.get("acumulado", 0)
        avance_pct = round(acumulado / objetivo * 100, 2) if objetivo else 0
        ccc_trad = ccc.get("TRADICIONAL", 0)
        ccc_auto = 0 if cn == "3" else ccc.get("AUTOSERVICIO", 0)   # V3 no AS
        ccc_op   = 0 if cn == "3" else ccc.get("ON_PREMISE_VTK", 0) # V3 no On Premise
        ccc_total = ccc_trad + ccc_auto + ccc_op + ccc.get("PROXIMITY", 0) + ccc.get("OTROS", 0)
        vendedores.append({
            "codigo": cn, "nombre": ob.get("nombre") or nombre_config,
            "objetivo": objetivo, "acumulado": acumulado, "avance_pct": avance_pct,
            "faltante": round(objetivo - acumulado, 2),
            "ccc_total": ccc_total, "ccc_tradicional": ccc_trad,
            "ccc_autoservicio": ccc_auto, "ccc_onpremise": ccc_op,
        })
    vendedores.sort(key=lambda x: x["avance_pct"], reverse=True)

    empresa = {k: sum(v[k] for v in vendedores) for k in
               ("objetivo", "acumulado", "ccc_total", "ccc_tradicional",
                "ccc_autoservicio", "ccc_onpremise")}
    empresa["avance_pct"] = round(empresa["acumulado"] / empresa["objetivo"] * 100, 2) if empresa["objetivo"] else 0
    empresa["faltante"]   = round(empresa["objetivo"] - empresa["acumulado"], 2)

    dias_habiles = None
    try:
        from calendar import monthrange
        y, m = int(files["mmaaaa"][2:]), int(files["mmaaaa"][:2])
        cal = contar_dias_habiles(fecha_corte=datetime(y, m, monthrange(y, m)[1], 23, 59, 59))
        dias_habiles = cal.get("total")
    except Exception:
        pass

    return {
        "dias_habiles":     dias_habiles,
        "empresa":          empresa,
        "vendedores":       vendedores,
        "fuente_objetivos": files["resultado"].name,
        "fuente_acumulado": files["resultado"].name,
        "fuente_ccc":       files["ventas_mes"].name,
        "congelado_desde":  "01_INPUTS/cierres mes/ (versionado por archivo)",
    }


_VACUM_CIERRE_CACHE = {}
def _leer_ventas_acum_cierre(path):
    """Lee ventas_acumulada_<MMAAAA>.csv del cierre (sep=';', latin1) con caché.
    Mismo criterio que /api/gerencia/once_titulares: NO filtra por Empresa (P&P Logística
    es nuestra segunda razón social), excluye V2/V5/V20 y filtra neto>0."""
    try:
        key = (str(path), os.path.getmtime(path))
    except OSError:
        key = (str(path), 0)
    df = _VACUM_CIERRE_CACHE.get(key)
    if df is None:
        df = pd.read_csv(path, sep=";", encoding="latin1", low_memory=False)
        df.columns = [c.strip() for c in df.columns]
        df["ImporteNetoItem"] = pd.to_numeric(
            df["ImporteNetoItem"].astype(str).str.replace(",", ".", regex=False), errors="coerce")
        cv = pd.to_numeric(df["CodVendedor"], errors="coerce")
        df = df[(~cv.isin(_VENDEDORES_EXCLUIDOS)) & (df["ImporteNetoItem"] > 0)].copy()
        _VACUM_CIERRE_CACHE[key] = df
    return df


def _cierre_once_titulares(files):
    """11T CCC vs Objetivo del cierre. CCC = clientes únicos (neto>0) por marca titular,
    MISMO criterio que /api/gerencia/once_titulares. Fuente: ventas_acumulada_<MMAAAA>.csv
    (bimestral — el 11T se mide en 2 meses); fallback a ventas_mes si no existe la acumulada.
    Objetivo: objetivo 11T_<MMAAAA>.xlsx."""
    obj_map = {}
    try:
        odf = pd.read_excel(files["objetivo_11t"], header=1)
        odf = odf.dropna(subset=odf.columns[1:2])
        for _, r in odf.iterrows():
            raw = str(r.iloc[1]).upper().strip()
            mk = _OBJ_ALIAS_11T_CIERRE.get(raw, raw)
            try:
                obj_map[mk] = int(float(r.iloc[2]))
            except (ValueError, TypeError):
                pass
    except Exception:
        pass

    ccc_map = {}
    # Fuente bimestral (ventas_acumulada) si está; si no, ventas_mes (1 mes). Ambos lectores
    # ya filtran neto>0 y excluyen V2/V5/V20. No se filtra por Empresa (P&P Logística es
    # nuestra segunda razón social) — igual criterio que /api/gerencia/once_titulares.
    src_acum = files.get("ventas_acumulada")
    df = _leer_ventas_acum_cierre(src_acum) if src_acum is not None else _leer_ventas_mes_cacheado(files["ventas_mes"])
    if not df.empty:
        df = df.copy()
        df = df[_mask_superficie_11t(df)]   # 11T mide solo AS + Almacén + Kiosco
        # 1) match por Código Art. exacto (matriz oficial); 2) fallback texto de Marca
        df["mo"] = _marca_11t_por_codigo(df)
        _falta_mo = df["mo"].isna()
        df.loc[_falta_mo, "mo"] = df.loc[_falta_mo, "Marca"].astype(str).str.upper().str.strip().map(_MARCA_LKP_CIERRE)
        for kw, mo in _ART_KW_11T_CIERRE:
            still = df["mo"].isna()
            if not still.any():
                break
            hits = df.loc[still, "Articulo"].astype(str).str.upper().str.contains(kw, regex=False, na=False)
            df.loc[still & hits, "mo"] = mo
        df["_cli"] = pd.to_numeric(df["Cliente"], errors="coerce")
        ccc_map = df[df["mo"].notna()].groupby("mo")["_cli"].nunique().to_dict()

    marcas, tot_ccc, tot_obj = [], 0, 0
    for mk in sorted(obj_map, key=lambda x: ccc_map.get(x, 0), reverse=True):
        ccc_v = int(ccc_map.get(mk, 0))
        obj_v = obj_map[mk]
        pct   = round(ccc_v / obj_v * 100, 1) if obj_v else None
        marcas.append({"marca": mk, "ccc": ccc_v, "objetivo": obj_v, "pct": pct})
        tot_ccc += ccc_v
        tot_obj += obj_v
    return {
        "marcas": marcas,
        "empresa": {
            "ccc_total": tot_ccc, "objetivo_total": tot_obj,
            "pct": round(tot_ccc / tot_obj * 100, 1) if tot_obj else 0,
            "marcas_sobre_objetivo": sum(1 for m in marcas if (m["pct"] or 0) >= 100),
            "total_marcas": len(marcas),
        },
    }


_GCM_MODULE = None
def _gcm():
    """Importa (cacheado) tools/generar_cierre_mensual.py para reutilizar el motor oficial
    de cálculo del cierre (litros, 11T, innovaciones, ranking) sobre un DataFrame."""
    global _GCM_MODULE
    if _GCM_MODULE is None:
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "generar_cierre_mensual", str(BASE / "tools" / "generar_cierre_mensual.py"))
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _GCM_MODULE = mod
    return _GCM_MODULE


_GDA_MODULE = None
def _gda():
    """Importa (cacheado) generar_datasets_acum.py para reutilizar el catálogo y el matching
    de acciones comerciales (_REGLA_CANAL_SEG_MAP, _filtrar_ventas_accion, INOV_PRODUCTOS,
    cargar_clientes). Solo define funciones/constantes al import (sin I/O ni main())."""
    global _GDA_MODULE
    if _GDA_MODULE is None:
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "generar_datasets_acum", str(BASE / "generar_datasets_acum.py"))
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _GDA_MODULE = mod
    return _GDA_MODULE


_GCM_VENTAS_CACHE = {}
def _gcm_leer_ventas_cacheado(path):
    """DataFrame de ventas_mes leído por el motor oficial (filtra activos, V3 solo trad), cacheado."""
    try:
        key = (str(path), os.path.getmtime(path))
    except OSError:
        key = (str(path), 0)
    df = _GCM_VENTAS_CACHE.get(key)
    if df is None:
        df = _gcm()._leer_ventas(path)
        _GCM_VENTAS_CACHE[key] = df
    return df


def _cierre_ranking_payload(rank, avance_map=None):
    """Transforma la lista de ranking en {ranking_top3, ranking, ganadores} para el portal.
    Si avance_map (codigo_str -> avance_pct) viene dado, el 'mejor en VOLUMEN' se determina por
    ALCANCE DEL OBJETIVO MENSUAL (avance vs objetivo), NO por litros+dinero. Solo cambia esa
    dimensión y su etiqueta; el score general y el resto del ranking quedan intactos."""
    rank = [dict(r) for r in rank]  # copia: no mutar el input cacheado

    if avance_map:
        def _av(r):
            v = avance_map.get(str(r.get("vendedor_codigo", "")).lstrip("Vv"))
            return float(v) if v is not None else -1.0
        for r in rank:
            a = _av(r)
            r["alcance_objetivo_pct"] = a if a >= 0 else None
        # Re-rankear la dimensión volumen por alcance de objetivo (desc) y rehacer su etiqueta.
        for i, r in enumerate(sorted(rank, key=lambda r: -_av(r))):
            r["ranking_volumen_dinero"] = i + 1
        for r in rank:
            etiq = [e for e in str(r.get("etiqueta_destacada", "")).split("|")
                    if e and e != "MEJOR_VOLUMEN_DINERO"]
            if r.get("ranking_volumen_dinero") == 1:
                etiq.append("MEJOR_VOLUMEN_DINERO")
            r["etiqueta_destacada"] = "|".join(etiq)

    top3 = sorted(rank, key=lambda r: r.get("ranking_general", 99))[:3]
    ranking_top3 = [{
        "ranking_general":       r.get("ranking_general"),
        "vendedor_codigo":       r.get("vendedor_codigo"),
        "vendedor_nombre":       r.get("vendedor_nombre"),
        "score_total":           r.get("score_total"),
        "clientes_11_titulares": r.get("clientes_11_titulares"),
        "clientes_innovaciones": r.get("clientes_innovaciones"),
        "etiqueta_destacada":    r.get("etiqueta_destacada", ""),
    } for r in top3]
    ranking = [{
        "vendedor_codigo":        r.get("vendedor_codigo"),
        "vendedor_nombre":        r.get("vendedor_nombre"),
        "dinero_vendido":         r.get("dinero_vendido"),
        "litros_vendidos":        r.get("litros_vendidos"),
        "clientes_11_titulares":  r.get("clientes_11_titulares"),
        "clientes_innovaciones":  r.get("clientes_innovaciones"),
        "alcance_objetivo_pct":   r.get("alcance_objetivo_pct"),
        "score_total":            r.get("score_total"),
        "ranking_general":        r.get("ranking_general"),
        "ranking_volumen_dinero": r.get("ranking_volumen_dinero"),
        "ranking_11_titulares":   r.get("ranking_11_titulares"),
        "ranking_innovaciones":   r.get("ranking_innovaciones"),
        "etiqueta_destacada":     r.get("etiqueta_destacada", ""),
    } for r in sorted(rank, key=lambda r: r.get("ranking_general", 99))]

    def _ganador(campo, metrica):
        for r in rank:
            if r.get(campo) == 1:
                return {"vendedor_codigo": r.get("vendedor_codigo"),
                        "vendedor_nombre": r.get("vendedor_nombre"),
                        "score_total": r.get("score_total"),
                        "metrica": r.get(metrica)}
        return None
    # Ganador de volumen: si hay avance_map, la métrica es el alcance del objetivo (%), no $.
    gan_vol = _ganador("ranking_volumen_dinero", "alcance_objetivo_pct" if avance_map else "dinero_vendido")
    if gan_vol is not None:
        gan_vol["base"] = "alcance_objetivo" if avance_map else "dinero_vendido"
    ganadores = {
        "general":        _ganador("ranking_general",        "score_total"),
        "volumen_dinero": gan_vol,
        "once_titulares": _ganador("ranking_11_titulares",   "clientes_11_titulares"),
        "innovaciones":   _ganador("ranking_innovaciones",   "clientes_innovaciones"),
    }
    return {"ranking_top3": ranking_top3, "ranking": ranking, "ganadores": ganadores}


def _cierre_extras_versionado(files):
    """Sell-Out, Innovaciones y Ranking del cierre desde ventas_mes_<MMAAAA>.csv,
    reutilizando el motor oficial. Catálogos compartidos: 04D, Innovaciones.xlsx, vendedores.
    Devuelve todo casteado a tipos nativos (jsonify-safe)."""
    out = {}
    # Sell-Out: TODA la venta agrupada (ruta + V20 Depósito) vs objetivo de EMPRESA, igual
    # criterio que el sell out vivo (/api/gerencia/sellout_litros). categorias ya incluye V20;
    # `deposito` queda como desglose informativo (ya está sumado en categorias).
    try:
        so_df = _leer_ventas_mes_cacheado(files["ventas_mes"], incluir_deposito=True)
        if so_df.empty:
            out["sellout"] = {"fuente": files["ventas_mes"].name, "filas_ventas_mes": 0,
                              "categorias": [], "deposito": [], "incluye_deposito": True}
        else:
            categorias = _sellout_desde_ventas(so_df)             # ruta + V20 vs objetivo
            dep_df = so_df[so_df["CodVendedor"] == 20]
            deposito = _sellout_desde_ventas(dep_df) if not dep_df.empty else []
            for c in deposito:                                    # depósito: sin objetivo propio
                c["objetivo"], c["alcance_pct"] = None, None
                for sub in c.get("subcategorias", []):
                    sub["objetivo"], sub["alcance_pct"] = None, None
            out["sellout"] = {
                "fuente": files["ventas_mes"].name,
                "filas_ventas_mes": int(len(so_df)),
                "categorias":       categorias,
                "total_litros":     round(sum(float(c["litros"]) for c in categorias), 1),
                "incluye_deposito": True,
                "deposito":         deposito,
                "total_deposito":   round(sum(float(c["litros"]) for c in deposito), 1),
            }
    except Exception as e:
        out["sellout"] = {"error": str(e), "categorias": []}

    # Innovaciones + Ranking (motor oficial sobre el ventas_mes versionado)
    try:
        gcm = _gcm()
        df = _gcm_leer_ventas_cacheado(files["ventas_mes"])
        vn = gcm._leer_vendedores(CONFIG / "vendedores_activos.csv")
        # Códigos de innovación desde el loader oficial (mismo que el dashboard: formato
        # "CODIGO - NOMBRE", 22 productos). El parser viejo gcm._leer_innovaciones quedó en 0
        # al cambiar el formato del xlsx → ranking marcaba "mejor innovaciones" con 0 clientes.
        cod_inov = set(_gda().INOV_PRODUCTOS.keys())

        ccc_mes = int(df["Cliente"].nunique()) if not df.empty else 0
        prod = {}
        for r in gcm._inov_detalle(df, cod_inov, vn):
            k = (int(r["producto_codigo"]), str(r["producto_nombre"]))
            prod[k] = prod.get(k, 0) + int(r["clientes_compraron"])
        por_producto = [{
            "producto_codigo": cod, "producto_nombre": nom,
            "compraron": int(comp), "pct": round(comp / ccc_mes * 100, 1) if ccc_mes else 0.0,
        } for (cod, nom), comp in prod.items()]
        por_producto.sort(key=lambda x: -x["pct"])
        out["innovaciones"] = {
            "resumen": {
                "productos":       len(por_producto),
                "compraron_total": int(sum(p["compraron"] for p in por_producto)),
                "ccc_mes":         ccc_mes,
                "pct_promedio":    round(sum(p["pct"] for p in por_producto) / len(por_producto), 1) if por_producto else 0,
            },
            "por_producto": por_producto[:20],
        }

        # Ranking — litros desde el maestro 04D CSV liviano (el xlsx de 19MB tarda ~40s).
        # castear litros/dinero a float nativo (groupby.to_dict() devuelve numpy)
        _, _, cod2lxu, _ = _cargar_maestro_04D()
        dfl = df.copy()
        dfl["_lxu"] = (dfl["Codigo"].astype(str).str.strip().str.upper()
                       .str.replace(r"\.0$", "", regex=True).map(cod2lxu).fillna(0.0))
        dfl["litros"] = dfl["CantBase"] * dfl["_lxu"]
        litros = {int(k): float(v) for k, v in dfl.groupby("CodVendedor")["litros"].sum().items()}
        dinero = {int(k): float(v) for k, v in df.groupby("CodVendedor")["ImporteNetoItem"].sum().items()}
        c11t   = {int(k): int(v) for k, v in gcm._11t_por_vend(df).items()}
        inov   = {int(k): int(v) for k, v in gcm._inov_por_vend(df, cod_inov).items()}
        out["ranking"] = gcm._ranking(litros, dinero, c11t, inov, vn)
    except Exception as e:
        out["_warn"] = "extras versionado (sellout/innov/ranking): " + str(e)

    return _to_native(out)


# Catálogo de reglas de acciones por período (esquema MAYO, en 09_CONFIG). Mayo usa este path.
# Junio en adelante: el catálogo viaja versionado en 01_INPUTS/cierres mes/acciones_<MMAAAA>.csv
# (esquema nuevo ACCIONES COMERCIALES) y lo procesa _cierre_acciones_junio_schema.
_ACC_REGLAS_POR_MMAAAA = {"052026": "reglas_acciones_mayo_2026_orbit.csv"}


def _cierre_acciones_junio_schema(files, reglas):
    """Acciones del cierre con el catálogo en esquema NUEVO (ACCIONES COMERCIALES/<mes>/,
    p.ej. junio 2026: canal_aplica/segmento_cliente_aplica/tipo_regla/productos_marcas).
    Reúsa el matching del motor live (_acc_seg_canon/_acc_subseg_filtro/_acc_product_pred)
    sobre el ventas_mes CONGELADO del cierre. Inversión = valorDescuento × CantBase.
    Totales sobre la UNIÓN de líneas (sin doble conteo entre acciones). Devuelve
    {resumen, detalle} jsonify-safe, o None si no hay ventas/catálogo."""
    v = _acc_preparar_ventas_mes_versionado(files["ventas_mes"])
    if v.empty or not reglas:
        return None
    all_lineas = set(l for l in v["_linea"].dropna().unique() if l and l != "NAN")
    plan_as_clientes = _acc_plan_as_clientes()

    def _match(r):
        vends_raw = str(r.get("vendedores_aplica", "")).upper()
        if "TODOS" in vends_raw:
            codes = {3, 4, 6, 7, 8, 9, 10}
        else:
            codes = {int(_re.sub(r"\D", "", x)) for x in _re.findall(r"V\s*\d+", vends_raw)} & {3, 4, 6, 7, 8, 9, 10}
        seg_set = _acc_seg_canon(r.get("segmento_cliente_aplica"), r.get("canal_aplica"))
        sub_allowed = _acc_subseg_filtro(r.get("segmento_cliente_aplica"), r.get("canal_aplica"))
        pred = _acc_product_pred(r, all_lineas)
        regla_txt = _acc_norm(" ".join(str(r.get(k, "")) for k in
                                       ("categoria", "canal_aplica", "segmento_cliente_aplica")))
        requiere_plan_as = "PLANES AASS" in regla_txt or "PLAN AASS" in regla_txt
        m = v["_vend"].isin(codes) & v["_seg"].isin(seg_set)
        if sub_allowed is not None:
            is_trad = v["_seg"].astype(str).str.upper().eq("TRADICIONAL")
            sub_ok = v["_subseg"].apply(lambda s: any(tok in s for tok in sub_allowed))
            m = m & (~is_trad | sub_ok)
        sub = v[m]
        if sub.empty:
            return sub
        keys = list(zip(sub["_cat"], sub["_linea"], sub["_art"], sub["_marca"], sub["_cod"]))
        predcache, keep = {}, []
        for k in keys:
            if k not in predcache:
                predcache[k] = bool(pred(*k))
            keep.append(predcache[k])
        sub = sub[pd.Series(keep, index=sub.index, dtype=bool)]
        if requiere_plan_as:
            sub = sub[pd.to_numeric(sub["_cli"], errors="coerce").isin(plan_as_clientes)]
        return sub

    detalle, matched_idx, cli_union = [], set(), set()
    for r in reglas:
        cur = _match(r)
        if cur.empty:
            continue
        matched_idx |= set(cur.index)
        cli_ids = set(cur["_cli"].dropna().astype(int))
        cli_union |= cli_ids
        cur_desc = cur[cur["_desc"] > 0]
        tipo_raw = str(r.get("tipo_regla", "")).upper()
        tipo = "Sin cargo" if ("SIN_CARGO" in tipo_raw or "BONIFIC" in tipo_raw) else "Descuento"
        nombre = (str(r.get("observaciones", "")).strip().split(".")[0]
                  or str(r.get("id_accion", "")).strip()
                  or f"{r.get('canal_aplica', '')}|{r.get('categoria', '')}")
        detalle.append({
            "nombre":       nombre[:80],
            "tipo":         tipo,
            "canal":        str(r.get("canal_aplica", "")).strip(),
            "categoria":    str(r.get("categoria", "")).strip().upper(),
            "descuento":    (str(r.get("descuento_pct", "")).strip() + "%") if str(r.get("descuento_pct", "")).strip() else "–",
            "litros":       round(float(cur["_litros"].sum()), 1),
            "inversion":    round(float(cur_desc["_desc"].sum()), 0),
            "importe_neto": round(float(cur["_imp_neto"].sum()), 0),
            "clientes":     int(len(cli_ids)),
        })
    if not detalle:
        return None
    detalle.sort(key=lambda a: -a["inversion"])
    uni = v.loc[v.index.isin(matched_idx)]
    uni_desc = uni[uni["_desc"] > 0]
    resumen = {
        "total_acciones":     len(detalle),
        "inversion_total":    round(float(uni_desc["_desc"].sum()), 0),
        "importe_neto":       round(float(uni["_imp_neto"].sum()), 0),
        "clientes_afectados": int(len(cli_union)),
    }
    return _to_native({"resumen": resumen, "detalle": detalle[:10],
                       "fuente": files["ventas_mes"].name, "catalogo": f"acciones_{files.get('mmaaaa', '')}.csv"})


def _cierre_acciones_versionado(files):
    """Acciones Comerciales del cierre desde ventas_mes_<MMAAAA>.csv.
    Inversión = valorDescuento × CantBase (descuento REAL; NO ImporteItem-ImporteNetoItem,
    que es IVA ~21%). Matching de catálogo idéntico a generar_acciones_ranking (reutiliza
    _REGLA_CANAL_SEG_MAP / _filtrar_ventas_accion / INOV_PRODUCTOS). Maestro vía CSV liviano
    (no el xlsx de 19MB). Devuelve {resumen, detalle} jsonify-safe, o None si no hay catálogo/ventas."""
    reglas_name = _ACC_REGLAS_POR_MMAAAA.get(files.get("mmaaaa", ""))
    if not reglas_name:
        # Sin catálogo mayo-schema registrado: usar el catálogo versionado del cierre
        # (esquema nuevo ACCIONES COMERCIALES, p.ej. junio) sobre el ventas_mes congelado.
        acc_path = CIERRES_MES_DIR / f"acciones_{files.get('mmaaaa', '')}.csv"
        if acc_path.exists():
            try:
                rdf = pd.read_csv(acc_path, sep=";", encoding="utf-8-sig", dtype=str)
                rdf.columns = [c.strip() for c in rdf.columns]
                if "canal_aplica" in rdf.columns:
                    return _cierre_acciones_junio_schema(files, rdf.to_dict("records"))
            except Exception:
                return None
        return None
    reglas_path = CONFIG / reglas_name
    if not reglas_path.exists():
        return None
    gda = _gda()
    df = _leer_ventas_mes_cacheado(files["ventas_mes"]).copy()   # neto>0, excl V2/V5/V20
    if df.empty:
        return None

    # Preparación (réplica de _preparar_ventas_acciones pero con inversión correcta = vd×CantBase)
    cod2cat, _cod2seg, cod2lxu, _cod2linea = _cargar_maestro_04D()
    clientes = gda.cargar_clientes()   # clientes.xlsx (liviano) → _seg oficial por cliente
    df["_cod"] = pd.to_numeric(df["Codigo"], errors="coerce")
    df["_cli"] = pd.to_numeric(df["Cliente"], errors="coerce")
    df["_codstr"] = (df["Codigo"].astype(str).str.strip().str.upper().str.replace(r"\.0$", "", regex=True))
    df["_vd"] = pd.to_numeric(df["valorDescuento"].astype(str).str.replace(",", ".", regex=False),
                              errors="coerce").fillna(0)
    df["Descuento_pct"] = pd.to_numeric(
        df["Descuento"].astype(str).str.replace("%", "", regex=False).str.replace(",", ".", regex=False),
        errors="coerce").fillna(0)
    v = df[df["Descuento_pct"] > 0].copy()
    if v.empty:
        return None
    v["_descuento_p"] = (v["_vd"] * v["CantBase"]).clip(lower=0)      # inversión real
    v["Categoria"] = v["_codstr"].map(cod2cat)
    v["_litros"] = v["CantBase"] * v["_codstr"].map(cod2lxu).fillna(0.0)
    cs = clientes[["Codigo", "_seg"]].copy()
    cs["Codigo"] = pd.to_numeric(cs["Codigo"], errors="coerce")
    v = v.merge(cs.rename(columns={"Codigo": "cli_id"}), left_on="_cli", right_on="cli_id", how="left")

    reglas = pd.read_csv(reglas_path, encoding="utf-8-sig")
    reglas.columns = [c.strip() for c in reglas.columns]

    seen, detalle, cli_union = set(), [], set()
    for _, r in reglas.iterrows():
        canal = str(r.get("canal", "")).strip()
        cat   = str(r.get("categoria", "")).strip().upper()
        grupo = str(r.get("accion_grupo", f"{canal}|{cat}")).strip()
        if grupo in seen:
            continue
        seen.add(grupo)
        if canal == "PLANES AASS" or cat == "RESTO SKU" or canal not in gda._REGLA_CANAL_SEG_MAP:
            continue
        vm = gda._filtrar_ventas_accion(v, canal, cat, gda._REGLA_CANAL_SEG_MAP[canal])
        if vm is None or vm.empty:
            continue
        desc_vals = sorted(vm["Descuento_pct"].dropna().unique())
        if desc_vals:
            lo, hi = round(min(desc_vals)), round(max(desc_vals))
            desc_disp = f"{int(lo)}%" if lo == hi else f"{int(lo)}-{int(hi)}%"
        else:
            desc_disp = "–"
        cli_ids = set(vm["_cli"].dropna().astype(int))
        cli_union |= cli_ids
        detalle.append({
            "nombre":       str(r.get("accion_nombre", grupo)).strip(),
            "tipo":         str(r.get("tipo_accion", "")).strip(),
            "canal":        canal,
            "categoria":    cat,
            "descuento":    desc_disp,
            "litros":       round(float(vm["_litros"].sum()), 1),
            "inversion":    round(float(vm["_descuento_p"].sum()), 0),
            "importe_neto": round(float(vm["ImporteNetoItem"].sum()), 0),
            "clientes":     int(len(cli_ids)),
        })

    if not detalle:
        return None
    detalle.sort(key=lambda a: -a["inversion"])
    resumen = {
        "total_acciones":     len(detalle),
        "inversion_total":    round(sum(a["inversion"] for a in detalle), 0),
        "importe_neto":       round(sum(a["importe_neto"] for a in detalle), 0),
        "clientes_afectados": int(len(cli_union)),
    }
    return _to_native({"resumen": resumen, "detalle": detalle[:10],
                       "fuente": files["ventas_mes"].name, "catalogo": reglas_name})


def _cierre_manifest_versionado(files):
    """Manifest minimo para cierres descubiertos por carpeta (01_INPUTS/cierres mes/) que NO
    tienen carpeta en 07_CIERRES_MENSUALES/. Reusa el motor oficial para filas/fechas/
    vendedores/V3, mismo criterio que generar_cierre_mensual.py (activos, V3 solo trad)."""
    gcm = _gcm()
    src = files["ventas_mes"]
    df  = _gcm_leer_ventas_cacheado(src)
    excl = sorted(gcm._raw_codigos(src) - gcm.VENDEDORES_ACTIVOS)
    fmin = df["FechaComprobante"].min() if not df.empty else None
    fmax = df["FechaComprobante"].max() if not df.empty else None
    vdet = sorted(int(x) for x in df["CodVendedor"].unique()) if not df.empty else []
    v3   = df[df["CodVendedor"] == 3] if not df.empty else df
    v3ok = bool((v3["segmento"] == "TRADICIONAL").all()) if not v3.empty else True
    return {
        "filas_leidas":             int(len(df)),
        "fecha_min":                fmin.strftime("%Y-%m-%d") if pd.notna(fmin) else None,
        "fecha_max":                fmax.strftime("%Y-%m-%d") if pd.notna(fmax) else None,
        "vendedores_detectados":    [f"V{c}" for c in vdet],
        "vendedores_excluidos_csv": [f"V{c}" for c in excl],
        "v3_solo_tradicional_pass": v3ok,
        "fuente_ventas":            "01_INPUTS/cierres mes/" + src.name,
        "estado":                   "PASS",
    }


# ====== CIERRES HISTORICOS ======
@app.route("/api/gerencia/cierres_historicos")
def gerencia_cierres_historicos():
    """Lista de cierres mensuales historicos generados en 07_CIERRES_MENSUALES/.
    Solo lectura. No genera cierres nuevos. No toca ventas_mes.csv ni ningun input.
    """
    cierres_dir = BASE / "07_CIERRES_MENSUALES"
    idx_path    = cierres_dir / "index_cierres_mensuales.json"

    indice = []
    if idx_path.exists():
        try:
            with open(idx_path, encoding="utf-8") as f:
                indice = json.load(f)
        except Exception as e:
            return jsonify({"cierres": [], "estado": "ERROR",
                            "error": "No se pudo leer el indice: " + str(e)}), 500

    # Descubrir cierres versionados por carpeta (01_INPUTS/cierres mes/) que NO esten en el
    # indice 07. Asi cada cierre que genera CIERRE_MES_ORBIT.bat aparece solo y el selector de
    # mes del portal crece con cada mes, sin depender de 07_CIERRES_MENSUALES/.
    periodos_idx = {str(e.get("periodo", "")) for e in indice}
    for vm in sorted(CIERRES_MES_DIR.glob("ventas_mes_*.csv")):
        mmaaaa = vm.stem.replace("ventas_mes_", "")
        if len(mmaaaa) != 6 or not mmaaaa.isdigit():
            continue
        periodo = f"{mmaaaa[2:]}-{mmaaaa[:2]}"
        if periodo in periodos_idx:
            continue
        try:
            ts = datetime.fromtimestamp(vm.stat().st_mtime).strftime("%Y-%m-%dT%H:%M:%S-03:00")
        except Exception:
            ts = ""
        indice.append({"periodo": periodo, "version": "version_001", "carpeta": "",
                       "timestamp_argentina": ts, "estado": "PASS", "_versionado_only": True})

    if not indice:
        return jsonify({"cierres": [], "estado": "SIN_CIERRES",
                        "nota": "No hay cierres en 07_CIERRES_MENSUALES/ ni en 01_INPUTS/cierres mes/"})

    cierres = []
    for entrada in indice:
        periodo  = entrada.get("periodo", "")
        version  = entrada.get("version", "")
        vonly    = bool(entrada.get("_versionado_only"))
        # Cierres descubiertos por carpeta no tienen carpeta en 07_CIERRES_MENSUALES/: apuntamos
        # a una ruta inexistente para que las lecturas legacy (.exists()) den False sin romper;
        # el manifest y los bloques se reconstruyen desde el cierre versionado mas abajo.
        carpeta  = (cierres_dir / periodo / "__versionado__") if vonly \
                   else (cierres_dir.parent / Path(entrada.get("carpeta", "").replace("\\", "/")))
        ts_ar    = entrada.get("timestamp_argentina", "")
        estado   = entrada.get("estado", "")

        cierre = {
            "periodo":             periodo,
            "version":             version,
            "timestamp_argentina": ts_ar,
            "estado":              estado,
            "manifest":            None,
            "empresa":             None,
            "ranking_top3":        [],
            "ranking":             [],
            "ganadores":           {},
            "objetivos_avance":    None,
            "ccc_segmentos":       None,
            "once_titulares":      None,
            "innovaciones":        None,
            "sellout":             None,
            "planes_as":           None,
            "acciones_comerciales": None,
            "warn":                [],
        }

        manifest_path = carpeta / "manifest.json"
        if manifest_path.exists():
            try:
                with open(manifest_path, encoding="utf-8") as f:
                    m = json.load(f)
                cierre["manifest"] = {
                    "filas_leidas":            m.get("filas_leidas"),
                    "fecha_min":               m.get("fecha_min"),
                    "fecha_max":               m.get("fecha_max"),
                    "vendedores_detectados":   m.get("vendedores_detectados", []),
                    "vendedores_excluidos_csv": m.get("vendedores_excluidos_csv", []),
                    "v3_solo_tradicional_pass": m.get("v3_solo_tradicional_pass"),
                    "fuente_ventas":           m.get("fuente_ventas"),
                    "fuente_ventas_hash_head": m.get("fuente_ventas_hash_head"),
                    "estado":                  m.get("estado"),
                }
            except Exception as e:
                cierre["warn"].append("manifest.json no legible: " + str(e))
        else:
            cierre["warn"].append("manifest.json no encontrado")

        ranking_path = carpeta / "ranking_vendedores_mes.json"
        if ranking_path.exists():
            try:
                with open(ranking_path, encoding="utf-8") as f:
                    rank = json.load(f)
                top3 = sorted(rank, key=lambda r: r.get("ranking_general", 99))[:3]
                cierre["ranking_top3"] = [
                    {
                        "ranking_general":    r.get("ranking_general"),
                        "vendedor_codigo":    r.get("vendedor_codigo"),
                        "vendedor_nombre":    r.get("vendedor_nombre"),
                        "score_total":        r.get("score_total"),
                        "clientes_11_titulares":  r.get("clientes_11_titulares"),
                        "clientes_innovaciones":  r.get("clientes_innovaciones"),
                        "etiqueta_destacada": r.get("etiqueta_destacada", ""),
                    }
                    for r in top3
                ]
                # Ranking completo (todos los vendedores del cierre), ordenado por ranking_general.
                # Solo campos del cierre versionado; sin CantBase ni botellas.
                cierre["ranking"] = [
                    {
                        "vendedor_codigo":        r.get("vendedor_codigo"),
                        "vendedor_nombre":        r.get("vendedor_nombre"),
                        "dinero_vendido":         r.get("dinero_vendido"),
                        "litros_vendidos":        r.get("litros_vendidos"),
                        "clientes_11_titulares":  r.get("clientes_11_titulares"),
                        "clientes_innovaciones":  r.get("clientes_innovaciones"),
                        "score_total":            r.get("score_total"),
                        "ranking_general":        r.get("ranking_general"),
                        "ranking_volumen_dinero": r.get("ranking_volumen_dinero"),
                        "ranking_11_titulares":   r.get("ranking_11_titulares"),
                        "ranking_innovaciones":   r.get("ranking_innovaciones"),
                        "etiqueta_destacada":     r.get("etiqueta_destacada", ""),
                    }
                    for r in sorted(rank, key=lambda r: r.get("ranking_general", 99))
                ]
                # Ganadores por categoria = vendedor con ranking_X == 1 en cada eje.
                def _ganador(campo, metrica):
                    for r in rank:
                        if r.get(campo) == 1:
                            return {
                                "vendedor_codigo": r.get("vendedor_codigo"),
                                "vendedor_nombre": r.get("vendedor_nombre"),
                                "score_total":     r.get("score_total"),
                                "metrica":         r.get(metrica),
                            }
                    return None
                cierre["ganadores"] = {
                    "general":        _ganador("ranking_general",        "score_total"),
                    "volumen_dinero": _ganador("ranking_volumen_dinero", "dinero_vendido"),
                    "once_titulares": _ganador("ranking_11_titulares",   "clientes_11_titulares"),
                    "innovaciones":   _ganador("ranking_innovaciones",   "clientes_innovaciones"),
                }
            except Exception as e:
                cierre["warn"].append("ranking_vendedores_mes.json no legible: " + str(e))
        else:
            cierre["warn"].append("ranking_vendedores_mes.json no encontrado")

        # Resumen empresa del cierre (solo lectura de cierre_mensual_resumen.json).
        resumen_path = carpeta / "cierre_mensual_resumen.json"
        if resumen_path.exists():
            try:
                with open(resumen_path, encoding="utf-8") as f:
                    res = json.load(f)
                emp = res.get("empresa", {}) or {}
                cierre["empresa"] = {
                    "importe_neto_total":   emp.get("importe_neto_total"),
                    "litros_total":         emp.get("litros_total"),
                    "ccc_total":            emp.get("ccc_total"),
                    "filas_ventas_mes":     emp.get("filas_ventas_mes"),
                    "vendedores_incluidos": emp.get("vendedores_incluidos", []),
                }
            except Exception as e:
                cierre["warn"].append("cierre_mensual_resumen.json no legible: " + str(e))
        else:
            cierre["warn"].append("cierre_mensual_resumen.json no encontrado")

        # ── Bloques de detalle congelados (artefactos versionados del cierre) ──
        # Solo lectura de JSON ya generados en la carpeta del cierre. No recalcula.
        def _load_art(nombre):
            ruta = carpeta / nombre
            if not ruta.exists():
                return None
            try:
                with open(ruta, encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                cierre["warn"].append(f"{nombre} no legible: {e}")
                return None

        oa = _load_art("cierre_objetivos_avance.json")
        if oa is not None:
            cierre["objetivos_avance"] = oa
            # ccc_segmentos derivado del mismo artefacto (empresa + por vendedor), sin duplicar archivo
            emp_oa = oa.get("empresa", {}) or {}
            cierre["ccc_segmentos"] = {
                "empresa": {
                    "ccc_total":        emp_oa.get("ccc_total"),
                    "ccc_tradicional":  emp_oa.get("ccc_tradicional"),
                    "ccc_autoservicio": emp_oa.get("ccc_autoservicio"),
                    "ccc_onpremise":    emp_oa.get("ccc_onpremise"),
                },
                "vendedores": [
                    {
                        "codigo":           v.get("codigo"),
                        "nombre":           v.get("nombre"),
                        "ccc_total":        v.get("ccc_total"),
                        "ccc_tradicional":  v.get("ccc_tradicional"),
                        "ccc_autoservicio": v.get("ccc_autoservicio"),
                        "ccc_onpremise":    v.get("ccc_onpremise"),
                    }
                    for v in (oa.get("vendedores", []) or [])
                ],
            }
        cierre["once_titulares"]       = _load_art("cierre_11_titulares_detalle.json")
        cierre["innovaciones"]         = _load_art("cierre_innovaciones_detalle.json")
        cierre["sellout"]              = _load_art("cierre_sellout.json")
        cierre["planes_as"]            = _load_art("cierre_planes_as.json")
        cierre["acciones_comerciales"] = _load_art("cierre_acciones_comerciales.json")

        # ── Cierre versionado por carpeta (01_INPUTS/cierres mes/) — FASE 1 ──
        # Si existen los 3 archivos del mes, objetivos/avance + 11T se calculan desde ellos
        # (fuente única y versionada), sustituyendo a los artefactos congelados.
        _files = _cierre_archivos_mes(periodo)
        if _files:
            # Manifest: para cierres descubiertos por carpeta (sin 07/) se reconstruye desde el
            # ventas_mes versionado; para los del indice 07 se respeta el manifest.json legacy.
            if cierre["manifest"] is None:
                try:
                    cierre["manifest"] = _cierre_manifest_versionado(_files)
                except Exception as e:
                    cierre["warn"].append("manifest versionado: " + str(e))
            try:
                oa = _cierre_objetivos_avance(_files)
                cierre["objetivos_avance"] = oa
                emp_oa = oa.get("empresa", {}) or {}
                cierre["ccc_segmentos"] = {
                    "empresa": {
                        "ccc_total":        emp_oa.get("ccc_total"),
                        "ccc_tradicional":  emp_oa.get("ccc_tradicional"),
                        "ccc_autoservicio": emp_oa.get("ccc_autoservicio"),
                        "ccc_onpremise":    emp_oa.get("ccc_onpremise"),
                    },
                    "vendedores": [
                        {"codigo": v.get("codigo"), "nombre": v.get("nombre"),
                         "ccc_total": v.get("ccc_total"), "ccc_tradicional": v.get("ccc_tradicional"),
                         "ccc_autoservicio": v.get("ccc_autoservicio"), "ccc_onpremise": v.get("ccc_onpremise")}
                        for v in oa.get("vendedores", [])
                    ],
                }
                cierre["once_titulares"] = _cierre_once_titulares(_files)
                cierre["fuente_cierre_versionado"] = "01_INPUTS/cierres mes/"
            except Exception as e:
                cierre["warn"].append("recálculo cierre versionado (objetivos/11T): " + str(e))
            # Sell-Out + Innovaciones + Ranking desde el ventas_mes versionado (FASE 2a)
            try:
                _ex = _cierre_extras_versionado(_files)
                if _ex.get("sellout") is not None:
                    cierre["sellout"] = _ex["sellout"]
                if _ex.get("innovaciones") is not None:
                    cierre["innovaciones"] = _ex["innovaciones"]
                if _ex.get("ranking"):
                    # Mejor en VOLUMEN = mayor alcance del objetivo mensual (avance vs objetivo),
                    # no litros+dinero. avance_map desde objetivos_avance ya calculado arriba.
                    _oa = cierre.get("objetivos_avance") or {}
                    _avance_map = {str(v.get("codigo")): v.get("avance_pct")
                                   for v in _oa.get("vendedores", []) if v.get("codigo") is not None}
                    _rk = _cierre_ranking_payload(_ex["ranking"], _avance_map or None)
                    cierre["ranking_top3"] = _rk["ranking_top3"]
                    cierre["ranking"]      = _rk["ranking"]
                    cierre["ganadores"]    = _rk["ganadores"]
                if _ex.get("_warn"):
                    cierre["warn"].append(_ex["_warn"])
            except Exception as e:
                cierre["warn"].append("recálculo cierre versionado (sellout/innov/ranking): " + str(e))
            # Acciones Comerciales desde ventas_mes versionado (FASE 2b) — inversión real (vd×CantBase),
            # NO IVA. Solo si hay catálogo del mes; si no, queda el artefacto congelado.
            try:
                _acc = _cierre_acciones_versionado(_files)
                if _acc is not None:
                    cierre["acciones_comerciales"] = _acc
            except Exception as e:
                cierre["warn"].append("recálculo cierre versionado (acciones): " + str(e))

        # Cierres descubiertos por carpeta no tienen artefactos 07/: descartar el ruido legacy
        # ("no encontrado"/"no legible"); ya se reconstruyo todo desde el cierre versionado.
        if vonly:
            cierre["warn"] = [w for w in cierre["warn"]
                              if "no encontrado" not in w and "no legible" not in w]

        if not cierre["warn"]:
            cierre.pop("warn")

        cierres.append(cierre)

    cierres.sort(key=lambda c: (c.get("periodo",""), c.get("version","")), reverse=True)

    return jsonify({
        "cierres":       cierres,
        "total_cierres": len(cierres),
        "estado":        "OK",
        "nota":          "Solo lectura. No recalcula ni modifica datos.",
    })


# ====== STARTUP (gunicorn + __main__) ======
# Se ejecuta cuando gunicorn importa el módulo, no solo en __main__.
# M1: bajo PENAFLOR_SKIP_BOOT=1 (mount embebido en Orbit Home) NO se ejecuta NADA de esto al
# importar -> importar `server_orbit` no escribe SQLite/CSV ni toca el filesystem. En standalone
# (variable ausente) el arranque es idéntico al de siempre.
if not _PENAFLOR_SKIP_BOOT:
    backup_orbit_db()         # 1. copia orbit.db antes de cualquier cambio
    init_db()                 # 2. crea/migra tablas
    restore_planificacion_if_empty()  # 3. recupera desde CSV si la tabla quedó vacía
    export_planificacion_csv()        # 4. actualiza CSV de seguridad con estado actual


def _warm_caches():
    """Precalienta en background el payload pesado de Acciones Comerciales (vista
    gerencia, sin filtro). En Render (0.5 vCPU) ese cómputo puede acercarse al timeout
    del worker; al calcularlo en un hilo al arranque (sin timeout HTTP) queda cacheado
    y la primera request del gerente cae en caché. No bloquea el arranque ni el request."""
    try:
        _acciones_mes_payload(None)
    except Exception as e:  # nunca tumbar el arranque por el warmup
        try:
            print(f"[ORBIT] warmup acciones_mes fallo (no fatal): {e}")
        except Exception:
            pass


# M1: el warmup (cómputo pandas pesado en background) tampoco corre en modo embebido.
if not _PENAFLOR_SKIP_BOOT:
    threading.Thread(target=_warm_caches, name="orbit-warmup", daemon=True).start()

if __name__ == "__main__":
    print("\n===== ORBIT SERVER v3 =====")
    print("Diagnóstico: http://localhost:8502/api/diagnostico")
    print("Dashboard:   http://localhost:8502/api/dashboard")
    print("Portal:      http://localhost:8502/index.html")
    print("===============================\n")
    port = int(os.environ.get("PORT", 8502))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug)
