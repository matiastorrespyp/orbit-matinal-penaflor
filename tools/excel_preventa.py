#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
excel_preventa.py — Genera un Excel de preventa, una HOJA por día de visita.

Cada hoja lista los clientes de ese día (ordenados por vendedor y orden de ruta) con:
  Vendedor / Código / Nombre / Dirección / Localidad / Segmento / Compró (Sí/No) /
  Titulares cubiertos (X/11) / una columna por cada uno de los 11 Titulares.

En cada columna de titular: estado de cobertura del cliente vs ventas_acumulada.csv:
  - "OK"              -> ya cubre la marca (>= umbral del segmento)
  - "b/umbral vender f" -> compró b botellas, faltan f para cubrir (umbral - b)
Umbral de cobertura (regla Peñaflor): TRADICIONAL (kiosco/almacén/despensa) = 3 botellas;
el resto de los segmentos = 6 botellas.

Fuentes: 01_INPUTS/clientes.xlsx (cartera + día de visita + segmento)
         01_INPUTS/ventas_acumulada.csv (compras del período, marca, botellas=CantBase)
Reglas (11 Titulares, alias de marca, umbrales, segmento) reutilizadas de generar_datasets_acum.py.

Uso:  python tools/excel_preventa.py
Salida: 03_OUTPUTS/PREVENTA_11T_<YYYYMMDD_HHMM>.xlsx
"""
import sys
import math
from datetime import datetime
from pathlib import Path

import pandas as pd

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
import generar_datasets_acum as G  # cargar_clientes, cargar_ventas_acumulada
import motor_11t                   # motor autoritativo del 11T (titulares, botellas, umbrales)

OUT_DIR = ROOT / "03_OUTPUTS"

# Día de visita: código en clientes.xlsx -> (orden, nombre de hoja)
DIAS = [("Lu", "Lunes"), ("Ma", "Martes"), ("Mi", "Miercoles"),
        ("Ju", "Jueves"), ("Vi", "Viernes"), ("Sa", "Sabado")]

# Los titulares salen de la matriz oficial de SKU, no de una lista propia de este script.
TITULARES = motor_11t.titulares_oficiales()


def _umbral(seg):
    """Mínimo de botellas del segmento, desde el motor (3 tradicional / 6 autoservicio)."""
    return motor_11t.UMBRALES_11T.get(str(seg).upper(), 6)


def main():
    print("Generando Excel de preventa (11 Titulares por día de visita)...")

    clientes = G.cargar_clientes()
    clientes = clientes.dropna(subset=["Codigo"]).copy()
    clientes["Codigo"] = clientes["Codigo"].astype(int)

    ventas = G.cargar_ventas_acumulada()

    # CCC comercial (para la columna "Compró"): clientes con alguna compra válida.
    # Es OTRA métrica que la cobertura 11T y se calcula aparte a propósito.
    _v = ventas[ventas["ImporteNetoItem"] > 0].copy()
    ccc_ids = set(pd.to_numeric(_v["Cliente"], errors="coerce").dropna().astype(int))

    # Botellas netas por (cliente, titular) desde el motor: match por código de artículo
    # contra la matriz oficial, no por texto de Marca.
    _desde, _hasta = motor_11t.periodo_trimestre_en_curso()
    det, _exc = motor_11t.cobertura_11t(ventas, desde=_desde, hasta=_hasta)
    bot_map = {(int(c), t): float(b) for c, t, b in
               zip(det["cliente_id"], det["titular"], det["botellas_netas"])}
    print(f"  periodo 11T medido: {_desde} -> {_hasta}")

    print(f"  clientes en cartera: {len(clientes)}  |  clientes con compra (CCC): {len(ccc_ids)}")

    # openpyxl para formato
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    HDR_FILL = PatternFill("solid", fgColor="9C1458")     # magenta Orbit
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    OK_FILL  = PatternFill("solid", fgColor="C6EFCE")     # verde
    MED_FILL = PatternFill("solid", fgColor="FFEB9C")     # amarillo (compró algo, falta)
    BAD_FILL = PatternFill("solid", fgColor="FFC7CE")     # rojo (no compró la marca)
    NO_FILL  = PatternFill("solid", fgColor="FFC7CE")
    SI_FILL  = PatternFill("solid", fgColor="C6EFCE")
    THIN = Side(style="thin", color="DDDDDD")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CEN = Alignment(horizontal="center", vertical="center")
    LEF = Alignment(horizontal="left", vertical="center", wrap_text=False)

    base_cols = ["Vendedor", "Código", "Nombre", "Dirección", "Localidad",
                 "Segmento", "Compró", "Titulares cubiertos"]
    headers = base_cols + TITULARES
    widths = [9, 9, 30, 28, 16, 14, 8, 16] + [13] * len(TITULARES)

    total_filas = 0
    for cod_dia, nombre_hoja in DIAS:
        sub = clientes[clientes["DiasVisita"].astype(str).str.strip().str.lower() == cod_dia.lower()].copy()
        ws = wb.create_sheet(title=nombre_hoja)

        # encabezado
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CEN; c.border = BORDER
            ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
        ws.freeze_panes = "C2"

        if sub.empty:
            ws.cell(row=2, column=1, value="(sin clientes este día)")
            continue

        # orden: vendedor, luego orden de ruta
        sub["_ord"] = pd.to_numeric(sub.get("Orden"), errors="coerce").fillna(99999)
        sub["_vc"] = pd.to_numeric(sub.get("codven"), errors="coerce").fillna(0).astype(int)
        sub = sub.sort_values(["_vc", "_ord"])

        r = 2
        for _, cli in sub.iterrows():
            cid = int(cli["Codigo"]); seg = str(cli["_seg"]); umb = _umbral(seg)
            compro = "Sí" if cid in ccc_ids else "No"
            cubiertos = 0
            row_vals = [
                f"V{cli['_vc']}", cid, str(cli.get("Razon_Social", "")),
                str(cli.get("Direccion", "")), str(cli.get("Localidad", "")),
                seg, compro, "",  # cubiertos se llena luego
            ]
            for j, val in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=j, value=val)
                c.border = BORDER
                c.alignment = LEF if j in (3, 4, 5) else CEN
            # estado Compró
            cc = ws.cell(row=r, column=7)
            cc.fill = SI_FILL if compro == "Sí" else NO_FILL

            # columnas de titulares
            for k, marca in enumerate(TITULARES):
                col = len(base_cols) + 1 + k
                b = bot_map.get((cid, marca), 0.0)
                bi = int(round(b))
                if b >= umb:
                    txt = "OK"; fill = OK_FILL; cubiertos += 1
                else:
                    falt = int(math.ceil(umb - b))
                    txt = f"{bi}/{umb} vender {falt}"
                    fill = MED_FILL if b > 0 else BAD_FILL
                c = ws.cell(row=r, column=col, value=txt)
                c.fill = fill; c.alignment = CEN; c.border = BORDER
                c.font = Font(size=9)
            ws.cell(row=r, column=8, value=f"{cubiertos}/{len(TITULARES)}").alignment = CEN
            r += 1
        total_filas += (r - 2)
        print(f"  {nombre_hoja:11}: {r - 2} clientes")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = OUT_DIR / f"PREVENTA_11T_{ts}.xlsx"
    wb.save(out)
    print(f"\nOK -> {out}  ({total_filas} filas, {len(DIAS)} hojas)")
    print("Umbral cobertura: TRADICIONAL=3 botellas, resto=6.  Fuente compras: ventas_acumulada.csv")
    return str(out)


if __name__ == "__main__":
    main()
