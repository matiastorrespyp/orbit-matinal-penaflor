# -*- coding: utf-8 -*-
"""Alta de la accion AGO26-TRAD-NC y split de los dos SKU Alma Mora Low en el libro de
Acciones Comerciales de agosto 2026.

QUE HACE
  1. ACCIONES            -> agrega la fila conceptual AGO26-TRAD-NC.
  2. ESCALAS             -> agrega la escala 3+3 / 15% de la accion.
  3. PRODUCTOS_Y_LINEAS  -> agrega las 10 marcas elegibles de AGO26-TRAD-NC y reemplaza la
                            entrada generica "Alma Mora Low" de AGO26-INNOV por los dos SKU
                            explicitos (74827 y 74887).

POR QUE UN SCRIPT Y NO EDICION A MANO: insertar filas en el medio de PRODUCTOS_Y_LINEAS
corre todo el bloque de abajo, y el libro tiene convenciones de formato (zebra por paridad
de fila, borde inferior solo en la ultima fila, alto 30, "0%" en la columna descuento) que
openpyxl no arrastra sola al insertar. El script reescribe el bloque de datos completo y
vuelve a aplicar la convencion, asi el libro queda identico salvo por las filas nuevas.

El script es idempotente: si la accion ya esta, no la duplica.

Uso:  python tools/actualizar_acciones_agosto_trad_nc.py
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE = Path(__file__).resolve().parent.parent
LIBRO = (BASE / "01_INPUTS" / "ACCIONES COMERCIALES" / "2026-08"
         / "ORBIT_Acciones_Comerciales_Agosto_2026.xlsx")

ACTION_ID = "AGO26-TRAD-NC"
FILA_HDR = 4          # encabezado real; arriba hay portada (titulo + bajada) y una fila vacia
FILA_DATOS = 5

# Convencion de estilo del libro, derivada de las hojas existentes (no inventada):
# fuente Carlito 9 / FF292B33, alineacion top + wrap, zebra FFF8F9FB en filas pares,
# borde inferior fino solo en la ultima fila de datos, alto de fila 30.
FUENTE = Font(name="Carlito", size=9, bold=False, color="FF292B33")
ALINEACION = Alignment(vertical="top", wrap_text=True)
ZEBRA = PatternFill(fill_type="solid", fgColor="FFF8F9FB")
SIN_FILL = PatternFill(fill_type=None)
BORDE_ULTIMA = Border(bottom=Side(style="thin"))
SIN_BORDE = Border()
ALTO_FILA = 30.0

# Las 10 marcas elegibles de la caja mixta, en el orden de la presentacion comercial.
MARCAS_TRAD_NC = [
    "Alma Mora",
    "Alaris",
    "Finca Las Moras",
    "Dadá",
    "Los Árboles",
    "Trapiche Reserva",
    "Don David",
    "Smirnoff botella",
    "Frizze",
    "Gordon's",
]

FILA_ACCIONES = [
    ACTION_ID,
    "Tradicionales no compradores",
    "Caja mixta 3+3",
    "Caja mixta",
    "Especiales",
    "Activa",
    "15% en caja mixta de 6 botellas: 3 de una marca y 3 de otra distinta, entre las 10 "
    "marcas elegibles. Solo para clientes del canal Tradicional que en agosto no compraron "
    "ninguna de esas marcas (pertenencia al mes por FechaComprobante).",
]

FILA_ESCALAS = [
    ACTION_ID,
    "Tradicional",
    "Tradicional | Almacenes | Kioscos",
    "botella",
    6,
    6,
    0.15,
    "descuento",
    "Caja mixta de 6 botellas · 3 + 3 de dos marcas distintas · 15%",
    "Elegibles: clientes Tradicionales sin compra de ninguna de las 10 marcas durante agosto "
    "(FechaComprobante, ImporteNetoItem > 0). La caja debe llevar 3 botellas de una marca y 3 "
    "de otra diferente: 6 de una sola marca no califica.",
]

# Reemplazo de la entrada generica "Alma Mora Low" de AGO26-INNOV por los dos SKU reales.
# Descripciones verificadas contra 09_CONFIG/mpa_codigos.csv, 09_CONFIG/maestro_04D_productos.csv
# y el maestro de articulos vigente 01_INPUTS/RAW_PRODUCTOS/productosjulio.xlsx.
INNOV_GENERICO = "Alma Mora Low"
FILAS_INNOV_LOW = [
    ["AGO26-INNOV", "sku", "74827 — Alma Mora Blanco Dulce Low 6x750", "Incluido",
     "SKU explicito (antes: entrada generica “Alma Mora Low”). ERP: ALMA MORA BLANCO "
     "DULCE LOW 6X750, linea comercial Alma Mora, vigente."],
    ["AGO26-INNOV", "sku", "74887 — Alma Mora Malbec Dulce Low 6x750", "Incluido",
     "SKU explicito (antes: entrada generica “Alma Mora Low”). ERP: ALMA MORA MALBEC "
     "DULCE LOW 6X750, linea comercial Alma Mora, en proceso de alta en el maestro."],
]


def leer_datos(ws):
    """Valores del bloque de datos (de FILA_DATOS a max_row), como lista de listas."""
    return [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            for r in range(FILA_DATOS, ws.max_row + 1)]


def escribir_datos(ws, filas, formatos_col=None):
    """Reescribe el bloque de datos aplicando la convencion de formato del libro.

    `formatos_col` es {indice_columna_1based: number_format} para las columnas que no son
    'General' (en ESCALAS, la columna 'descuento' va como '0%')."""
    formatos_col = formatos_col or {}
    ncols = ws.max_column
    ultima = FILA_DATOS + len(filas) - 1
    for i, valores in enumerate(filas):
        r = FILA_DATOS + i
        ws.row_dimensions[r].height = ALTO_FILA
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = valores[c - 1] if c - 1 < len(valores) else None
            cell.font = FUENTE
            cell.alignment = ALINEACION
            # Zebra por paridad de fila: pares con fondo, impares sin fondo.
            cell.fill = ZEBRA if r % 2 == 0 else SIN_FILL
            cell.border = BORDE_ULTIMA if r == ultima else SIN_BORDE
            cell.number_format = formatos_col.get(c, "General")
    # Sobrantes de una version anterior mas larga: se limpian valores y formato.
    for r in range(ultima + 1, ws.max_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = SIN_FILL
            cell.border = SIN_BORDE
        if r in ws.row_dimensions:
            del ws.row_dimensions[r]


def main():
    if not LIBRO.exists():
        raise SystemExit(f"No existe el libro: {LIBRO}")
    wb = openpyxl.load_workbook(LIBRO)

    # ── ACCIONES ──
    ws = wb["ACCIONES"]
    filas = leer_datos(ws)
    if any(str(f[0]).strip() == ACTION_ID for f in filas):
        print("  ACCIONES: ya estaba, no se duplica")
    else:
        filas.append(FILA_ACCIONES)
        print(f"  ACCIONES: +1 fila -> {len(filas)} acciones")
    escribir_datos(ws, filas)

    # ── ESCALAS ──
    ws = wb["ESCALAS"]
    filas = leer_datos(ws)
    if any(str(f[0]).strip() == ACTION_ID for f in filas):
        print("  ESCALAS: ya estaba, no se duplica")
    else:
        filas.append(FILA_ESCALAS)
        print(f"  ESCALAS: +1 fila -> {len(filas)} escalas")
    escribir_datos(ws, filas, formatos_col={7: "0%"})   # G = descuento

    # ── PRODUCTOS_Y_LINEAS ──
    ws = wb["PRODUCTOS_Y_LINEAS"]
    filas = leer_datos(ws)

    # 1) Split del generico "Alma Mora Low" en su lugar exacto, para no romper el orden
    #    por action_id del bloque de Innovaciones.
    salida, reemplazado = [], False
    for f in filas:
        if (str(f[0]).strip() == "AGO26-INNOV"
                and str(f[2] or "").strip() == INNOV_GENERICO):
            salida.extend(FILAS_INNOV_LOW)
            reemplazado = True
        else:
            salida.append(f)
    if reemplazado:
        print("  PRODUCTOS_Y_LINEAS: “Alma Mora Low” -> 74827 + 74887")
    elif any(str(f[2] or "").startswith("74827") for f in salida):
        print("  PRODUCTOS_Y_LINEAS: los SKU Low ya estaban separados")
    else:
        raise SystemExit("No se encontro la entrada generica 'Alma Mora Low' en AGO26-INNOV")

    # 2) Marcas elegibles de la accion nueva, al final del bloque. Se reescribe entero en
    #    vez de saltearse si ya existe: asi la lista de marcas del script es siempre la que
    #    manda y correr el script dos veces no deja una version vieja a medias.
    previas = sum(1 for f in salida if str(f[0]).strip() == ACTION_ID)
    salida = [f for f in salida if str(f[0]).strip() != ACTION_ID]
    for marca in MARCAS_TRAD_NC:
        salida.append([ACTION_ID, "marca", marca,
                       "Elegible para la caja mixta 3+3 · 15%",
                       "Comprar esta marca durante agosto deja al cliente fuera de la "
                       "accion (deja de ser no comprador)."])
    print(f"  PRODUCTOS_Y_LINEAS: {len(MARCAS_TRAD_NC)} marcas de {ACTION_ID} "
          f"({'reescritas, antes ' + str(previas) if previas else 'nuevas'})")

    escribir_datos(ws, salida)
    print(f"  PRODUCTOS_Y_LINEAS: {len(salida)} filas totales")

    wb.save(LIBRO)
    print(f"\nOK: {LIBRO}")


if __name__ == "__main__":
    main()
